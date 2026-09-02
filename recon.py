"""Punkt 5: Bilanz, GuV und die beiden Recon-Tabs.

**Bilanz und GuV kommen aus der Bilanzposten-Spalte**, nicht aus unserer
Klassifizierung. Das ist der Kern: eine Recon gegen den Jahresabschluss darf
nicht über dieselbe Zuordnung laufen, die sie prüfen soll. Der Export führt
den Posten selbst (``ClassDescription``), und genau der wird gezeigt —
Current Assets, Non-Current Assets, Current Liabilities, Non-Current
Liabilities, Equity beziehungsweise Income, Cost of Goods Sold,
Manufacturing, Expenses, Other Income, Other Expenses.

Sortiert wird nach ``ClassCode``, nicht nach Zeilenreihenfolge: im
GuV-Export steht ``Other Income`` (Code 90) an zweiter Stelle.

**Vorzeichen: es wird nirgends gedreht.** Soll positiv, Haben negativ —
Aktiva positiv, Passiva und Eigenkapital negativ, Erträge negativ,
Aufwendungen positiv. Ein Gewinn ist damit eine negative Summe. Die
Beispieldatei ``abschluss.json`` führt dieselbe Konvention. Eine
Vorzeichenumkehr an einer Stelle und nicht an der anderen ist der häufigste
stille Fehler in einer Recon: die Differenz ist dann doppelt so groß wie der
Sachverhalt und zeigt in die falsche Richtung.

**Die Recon zeigt drei Blöcke nebeneinander** — links Jahresabschluss, Mitte
Saldenliste, rechts Differenz. Je Zeile ein Bilanzposten.

**Keine Gesamtsumme über alle Zeilen.** Sie mischte Aktiva, Passiva und
Eigenkapital und wäre bedeutungslos — in der Saldenliste ist sie außerdem
konstruktionsbedingt null, was Übereinstimmung vortäuschte, wo keine ist. Die
einzige sinnvolle Gesamtgröße der Bilanz-Recon ist die **Differenz im
Nettovermögen**: Aktiva plus Passiva ohne Eigenkapital, auf beiden Seiten
gleich gerechnet.

In der GuV-Recon ist die Summe über alle Zeilen dagegen sehr wohl sinnvoll —
sie ist das Periodenergebnis, und alle Zeilen tragen dieselbe
Vorzeichenkonvention. Dort steht sie deshalb.
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass, field
from typing import Optional

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from aufrisse import (BLAU, GELB, GRAU, GROESSE, GROESSE_KONTROLLE, GRUEN,
                      MS_ERSTE_PERIODE, SCHWARZ,
                      SP_DE, SP_EN, SP_GRUPPE, SP_TICKER, SP_TYP, TEAL,
                      TEAL_DK, TINT1, TINT2, ZAHL, Z_EINHEIT, Z_ERSTE, Z_KOPF,
                      Z_PROJEKT, Z_TITEL, _RAHMEN_OBEN, _RAHMEN_SUMME,
                      _schrift, _zahlzelle)

#: Standardpfad der Abschlussdatei.
ABSCHLUSS = "abschluss.json"

#: Das Eigenkapital zaehlt NICHT ins Nettovermoegen — es ist dessen
#: Gegenprobe. Der Posten heisst in der Saldenliste so.
EIGENKAPITALPOSTEN = "Equity"

#: Konto, das die GuV in einer Zahl traegt (Gegenprobe fuer den GuV-Tab).
ERGEBNISKONTO = "3-90000"


# --------------------------------------------------------------------------
# Abschlussdatei
# --------------------------------------------------------------------------

@dataclass
class Abschluss:
    """Die von Hand gepflegte zweite Quelle."""

    mandant: str = ""
    quelle: str = ""
    waehrung: str = ""
    perioden: list[str] = field(default_factory=list)
    bilanz: dict[str, dict[str, float]] = field(default_factory=dict)
    guv: dict[str, dict[str, float]] = field(default_factory=dict)
    nettovermoegen: dict[str, float] = field(default_factory=dict)
    jahresergebnis: dict[str, float] = field(default_factory=dict)
    beispiel: bool = False
    vorhanden: bool = False

    def wert(self, werk: str, posten: str, periode: str) -> Optional[float]:
        quelle = self.bilanz if werk == "bilanz" else self.guv
        return quelle.get(posten, {}).get(periode)


def lies_abschluss(pfad: str = ABSCHLUSS) -> Abschluss:
    """Liest ``abschluss.json``. Fehlt sie, entsteht eine leere Recon.

    Kein PDF-Parser in dieser Stufe: die Zuordnung der Abschlusspositionen zu
    den Bilanzposten des Databooks ist eine fachliche Entscheidung. Ein
    Parser, der sie raet, macht die Recon wertlos, weil niemand mehr weiss,
    ob eine Differenz aus der Sache oder aus der Zuordnung stammt.
    """
    if not os.path.exists(pfad):
        return Abschluss()
    with open(pfad, encoding="utf-8") as f:
        d = json.load(f)
    return Abschluss(
        mandant=d.get("mandant", ""), quelle=d.get("quelle", ""),
        waehrung=d.get("waehrung", ""), perioden=d.get("perioden", []),
        bilanz=d.get("bilanz", {}), guv=d.get("guv", {}),
        nettovermoegen=d.get("nettovermoegen", {}),
        jahresergebnis=d.get("jahresergebnis", {}),
        beispiel=str(d.get("_status", "")).upper().startswith("BEISPIEL"),
        vorhanden=True)


# --------------------------------------------------------------------------
# Bilanz und GuV aus der Bilanzposten-Spalte
# --------------------------------------------------------------------------

def _posten(konten, perioden) -> list[tuple[str, list[str]]]:
    """Bilanzposten in der Reihenfolge ihres ``ClassCode``."""
    je: dict[str, list[str]] = {}
    code: dict[str, str] = {}
    for no, k in konten.items():
        je.setdefault(k.posten or "ohne Posten", []).append(no)
        code[k.posten or "ohne Posten"] = k.postencode or "zz"
    return [(p, sorted(je[p], key=lambda x: abs(konten[x].saldo(perioden[-1])),
                       reverse=True))
            for p in sorted(je, key=lambda p: (code[p], p))]


def _kopf(ws, projekt, waehrung, titel_de, titel_en, spalten, perioden,
          quelle_de: str) -> None:
    ws.cell(Z_PROJEKT, SP_DE, projekt).font = _schrift(GROESSE + 2, BLAU,
                                                       fett=True)
    ws.cell(Z_TITEL, SP_DE, titel_de).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_TITEL, SP_EN, titel_en).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_EINHEIT, SP_DE, f"in T {waehrung}").font = _schrift(farbe=GRAU)
    ws.cell(Z_EINHEIT, SP_EN, quelle_de).font = _schrift(farbe=GRAU)
    for spalte, titel in ((SP_TYP, "Typ"), (SP_TICKER, "Ticker"),
                          (SP_DE, "Bezeichnung"), (SP_EN, "Description"),
                          (SP_GRUPPE, "Herkunft")):
        z = ws.cell(Z_KOPF, spalte, titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
    for i, p in enumerate(perioden):
        z = ws.cell(Z_KOPF, spalten[i], p)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
        z.alignment = Alignment(horizontal="right")


def schreibe_rechenwerk(wb, blatt: str, titel_de: str, titel_en: str,
                        konten, perioden, ms_zeilen: int, projekt: str,
                        waehrung: str, rechenwerk: str,
                        schlusszeile: tuple[str, str]) -> dict:
    """Ein Rechenwerk aus der Bilanzposten-Spalte, mit Kontozeilen darunter."""
    ws = wb.create_sheet(blatt)
    spalten = [6 + i for i in range(len(perioden))]
    _kopf(ws, projekt, waehrung, titel_de, titel_en, spalten, perioden,
          "Gliederung: Bilanzposten des Exports (ClassDescription), "
          "sortiert nach ClassCode")

    zeile = Z_ERSTE
    erste = zeile
    postenzeilen = []
    for i, (posten, nos) in enumerate(_posten(konten, perioden)):
        pos_zeile = zeile
        postenzeilen.append((posten, pos_zeile))
        ws.cell(pos_zeile, SP_TYP, "POS").font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_TICKER, posten).font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_DE, posten).font = _schrift(fett=True)
        ws.cell(pos_zeile, SP_EN, posten).font = _schrift(fett=True)
        ws.cell(pos_zeile, SP_GRUPPE, f"{len(nos)} Konten").font = _schrift(
            farbe=GRAU)
        fuellung = PatternFill("solid", fgColor=TINT1 if i % 2 else TINT2)
        for spalte in range(SP_TYP, spalten[-1] + 1):
            ws.cell(pos_zeile, spalte).fill = fuellung

        zeile += 1
        erste_kto = zeile
        for no in nos:
            k = konten[no]
            ws.cell(zeile, SP_TYP, "KTO").font = _schrift(farbe=BLAU)
            ws.cell(zeile, SP_TICKER, k.konto).font = _schrift(farbe=BLAU)
            for spalte, wert in ((SP_DE, k.bezeichnung),
                                 (SP_EN, k.bezeichnung),
                                 (SP_GRUPPE, k.gruppe)):
                z = ws.cell(zeile, spalte, wert)
                z.font = _schrift(farbe=GRAU)
                z.alignment = Alignment(indent=2)
            for j, spalte in enumerate(spalten):
                ms = get_column_letter(MS_ERSTE_PERIODE + j)
                _zahlzelle(ws, zeile, spalte,
                           f"=SUMIFS(Mastersheet!${ms}$2:${ms}${ms_zeilen},"
                           f"Mastersheet!$A$2:$A${ms_zeilen},$B{zeile})/1000",
                           GRUEN)
            ws.row_dimensions[zeile].outlineLevel = 2
            ws.row_dimensions[zeile].hidden = True
            zeile += 1
        for spalte in spalten:
            sp = get_column_letter(spalte)
            _zahlzelle(ws, pos_zeile, spalte,
                       f"=SUM({sp}{erste_kto}:{sp}{zeile - 1})", SCHWARZ,
                       fett=True)
    letzte = zeile - 1
    typ = get_column_letter(SP_TYP)

    summenzeile = zeile
    ws.cell(summenzeile, SP_TYP, "SUM").font = _schrift(farbe=BLAU)
    for spalte, wert in ((SP_DE, schlusszeile[0]), (SP_EN, schlusszeile[1])):
        z = ws.cell(summenzeile, spalte, wert)
        z.font = _schrift(GROESSE + 1, SCHWARZ, fett=True)
        z.border = _RAHMEN_SUMME
    for spalte in spalten:
        sp = get_column_letter(spalte)
        _zahlzelle(ws, summenzeile, spalte,
                   f'=SUMIF(${typ}${erste}:${typ}${letzte},"<>KTO",'
                   f'{sp}{erste}:{sp}{letzte})', SCHWARZ, fett=True)
        ws.cell(summenzeile, spalte).border = _RAHMEN_SUMME
    zeile += 2

    # Nettovermoegen bzw. Ergebnis mit gedrehtem Vorzeichen — als Merkposten,
    # damit der Leser die gewohnte Zahl sieht, ohne dass irgendeine Rechnung
    # die Konvention verlaesst.
    netto_zeile = None
    if rechenwerk == "Bilanz":
        netto_zeile = zeile
        ws.cell(zeile, SP_TYP, "NET").font = _schrift(farbe=BLAU)
        ws.cell(zeile, SP_DE, "Nettovermögen (ohne Eigenkapital)").font = \
            _schrift(GROESSE + 1, SCHWARZ, fett=True)
        ws.cell(zeile, SP_EN, "Net assets (excluding equity)").font = \
            _schrift(GROESSE + 1, SCHWARZ, fett=True)
        ek = next((z for p, z in postenzeilen if p == EIGENKAPITALPOSTEN), None)
        for spalte in spalten:
            sp = get_column_letter(spalte)
            formel = f'=SUMIF(${typ}${erste}:${typ}${letzte},"<>KTO",' \
                     f'{sp}{erste}:{sp}{letzte})'
            if ek:
                formel += f"-{sp}{ek}"
            _zahlzelle(ws, zeile, spalte, formel, SCHWARZ, fett=True)
        zeile += 1
    else:
        netto_zeile = zeile
        ws.cell(zeile, SP_TYP, "MEM").font = _schrift(farbe=BLAU)
        ws.cell(zeile, SP_DE, "Jahresüberschuss / -fehlbetrag "
                              "(Vorzeichen gedreht)").font = _schrift(farbe=GRAU)
        ws.cell(zeile, SP_EN, "Profit / loss for the period "
                              "(sign reversed)").font = _schrift(farbe=GRAU)
        for spalte in spalten:
            sp = get_column_letter(spalte)
            _zahlzelle(ws, zeile, spalte, f"=-{sp}{summenzeile}", SCHWARZ)
        zeile += 1

    kontrollzeile = zeile + 1
    ws.cell(kontrollzeile, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
    if rechenwerk == "Bilanz":
        text = ("Kontrolle: Summe aller Bilanzposten (Bilanzidentität, "
                "muss null sein)")
        text_en = "Control: sum of all line items (balance identity, must be nil)"
        formeln = [f"={get_column_letter(s)}{summenzeile}" for s in spalten]
    else:
        # Zweiter Weg: das Ergebniskonto der Bilanz traegt dieselbe Zahl.
        text = (f"Kontrolle: Ergebnis ./. Konto {ERGEBNISKONTO} im "
                "Mastersheet (muss null sein)")
        text_en = (f"Control: result less account {ERGEBNISKONTO} in the "
                   "mastersheet (must be nil)")
        formeln = []
        for j, s in enumerate(spalten):
            ms = get_column_letter(MS_ERSTE_PERIODE + j)
            formeln.append(
                f"={get_column_letter(s)}{summenzeile}-"
                f'SUMIFS(Mastersheet!${ms}$2:${ms}${ms_zeilen},'
                f'Mastersheet!$A$2:$A${ms_zeilen},"{ERGEBNISKONTO}")/1000')
    for spalte, wert in ((SP_DE, text), (SP_EN, text_en)):
        z = ws.cell(kontrollzeile, spalte, wert)
        z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
        z.border = _RAHMEN_OBEN
    for spalte, formel in zip(spalten, formeln):
        _zahlzelle(ws, kontrollzeile, spalte, formel, SCHWARZ, ZAHL,
                   fett=True, groesse=GROESSE_KONTROLLE)
        ws.cell(kontrollzeile, spalte).border = _RAHMEN_OBEN

    _breiten(ws, spalten)
    return {"blatt": blatt, "posten": postenzeilen, "summenzeile": summenzeile,
            "nettozeile": netto_zeile, "kontrollzeile": kontrollzeile}


# --------------------------------------------------------------------------
# Recon: drei Bloecke nebeneinander
# --------------------------------------------------------------------------

#: Spaltenaufbau der Recon. Drei Bloecke a vier Perioden, dazwischen je eine
#: Leerspalte, damit die Bloecke optisch nicht ineinanderlaufen.
RECON_ERSTE = 6


def _recon_spalten(n: int) -> tuple[list[int], list[int], list[int]]:
    ja = [RECON_ERSTE + i for i in range(n)]
    susa = [ja[-1] + 2 + i for i in range(n)]
    diff = [susa[-1] + 2 + i for i in range(n)]
    return ja, susa, diff


def schreibe_recon(wb, blatt: str, titel_de: str, titel_en: str, werk: str,
                   abschluss: Abschluss, rechenwerk: dict, konten, perioden,
                   ms_zeilen: int, projekt: str, waehrung: str) -> dict:
    """Recon mit drei Bloecken nebeneinander."""
    ws = wb.create_sheet(blatt)
    n = len(perioden)
    ja, susa, diff = _recon_spalten(n)

    ws.cell(Z_PROJEKT, SP_DE, projekt).font = _schrift(GROESSE + 2, BLAU,
                                                       fett=True)
    ws.cell(Z_TITEL, SP_DE, titel_de).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_TITEL, SP_EN, titel_en).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_EINHEIT, SP_DE, f"in T {waehrung}").font = _schrift(farbe=GRAU)
    ws.cell(Z_EINHEIT, SP_EN,
            (f"Abschluss: {abschluss.quelle}" if abschluss.vorhanden
             else "abschluss.json fehlt — der linke Block bleibt leer")
            ).font = _schrift(farbe=GRAU)
    if abschluss.beispiel:
        z = ws.cell(Z_EINHEIT + 1, SP_DE,
                    "ACHTUNG: abschluss.json ist als BEISPIEL gekennzeichnet. "
                    "Die Zahlen im linken Block sind keine Abschlusszahlen.")
        z.font = _schrift(fett=True)
        z.fill = PatternFill("solid", fgColor=GELB)

    # Blockueberschriften ueber die Periodenspalten.
    for spalten, titel in ((ja, "Jahresabschluss"), (susa, "Saldenliste"),
                           (diff, "Differenz (Abschluss ./. Saldenliste)")):
        z = ws.cell(Z_KOPF - 1, spalten[0], titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        for spalte in spalten:
            ws.cell(Z_KOPF - 1, spalte).fill = PatternFill("solid",
                                                           fgColor=TEAL_DK)
    for spalte, titel in ((SP_TYP, "Typ"), (SP_TICKER, "Ticker"),
                          (SP_DE, "Bilanzposten"), (SP_EN, "Line item"),
                          (SP_GRUPPE, "Status")):
        z = ws.cell(Z_KOPF, spalte, titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
    for spalten in (ja, susa, diff):
        for i, p in enumerate(perioden):
            z = ws.cell(Z_KOPF, spalten[i], p)
            z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
            z.fill = PatternFill("solid", fgColor=TEAL)
            z.alignment = Alignment(horizontal="right")

    quelle = abschluss.bilanz if werk == "bilanz" else abschluss.guv
    susa_posten = dict(rechenwerk["posten"])
    # Reihenfolge: erst die Posten der Saldenliste, dann die, die es nur im
    # Abschluss gibt. Beides ist ein Befund und wird beschriftet.
    reihenfolge = [p for p, _ in rechenwerk["posten"]]
    nur_abschluss = [p for p in quelle if p not in susa_posten]
    reihenfolge += nur_abschluss

    zeile = Z_ERSTE
    zeilen_je_posten: dict[str, int] = {}
    nur_susa = []
    for i, posten in enumerate(reihenfolge):
        zeilen_je_posten[posten] = zeile
        im_abschluss = posten in quelle
        im_susa = posten in susa_posten
        if not im_abschluss:
            nur_susa.append(posten)
        ws.cell(zeile, SP_TYP, "POS").font = _schrift(farbe=BLAU)
        ws.cell(zeile, SP_TICKER, posten).font = _schrift(farbe=BLAU)
        ws.cell(zeile, SP_DE, posten).font = _schrift(fett=True)
        ws.cell(zeile, SP_EN, posten).font = _schrift(fett=True)
        status = ("" if im_abschluss and im_susa else
                  "nur im Abschluss" if im_abschluss else
                  "nur in der Saldenliste")
        ws.cell(zeile, SP_GRUPPE, status).font = _schrift(fett=bool(status))
        fuellung = PatternFill("solid", fgColor=GELB if status else
                               (TINT1 if i % 2 else TINT2))
        for spalte in range(SP_TYP, diff[-1] + 1):
            ws.cell(zeile, spalte).fill = fuellung

        for j, p in enumerate(perioden):
            wert = quelle.get(posten, {}).get(p)
            if wert is not None:
                # Blau: der Abschluss ist hartkodierter Input aus abschluss.json.
                _zahlzelle(ws, zeile, ja[j], round(wert / 1000.0, 6), BLAU)
            if im_susa:
                _zahlzelle(ws, zeile, susa[j],
                           f"='{rechenwerk['blatt']}'!"
                           f"{get_column_letter(6 + j)}{susa_posten[posten]}",
                           GRUEN)
            a, b = get_column_letter(ja[j]), get_column_letter(susa[j])
            _zahlzelle(ws, zeile, diff[j],
                       f"=N({a}{zeile})-N({b}{zeile})", SCHWARZ, fett=True)
        zeile += 1

    letzte = zeile - 1
    zeile += 1

    # Die EINZIGE sinnvolle Gesamtgroesse.
    gesamt = zeile
    ws.cell(gesamt, SP_TYP, "NET").font = _schrift(farbe=BLAU)
    if werk == "bilanz":
        de, en = "Nettovermögen (ohne Eigenkapital)", "Net assets (excluding equity)"
    else:
        de, en = "Ergebnis der Periode", "Result for the period"
    for spalte, wert in ((SP_DE, de), (SP_EN, en)):
        z = ws.cell(gesamt, spalte, wert)
        z.font = _schrift(GROESSE + 1, SCHWARZ, fett=True)
        z.border = _RAHMEN_SUMME
    typ = get_column_letter(SP_TYP)
    ek_zeile = zeilen_je_posten.get(EIGENKAPITALPOSTEN)
    for j in range(n):
        for spalten in (ja, susa):
            sp = get_column_letter(spalten[j])
            formel = f'=SUMIF(${typ}${Z_ERSTE}:${typ}${letzte},"<>KTO",' \
                     f'{sp}{Z_ERSTE}:{sp}{letzte})'
            if werk == "bilanz" and ek_zeile:
                formel += f"-N({sp}{ek_zeile})"
            _zahlzelle(ws, gesamt, spalten[j], formel, SCHWARZ, fett=True)
            ws.cell(gesamt, spalten[j]).border = _RAHMEN_SUMME
        a, b = get_column_letter(ja[j]), get_column_letter(susa[j])
        _zahlzelle(ws, gesamt, diff[j], f"={a}{gesamt}-{b}{gesamt}", SCHWARZ,
                   fett=True)
        ws.cell(gesamt, diff[j]).border = _RAHMEN_SUMME
    zeile += 2

    # Kontrolle: der SuSa-Block muss dem Rechenwerk entsprechen.
    kontrollzeile = zeile
    ws.cell(kontrollzeile, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
    bezug = rechenwerk["nettozeile"] if werk == "bilanz" \
        else rechenwerk["summenzeile"]
    text = (f"Kontrolle: Saldenliste-Block ./. Tab {rechenwerk['blatt']} "
            "(muss null sein)")
    for spalte, wert in ((SP_DE, text),
                         (SP_EN, f"Control: trial balance block less tab "
                                 f"{rechenwerk['blatt']} (must be nil)")):
        z = ws.cell(kontrollzeile, spalte, wert)
        z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
        z.border = _RAHMEN_OBEN
    for j in range(n):
        sp = get_column_letter(susa[j])
        _zahlzelle(ws, kontrollzeile, diff[j],
                   f"={sp}{gesamt}-'{rechenwerk['blatt']}'!"
                   f"{get_column_letter(6 + j)}{bezug}", SCHWARZ, ZAHL,
                   fett=True, groesse=GROESSE_KONTROLLE)
        ws.cell(kontrollzeile, diff[j]).border = _RAHMEN_OBEN
    zeile += 1

    # Optionale Gegenprobe gegen die ausgewiesene Gesamtgroesse des Abschlusses.
    gemeldet = (abschluss.nettovermoegen if werk == "bilanz"
                else abschluss.jahresergebnis)
    if gemeldet:
        zeile += 1
        ws.cell(zeile, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
        was = "NET ASSETS" if werk == "bilanz" else "Jahresergebnis"
        for spalte, wert in (
                (SP_DE, f"Kontrolle: Summe der Posten ./. ausgewiesenes "
                        f"{was} lt. Abschluss (muss null sein)"),
                (SP_EN, f"Control: sum of line items less reported {was} "
                        "(must be nil)")):
            z = ws.cell(zeile, spalte, wert)
            z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
            z.border = _RAHMEN_OBEN
        for j, p in enumerate(perioden):
            sp = get_column_letter(ja[j])
            _zahlzelle(ws, zeile, ja[j],
                       f"={sp}{gesamt}-{round(gemeldet.get(p, 0.0) / 1000.0, 6)}",
                       SCHWARZ, ZAHL, fett=True, groesse=GROESSE_KONTROLLE)
            ws.cell(zeile, ja[j]).border = _RAHMEN_OBEN
        zeile += 1

    zeile += 1
    z = ws.cell(zeile, SP_DE,
                "Vermerk: Es gibt hier bewusst KEINE Summe über alle Zeilen. "
                "Sie mischte Aktiva, Passiva und Eigenkapital und wäre "
                "bedeutungslos; in der Saldenliste ist sie zudem "
                "konstruktionsbedingt null und täuschte Übereinstimmung vor. "
                "Die einzige sinnvolle Gesamtgröße ist die Differenz im "
                "Nettovermögen."
                if werk == "bilanz" else
                "Vermerk: In der GuV ist die Summe über alle Zeilen sehr wohl "
                "sinnvoll — sie ist das Periodenergebnis, und alle Zeilen "
                "tragen dieselbe Vorzeichenkonvention (Erträge im Haben, "
                "also negativ).")
    z.font = _schrift(farbe=GRAU)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=zeile, start_column=SP_DE, end_row=zeile,
                   end_column=diff[-1])
    ws.row_dimensions[zeile].height = 40

    _breiten(ws, ja + susa + diff)
    ws.column_dimensions[get_column_letter(ja[-1] + 1)].width = 3
    ws.column_dimensions[get_column_letter(susa[-1] + 1)].width = 3
    return {"blatt": blatt, "zeilen": zeilen_je_posten, "gesamtzeile": gesamt,
            "kontrollzeile": kontrollzeile, "nur_abschluss": nur_abschluss,
            "nur_susa": nur_susa, "werk": werk}


def _breiten(ws, spalten) -> None:
    for spalte, breite in (("A", 5), ("B", 24), ("C", 40), ("D", 40),
                           ("E", 20)):
        ws.column_dimensions[spalte].width = breite
    for spalte in spalten:
        ws.column_dimensions[get_column_letter(spalte)].width = 14
    ws.column_dimensions["A"].hidden = True
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = ws.cell(Z_ERSTE, RECON_ERSTE)


def schreibe_rechenwerke(wb, bs_konten, guv_konten, perioden, ms_zeilen,
                         projekt, waehrung, abschlusspfad=ABSCHLUSS) -> dict:
    """Bilanz, GuV, Recon Bilanz, Recon GuV."""
    abschluss = lies_abschluss(abschlusspfad)
    bilanz = schreibe_rechenwerk(
        wb, "Bilanz", "Bilanz", "Balance sheet", bs_konten, perioden,
        ms_zeilen, projekt, waehrung, "Bilanz",
        ("Summe aller Bilanzposten", "Sum of all line items"))
    guv = schreibe_rechenwerk(
        wb, "GuV", "Gewinn- und Verlustrechnung", "Profit and loss",
        guv_konten, perioden, ms_zeilen, projekt, waehrung, "GuV",
        ("Saldo der GuV-Konten (negativ = Gewinn)",
         "Balance of P&L accounts (negative = profit)"))
    recon_b = schreibe_recon(wb, "Recon Bilanz", "Recon Bilanz",
                             "Reconciliation balance sheet", "bilanz",
                             abschluss, bilanz, bs_konten, perioden, ms_zeilen,
                             projekt, waehrung)
    recon_g = schreibe_recon(wb, "Recon GuV", "Recon GuV",
                             "Reconciliation profit and loss", "guv",
                             abschluss, guv, guv_konten, perioden, ms_zeilen,
                             projekt, waehrung)
    return {"abschluss": abschluss, "bilanz": bilanz, "guv": guv,
            "recon_bilanz": recon_b, "recon_guv": recon_g}
