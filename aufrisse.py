"""Punkt 3: die sieben Aufriss-Tabs.

Ein Aufriss ist die mittlere Schicht: das Mastersheet traegt jede Kontozahl
genau einmal, der Aufriss gliedert eine Position und zeigt per **sichtbarer
Formel**, aus welchen Konten sie besteht. Wer eine Zahl im Databook nicht
zurueckverfolgen kann, hat kein Databook, sondern eine Behauptung.

Daraus folgt der Aufbau jedes Tabs:

* Die Kontozeilen holen ihren Wert per ``SUMIF`` aus dem Mastersheet. Sie
  stehen eingeklappt unter ihrer Position und sind gruen — Querverweis auf ein
  anderes Blatt.
* Die Positionszeile summiert ihre eigenen Kontozeilen. Schwarz, weil die
  Formel im Blatt bleibt.
* Die Summenzeile summiert **ueber die Zeilentyp-Spalte**
  (``SUMIF(...;"<>KTO";...)``). Ein blankes ``SUM`` zaehlte die Kontozeilen
  doppelt, weil sie innerhalb des Bereichs liegen.
* Die Kontrollzeile stellt der Summe einen **zweiten, unabhaengigen Weg**
  gegenueber: ein ``SUMIFS`` ueber das Mastersheet, das nicht die einzelnen
  Konten nennt, sondern das Merkmal (Klasse, Kategorie, WC-Seite). Beide Wege
  muessen dieselbe Zahl liefern, die Differenz muss null sein. Eine
  Kontrollzeile prueft — sie gleicht nicht aus.

**Ein Aufriss ohne Datengrundlage wird trotzdem angelegt.** Leer, mit Vermerk,
und der Vermerk nennt die fehlende Unterlage. Ein Tab, der gar nicht erst
entsteht, verschwindet aus der Wahrnehmung; ein leerer Tab mit der Zeile
"Anlagenspiegel fehlt" ist eine Datenanforderung, die jeder sieht.

Aufruf::

    python3 aufrisse.py [ziel.xlsx] [referenz] [klassifizierung_v1.json]
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from abnahme import lies_export
from klassifizierung import Classifier

# ==========================================================================
# LAYOUT — das, was FORMAT.md regelt
#
# FORMAT.md liegt nicht im Repository. Gebaut ist deshalb nach dem
# Excel-Hausformat: Teal-Palette fuer Schedules, Arial 8 im Datenbereich und
# Arial 9 in Kontrollzeilen, Buchhaltungsformat mit Klammern fuer negative
# Werte und Bindestrich fuer null, Betraege in Tausend mit der Division IN
# der Formel, und die Schriftfarbe als Herkunftsangabe: blau hartkodierter
# Input, schwarz Formel im selben Blatt, gruen Querverweis, gelbe Fuellung
# als Flag.
#
# Alles Sichtbare steht in diesem Block. Taucht FORMAT.md auf, ist das die
# eine Stelle, die sich aendert.
# ==========================================================================

TEAL_DK = "FF005858"        # Titel
TEAL = "FF008888"           # Kopfbaender
TINT1 = "FFE0F8F8"          # Baenderung hell
TINT2 = "FFC8F0F0"          # Baenderung dunkler
GELB = "FFFFF2CC"           # Flag

BLAU = "FF0000FF"           # hartkodierter Input, dazu die Ticker
SCHWARZ = "FF000000"        # Formel im selben Blatt
GRUEN = "FF008000"          # Querverweis auf ein anderes Blatt
GRAU = "FF808080"           # Kontozeilen

SCHRIFT = "Arial"
GROESSE = 8
GROESSE_KONTROLLE = 9

#: Buchhaltungsformat: negative Werte in Klammern, Bindestrich fuer null.
ZAHL = '#,##0.0;(#,##0.0);"-"'
PROZENT = '0.0%;(0.0%);"-"'

#: Spalten des Aufrisses. A traegt den Zeilentyp und wird ausgeblendet — er
#: ist die Grundlage der Summenformel, nicht Information fuer den Leser.
SP_TYP, SP_TICKER, SP_DE, SP_EN, SP_GRUPPE = 1, 2, 3, 4, 5
SP_ERSTE_PERIODE = 6
BREITEN = {"A": 5, "B": 13, "C": 44, "D": 44, "E": 26}

#: Kopfzeilen. Zeile 1 Projekt, 2 Titel, 3 Einheit und Quelle, 5 Spaltenkopf.
Z_PROJEKT, Z_TITEL, Z_EINHEIT, Z_KOPF, Z_ERSTE = 1, 2, 3, 5, 6

#: Mastersheet — Arbeitsblatt, vom Formatgrundsatz ausgenommen.
#:
#: Es traegt BEIDE Rechenwerke. Die GuV-Konten sind NICHT klassifiziert: die
#: Entscheidungsdatei kennt keine GuV-Kategorien, und wuerde man sie trotzdem
#: durch die Stichwortregeln schicken, landete ein Zinsaufwand im Net Debt.
#: Ihre Spalten Klasse, Kategorie und WC-Seite bleiben deshalb leer, und die
#: Kriterien der Aufrisse greifen sie nicht.
MS_SPALTEN = ["Konto", "Bezeichnung", "Kontogruppe", "Bilanzseite", "Klasse",
              "Kategorie", "WC-Seite", "Quelle", "Regel", "Review",
              "Rechenwerk", "Bilanzposten", "Postencode"]
MS_KONTO, MS_KLASSE, MS_KATEGORIE, MS_SEITE = 1, 5, 6, 7
MS_RECHENWERK, MS_POSTEN, MS_CODE = 11, 12, 13
MS_ERSTE_PERIODE = len(MS_SPALTEN) + 1

_RAHMEN_OBEN = Border(top=Side(style="thin", color=TEAL_DK))
_RAHMEN_SUMME = Border(top=Side(style="thin", color=TEAL_DK),
                       bottom=Side(style="double", color=TEAL_DK))


# ==========================================================================
# Die sieben Aufrisse
# ==========================================================================

@dataclass
class Aufriss:
    """Ein Aufriss-Tab: was er zeigt, wie er gliedert, wogegen er prueft."""

    blatt: str
    titel_de: str
    titel_en: str
    #: Welche Konten der Aufriss zeigt.
    auswahl: Optional[Callable[[dict], bool]] = None
    #: Gliederung: nach der FDD-Kategorie oder nach der Kontogruppe des
    #: Kontenplans. Die Vorrats- und Anlagenaufrisse gliedern nach der
    #: Kontogruppe, weil dort die Bestandteile stehen.
    gliederung: str = "kategorie"
    #: Der zweite, unabhaengige Weg: Merkmal und Wert im Mastersheet.
    kontrolle: list[tuple[int, str]] = field(default_factory=list)
    #: Zusatzzeilen unter der Summe (Bezeichnung, Kriterien) — Merkposten.
    memo: list[tuple[str, str, list[tuple[int, str]]]] = field(default_factory=list)
    #: Abnutzungsgrad aus zwei Memo-Zeilen, falls gesetzt (Index, Index).
    quote: Optional[tuple[str, str, int, int]] = None
    #: Eine Kategorie, die nicht allein in einer Position stehen darf.
    #: Kumulierte Abschreibung ist ein Korrekturposten und gehoert neben die
    #: Anschaffungskosten derselben Anlagenklasse. Steht sie allein, fehlt
    #: entweder die Klasse im Aufriss oder die Zuordnung im Kontenplan — und
    #: die Position zeigt einen negativen "Buchwert".
    unselbstaendig: str = ""
    vermerk: str = ""


def aufrisse() -> list[Aufriss]:
    """Die sieben Tabs. Sechs tragen Daten, einer nicht — und wird trotzdem
    angelegt."""
    return [
        Aufriss("NA_OA", "Operatives Vermögen", "Operating assets",
                auswahl=lambda r: r["wc_seite"] == "OA",
                kontrolle=[(MS_SEITE, "OA")],
                vermerk="Working Capital der Aktivseite. OA ist keine eigene "
                        "Klasse, sondern die Bilanzseite: aktive "
                        "Rechnungsabgrenzung bleibt OWC und liegt hier."),
        Aufriss("NA_OL", "Operative Verbindlichkeiten",
                "Operating liabilities",
                auswahl=lambda r: r["wc_seite"] == "OL",
                kontrolle=[(MS_SEITE, "OL")],
                vermerk="Working Capital der Passivseite. Passive "
                        "Rechnungsabgrenzung bleibt OWC und liegt hier."),
        Aufriss("NA_TWC", "Trade Working Capital", "Trade working capital",
                auswahl=lambda r: r["klasse"] == "TWC",
                kontrolle=[(MS_KLASSE, "TWC")],
                vermerk="Nur der Handelsteil: Forderungen, Vorräte, "
                        "Verbindlichkeiten aus L+L. Die übrigen "
                        "Working-Capital-Positionen stehen in NA_OA und "
                        "NA_OL."),
        Aufriss("NA_Net Debt", "Net Debt", "Net debt",
                auswahl=lambda r: r["klasse"] == "ND",
                kontrolle=[(MS_KLASSE, "ND")],
                vermerk="Net Debt gegen Working Capital ist eine "
                        "Charakterfrage — finanzierungsartig gegen operativ. "
                        "Höhe und Wiederkehr führen zur Normalisierung, nie "
                        "zur Umklassifizierung."),
        Aufriss("NA_Vorräte", "Vorräte", "Inventories",
                auswahl=lambda r: r["category"] == "TWC/Inventories",
                gliederung="kontogruppe",
                kontrolle=[(MS_KATEGORIE, "TWC/Inventories")],
                vermerk="Gegliedert nach Bestandteilen des Kontenplans, nicht "
                        "nach Produkten — eine Saldenliste kennt keine "
                        "Produktdimension. Wertberichtigungen bleiben in "
                        "derselben Position wie der Bestand."),
        Aufriss("NA_Sachanlagen", "Sachanlagen", "Tangible assets",
                auswahl=lambda r: r["category"] in ("FA/PPE",
                                                    "FA/Accumulated depreciation"),
                gliederung="kontogruppe",
                kontrolle=[(MS_KATEGORIE, "FA/PPE"),
                           (MS_KATEGORIE, "FA/Accumulated depreciation")],
                memo=[("Anschaffungs- und Herstellungskosten",
                       "Historical costs", [(MS_KATEGORIE, "FA/PPE")]),
                      ("Kumulierte Abschreibung", "Accumulated depreciation",
                       [(MS_KATEGORIE, "FA/Accumulated depreciation")])],
                quote=("Abnutzungsgrad", "Degree of wear", 1, 0),
                unselbstaendig="FA/Accumulated depreciation",
                vermerk="Buchwert je Anlagenklasse, darunter die Konten mit "
                        "Anschaffungskosten und kumulierter Abschreibung. "
                        "BEFUND: die Kategorie 'FA/Accumulated depreciation' "
                        "trägt keine Anlagenklasse. Sie enthält deshalb auch "
                        "die Abschreibung auf immaterielle Vermögensgegen"
                        "stände und Nutzungsrechte, deren Anschaffungskosten "
                        "in anderen Klassen stehen. Gelb markierte Positionen "
                        "zeigen Abschreibung ohne zugehörige Anschaffungs"
                        "kosten; die Summe dieses Tabs ist deshalb KEIN "
                        "Sachanlagen-Buchwert. Die Kontrollzeile stimmt "
                        "trotzdem — sie prüft die Auswahl, nicht die "
                        "Sachlogik."),
        Aufriss("NA_CAPEX", "Investitionen", "Capital expenditure",
                vermerk="KEINE DATENGRUNDLAGE. Investitionen sind die Zugänge "
                        "einer Periode. Die Saldenliste führt Schlussbestände; "
                        "die Veränderung der Anschaffungskosten ist Zugänge "
                        "minus Abgänge minus Währungseffekte und darf nicht "
                        "als CAPEX beschriftet werden. Benötigt wird ein "
                        "Anlagenspiegel oder ein Anlagengitter mit Zugängen, "
                        "Abgängen und Umbuchungen je Anlagenklasse und "
                        "Periode. Der Tab bleibt bis dahin leer."),
    ]


# ==========================================================================
# Bauen
# ==========================================================================

def _schrift(groesse: int = GROESSE, farbe: str = SCHWARZ,
             fett: bool = False) -> Font:
    return Font(name=SCHRIFT, size=groesse, color=farbe, bold=fett)


def _zahlzelle(ws, zeile: int, spalte: int, wert, farbe: str,
               format_: str = ZAHL, fett: bool = False,
               groesse: int = GROESSE) -> None:
    z = ws.cell(zeile, spalte, wert)
    z.font = _schrift(groesse, farbe, fett)
    z.number_format = format_
    z.alignment = Alignment(horizontal="right")


def _bereich(erste_spalte: int, anzahl: int) -> list[int]:
    return [erste_spalte + i for i in range(anzahl)]


def schreibe_mastersheet(wb: Workbook, konten, perioden, ergebnisse,
                        guv_konten=None) -> int:
    """Das Mastersheet. Jede Kontozahl wohnt hier und nirgends sonst.

    Arbeitsblatt, deshalb schlicht: eine Kopfzeile, eine Zeile je Konto, die
    Merkmale der Klassifizierung als eigene Spalten. Genau diese Spalten
    machen die Kontrollzeile der Aufrisse moeglich — sie summiert ueber ein
    Merkmal statt ueber eine Kontenliste und ist damit vom Aufriss unabhaengig.
    """
    ws = wb.create_sheet("Mastersheet")
    kopf = MS_SPALTEN + list(perioden)
    for i, titel in enumerate(kopf, start=1):
        z = ws.cell(1, i, titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
    zeile = 2
    for no in sorted(konten):
        k, e = konten[no], ergebnisse[no]
        _mastersheetzeile(ws, zeile, k, perioden,
                          [e["klasse"], e["category"], e["wc_seite"],
                           e["source"], e["rule_id"],
                           "ja" if e["review"] else "", "Bilanz"])
        zeile += 1
    for no in sorted(guv_konten or {}):
        # GuV-Konten: unklassifiziert, siehe Kommentar an MS_SPALTEN.
        _mastersheetzeile(ws, zeile, (guv_konten or {})[no], perioden,
                          ["", "", "", "", "", "", "GuV"])
        zeile += 1
    ws.freeze_panes = "B2"
    for spalte, breite in (("A", 12), ("B", 42), ("C", 30), ("D", 11),
                           ("E", 8), ("F", 26), ("G", 10), ("H", 13),
                           ("I", 24), ("J", 8), ("K", 11), ("L", 22),
                           ("M", 12)):
        ws.column_dimensions[spalte].width = breite
    return ws.max_row


def _mastersheetzeile(ws, zeile: int, k, perioden, merkmale: list) -> None:
    for i, wert in enumerate([k.konto, k.bezeichnung, k.gruppe, k.seite]
                             + merkmale[:6] + [merkmale[6], k.posten,
                                               k.postencode], start=1):
        ws.cell(zeile, i, wert).font = _schrift()
    for i, p in enumerate(perioden):
        _zahlzelle(ws, zeile, MS_ERSTE_PERIODE + i, round(k.saldo(p), 2), BLAU,
                   '#,##0.00;(#,##0.00);"-"')


def _kontrollformel(kriterien: list[tuple[int, str]], ms_zeilen: int,
                    ms_spalte: int) -> str:
    """SUMIFS ueber ein MERKMAL des Mastersheets, nicht ueber Konten.

    Das ist der Punkt: haette die Kontrolle dieselbe Kontenliste wie der
    Aufriss, pruefte sie nur das Addieren. So prueft sie die Auswahl.
    """
    wert = get_column_letter(ms_spalte)
    teile = [f'SUMIFS(Mastersheet!${wert}$2:${wert}${ms_zeilen},'
             f'Mastersheet!${get_column_letter(sp)}$2:'
             f'${get_column_letter(sp)}${ms_zeilen},"{w}")'
             for sp, w in kriterien]
    return "(" + "+".join(teile) + ")/1000"


def schreibe_aufriss(wb: Workbook, a: Aufriss, konten, perioden, ergebnisse,
                     ms_zeilen: int, projekt: str, waehrung: str) -> dict:
    """Ein Aufriss-Tab. Gibt zurueck, was der Abnahmetest pruefen muss."""
    ws = wb.create_sheet(a.blatt)
    n = len(perioden)
    spalten = _bereich(SP_ERSTE_PERIODE, n)

    # -- Kopf ------------------------------------------------------------
    # Blau: der Projektname ist hartkodierter Input.
    ws.cell(Z_PROJEKT, SP_DE, projekt).font = _schrift(GROESSE + 2, BLAU,
                                                       fett=True)
    ws.cell(Z_TITEL, SP_DE, a.titel_de).font = _schrift(GROESSE + 1, TEAL_DK,
                                                        fett=True)
    ws.cell(Z_TITEL, SP_EN, a.titel_en).font = _schrift(GROESSE + 1, TEAL_DK,
                                                        fett=True)
    ws.cell(Z_EINHEIT, SP_DE, f"in T {waehrung}").font = _schrift(farbe=GRAU)
    ws.cell(Z_EINHEIT, SP_EN,
            "Quelle: Saldenliste je Geschäftsjahr / Source: annual trial "
            "balance").font = _schrift(farbe=GRAU)

    for spalte, titel in ((SP_TYP, "Typ"), (SP_TICKER, "Ticker"),
                          (SP_DE, "Bezeichnung"), (SP_EN, "Description"),
                          (SP_GRUPPE, "Kontogruppe")):
        z = ws.cell(Z_KOPF, spalte, titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
    for i, p in enumerate(perioden):
        z = ws.cell(Z_KOPF, spalten[i], p)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
        z.alignment = Alignment(horizontal="right")

    # -- Inhalt ----------------------------------------------------------
    auswahl = {no: k for no, k in konten.items()
               if a.auswahl and a.auswahl(ergebnisse[no])}
    gruppen: dict[str, list[str]] = {}
    for no, k in auswahl.items():
        schluessel = (ergebnisse[no]["category"] if a.gliederung == "kategorie"
                      else (k.gruppe or "ohne Kontogruppe"))
        gruppen.setdefault(schluessel, []).append(no)

    def groesse_von(nos):
        return max((abs(konten[x].saldo(perioden[-1])) for x in nos), default=0.0)

    zeile = Z_ERSTE
    positionszeilen: list[dict] = []
    geflaggt: list[str] = []
    erste_inhaltszeile = zeile
    for i, schluessel in enumerate(sorted(gruppen, key=lambda g:
                                          groesse_von(gruppen[g]), reverse=True)):
        nos = sorted(gruppen[schluessel],
                     key=lambda x: abs(konten[x].saldo(perioden[-1])),
                     reverse=True)
        pos_zeile = zeile
        positionszeilen.append({"zeile": pos_zeile, "schluessel": schluessel,
                                "konten": nos})
        # Traegt die Position NUR den unselbstaendigen Korrekturposten, fehlt
        # ihr das Gegenstueck. Gelb, damit es niemand ueberliest.
        allein = bool(a.unselbstaendig) and all(
            ergebnisse[x]["category"] == a.unselbstaendig for x in nos)
        if allein:
            geflaggt.append(schluessel)
        ws.cell(pos_zeile, SP_TYP, "POS").font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_TICKER, schluessel).font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_DE, schluessel).font = _schrift(fett=True)
        ws.cell(pos_zeile, SP_EN, schluessel).font = _schrift(fett=True)
        if allein:
            ws.cell(pos_zeile, SP_GRUPPE,
                    "Abschreibung ohne zugehörige Anschaffungskosten"
                    ).font = _schrift(fett=True)
        fuellung = PatternFill("solid", fgColor=GELB if allein
                               else (TINT1 if i % 2 else TINT2))
        for spalte in range(SP_TYP, spalten[-1] + 1):
            ws.cell(pos_zeile, spalte).fill = fuellung

        zeile += 1
        erste_kto, letzte_kto = zeile, zeile + len(nos) - 1
        for no in nos:
            k = konten[no]
            ws.cell(zeile, SP_TYP, "KTO").font = _schrift(farbe=BLAU)
            ws.cell(zeile, SP_TICKER, k.konto).font = _schrift(farbe=BLAU)
            for spalte, wert in ((SP_DE, k.bezeichnung), (SP_EN, k.bezeichnung),
                                 (SP_GRUPPE, k.gruppe)):
                z = ws.cell(zeile, spalte, wert)
                z.font = _schrift(farbe=GRAU)
                z.alignment = Alignment(indent=2)
            for j in range(n):
                ms = get_column_letter(MS_ERSTE_PERIODE + j)
                # Gruen: der Wert kommt aus einem anderen Blatt.
                _zahlzelle(ws, zeile, spalten[j],
                           f"=SUMIF(Mastersheet!$A$2:$A${ms_zeilen},"
                           f"$B{zeile},Mastersheet!${ms}$2:${ms}${ms_zeilen})"
                           f"/1000", GRUEN)
            ws.row_dimensions[zeile].outlineLevel = 2
            ws.row_dimensions[zeile].hidden = True
            zeile += 1

        for j in range(n):
            sp = get_column_letter(spalten[j])
            # Schwarz: die Position summiert ihre eigenen Kontozeilen.
            _zahlzelle(ws, pos_zeile, spalten[j],
                       f"=SUM({sp}{erste_kto}:{sp}{letzte_kto})", SCHWARZ,
                       fett=True)

    letzte_inhaltszeile = zeile - 1

    # -- Summe, Kontrolle, Vermerk ---------------------------------------
    if not auswahl:
        # Ein Aufriss ohne Datengrundlage wird trotzdem angelegt.
        z = ws.cell(Z_ERSTE, SP_DE, "Kein Aufriss möglich — siehe Vermerk")
        z.font = _schrift(fett=True)
        z.fill = PatternFill("solid", fgColor=GELB)
        for spalte in range(SP_TYP, spalten[-1] + 1):
            ws.cell(Z_ERSTE, spalte).fill = PatternFill("solid", fgColor=GELB)
        summenzeile = kontrollzeile = None
        zeile = Z_ERSTE + 2
    else:
        summenzeile = zeile + 1
        ws.cell(summenzeile, SP_TYP, "SUM").font = _schrift(farbe=BLAU)
        for spalte, wert in ((SP_DE, a.titel_de), (SP_EN, a.titel_en)):
            z = ws.cell(summenzeile, spalte, wert)
            z.font = _schrift(GROESSE, SCHWARZ, fett=True)
            z.border = _RAHMEN_SUMME
        for j in range(n):
            sp = get_column_letter(spalten[j])
            typ = get_column_letter(SP_TYP)
            # Die Kontoslots liegen IM Bereich. Ohne "<>KTO" doppelt gezaehlt.
            _zahlzelle(ws, summenzeile, spalten[j],
                       f'=SUMIF(${typ}${erste_inhaltszeile}:'
                       f'${typ}${letzte_inhaltszeile},"<>KTO",'
                       f'{sp}{erste_inhaltszeile}:{sp}{letzte_inhaltszeile})',
                       SCHWARZ, fett=True)
            ws.cell(summenzeile, spalten[j]).border = _RAHMEN_SUMME

        zeile = summenzeile + 2
        for de, en, kriterien in a.memo:
            ws.cell(zeile, SP_TYP, "MEM").font = _schrift(farbe=BLAU)
            ws.cell(zeile, SP_DE, de).font = _schrift(farbe=GRAU)
            ws.cell(zeile, SP_EN, en).font = _schrift(farbe=GRAU)
            for j in range(n):
                _zahlzelle(ws, zeile, spalten[j],
                           "=" + _kontrollformel(kriterien, ms_zeilen,
                                                 MS_ERSTE_PERIODE + j), GRUEN)
            zeile += 1
        if a.quote and a.memo:
            de, en, oben, unten = a.quote
            ws.cell(zeile, SP_TYP, "MEM").font = _schrift(farbe=BLAU)
            ws.cell(zeile, SP_DE, de + " (in %)").font = _schrift(farbe=GRAU)
            ws.cell(zeile, SP_EN, en + " (in %)").font = _schrift(farbe=GRAU)
            for j in range(n):
                sp = get_column_letter(spalten[j])
                a_zeile = summenzeile + 2 + oben
                b_zeile = summenzeile + 2 + unten
                _zahlzelle(ws, zeile, spalten[j],
                           f'=IFERROR(-{sp}{a_zeile}/{sp}{b_zeile},"n/a")',
                           SCHWARZ, PROZENT)
            zeile += 1

        kontrollzeile = zeile + 1
        ws.cell(kontrollzeile, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
        for spalte, wert in ((SP_DE, "Kontrolle gegen das Mastersheet (muss "
                                     "null sein)"),
                             (SP_EN, "Control against the mastersheet (must "
                                     "be nil)")):
            z = ws.cell(kontrollzeile, spalte, wert)
            z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
            z.border = _RAHMEN_OBEN
        for j in range(n):
            sp = get_column_letter(spalten[j])
            _zahlzelle(ws, kontrollzeile, spalten[j],
                       f"={sp}{summenzeile}-"
                       + _kontrollformel(a.kontrolle, ms_zeilen,
                                         MS_ERSTE_PERIODE + j),
                       SCHWARZ, ZAHL, fett=True, groesse=GROESSE_KONTROLLE)
            ws.cell(kontrollzeile, spalten[j]).border = _RAHMEN_OBEN
        zeile = kontrollzeile + 2

    vermerkzeile = zeile
    z = ws.cell(vermerkzeile, SP_DE, "Vermerk: " + a.vermerk)
    z.font = _schrift(farbe=GRAU)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    if not auswahl:
        z.fill = PatternFill("solid", fgColor=GELB)
    ws.merge_cells(start_row=vermerkzeile, start_column=SP_DE,
                   end_row=vermerkzeile, end_column=spalten[-1])
    ws.row_dimensions[vermerkzeile].height = 46

    for spalte, breite in BREITEN.items():
        ws.column_dimensions[spalte].width = breite
    for spalte in spalten:
        ws.column_dimensions[get_column_letter(spalte)].width = 15
    ws.column_dimensions["A"].hidden = True
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = ws.cell(Z_ERSTE, SP_ERSTE_PERIODE)

    return {"blatt": a.blatt, "konten": sorted(auswahl),
            "positionen": len(gruppen), "summenzeile": summenzeile,
            "kontrollzeile": kontrollzeile, "vermerkzeile": vermerkzeile,
            "leer": not auswahl, "geflaggt": geflaggt,
            "kontrolle": a.kontrolle, "vermerk": a.vermerk,
            # Der Lead zieht seine Positionssummen aus genau diesen Zeilen.
            "positionszeilen": positionszeilen}


def baue(ordner: str, spec: str, ziel: str, projekt: str = "Referenzfall",
         waehrung: str = "AUD", abschluss: str = "abschluss.json") -> dict:
    """Liest den Referenzfall, klassifiziert und schreibt die Arbeitsmappe."""
    bs = lies_export(os.path.join(ordner, "referenz_BS.xlsx"))
    guv = lies_export(os.path.join(ordner, "referenz_PL.xlsx"))
    c = Classifier(spec)
    ergebnisse = {no: c.classify(k.konto, k.bezeichnung, k.gruppe,
                                 seite=k.seite or None)
                  for no, k in bs.konten.items()}

    wb = Workbook()
    wb.remove(wb.active)
    ms_zeilen = schreibe_mastersheet(wb, bs.konten, bs.perioden, ergebnisse,
                                     guv.konten)
    befunde = [schreibe_aufriss(wb, a, bs.konten, bs.perioden, ergebnisse,
                                ms_zeilen, projekt, waehrung)
               for a in aufrisse()]

    # Die Lead-Schicht setzt auf den fertigen Aufrissen auf und braucht deren
    # Zeilennummern. Der Import steht hier und nicht oben, weil leads.py
    # umgekehrt das Layout dieses Moduls verwendet — zwei Module, ein Format,
    # kein Ringschluss beim Laden.
    from leads import schreibe_leads
    leadbefunde = schreibe_leads(wb, {b["blatt"]: b for b in befunde},
                                 bs.konten, bs.perioden, ergebnisse,
                                 ms_zeilen, projekt, waehrung)

    # Bilanz, GuV und die beiden Recon-Tabs. Sie stehen NEBEN der
    # Klassifizierung, nicht auf ihr: sie gliedern nach dem Bilanzposten des
    # Exports. Eine Recon, die ueber dieselbe Zuordnung liefe, die sie pruefen
    # soll, pruefte nichts.
    from recon import schreibe_rechenwerke
    rechenwerke = schreibe_rechenwerke(wb, bs.konten, guv.konten, bs.perioden,
                                       ms_zeilen, projekt, waehrung, abschluss)

    os.makedirs(os.path.dirname(os.path.abspath(ziel)), exist_ok=True)
    wb.save(ziel)
    return {"pfad": ziel, "perioden": bs.perioden, "konten": bs.konten,
            "guv_konten": guv.konten, "ergebnisse": ergebnisse,
            "befunde": befunde, "leads": leadbefunde,
            "rechenwerke": rechenwerke, "ms_zeilen": ms_zeilen}


def main() -> int:
    ziel = sys.argv[1] if len(sys.argv) > 1 else "out/Referenzfall_Aufrisse.xlsx"
    ordner = sys.argv[2] if len(sys.argv) > 2 else "referenz"
    spec = sys.argv[3] if len(sys.argv) > 3 else "klassifizierung_v1.json"
    e = baue(ordner, spec, ziel)
    print(f"  Mastersheet: {e['ms_zeilen'] - 1} Konten, "
          f"{len(e['perioden'])} Perioden")
    for b in e["befunde"]:
        art = "LEER (Vermerk)" if b["leer"] else \
            f"{b['positionen']} Positionen, {len(b['konten'])} Konten"
        print(f"  {b['blatt']:16}{art}")
    for b in e["leads"]:
        art = "LEER (Vermerk)" if b["leer"] else \
            f"{len(b['bloecke'])} Blöcke, " \
            f"{sum(1 for x in b['bloecke'] if x['ohne_aufriss'])} ohne Aufriss"
        print(f"  {b['blatt']:16}{art}")
    r = e["rechenwerke"]
    for name in ("Bilanz", "GuV"):
        w = r["bilanz"] if name == "Bilanz" else r["guv"]
        print(f"  {name:16}{len(w['posten'])} Bilanzposten")
    for name, w in (("Recon Bilanz", r["recon_bilanz"]),
                    ("Recon GuV", r["recon_guv"])):
        print(f"  {name:16}{len(w['zeilen'])} Zeilen, "
              f"{len(w['nur_abschluss'])} nur im Abschluss, "
              f"{len(w['nur_susa'])} nur in der Saldenliste")
    print(f"  geschrieben: {e['pfad']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
