"""Punkt 4: die Lead-Tabs — die dritte Schicht.

Hier entscheidet sich, ob die Architektur dreischichtig ist oder nur so
aussieht. Die Regel:

* **Die Positionssumme im Lead zieht aus dem Aufriss**, nicht aus dem
  Mastersheet. ``='NA_OA'!F12`` — die Positionszeile des Aufrisses, nicht ein
  ``SUMIFS`` auf die Konten.
* **Nur die eingeklappten Kontozeilen ziehen weiter per ``SUMIFS`` aus dem
  Mastersheet.**

Das ist kein Formalismus. Zöge die Positionssumme ebenfalls aus dem
Mastersheet, gäbe es zwei Wege, die beide am selben Ort beginnen — der
Aufriss wäre Dekoration, und eine Position, die im Aufriss fehlt, fiele
niemandem auf. So dagegen laufen zwei **unabhängige** Wege auf dieselbe Zahl:
oben über den Aufriss, unten über die Konten. Genau das prüft die
Pflicht-Kontrollzeile je Block:

    Aufrisssumme  −  Summe der Kontozeilen  =  0

Ein Konto, das der Aufriss nicht führt, bricht sie. Eine Position, die der
Aufriss doppelt zeigt, ebenso.

**Blöcke ohne Aufriss.** Vier Klassen tragen keinen der sieben Aufrisse:
immaterielle Vermögensgegenstände, Nutzungsrechte, latente Steuern und
Eigenkapital. Sie werden trotzdem gezeigt — sonst ginge das Nettovermögen
nicht auf — aber gelb markiert, und ihre Kontrollzeile prüft gegen das
Mastersheet statt gegen einen Aufriss. Die Markierung ist die Aussage: hier
fehlt eine Schicht.

``Lead PL`` entsteht leer mit Vermerk. Die GuV-Datei trägt 237 Konten, aber
die Entscheidungsdatei kennt keine GuV-Kategorien und es gibt keinen
GuV-Aufriss — ein Lead PL, der trotzdem Zahlen zeigte, käme direkt aus dem
Mastersheet und bräche genau die Regel, um die es in diesem Punkt geht.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Callable, Optional

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from aufrisse import (BLAU, BREITEN, GELB, GRAU, GROESSE, GROESSE_KONTROLLE,
                      GRUEN, MS_ERSTE_PERIODE, MS_KATEGORIE, MS_KLASSE,
                      SCHWARZ, SP_DE, SP_EN, SP_ERSTE_PERIODE, SP_GRUPPE,
                      SP_TICKER, SP_TYP, TEAL, TEAL_DK, TINT1, TINT2, ZAHL,
                      Z_EINHEIT, Z_ERSTE, Z_KOPF, Z_PROJEKT, Z_TITEL,
                      _RAHMEN_OBEN, _RAHMEN_SUMME, _kontrollformel, _schrift,
                      _zahlzelle)


@dataclass
class Block:
    """Ein Block des Leads: eine Gruppe von Positionen mit einer Kontrolle."""

    titel_de: str
    titel_en: str
    #: Blattname des Aufrisses, aus dem die Positionssummen ziehen.
    aufriss: str = ""
    #: Nur fuer Bloecke OHNE Aufriss: welche Konten hierher gehoeren.
    auswahl: Optional[Callable[[dict], bool]] = None
    #: Nur fuer Bloecke ohne Aufriss: das Merkmal, gegen das geprueft wird.
    kontrolle: Optional[list[tuple[int, str]]] = None
    vermerk: str = ""


def bloecke() -> list[Block]:
    """Die Blöcke des Lead NA — zusammen genau einmal jedes Bilanzkonto.

    Die vier Aufrisse ``NA_Sachanlagen``, ``NA_OA``, ``NA_OL`` und
    ``NA_Net Debt`` decken sich nicht: FA/PPE mit kumulierter Abschreibung,
    das Working Capital der Aktiv- und der Passivseite, das Net Debt. Was
    keiner von ihnen führt, steht in den Blöcken ohne Aufriss.

    ``NA_TWC`` und ``NA_Vorräte`` sind Ausschnitte aus ``NA_OA``/``NA_OL``
    und dürfen deshalb kein eigener Block sein — sie erscheinen nachrichtlich
    unter der Summe.
    """
    return [
        Block("Sachanlagen", "Tangible assets", aufriss="NA_Sachanlagen"),
        Block("Übriges Anlagevermögen", "Other fixed assets",
              auswahl=lambda r: r["category"] in ("FA/Intangibles",
                                                  "FA/Right-of-use assets"),
              kontrolle=[(MS_KATEGORIE, "FA/Intangibles"),
                         (MS_KATEGORIE, "FA/Right-of-use assets")],
              vermerk="OHNE AUFRISS. Immaterielle Vermögensgegenstände und "
                      "Nutzungsrechte tragen keinen der sieben Aufriss-Tabs. "
                      "Die Positionssummen kommen deshalb aus den Kontozeilen "
                      "desselben Blattes und nicht aus einer eigenen Schicht; "
                      "die Kontrollzeile prüft gegen das Mastersheet. Ein "
                      "Aufriss NA_Immaterielle und ein Aufriss "
                      "NA_Nutzungsrechte würden das schließen."),
        Block("Operatives Vermögen", "Operating assets", aufriss="NA_OA"),
        Block("Operative Verbindlichkeiten", "Operating liabilities",
              aufriss="NA_OL"),
        Block("Net Debt", "Net debt", aufriss="NA_Net Debt"),
        Block("Latente Steuern", "Deferred tax",
              auswahl=lambda r: r["klasse"] == "DT",
              kontrolle=[(MS_KLASSE, "DT")],
              vermerk="OHNE AUFRISS. Latente Steuern tragen keinen der "
                      "sieben Aufriss-Tabs; die Kontrollzeile prüft gegen das "
                      "Mastersheet."),
    ]


#: Der Block unterhalb des Nettovermögens. Er zählt nicht in die Summe.
EIGENKAPITAL = Block(
    "Eigenkapital", "Equity",
    auswahl=lambda r: r["klasse"] == "EQ",
    kontrolle=[(MS_KLASSE, "EQ")],
    vermerk="OHNE AUFRISS. Das Eigenkapital steht unterhalb des "
            "Nettovermögens und zählt nicht in dessen Summe — es ist die "
            "Gegenprobe. Die Schlusskontrolle darunter muss null sein.")

#: Nachrichtlich unter der Summe: Ausschnitte, die schon in einem Block
#: stecken und deshalb nicht mitzählen dürfen.
NACHRICHTLICH = [("Trade Working Capital", "Trade working capital", "NA_TWC"),
                 ("davon Vorräte", "thereof inventories", "NA_Vorräte"),
                 ("Investitionen", "Capital expenditure", "NA_CAPEX")]


# --------------------------------------------------------------------------

def _kopf(ws, projekt: str, waehrung: str, titel_de: str, titel_en: str,
          perioden: list[str], spalten: list[int]) -> None:
    ws.cell(Z_PROJEKT, SP_DE, projekt).font = _schrift(GROESSE + 2, BLAU,
                                                       fett=True)
    ws.cell(Z_TITEL, SP_DE, titel_de).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_TITEL, SP_EN, titel_en).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(Z_EINHEIT, SP_DE, f"in T {waehrung}").font = _schrift(farbe=GRAU)
    ws.cell(Z_EINHEIT, SP_EN,
            "Positionen aus den Aufrissen, Konten aus dem Mastersheet"
            ).font = _schrift(farbe=GRAU)
    for spalte, titel in ((SP_TYP, "Typ"), (SP_TICKER, "Ticker"),
                          (SP_DE, "Bezeichnung"), (SP_EN, "Description"),
                          (SP_GRUPPE, "Herkunft")):
        z = ws.cell(Z_KOPF, spalte, titel)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
    for i, p in enumerate(perioden):
        z = ws.cell(Z_KOPF, spalten[i], p)
        z.font = _schrift(GROESSE, "FFFFFFFF", fett=True)
        z.fill = PatternFill("solid", fgColor=TEAL)
        z.alignment = Alignment(horizontal="right")


def _kontozeile(ws, zeile: int, konto, spalten: list[int],
                ms_zeilen: int) -> None:
    """Eine eingeklappte Kontozeile — der einzige Weg zum Mastersheet."""
    ws.cell(zeile, SP_TYP, "KTO").font = _schrift(farbe=BLAU)
    ws.cell(zeile, SP_TICKER, konto.konto).font = _schrift(farbe=BLAU)
    for spalte, wert in ((SP_DE, konto.bezeichnung), (SP_EN, konto.bezeichnung),
                         (SP_GRUPPE, "Mastersheet")):
        z = ws.cell(zeile, spalte, wert)
        z.font = _schrift(farbe=GRAU)
        z.alignment = Alignment(indent=2)
    for j, spalte in enumerate(spalten):
        ms = get_column_letter(MS_ERSTE_PERIODE + j)
        _zahlzelle(ws, zeile, spalte,
                   f"=SUMIFS(Mastersheet!${ms}$2:${ms}${ms_zeilen},"
                   f"Mastersheet!$A$2:$A${ms_zeilen},$B{zeile})/1000", GRUEN)
    ws.row_dimensions[zeile].outlineLevel = 2
    ws.row_dimensions[zeile].hidden = True


def _schreibe_block(ws, b: Block, zeile: int, spalten: list[int], konten,
                    ergebnisse, befunde: dict, ms_zeilen: int,
                    faerbung: int) -> tuple[int, dict]:
    """Ein Block: Positionen, Kontozeilen, Summe, Pflicht-Kontrollzeile."""
    erste = zeile
    if b.aufriss:
        quelle = befunde[b.aufriss]
        positionen = [(p["schluessel"], p["konten"], p["zeile"])
                      for p in quelle["positionszeilen"]]
    else:
        gruppen: dict[str, list[str]] = {}
        for no, r in ergebnisse.items():
            if b.auswahl and b.auswahl(r):
                gruppen.setdefault(r["category"], []).append(no)
        positionen = [(k, sorted(v, key=lambda x: abs(
            konten[x].salden[list(konten[x].salden)[-1]]), reverse=True), None)
            for k, v in sorted(gruppen.items())]

    # Blocktitel
    z = ws.cell(zeile, SP_DE, b.titel_de)
    z.font = _schrift(GROESSE + 1, TEAL_DK, fett=True)
    ws.cell(zeile, SP_EN, b.titel_en).font = _schrift(GROESSE + 1, TEAL_DK,
                                                      fett=True)
    ws.cell(zeile, SP_TYP, "TIT").font = _schrift(farbe=BLAU)
    zeile += 1

    for schluessel, nos, aufrisszeile in positionen:
        pos_zeile = zeile
        ws.cell(pos_zeile, SP_TYP, "POS").font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_TICKER, schluessel).font = _schrift(farbe=BLAU)
        ws.cell(pos_zeile, SP_DE, schluessel).font = _schrift(fett=True)
        ws.cell(pos_zeile, SP_EN, schluessel).font = _schrift(fett=True)
        ws.cell(pos_zeile, SP_GRUPPE,
                b.aufriss or "Kontozeilen (ohne Aufriss)").font = _schrift(
                    farbe=GRUEN if b.aufriss else SCHWARZ)
        fuellung = PatternFill("solid", fgColor=GELB if not b.aufriss
                               else (TINT1 if faerbung % 2 else TINT2))
        for spalte in range(SP_TYP, spalten[-1] + 1):
            ws.cell(pos_zeile, spalte).fill = fuellung

        zeile += 1
        erste_kto = zeile
        for no in nos:
            _kontozeile(ws, zeile, konten[no], spalten, ms_zeilen)
            zeile += 1
        letzte_kto = zeile - 1

        for j, spalte in enumerate(spalten):
            sp = get_column_letter(spalte)
            if aufrisszeile is not None:
                # DER KERN: die Positionssumme kommt aus dem Aufriss.
                _zahlzelle(ws, pos_zeile, spalte,
                           f"='{b.aufriss}'!{sp}{aufrisszeile}", GRUEN,
                           fett=True)
            else:
                # Ohne Aufriss bleibt nur der Weg ueber die eigenen
                # Kontozeilen. Gelb markiert, weil eine Schicht fehlt.
                _zahlzelle(ws, pos_zeile, spalte,
                           f"=SUM({sp}{erste_kto}:{sp}{letzte_kto})", SCHWARZ,
                           fett=True)
        faerbung += 1

    letzte = zeile - 1
    typ = get_column_letter(SP_TYP)

    summenzeile = zeile
    ws.cell(summenzeile, SP_TYP, "SUM").font = _schrift(farbe=BLAU)
    for spalte, wert in ((SP_DE, b.titel_de), (SP_EN, b.titel_en)):
        z = ws.cell(summenzeile, spalte, wert)
        z.font = _schrift(GROESSE, SCHWARZ, fett=True)
        z.border = _RAHMEN_SUMME
    for spalte in spalten:
        sp = get_column_letter(spalte)
        _zahlzelle(ws, summenzeile, spalte,
                   f'=SUMIF(${typ}${erste}:${typ}${letzte},"<>KTO",'
                   f'{sp}{erste}:{sp}{letzte})', SCHWARZ, fett=True)
        ws.cell(summenzeile, spalte).border = _RAHMEN_SUMME
    zeile += 1

    # Pflicht-Kontrollzeile: Aufrisssumme gegen Summe der Kontozeilen.
    kontrollzeile = zeile
    ws.cell(kontrollzeile, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
    if b.aufriss:
        text_de = (f"Kontrolle: Summe {b.aufriss} ./. Summe der Kontozeilen "
                   "(muss null sein)")
        text_en = (f"Control: total per {b.aufriss} less sum of account rows "
                   "(must be nil)")
    else:
        text_de = ("Kontrolle ohne Aufriss: Summe ./. Mastersheet "
                   "(muss null sein)")
        text_en = "Control without schedule: total less mastersheet (must be nil)"
    for spalte, wert in ((SP_DE, text_de), (SP_EN, text_en)):
        z = ws.cell(kontrollzeile, spalte, wert)
        z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
        z.border = _RAHMEN_OBEN
        if not b.aufriss:
            z.fill = PatternFill("solid", fgColor=GELB)
    for j, spalte in enumerate(spalten):
        sp = get_column_letter(spalte)
        if b.aufriss:
            gegen = f"'{b.aufriss}'!{sp}{befunde[b.aufriss]['summenzeile']}"
        else:
            gegen = _kontrollformel(b.kontrolle or [], ms_zeilen,
                                    MS_ERSTE_PERIODE + j)
        _zahlzelle(ws, kontrollzeile, spalte,
                   f'={gegen}-SUMIF(${typ}${erste}:${typ}${letzte},"KTO",'
                   f'{sp}{erste}:{sp}{letzte})', SCHWARZ, ZAHL, fett=True,
                   groesse=GROESSE_KONTROLLE)
        ws.cell(kontrollzeile, spalte).border = _RAHMEN_OBEN
        if not b.aufriss:
            ws.cell(kontrollzeile, spalte).fill = PatternFill("solid",
                                                              fgColor=GELB)
    zeile += 1

    if b.vermerk:
        z = ws.cell(zeile, SP_DE, "Vermerk: " + b.vermerk)
        z.font = _schrift(farbe=GRAU)
        z.alignment = Alignment(wrap_text=True, vertical="top")
        z.fill = PatternFill("solid", fgColor=GELB)
        ws.merge_cells(start_row=zeile, start_column=SP_DE, end_row=zeile,
                       end_column=spalten[-1])
        ws.row_dimensions[zeile].height = 40
        zeile += 1

    return zeile + 1, {"titel": b.titel_de, "aufriss": b.aufriss,
                       "summenzeile": summenzeile,
                       "kontrollzeile": kontrollzeile,
                       "konten": [no for _, nos, _ in positionen for no in nos],
                       "ohne_aufriss": not b.aufriss, "faerbung": faerbung}


def schreibe_lead_na(wb, befunde: dict, konten, perioden, ergebnisse,
                     ms_zeilen: int, projekt: str, waehrung: str) -> dict:
    """Lead NA — die Überleitung zum Nettovermögen."""
    ws = wb.create_sheet("Lead NA")
    spalten = [SP_ERSTE_PERIODE + i for i in range(len(perioden))]
    _kopf(ws, projekt, waehrung, "Nettovermögen", "Net assets", perioden,
          spalten)

    zeile, faerbung = Z_ERSTE, 0
    blockbefunde = []
    for b in bloecke():
        zeile, befund = _schreibe_block(ws, b, zeile, spalten, konten,
                                        ergebnisse, befunde, ms_zeilen,
                                        faerbung)
        faerbung = befund["faerbung"]
        blockbefunde.append(befund)

    # Nettovermoegen: die Summenzeilen der Bloecke, nicht der Bereich.
    netto_zeile = zeile
    ws.cell(netto_zeile, SP_TYP, "NET").font = _schrift(farbe=BLAU)
    for spalte, wert in ((SP_DE, "Nettovermögen"), (SP_EN, "Net assets")):
        z = ws.cell(netto_zeile, spalte, wert)
        z.font = _schrift(GROESSE + 1, SCHWARZ, fett=True)
        z.border = _RAHMEN_SUMME
    for spalte in spalten:
        sp = get_column_letter(spalte)
        _zahlzelle(ws, netto_zeile, spalte,
                   "=" + "+".join(f"{sp}{b['summenzeile']}"
                                  for b in blockbefunde), SCHWARZ, fett=True)
        ws.cell(netto_zeile, spalte).border = _RAHMEN_SUMME
    zeile += 2

    zeile, ek = _schreibe_block(ws, EIGENKAPITAL, zeile, spalten, konten,
                                ergebnisse, befunde, ms_zeilen, faerbung)
    blockbefunde.append(ek)

    # Schlusskontrolle: Nettovermoegen und Eigenkapital spiegeln sich.
    schluss = zeile
    ws.cell(schluss, SP_TYP, "CHK").font = _schrift(farbe=BLAU)
    for spalte, wert in ((SP_DE, "Schlusskontrolle: Nettovermögen + "
                                 "Eigenkapital (muss null sein)"),
                         (SP_EN, "Final control: net assets plus equity "
                                 "(must be nil)")):
        z = ws.cell(schluss, spalte, wert)
        z.font = _schrift(GROESSE_KONTROLLE, SCHWARZ, fett=True)
        z.border = _RAHMEN_OBEN
    for spalte in spalten:
        sp = get_column_letter(spalte)
        _zahlzelle(ws, schluss, spalte,
                   f"={sp}{netto_zeile}+{sp}{ek['summenzeile']}", SCHWARZ,
                   ZAHL, fett=True, groesse=GROESSE_KONTROLLE)
        ws.cell(schluss, spalte).border = _RAHMEN_OBEN
    zeile += 2

    # Nachrichtlich: Ausschnitte, die schon in einem Block stecken.
    ws.cell(zeile, SP_DE, "Nachrichtlich (nicht in der Summe)").font = \
        _schrift(GROESSE + 1, TEAL_DK, fett=True)
    ws.cell(zeile, SP_EN, "Memorandum (not included above)").font = \
        _schrift(GROESSE + 1, TEAL_DK, fett=True)
    zeile += 1
    for de, en, blatt in NACHRICHTLICH:
        ws.cell(zeile, SP_TYP, "MEM").font = _schrift(farbe=BLAU)
        ws.cell(zeile, SP_DE, de).font = _schrift(farbe=GRAU)
        ws.cell(zeile, SP_EN, en).font = _schrift(farbe=GRAU)
        ws.cell(zeile, SP_GRUPPE, blatt).font = _schrift(farbe=GRUEN)
        quelle = befunde[blatt]
        for spalte in spalten:
            sp = get_column_letter(spalte)
            if quelle["leer"]:
                z = ws.cell(zeile, spalte, "kein Aufriss")
                z.font = _schrift(farbe=GRAU)
                z.alignment = Alignment(horizontal="right")
                z.fill = PatternFill("solid", fgColor=GELB)
            else:
                _zahlzelle(ws, zeile, spalte,
                           f"='{blatt}'!{sp}{quelle['summenzeile']}", GRUEN)
        zeile += 1

    _spaltenbreiten(ws, spalten)
    return {"blatt": "Lead NA", "bloecke": blockbefunde,
            "nettozeile": netto_zeile, "schlusskontrolle": schluss,
            "leer": False}


def schreibe_lead_pl(wb, perioden, projekt: str, waehrung: str) -> dict:
    """Lead PL — angelegt, leer, mit Vermerk.

    Die GuV-Datei traegt 237 Konten. Die Entscheidungsdatei kennt aber keine
    GuV-Kategorien, und es gibt keinen GuV-Aufriss. Ein Lead PL mit Zahlen
    muesste sie direkt aus dem Mastersheet holen — und genau das ist die
    Regel, um die es in diesem Punkt geht.
    """
    ws = wb.create_sheet("Lead PL")
    spalten = [SP_ERSTE_PERIODE + i for i in range(len(perioden))]
    _kopf(ws, projekt, waehrung, "Ergebnisrechnung", "Profit and loss",
          perioden, spalten)
    z = ws.cell(Z_ERSTE, SP_DE, "Kein Lead möglich — siehe Vermerk")
    z.font = _schrift(fett=True)
    for spalte in range(SP_TYP, spalten[-1] + 1):
        ws.cell(Z_ERSTE, spalte).fill = PatternFill("solid", fgColor=GELB)
    vermerkzeile = Z_ERSTE + 2
    z = ws.cell(vermerkzeile, SP_DE,
                "Vermerk: KEINE SCHICHT DARUNTER. Die GuV-Datei trägt 237 "
                "Konten, aber die Entscheidungsdatei kennt keine "
                "GuV-Kategorien (ihre Klassen sind FA, TWC, OWC, ND, DT, EQ "
                "und TECH), und es gibt keinen GuV-Aufriss. Ein Lead PL mit "
                "Zahlen müsste sie unmittelbar aus dem Mastersheet holen und "
                "bräche damit die Regel dieser Schicht. Benötigt werden "
                "GuV-Kategorien in klassifizierung_v1.json und darauf "
                "aufbauend ein Aufriss je Ergebnisposition.")
    z.font = _schrift(farbe=GRAU)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    z.fill = PatternFill("solid", fgColor=GELB)
    ws.merge_cells(start_row=vermerkzeile, start_column=SP_DE,
                   end_row=vermerkzeile, end_column=spalten[-1])
    ws.row_dimensions[vermerkzeile].height = 60
    _spaltenbreiten(ws, spalten)
    return {"blatt": "Lead PL", "bloecke": [], "leer": True,
            "vermerkzeile": vermerkzeile}


def _spaltenbreiten(ws, spalten: list[int]) -> None:
    for spalte, breite in BREITEN.items():
        ws.column_dimensions[spalte].width = breite
    for spalte in spalten:
        ws.column_dimensions[get_column_letter(spalte)].width = 15
    ws.column_dimensions["A"].hidden = True
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = ws.cell(Z_ERSTE, SP_ERSTE_PERIODE)


def schreibe_leads(wb, befunde: dict, konten, perioden, ergebnisse,
                   ms_zeilen: int, projekt: str, waehrung: str) -> list[dict]:
    return [schreibe_lead_na(wb, befunde, konten, perioden, ergebnisse,
                             ms_zeilen, projekt, waehrung),
            schreibe_lead_pl(wb, perioden, projekt, waehrung)]
