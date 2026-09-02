"""Liest alle Kontrollzeilen der erzeugten Arbeitsmappe — mit ihren Werten.

openpyxl rechnet keine Formeln. Eine Kontrollzeile, die im Code richtig
aussieht, kann in der Datei trotzdem etwas anderes rechnen: ein falscher
Bereich, ein verrutschter Verweis, ein Blattname mit Leerzeichen ohne
Anführungszeichen. Deshalb wird die fertige Datei durchgerechnet und
ausgelesen, statt die Absicht zu glauben.

Gerechnet wird mit dem Paket ``formulas``. Der LibreOffice-Recalc des
Hausformats bleibt für die Auslieferung maßgeblich; er läuft in dieser
Umgebung in den Timeout.

Aufruf::

    python3 kontrollen.py [out/Referenzfall_Aufrisse.xlsx]
"""

from __future__ import annotations

import sys

import openpyxl
from openpyxl.utils import get_column_letter

#: Zeilentyp der Kontrollzeilen.
TYP = "CHK"

#: Toleranz in T — vier Nachkommastellen sind ein Zehntel Cent.
TOLERANZ = 0.0005

MAPPE = "out/Referenzfall_Aufrisse.xlsx"


def gerechnete_werte(pfad: str) -> dict[str, float]:
    """Alle gerechneten Zahlenwerte, nach ``BLATT!A1`` (Blatt in Grossbuchstaben)."""
    import formulas

    loesung = formulas.ExcelModel().loads(pfad).finish().calculate()
    werte: dict[str, float] = {}
    for schluessel, zelle in loesung.items():
        if "]" not in schluessel or "!" not in schluessel:
            continue
        blatt = schluessel.split("]", 1)[1].split("'!")[0]
        try:
            werte[f"{blatt}!{schluessel.split('!', 1)[1]}"] = float(
                zelle.value[0, 0])
        except Exception:
            continue
    return werte


def lies_kontrollzeilen(pfad: str = MAPPE, werte: dict | None = None
                        ) -> list[dict]:
    """Jede Kontrollzeile der Mappe mit ihren gerechneten Werten."""
    if werte is None:
        werte = gerechnete_werte(pfad)
    wb = openpyxl.load_workbook(pfad)
    # Die Periodenspalten stehen ab Spalte F; die Kopfzeile nennt sie.
    zeilen: list[dict] = []
    for blatt in wb.sheetnames:
        ws = wb[blatt]
        perioden, spalten = [], []
        for c in range(6, ws.max_column + 1):
            kopf = ws.cell(5, c).value
            if kopf:
                perioden.append(str(kopf))
                spalten.append(get_column_letter(c))
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value != TYP:
                continue
            w = [werte.get(f"{blatt.upper()}!{c}{r}") for c in spalten]
            # Leere Zellen sind nicht "nicht null", sondern nicht belegt: die
            # Recon-Tabs fuehren drei Bloecke nebeneinander, und eine
            # Kontrollzeile gehoert immer nur zu einem davon. Geprueft wird,
            # was dasteht — aber es muss etwas dastehen.
            belegt = [x for x in w if x is not None]
            zeilen.append({
                "blatt": blatt, "zeile": r,
                "text": str(ws.cell(r, 3).value or ""),
                "perioden": [p for p, x in zip(perioden, w) if x is not None],
                "werte": belegt,
                "null": bool(belegt) and all(abs(x) <= TOLERANZ
                                             for x in belegt)})
    return zeilen


def main() -> int:
    pfad = sys.argv[1] if len(sys.argv) > 1 else MAPPE
    zeilen = lies_kontrollzeilen(pfad)
    if not zeilen:
        print("Keine Kontrollzeilen gefunden.")
        return 1
    perioden = zeilen[0]["perioden"]
    print("=" * 78)
    print(f"  KONTROLLZEILEN — {pfad}")
    print("=" * 78)
    print(f"  {'Blatt':16}{'Zeile':6}{'Kontrolle':58}"
          + "".join(f"{p:>12}" for p in perioden))
    print("  " + "-" * 134)
    for z in zeilen:
        print(f"  {z['blatt']:16}{z['zeile']:<6}{z['text'][:56]:58}"
              + "".join(f"{w:>12,.4f}" if w is not None else f"{'?':>12}"
                        for w in z["werte"])
              + ("" if z["null"] else "   NICHT NULL"))
    print("  " + "-" * 134)
    gut = sum(1 for z in zeilen if z["null"])
    print(f"  {gut} von {len(zeilen)} Kontrollzeilen stehen auf null")
    print("=" * 78)
    return 0 if gut == len(zeilen) else 1


if __name__ == "__main__":
    sys.exit(main())
