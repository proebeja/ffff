"""Reconciliation der gemappten Databook-Positionen gegen den Jahresabschluss.

Anders als ``reconciliation.py`` (SuSa gegen Kontennachweis, auf Kontoebene)
vergleicht dieses Modul **Positionssummen** gegen eine externe Bilanz/GuV.
Der Jahresabschluss ist hier ausdrücklich **keine Strukturquelle** — er ändert
kein Mapping, sondern ist nur Ziel der Abstimmung.

Die Zuordnung der JA-Zeilen auf kanonische HGB-Pfade nutzt für die Bilanz den
Überschriften-Crosswalk der Kontennachweis-Schicht wieder (dieselben §266-
Bezeichnungen, dieselbe Normalisierung); für die GuV steht der Crosswalk
unten, inklusive der Ertrag/Aufwand-Richtung: Der Abschluss weist Erträge und
Aufwendungen positiv aus, das Databook führt Erträge vorzeichenrichtig negativ.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from ..core.model import MappedAccount
from ..core.hausconvention import normalisiere
from ..readers.kontennachweis import _match_ueberschrift

# GuV-Position im Abschluss -> (kanonischer Pfad, Richtung).
# Richtung "E" = Ertrag (im Abschluss positiv, im Databook negativ).
_G = "/GuV"
_GUV_CROSSWALK: list[tuple[str, str, str]] = [
    ("umsatzerloese", f"{_G}/Umsatzerloese", "E"),
    ("erhoehung des bestandes", f"{_G}/Bestandsveraenderung", "E"),
    ("verminderung des bestandes", f"{_G}/Bestandsveraenderung", "E"),
    ("andere aktivierte eigenleistungen", f"{_G}/Andere aktivierte Eigenleistungen", "E"),
    ("sonstige betriebliche ertraege", f"{_G}/Sonstige betriebliche Ertraege", "E"),
    ("materialaufwand", f"{_G}/Materialaufwand", "A"),
    ("loehne und gehaelter", f"{_G}/Personalaufwand/Loehne und Gehaelter", "A"),
    ("soziale abgaben", f"{_G}/Personalaufwand/Soziale Abgaben und Altersversorgung", "A"),
    ("abschreibungen", f"{_G}/Abschreibungen", "A"),
    ("sonstige betriebliche aufwendungen", f"{_G}/Sonstige betriebliche Aufwendungen", "A"),
    ("ertraege aus beteiligungen", f"{_G}/Ertraege aus Beteiligungen", "E"),
    ("sonstige zinsen und aehnliche ertraege", f"{_G}/Sonstige Zinsen und aehnliche Ertraege", "E"),
    ("zinsen und aehnliche aufwendungen", f"{_G}/Zinsen und aehnliche Aufwendungen", "A"),
    ("steuern vom einkommen und vom ertrag", f"{_G}/Steuern vom Einkommen und vom Ertrag", "A"),
    ("sonstige steuern", f"{_G}/Sonstige Steuern", "A"),
    ("ertraege aus verlustuebernahme", f"{_G}/Ertraege aus Verlustuebernahme", "E"),
]

#: Zeilen, die im Abschluss Zwischensummen oder Kennzahlen sind und daher
#: nicht als Position abgestimmt werden.
_KEINE_POSITION = (
    "% wachstum", "% growth", "gesamtleistung", "total output", "summe",
    "ergebnis", "jahresueberschuss", "jahresfehlbetrag", "ebit", "ebitda",
    "aktiva", "passiva", "bilanzsumme", "total",
)


@dataclass
class JAZeile:
    label: str
    hgb_pfad: str
    quelle_tab: str
    ja: dict[str, float] = field(default_factory=dict)
    databook: dict[str, float] = field(default_factory=dict)

    def differenz(self, p: str) -> float:
        return self.databook.get(p, 0.0) - self.ja.get(p, 0.0)


@dataclass
class JAReconciliation:
    zeilen: list[JAZeile]
    perioden: list[str]
    quelle: str
    nur_im_ja: list[str] = field(default_factory=list)
    nur_im_databook: list[str] = field(default_factory=list)
    hinweise: list[str] = field(default_factory=list)

    @property
    def zeilen_mit_differenz(self) -> list[JAZeile]:
        return [z for z in self.zeilen
                if any(abs(z.differenz(p)) > 0.005 for p in self.perioden)]

    def gesamtdifferenz(self, p: str) -> float:
        return sum(z.differenz(p) for z in self.zeilen)


def _jahr(text: str) -> Optional[str]:
    m = re.search(r"(20\d{2})", str(text))
    return m.group(1) if m else None


def _zahl(v) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(".", "").replace(",", ".")) if "," in v else float(v)
        except ValueError:
            return None
    return None


def _ist_position(label: str) -> bool:
    low = normalisiere(label)
    return bool(low) and not any(low.startswith(k) or low == k for k in _KEINE_POSITION)


def _guv_pfad(label: str) -> tuple[Optional[str], str]:
    low = normalisiere(label)
    # führende Nummerierung ("1. ", "4. ", "a) ") abschneiden
    low = re.sub(r"^[0-9]+\.?\s*", "", low)
    low = re.sub(r"^[a-z]\)\s*", "", low)
    treffer = [(m, p, r) for m, p, r in _GUV_CROSSWALK if low.startswith(m)]
    if not treffer:
        return None, "A"
    m, p, r = max(treffer, key=lambda t: len(t[0]))
    return p, r


def _lies_ja(pfad: str) -> tuple[dict[tuple[str, str], dict[str, float]], list[str]]:
    """Liest Bilanz- und GuV-Tab. Gibt {(hgb_pfad, label): {jahr: wert}} in
    Databook-Vorzeichen sowie die nicht zuordenbaren Labels zurück."""
    wb = openpyxl.load_workbook(pfad, data_only=True)
    positionen: dict[tuple[str, str], dict[str, float]] = {}
    unbekannt: list[str] = []

    for tab in wb.sheetnames:
        low_tab = normalisiere(tab)
        ist_bilanz = "balance sheet" in low_tab or "bilanz" in low_tab
        ist_guv = "p&l" in tab.lower() or "guv" in low_tab or "profit" in low_tab
        if not (ist_bilanz or ist_guv):
            continue
        ws = wb[tab]
        jahr_spalten = {c: _jahr(ws.cell(1, c).value)
                        for c in range(1, ws.max_column + 1)
                        if _jahr(ws.cell(1, c).value)}
        if not jahr_spalten:
            continue
        sektion = "AKTIVA"
        for r in range(2, ws.max_row + 1):
            label = str(ws.cell(r, 2).value or "").strip()
            if not label:
                continue
            low = normalisiere(label)
            if ist_bilanz and low in ("aktiva", "passiva"):
                sektion = low.upper()
                continue
            if not _ist_position(label):
                continue
            if ist_bilanz:
                pfad_c = _match_ueberschrift(label, sektion)
                vorzeichen = -1.0 if sektion == "PASSIVA" else 1.0
            else:
                pfad_c, richtung = _guv_pfad(label)
                vorzeichen = -1.0 if richtung == "E" else 1.0
            if pfad_c is None:
                unbekannt.append(f"[{tab}] {label[:60]}")
                continue
            ziel = positionen.setdefault((pfad_c, label.strip()), {})
            for c, j in jahr_spalten.items():
                w = _zahl(ws.cell(r, c).value)
                if w is not None:
                    ziel[j] = ziel.get(j, 0.0) + w * vorzeichen
    wb.close()
    return positionen, unbekannt


def reconcile_gegen_ja(mapped: list[MappedAccount], ja_pfad: str,
                       perioden: list[str]) -> JAReconciliation:
    ja_positionen, unbekannt = _lies_ja(ja_pfad)

    # Databook-Positionssummen je HGB-Pfad
    db: dict[str, dict[str, float]] = {}
    for m in mapped:
        if m.hgb_pfad.startswith("("):
            continue
        ziel = db.setdefault(m.hgb_pfad, {})
        for p in perioden:
            ziel[p] = ziel.get(p, 0.0) + m.saldo(p)

    # Perioden, die beide Seiten führen (Jahr als gemeinsamer Nenner)
    ja_jahre = {j for w in ja_positionen.values() for j in w}
    gemeinsam = [p for p in perioden if (_jahr(p) or "") in ja_jahre]

    zeilen: list[JAZeile] = []
    getroffen: set[str] = set()
    for (pfad_c, label), werte in sorted(ja_positionen.items()):
        z = JAZeile(label=label, hgb_pfad=pfad_c, quelle_tab="JA")
        for p in gemeinsam:
            j = _jahr(p) or ""
            z.ja[p] = werte.get(j, 0.0)
            z.databook[p] = db.get(pfad_c, {}).get(p, 0.0)
        zeilen.append(z)
        getroffen.add(pfad_c)

    nur_db = sorted(p for p in db if p not in getroffen and not p.startswith("/GuV/Bestand"))
    hinweise: list[str] = []
    if not gemeinsam:
        hinweise.append(
            "Keine gemeinsame Periode: der Jahresabschluss deckt "
            f"{sorted(ja_jahre)} ab, das Databook {perioden}.")
    return JAReconciliation(
        zeilen=zeilen, perioden=gemeinsam, quelle=ja_pfad,
        nur_im_ja=unbekannt, nur_im_databook=nur_db, hinweise=hinweise)
