"""Reader für das Namur-Databook (Excel mit vorhandener Kategorisierung).

Namur ist ein fertiges Databook, kein roher Export. Sein Blatt
"Trial Balance (Susa)" enthält jedoch eine saubere SuSa mit Konto,
Bezeichnung und Jahressalden — plus eine bestehende, analytische
"Categorization" (eigene Taxonomie: /net debt/…, /working capital/…,
/fixed assets/…). Für die Pipeline lesen wir Konto/Bezeichnung/Salden; die
bestehende Kategorisierung dient in den Tests als Regressions-Benchmark
(siehe ``kategorie_map``), nicht als HGB-Struktur — sie ist analytisch, kein
§266-Pfad. Daher ``hat_kontennachweis = False`` (die Engine mappt über
Typ-1 -> SKR-Default).
"""

from __future__ import annotations

import re

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint, parse_deutsche_zahl

_SHEET = "Trial Balance (Susa)"


class NamurDatabookReader(Reader):
    name = "namur_databook"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            return _SHEET in wb.sheetnames
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        wb = openpyxl.load_workbook(pfad, data_only=True)
        ws = wb[_SHEET]
        hdr, cols = self._spalten(ws)
        perioden = [name for name, _ in cols["perioden"]]

        accounts: list[Account] = []
        for r in range(hdr + 1, ws.max_row + 1):
            konto = ws.cell(r, cols["konto"]).value
            if konto is None:
                continue
            konto = str(konto).strip()
            if not konto or konto.lower() == "none":
                continue
            bez = str(ws.cell(r, cols["bez"]).value or "").strip()
            salden = tuple(
                PeriodBalance(name, parse_deutsche_zahl(ws.cell(r, c).value))
                for name, c in cols["perioden"]
            )
            accounts.append(Account(konto=konto, bezeichnung=bez, salden=salden,
                                    entity="Projekt Namur", fs_pfad=None, kontotyp=None))
        wb.close()
        return NormalizedLedger(
            accounts=accounts, perioden=perioden, entity="Projekt Namur",
            quelle_datei=pfad, hat_kontennachweis=False,
            fingerprint=fingerprint(pfad), warnungen=[],
        )

    @staticmethod
    def _spalten(ws) -> tuple[int, dict]:
        """Kopfzeile finden und die Jahres-Saldospalten (FY…/YTD…) lokalisieren."""
        for r in range(1, 6):
            labels = {c: str(ws.cell(r, c).value).strip()
                      for c in range(1, min(ws.max_column, 60) + 1)
                      if ws.cell(r, c).value is not None}
            if any(v == "Categorization" for v in labels.values()):
                # Nur echte Berichtsperioden (FY<Jahr> / YTD <Jahr>) — keine
                # abgeleiteten Spalten wie 'FY23-FY22' oder 'FY Growth %'.
                per_re = re.compile(r"^(FY\s?\d{4}|YTD\s?\d{4})$")
                perioden = [(v, c) for c, v in labels.items() if per_re.match(v)]
                perioden.sort(key=lambda t: t[1])
                return r, {"konto": 1, "bez": 2, "kat": 4, "perioden": perioden}
        return 1, {"konto": 1, "bez": 2, "kat": 4, "perioden": [("FY", 7)]}


def kategorie_map(pfad: str) -> dict[str, str]:
    """Konto -> bestehende Namur-Kategorisierung (für Regressionstests)."""
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb[_SHEET]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        konto = ws.cell(r, 1).value
        kat = ws.cell(r, 4).value
        if konto is not None and kat:
            out[str(konto).strip()] = str(kat).strip()
    wb.close()
    return out
