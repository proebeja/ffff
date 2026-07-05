"""Reader für rohe DATEV-SuSa (Saldenliste Sachkonten), z.B. Eckart.

Fallstrick (verifiziert): Spaltenversatz. Die Kopfzeile beschriftet "Saldo" in
Spalte T, der Zahlenwert steht aber eine Spalte links davon (S), und das
S/H-Vorzeichen mehrere Spalten rechts (W). Der Reader lokalisiert deshalb die
"Saldo"-Kopfspalte dynamisch und liest Wert und Vorzeichen versatztolerant.

Kein Kontennachweis eingebettet -> ``hat_kontennachweis = False``; die Engine
fällt auf Typ-1 -> SKR-Default -> Review zurück.
"""

from __future__ import annotations

import re
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint, parse_deutsche_zahl

_KONTO_RE = re.compile(r"^\d{3,6}$")


class DatevSusaReader(Reader):
    name = "datev_susa"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            for ws in wb.worksheets:
                if str(ws.title).startswith("SaldenlisteSachkonten"):
                    return True
                for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                    if any(str(c).strip() == "Kontonummer" for c in row if c):
                        return True
            return False
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        wb = openpyxl.load_workbook(pfad, data_only=True)
        # konto -> {periode: betrag}, plus Bezeichnung
        konten: dict[str, dict[str, float]] = {}
        bezeichnungen: dict[str, str] = {}
        perioden: list[str] = []
        warnungen: list[str] = []
        entity = "Unbekannt"

        for ws in wb.worksheets:
            if not str(ws.title).startswith("SaldenlisteSachkonten"):
                continue
            periode, ent = self._meta(ws)
            if ent:
                entity = ent
            if periode in perioden:
                periode = f"{periode}#{ws.title}"
            perioden.append(periode)

            hdr_row, spalten = self._finde_spalten(ws)
            saldo_col = spalten["saldo"]
            if saldo_col is None:
                warnungen.append(f"Sheet '{ws.title}': keine 'Saldo'-Spalte gefunden.")
                continue
            konto_col, bez_col = spalten["konto"], spalten["bez"]

            for r in range(hdr_row + 1, ws.max_row + 1):
                konto = ws.cell(r, konto_col).value
                if konto is None:
                    continue
                konto = str(konto).strip()
                if not _KONTO_RE.match(konto):
                    continue
                bez = ws.cell(r, bez_col).value or ""
                betrag = self._lies_saldo(ws, r, saldo_col)
                konten.setdefault(konto, {})[periode] = betrag
                bezeichnungen.setdefault(konto, str(bez).strip())

        wb.close()

        accounts: list[Account] = []
        for konto in sorted(konten):
            salden = tuple(
                PeriodBalance(p, konten[konto].get(p, 0.0)) for p in perioden
            )
            accounts.append(Account(
                konto=konto, bezeichnung=bezeichnungen.get(konto, ""),
                salden=salden, entity=entity, fs_pfad=None, kontotyp=None,
            ))

        return NormalizedLedger(
            accounts=accounts, perioden=perioden, entity=entity,
            quelle_datei=pfad, hat_kontennachweis=False,
            fingerprint=fingerprint(pfad), warnungen=warnungen,
        )

    # ---- Hilfen -----------------------------------------------------------
    @staticmethod
    def _meta(ws) -> tuple[str, Optional[str]]:
        """Periode und Entity aus dem Blattkopf (erste ~8 Zeilen)."""
        periode = ws.title
        entity = None
        for r in range(1, 9):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                m = re.search(r"\b(20\d{2}/[0-9A-Za-z]{1,3})\b", s)
                if m:
                    periode = m.group(1)
                if s.startswith("Firma:"):
                    entity = s.split("Firma:", 1)[1].strip()
        return periode, entity

    @staticmethod
    def _finde_spalten(ws) -> tuple[int, dict]:
        """Findet die Kopfzeile ('Kontonummer') und die Spalten von Kontonummer,
        Bezeichnung und Saldo — aus den Kopf-Beschriftungen, nicht hartkodiert,
        damit ein evtl. Links-Offset nicht zu leeren Konten führt."""
        for r in range(1, min(15, ws.max_row) + 1):
            row_vals = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            labels = {c: str(v).strip() for c, v in row_vals.items() if v}
            if "Kontonummer" in labels.values():
                def spalte(name: str, default):
                    for c, v in labels.items():
                        if v == name:
                            return c
                    return default
                return r, {
                    "konto": spalte("Kontonummer", 1),
                    "bez": spalte("Bezeichnung", 3),
                    "saldo": spalte("Saldo", None),
                }
        return 7, {"konto": 1, "bez": 3, "saldo": None}

    @staticmethod
    def _lies_saldo(ws, r: int, saldo_col: int) -> float:
        """Versatztolerant: Zahlenwert im Fenster [saldo_col-3 .. saldo_col]
        (nächste Zahl links der Kopfspalte), Vorzeichen aus rechtsstehendem
        'S'/'H' in derselben Zeile."""
        wert = None
        for c in range(saldo_col, max(0, saldo_col - 4), -1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                wert = float(v)
                break
        if wert is None:
            return 0.0
        # Vorzeichen: rechtsstehendstes S/H (Soll positiv, Haben negativ)
        sign = "S"
        for c in range(ws.max_column, saldo_col - 1, -1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() in ("S", "H"):
                sign = v.strip()
                break
        return wert if sign == "S" else -abs(wert)
