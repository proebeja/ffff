"""Kontennachweis-Reader — die vorrangige Strukturquelle (Architektur v1.3).

Ein Kontennachweis ordnet Einzelkonten unter die §266-Positionen des
testierten Abschlusses. Genau das macht ihn zur maßgeblichen Quelle für den
HGB-Pfad: die Grundgliederung des Databooks entspricht dann der geprüften
Bilanz und lässt sich lückenlos darauf überleiten (Abschlusstreue).

Der Reader liefert drei Dinge:
  1. ``zuordnung``  Konto -> HGB-Pfad (Strukturquelle für Kaskadenstufe 1)
  2. ``konten``     Konto -> Bezeichnung + Salden je Periode (Konten, die in
                    der SuSa fehlen, kommen so trotzdem ins Databook)
  3. ``positionen`` HGB-Pfad -> Positionssumme je Periode (Ziel der
                    Reconciliation gegen die SuSa)

Unterstützt PDF (DATEV-Druckbild) und Excel. Die Überschriften-Erkennung ist
gekapselt; die Engine sieht nur fertige HGB-Pfade.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

from ..core.hausconvention import normalisiere
from .base import fingerprint, parse_deutsche_zahl

# ---- Überschrift -> kanonischer HGB-Pfad ---------------------------------
_AKT = "/Aktiva"
_AV = f"{_AKT}/A Anlagevermoegen"
_UV = f"{_AKT}/B Umlaufvermoegen"
_FORD = f"{_UV}/II Forderungen und sonstige Vermoegensgegenstaende"
_PAS = "/Passiva"

#: (normalisierte Überschrift, Sektion, HGB-Pfad). Sektion schränkt ein, wo
#: dieselbe Wortmarke auf beiden Bilanzseiten vorkommt ("geleistete
#: Anzahlungen", "sonstige ..."). None = beide Seiten.
_CROSSWALK: list[tuple[str, Optional[str], str]] = [
    # --- Anlagevermögen
    ("entgeltlich erworbene konzessionen", "AKTIVA", f"{_AV}/I Immaterielle Vermoegensgegenstaende/Entgeltlich erworbene Konzessionen"),
    ("selbst geschaffene schutzrechte", "AKTIVA", f"{_AV}/I Immaterielle Vermoegensgegenstaende/Selbst geschaffene Schutzrechte"),
    ("geschaefts- oder firmenwert", "AKTIVA", f"{_AV}/I Immaterielle Vermoegensgegenstaende/Geschaefts- oder Firmenwert"),
    ("grundstuecke", "AKTIVA", f"{_AV}/II Sachanlagen/Grundstuecke und Gebaeude"),
    ("technische anlagen und maschinen", "AKTIVA", f"{_AV}/II Sachanlagen/Technische Anlagen und Maschinen"),
    ("andere anlagen, betriebs- und geschaeftsausstattung", "AKTIVA", f"{_AV}/II Sachanlagen/Andere Anlagen Betriebs- und Geschaeftsausstattung"),
    ("andere anlagen betriebs- und geschaeftsausstattung", "AKTIVA", f"{_AV}/II Sachanlagen/Andere Anlagen Betriebs- und Geschaeftsausstattung"),
    ("geleistete anzahlungen und anlagen im bau", "AKTIVA", f"{_AV}/II Sachanlagen/Geleistete Anzahlungen und Anlagen im Bau"),
    ("anteile an verbundenen unternehmen", "AKTIVA", f"{_AV}/III Finanzanlagen/Anteile an verbundenen Unternehmen"),
    ("ausleihungen an verbundene unternehmen", "AKTIVA", f"{_AV}/III Finanzanlagen/Ausleihungen an verbundene Unternehmen"),
    ("beteiligungen", "AKTIVA", f"{_AV}/III Finanzanlagen/Beteiligungen"),
    ("wertpapiere des anlagevermoegens", "AKTIVA", f"{_AV}/III Finanzanlagen/Wertpapiere des Anlagevermoegens"),
    # Druckbild-Variante: Wasserzeichen kappt die Überschrift mitten im Wort.
    ("wertpapiere des anla", "AKTIVA", f"{_AV}/III Finanzanlagen/Wertpapiere des Anlagevermoegens"),
    ("sonstige ausleihungen", "AKTIVA", f"{_AV}/III Finanzanlagen/Sonstige Ausleihungen"),
    # --- Umlaufvermögen
    ("roh-, hilfs- und betriebsstoffe", "AKTIVA", f"{_UV}/I Vorraete/Roh- Hilfs- und Betriebsstoffe"),
    ("unfertige erzeugnisse", "AKTIVA", f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen"),
    ("fertige erzeugnisse und waren", "AKTIVA", f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren"),
    ("geleistete anzahlungen", "AKTIVA", f"{_UV}/I Vorraete/Geleistete Anzahlungen"),
    ("forderungen aus lieferungen und leistungen", "AKTIVA", f"{_FORD}/Forderungen aus Lieferungen und Leistungen"),
    ("forderungen gegen verbundene unternehmen", "AKTIVA", f"{_FORD}/Forderungen gegen verbundene Unternehmen"),
    ("forderungen gegen unternehmen mit beteiligungsverhaeltnis", "AKTIVA", f"{_FORD}/Forderungen gegen Unternehmen mit Beteiligungsverhaeltnis"),
    ("sonstige vermoegensgegenstaende", "AKTIVA", f"{_FORD}/Sonstige Vermoegensgegenstaende"),
    ("sonstige wertpapiere", "AKTIVA", f"{_UV}/III Wertpapiere"),
    ("wertpapiere", "AKTIVA", f"{_UV}/III Wertpapiere"),
    ("kassenbestand", "AKTIVA", f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten"),
    ("guthaben bei kreditinstituten", "AKTIVA", f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten"),
    ("rechnungsabgrenzungsposten", "AKTIVA", f"{_AKT}/C Rechnungsabgrenzungsposten"),
    ("aktive latente steuern", "AKTIVA", f"{_AKT}/D Aktive latente Steuern"),
    # --- Passiva
    ("gezeichnetes kapital", "PASSIVA", f"{_PAS}/A Eigenkapital/I Gezeichnetes Kapital"),
    ("kapitalruecklage", "PASSIVA", f"{_PAS}/A Eigenkapital/II Kapitalruecklage"),
    ("gewinnruecklagen", "PASSIVA", f"{_PAS}/A Eigenkapital/III Gewinnruecklagen"),
    ("gewinnvortrag", "PASSIVA", f"{_PAS}/A Eigenkapital/IV Gewinnvortrag Verlustvortrag"),
    ("verlustvortrag", "PASSIVA", f"{_PAS}/A Eigenkapital/IV Gewinnvortrag Verlustvortrag"),
    ("jahresueberschuss", "PASSIVA", f"{_PAS}/A Eigenkapital/V Jahresueberschuss Jahresfehlbetrag"),
    ("rueckstellungen fuer pensionen", "PASSIVA", f"{_PAS}/B Rueckstellungen/Rueckstellungen fuer Pensionen und aehnliche Verpflichtungen"),
    ("steuerrueckstellungen", "PASSIVA", f"{_PAS}/B Rueckstellungen/Steuerrueckstellungen"),
    ("sonstige rueckstellungen", "PASSIVA", f"{_PAS}/B Rueckstellungen/Sonstige Rueckstellungen"),
    ("anleihen", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Anleihen"),
    ("verbindlichkeiten gegenueber kreditinstituten", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten"),
    ("erhaltene anzahlungen", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Erhaltene Anzahlungen auf Bestellungen"),
    ("verbindlichkeiten aus lieferungen und leistungen", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Verbindlichkeiten aus Lieferungen und Leistungen"),
    ("verbindlichkeiten gegenueber verbundenen unternehmen", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen"),
    # Gesellschafter-Positionen: §266 kennt keine eigene Zeile dafür. Die
    # Hausconvention führt sie (Regeln hgb-gesellschafter-*) unter den
    # verbundenen Unternehmen — hier explizit, damit die Zuordnung nicht
    # stillschweigend von der Vorgängerposition geerbt wird.
    ("verbindlichkeiten gegenueber gesellschaftern", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen"),
    ("verbindlichkeiten gegenueber gmbh-gesellschaftern", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen"),
    ("forderungen gegen gesellschafter", "AKTIVA", f"{_FORD}/Forderungen gegen verbundene Unternehmen"),
    ("sonstige verbindlichkeiten", "PASSIVA", f"{_PAS}/C Verbindlichkeiten/Sonstige Verbindlichkeiten"),
    ("rechnungsabgrenzungsposten", "PASSIVA", f"{_PAS}/D Rechnungsabgrenzungsposten"),
    ("passive latente steuern", "PASSIVA", f"{_PAS}/E Passive latente Steuern"),
    # --- GuV
    ("umsatzerloese", "GUV", "/GuV/Umsatzerloese"),
    ("bestandsveraenderung", "GUV", "/GuV/Bestandsveraenderungen Erzeugnisse"),
    ("andere aktivierte eigenleistungen", "GUV", "/GuV/Andere aktivierte Eigenleistungen"),
    ("sonstige betriebliche ertraege", "GUV", "/GuV/Sonstige betriebliche Ertraege"),
    ("aufwendungen fuer roh-, hilfs- und betriebsstoffe", "GUV", "/GuV/Materialaufwand/Roh- Hilfs- und Betriebsstoffe und Waren"),
    ("aufwendungen fuer bezogene leistungen", "GUV", "/GuV/Materialaufwand/Bezogene Leistungen"),
    ("loehne und gehaelter", "GUV", "/GuV/Personalaufwand/Loehne und Gehaelter"),
    ("soziale abgaben", "GUV", "/GuV/Personalaufwand/Soziale Abgaben und Altersversorgung"),
    ("abschreibungen", "GUV", "/GuV/Abschreibungen"),
    ("sonstige betriebliche aufwendungen", "GUV", "/GuV/Sonstige betriebliche Aufwendungen"),
    ("sonstige zinsen und aehnliche ertraege", "GUV", "/GuV/Sonstige Zinsen und aehnliche Ertraege"),
    ("zinsen und aehnliche aufwendungen", "GUV", "/GuV/Zinsen und aehnliche Aufwendungen"),
    ("steuern vom einkommen und vom ertrag", "GUV", "/GuV/Steuern vom Einkommen und vom Ertrag"),
    ("sonstige steuern", "GUV", "/GuV/Sonstige Steuern"),
]

_BETRAG = r"-?\d{1,3}(?:\.\d{3})*,\d{2}-?"
_ACCT_RE = re.compile(rf"^(?P<konto>\d{{3,5}})\s+(?P<bez>.+?)"
                      rf"(?P<betraege>(?:\s+{_BETRAG}){{1,3}})\s*$")
_BETRAG_RE = re.compile(_BETRAG)
#: Zeilen, die nie Position oder Konto sind (Druckbild-Beiwerk).
_NOISE = ("blatt", "kontennachweis", "uebertrag", "übertrag", "summe aktiva",
          "summe passiva", "konto bezeichnung", "geschaeftsjahr", "geschäftsjahr",
          "vorjahr", "seite", "eur", "€")


@dataclass
class KNKonto:
    konto: str
    bezeichnung: str
    hgb_pfad: str
    salden: dict[str, float]


@dataclass
class Kontennachweis:
    """Eingelesener Kontennachweis — Strukturquelle + Reconciliation-Ziel."""

    konten: dict[str, KNKonto]
    perioden: list[str]
    entity: str
    quelle_datei: str
    fingerprint: str = ""
    warnungen: list[str] = field(default_factory=list)

    @property
    def zuordnung(self) -> dict[str, str]:
        """Konto -> HGB-Pfad (was die Kaskade in Stufe 1 verwendet)."""
        return {k: v.hgb_pfad for k, v in self.konten.items()}

    def positionen(self) -> dict[str, dict[str, float]]:
        """HGB-Pfad -> Summe je Periode (Reconciliation-Ziel)."""
        out: dict[str, dict[str, float]] = {}
        for kk in self.konten.values():
            ziel = out.setdefault(kk.hgb_pfad, {p: 0.0 for p in self.perioden})
            for p in self.perioden:
                ziel[p] += kk.salden.get(p, 0.0)
        return out


def _match_ueberschrift(text: str, sektion: Optional[str]) -> Optional[str]:
    """Überschrift -> kanonischer HGB-Pfad.

    Zuerst exakte Wortmarke (längste gewinnt). Greift keine, wird tokenweise
    gewertet: das DATEV-Druckbild trägt ein diagonales Wasserzeichen, das
    einzelne Wörter zerhackt ("fertige Erzeugnisse und WJaren"). Eine Marke
    gilt als getroffen, wenn mindestens drei Viertel ihrer Tokens vorkommen.
    """
    t = normalisiere(text)
    t = re.sub(r"[^a-z0-9,\- ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    if not t or len(t) < 4:
        return None

    exakt: list[tuple[int, str]] = []
    for marke, sekt, pfad in _CROSSWALK:
        if sekt and sektion and sekt != sektion:
            continue
        if marke in t:
            exakt.append((len(marke), pfad))
    if exakt:
        return max(exakt, key=lambda x: x[0])[1]

    beste: Optional[tuple[int, int, str]] = None   # (Treffer, Markenlänge, Pfad)
    for marke, sekt, pfad in _CROSSWALK:
        if sekt and sektion and sekt != sektion:
            continue
        tokens = [tok for tok in marke.split() if len(tok) > 2]
        if len(tokens) < 2:
            continue
        hits = sum(1 for tok in tokens if tok in t)
        if hits / len(tokens) < 0.75:
            continue
        kandidat = (hits, len(marke), pfad)
        if beste is None or kandidat > beste:
            beste = kandidat
    return beste[2] if beste else None


class KontennachweisPdfReader:
    """Liest einen Kontennachweis aus dem DATEV-PDF-Druckbild."""

    name = "kontennachweis_pdf"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        return pfad.lower().endswith(".pdf")

    def lesen(self, pfad: str, perioden: Optional[list[str]] = None) -> Kontennachweis:
        warnungen: list[str] = []
        konten: dict[str, KNKonto] = {}
        entity = "Unbekannt"
        p_cur, p_prev = self._perioden(pfad)
        perioden = perioden or [p_cur, p_prev]

        try:
            import pdfplumber
        except Exception as e:                     # pragma: no cover
            return Kontennachweis({}, perioden, entity, pfad, "",
                                  [f"pdfplumber nicht verfügbar: {e}"])

        sektion: Optional[str] = None
        aktueller_pfad: Optional[str] = None
        puffer: list[str] = []

        try:
            with __import__("pdfplumber").open(pfad) as pdf:
                for page in pdf.pages:
                    for raw in (page.extract_text() or "").splitlines():
                        line = raw.strip()
                        if not line:
                            continue
                        low = normalisiere(line)
                        if "gmbh" in low and entity == "Unbekannt":
                            entity = line.strip()
                        # Sektionswechsel
                        if low.startswith("aktiva") or "zur handelsbilanz" in low:
                            sektion = "AKTIVA" if low.startswith("aktiva") else sektion
                            if "zur handelsbilanz" in low:
                                sektion = sektion or "AKTIVA"
                            puffer.clear()
                            continue
                        if low.startswith("passiva"):
                            sektion = "PASSIVA"; puffer.clear(); continue
                        if "zur g.u.v" in low or "zur guv" in low:
                            sektion = "GUV"; puffer.clear(); continue
                        if any(low.startswith(n) for n in _NOISE):
                            continue

                        m = _ACCT_RE.match(line)
                        if m:
                            konto = m.group("konto")
                            bez = re.sub(r"\s+", " ", m.group("bez")).strip()
                            betraege = _BETRAG_RE.findall(m.group("betraege"))
                            if self._korrupt(bez):
                                warnungen.append(
                                    f"Konto {konto}: Zeile unlesbar (Wasserzeichen) — übersprungen.")
                                puffer.clear()
                                continue
                            if aktueller_pfad is None:
                                warnungen.append(
                                    f"Konto {konto} ohne erkannte Positionsüberschrift — "
                                    "fällt auf die übrige Kaskade zurück.")
                                puffer.clear()
                                continue
                            werte = [parse_deutsche_zahl(b) for b in betraege]
                            salden = {perioden[0]: werte[0] if werte else 0.0}
                            if len(perioden) > 1:
                                salden[perioden[1]] = werte[-1] if len(werte) > 1 else 0.0
                            konten.setdefault(konto, KNKonto(
                                konto=konto, bezeichnung=bez,
                                hgb_pfad=aktueller_pfad, salden=salden))
                            puffer.clear()
                            continue

                        # Kein Konto -> Überschriftenkandidat (mehrzeilig möglich)
                        if _BETRAG_RE.search(line):
                            puffer.clear()          # Summenzeile o.ä.
                            continue
                        puffer.append(line)
                        if len(puffer) > 3:
                            puffer.pop(0)
                        for laenge in range(len(puffer), 0, -1):
                            kandidat = " ".join(puffer[-laenge:])
                            treffer = _match_ueberschrift(kandidat, sektion)
                            if treffer:
                                aktueller_pfad = treffer
                                break
        except Exception as e:
            warnungen.append(f"PDF-Parsing abgebrochen: {e}")

        return Kontennachweis(konten=konten, perioden=perioden, entity=entity,
                              quelle_datei=pfad, fingerprint=fingerprint(pfad),
                              warnungen=warnungen)

    @staticmethod
    def _korrupt(bez: str) -> bool:
        tokens = bez.split()
        if not tokens:
            return True
        einzeln = sum(1 for t in tokens if len(t) == 1)
        return len(tokens) >= 6 and einzeln / len(tokens) > 0.6

    @staticmethod
    def _perioden(pfad: str) -> tuple[str, str]:
        m = re.search(r"(20\d{2})", pfad)
        if m:
            j = int(m.group(1))
            return f"31.12.{j}", f"31.12.{j - 1}"
        return "Berichtsjahr", "Vorjahr"


class KontennachweisExcelReader:
    """Liest einen Kontennachweis aus Excel.

    Reales Layout (Eckart): **ein Blatt je Geschäftsjahr**; in der ersten
    Spalte stehen abwechselnd §266-Positionsüberschriften ("3. Andere Anlagen,
    Betriebs- und Geschäftsausstattung") und Kontozeilen, bei denen
    **Kontonummer und Bezeichnung in derselben Zelle** stehen ("0410
    Geschäftsausstattung"); der Wert steht in der Wertspalte daneben.

    Zwischenüberschriften ohne §266-Charakter ("Bankguthaben", "davon aus
    Steuern") greifen bewusst nicht — die zuletzt erkannte Position bleibt
    stehen, sodass die darunter stehenden Konten korrekt zugeordnet werden.

    Vorzeichen: Kontennachweise weisen die Passivseite positiv aus. Für die
    Reconciliation gegen die SuSa (Soll positiv / Haben negativ) werden
    Passiv-Werte negiert, damit beide Seiten dieselbe Sprache sprechen.
    """

    name = "kontennachweis_excel"

    #: Kontozeile: führende Kontonummer, dann die Bezeichnung.
    _KONTO_ZEILE = re.compile(r"^(?P<konto>\d{3,5})\s+(?P<bez>\S.*)$")

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            import openpyxl
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            for ws in wb.worksheets:
                treffer = 0
                for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
                    txt = " ".join(str(c) for c in row if c is not None)
                    low = normalisiere(txt)
                    if "kontennachweis" in low:
                        return True
                    if cls._KONTO_ZEILE.match(str(row[0]).strip() if row and row[0] else ""):
                        treffer += 1
                # Ein Blatt mit vielen "NNNN Bezeichnung"-Zeilen ist ein
                # Kontennachweis, auch ohne das Wort im Kopf.
                if treffer >= 8:
                    return True
            return False
        finally:
            wb.close()

    def lesen(self, pfad: str, perioden: Optional[list[str]] = None) -> Kontennachweis:
        import openpyxl
        wb = openpyxl.load_workbook(pfad, data_only=True)
        warnungen: list[str] = []
        konten: dict[str, KNKonto] = {}
        entity = "Unbekannt"
        blatt_perioden: list[str] = []

        for ws in wb.worksheets:
            periode = self._periode_des_blatts(ws.title, perioden)
            if periode is None:
                continue
            blatt_perioden.append(periode)
            sektion: Optional[str] = None
            aktueller_pfad: Optional[str] = None

            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                label = str(row[0]).strip()
                if not label:
                    continue
                low = normalisiere(label)
                if entity == "Unbekannt":
                    for zelle in row[1:3]:
                        if zelle and "gmbh" in normalisiere(str(zelle)):
                            entity = str(zelle).strip()
                if low in ("aktiva", "passiva"):
                    sektion = low.upper()
                    continue
                if any(low.startswith(n) for n in _NOISE):
                    continue

                wert = self._erster_wert(row)
                m = self._KONTO_ZEILE.match(label)
                if m:
                    if aktueller_pfad is None:
                        warnungen.append(
                            f"Konto {m.group('konto')} ohne erkannte "
                            "Positionsüberschrift — fällt auf die übrige Kaskade zurück.")
                        continue
                    if wert is None:
                        continue          # Konto ohne Saldo in dieser Periode
                    if sektion == "PASSIVA":
                        wert = -wert
                    konto = m.group("konto")
                    bez = re.sub(r"\s+", " ", m.group("bez")).strip()
                    vorhanden = konten.get(konto)
                    if vorhanden is None:
                        konten[konto] = KNKonto(konto=konto, bezeichnung=bez,
                                                hgb_pfad=aktueller_pfad,
                                                salden={periode: wert})
                    else:
                        # Dieselbe Kontonummer kann mehrfach vorkommen
                        # (z.B. drei Bankkonten unter 1210) -> aufsummieren.
                        vorhanden.salden[periode] = vorhanden.salden.get(periode, 0.0) + wert
                    continue

                treffer = _match_ueberschrift(label, sektion)
                if treffer:
                    aktueller_pfad = treffer

        wb.close()
        erkannt = perioden if perioden else blatt_perioden
        return Kontennachweis(
            konten=konten, perioden=list(erkannt), entity=entity,
            quelle_datei=pfad, fingerprint=fingerprint(pfad),
            warnungen=warnungen)

    @staticmethod
    def _erster_wert(row) -> Optional[float]:
        """Erster numerischer Wert der Zeile (die Wertspalte des Blatts)."""
        for zelle in row[1:]:
            if isinstance(zelle, (int, float)):
                return float(zelle)
            if isinstance(zelle, str) and _BETRAG_RE.fullmatch(zelle.strip()):
                return parse_deutsche_zahl(zelle)
        return None

    @staticmethod
    def _periode_des_blatts(titel: str, perioden: Optional[list[str]]) -> Optional[str]:
        """Blattname -> Periode. Das Jahr im Blattnamen ("Bilanz Konso 2023")
        wird auf die passende Ledger-Periode ("2023/12") abgebildet, damit
        Kontennachweis und SuSa dieselbe Zeitachse benutzen."""
        m = re.search(r"(20\d{2})", titel)
        if not m:
            return None
        jahr = m.group(1)
        if perioden:
            for p in perioden:
                if jahr in str(p):
                    return p
            return None
        return jahr


def lies_kontennachweis(pfad: str, perioden: Optional[list[str]] = None) -> Kontennachweis:
    """Formaterkennung für Kontennachweise (Excel vor PDF)."""
    if KontennachweisExcelReader.kann_lesen(pfad):
        return KontennachweisExcelReader().lesen(pfad, perioden)
    if KontennachweisPdfReader.kann_lesen(pfad):
        return KontennachweisPdfReader().lesen(pfad, perioden)
    raise ValueError(f"Kein Kontennachweis-Reader für: {pfad}")
