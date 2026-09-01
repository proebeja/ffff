"""Projekt Luma — jährliche Saldenliste aus einem MYOB-Kontenplan.

Diese Quelle unterscheidet sich in vier Punkten von allem bisherigen, und
jeder davon entscheidet über das Databook:

* **Der Kontenplan ist kein SKR.** Kontonummern wie ``1-12300`` und
  Bezeichnungen wie ``Trade Debtors - USD`` treffen weder die
  SKR03-Bereichstabelle noch die deutschen Stichworte der Typ-1-Regeln. Der
  Export bringt aber seine eigene Gliederung mit (``ClassDescription`` und
  ``AccountGroupDesc``), und die wird hier auf HGB-Pfade übersetzt — genau
  wie beim SAP-BW-Export, dessen FS-Hierarchie denselben Dienst tut. Damit
  greift Stufe 1 der Kaskade, und Klasse und NA-Zeile leitet die
  Reklassifizierung wie gewohnt ab.

* **Die Wertespalte ist nicht die erste Zahlenspalte.** Der Export führt in
  Spalte L die Bewegung der Periode und in Spalte O den Schlussbestand. Die
  Kopfzeile nennt Spalte O irreführend ``Year``. Wer L nimmt, baut ein
  Databook aus Veränderungen: die Aktivseite läge in FY2023 bei 260 TAUD
  statt bei 10,4 Mio. Die Probe steht in :meth:`MyobDiagnose.spaltenprobe`.

* **Das Geschäftsjahr endet am 31. März.** FY2023 ist zudem ein Rumpfjahr
  von neun Monaten (01.07.2022 bis 31.03.2023). Periodenlänge und Stichtag
  werden deshalb aus den Spalten ``PeriodFrom``/``PeriodTo`` gelesen und
  nicht aus dem Tabellennamen geraten.

* **Es gibt keine GuV.** Der Export führt nur die Klassen 10 bis 50, also
  Bilanz und Eigenkapital. Das Periodenergebnis steht als ein Betrag auf dem
  Konto ``3-90000 Current Earnings``. Ein Lead PL lässt sich daraus nicht
  füllen, und das ist eine Aussage über die Datenlage, kein Mangel des Laufs.
"""

from __future__ import annotations

import collections
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint

#: Spalten des Exports (1-basiert).
SP_KLASSE = 5           # ClassDescription
SP_GRUPPE = 7           # AccountGroupDesc
SP_KONTO = 8            # AccountNo
SP_KOSTENSTELLE = 9
SP_BEZEICHNUNG = 11     # GLDescription
SP_BEWEGUNG = 12        # PeriodMovementTYD
SP_SALDO = 15           # laut Kopfzeile "Year", tatsächlich der Schlussbestand
SP_VON = 16
SP_BIS = 17

_A = "/Aktiva/A Anlagevermoegen"
_UV = "/Aktiva/B Umlaufvermoegen"
_FORD = f"{_UV}/II Forderungen und sonstige Vermoegensgegenstaende"
_P = "/Passiva"

#: Kontogruppe des Exports -> HGB-Pfad. Die Zuordnung ist eine fachliche
#: Entscheidung je Gruppe und steht deshalb an einer Stelle, nicht verstreut
#: im Code.
#:
#: Zwei Grundsätze:
#:
#: * Der **Inhalt** entscheidet, nicht die Rubrik des Exports. ``Trade
#:   Facility``, ``R&D finance``, ``Finance Leases`` und die Firmenkreditkarte
#:   stehen im Export unter den kurzfristigen Verbindlichkeiten; es sind
#:   Finanzierungen und damit Net Debt. Sie werden direkt auf den
#:   Kreditinstitute-Pfad gelegt, statt über die gemischte Position
#:   ``Sonstige Verbindlichkeiten`` zu laufen — dort entschiede eine
#:   Typ-2-Regel über deutsche Stichworte, die auf englischen Bezeichnungen
#:   ohnehin nicht greifen.
#: * Wo der Inhalt **nicht** eindeutig ist, wird bewusst auf eine der drei
#:   gemischten Positionen gelegt. Dann greift die Typ-2-Logik, und was sie
#:   nicht auflöst, landet sichtbar in der Review-Queue — statt unsichtbar in
#:   einer Entscheidung, die hier niemand belegen kann.
GRUPPEN: dict[str, str] = {
    # --- Zahlungsmittel -----------------------------------------------------
    "efine": f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten",
    "petty cash": f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten",

    # --- Forderungen --------------------------------------------------------
    "trade debtors": f"{_FORD}/Forderungen aus Lieferungen und Leistungen",
    "provision for bad & doubtful debts":
        f"{_FORD}/Forderungen aus Lieferungen und Leistungen",
    "accrued income": f"{_FORD}/Forderungen aus Lieferungen und Leistungen",
    "sundry debtors": f"{_FORD}/Sonstige Vermoegensgegenstaende",
    "employee advance": f"{_FORD}/Sonstige Vermoegensgegenstaende",
    "income tax asset": f"{_FORD}/Sonstige Vermoegensgegenstaende",
    "trade-in clearing account": f"{_FORD}/Sonstige Vermoegensgegenstaende",
    "bt imaging product": f"{_FORD}/Sonstige Vermoegensgegenstaende",

    # Konzernforderungen. Zins tragende Ausleihung an eine Schwestergesellschaft,
    # in der Hausconvention Net Debt.
    "inter company loan - china":
        f"{_FORD}/Forderungen gegen verbundene Unternehmen",
    "inter company loan - aurora":
        f"{_FORD}/Forderungen gegen verbundene Unternehmen",
    "accrued interest - aurora":
        f"{_FORD}/Forderungen gegen verbundene Unternehmen",

    # --- Abgrenzung ---------------------------------------------------------
    "prepayments": "/Aktiva/C Rechnungsabgrenzungsposten",
    "rent in advance": "/Aktiva/C Rechnungsabgrenzungsposten",
    "deposits paid": "/Aktiva/C Rechnungsabgrenzungsposten",
    "loan borrowing costs": "/Aktiva/C Rechnungsabgrenzungsposten",

    # Ein Aktivkonto mit Habensaldo, das abgegrenzte Umsatzerlöse trägt. Der
    # Inhalt ist eine erhaltene Anzahlung, deshalb steht es auf der Passivseite.
    "unearned revenue - revenue recognition":
        f"{_P}/C Verbindlichkeiten/Erhaltene Anzahlungen auf Bestellungen",

    # --- Vorräte ------------------------------------------------------------
    "total inventory": f"{_UV}/I Vorraete/Roh- Hilfs- und Betriebsstoffe",
    "wip": f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen",
    "finished goods": f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren",
    "finished goods-labour": f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren",

    # --- Anlagevermögen -----------------------------------------------------
    "furniture & fittings": f"{_A}/II Sachanlagen",
    "office equipment": f"{_A}/II Sachanlagen",
    "computer equipment": f"{_A}/II Sachanlagen",
    "tools & lab equipment": f"{_A}/II Sachanlagen",
    "warehouse equipment": f"{_A}/II Sachanlagen",
    "prototype": f"{_A}/II Sachanlagen",
    "leasehold property improvements": f"{_A}/II Sachanlagen",
    "right of use assets": f"{_A}/II Sachanlagen",
    "property plant & equipment": f"{_A}/II Sachanlagen",
    "us office fixed assets": f"{_A}/II Sachanlagen",
    # Die Demo- und Entwicklungswerkzeuge führt der Export je Werkzeug mit
    # eigener Gruppe für die kumulierte Abschreibung. Alle gehören zum
    # Sachanlagevermögen; die Trennung Anschaffung/Abschreibung ist eine
    # Frage des Aufrisses, nicht der Position.
    "r&d/demo tools - waterloo": f"{_A}/II Sachanlagen",
    "r&d r3-169 & 181 - acc dep": f"{_A}/II Sachanlagen",
    "r&d b3-106 acc dep": f"{_A}/II Sachanlagen",
    "r&d m1-ct-102 acc dep": f"{_A}/II Sachanlagen",
    "intellectual property": f"{_A}/I Immaterielle Vermoegensgegenstaende",
    "goodwill": f"{_A}/I Immaterielle Vermoegensgegenstaende",
    "other intangibles": f"{_A}/I Immaterielle Vermoegensgegenstaende",
    "intangibles": f"{_A}/I Immaterielle Vermoegensgegenstaende",

    # --- Latente Steuern ----------------------------------------------------
    "deferred tax asset": "/Aktiva/D Aktive latente Steuern",

    # --- Verbindlichkeiten aus L+L -----------------------------------------
    "trade creditors":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten aus Lieferungen und Leistungen",
    "accrued trade creditors":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten aus Lieferungen und Leistungen",

    # --- Finanzierung (Net Debt) -------------------------------------------
    "credit cards":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten",
    "trade facility":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten",
    "r&d finance":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten",
    "finance leases":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten",
    # Partners for Growth ist der Wagnisfremdkapitalgeber des Unternehmens —
    # die Warrants im Eigenkapital tragen denselben Namen. Ein Darlehen,
    # keine Lieferantenverbindlichkeit.
    "loan pfg":
        f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten",
    "financial liability (convertible note)": f"{_P}/C Verbindlichkeiten/Anleihen",

    # --- Erhaltene Anzahlungen ---------------------------------------------
    "sales in advance":
        f"{_P}/C Verbindlichkeiten/Erhaltene Anzahlungen auf Bestellungen",

    # --- Rückstellungen -----------------------------------------------------
    "accrued expenses": f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen",
    "provision for annual leave": f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen",
    "provision for lsl": f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen",

    # --- Sonstige Verbindlichkeiten ----------------------------------------
    "ato liabilities": f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten",
    "payroll liabilities": f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten",
    "sundry creditors": f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten",
    "interest withholding payable":
        f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten",

    # --- Eigenkapital -------------------------------------------------------
    "share capital": f"{_P}/A Eigenkapital",
    "series a pref shares- redemption payments": f"{_P}/A Eigenkapital",
    "warrant - partners for growth iii lp": f"{_P}/A Eigenkapital",
    "warrant - silicon valley bank": f"{_P}/A Eigenkapital",
    "ast equity interest": f"{_P}/A Eigenkapital",
    "fundraising equity costs": f"{_P}/A Eigenkapital",
    "retained earnings": f"{_P}/A Eigenkapital",
    "current earnings": f"{_P}/A Eigenkapital",
}

#: Das Eigenkapitalkonto, auf dem der Bilanzexport das laufende Ergebnis in
#: EINER Zahl führt. Es ist eine Rechengröße, kein bebuchtes Konto — sobald
#: die GuV daneben liegt, steht dasselbe Ergebnis zweimal in der Mappe. Wird
#: die GuV geladen, fällt dieses Konto deshalb heraus, und die Summe aller
#: Konten geht wieder auf null.
ERGEBNIS_EIGENKAPITAL = "3-90000"

#: Klassen des Exports, die zur GuV gehören.
GUV_KLASSEN = ("income", "cost of goods sold", "manufacturing", "expenses",
               "other income", "other expenses")

_G = "/GuV"

#: Kontogruppe der GuV -> GuV-Pfad.
#:
#: Die Blätter dieser Pfade sind bewusst die **Kostenartengliederung der
#: Vorlage** und nicht die § 275-Gliederung: die Quelle ist eine
#: Kostenartenrechnung, die Vorlage führt dieselben Zeilen, und der Pfad-Blatt
#: wird in der Kaskade unmittelbar zur NA-Zeile. Damit trifft der Ticker ohne
#: Umweg, und der Lead PL zeigt die Kostenarten statt eines einzigen Blocks
#: "sonstige betriebliche Aufwendungen".
GUV_GRUPPEN: dict[str, str] = {
    # --- Erlöse -------------------------------------------------------------
    "product sales": f"{_G}/Umsatzerloese",
    "consulting income": f"{_G}/Umsatzerloese",
    "sales - miscellaneous": f"{_G}/Umsatzerloese",

    # --- Herstellkosten -----------------------------------------------------
    "sales": f"{_G}/Materialaufwand",              # Klasse "Cost of Goods Sold"
    "install": f"{_G}/Materialaufwand",
    "cogs licences": f"{_G}/Materialaufwand",
    "sales bid/service fee": f"{_G}/Materialaufwand",
    "shipping": f"{_G}/Kosten der Warenabgabe",

    # --- Fertigung ----------------------------------------------------------
    "consumables": f"{_G}/Materialaufwand",
    "consumables-r&d": f"{_G}/Materialaufwand",
    "materials": f"{_G}/Materialaufwand",
    "overhead": f"{_G}/Materialaufwand",
    "manufacturing variance": f"{_G}/Materialaufwand",
    "purchasing exchange gains/losses": f"{_G}/Materialaufwand",
    "mft external labour": f"{_G}/Materialaufwand",
    "mft losses - scrap,loss,exch": f"{_G}/Materialaufwand",
    "prototye revaluation": f"{_G}/Materialaufwand",
    "inventory write downs": f"{_G}/Materialaufwand",
    "inventory write offs": f"{_G}/Materialaufwand",
    "inventory write offs - 2018": f"{_G}/Materialaufwand",
    "inventory write offs - pre 2018": f"{_G}/Materialaufwand",
    "inventory write offs - trade-in": f"{_G}/Materialaufwand",
    "stock write-down": f"{_G}/Materialaufwand",
    "freight by dhl": f"{_G}/Kosten der Warenabgabe",
    "freight in manufacturing": f"{_G}/Kosten der Warenabgabe",
    "freight recoveries": f"{_G}/Kosten der Warenabgabe",
    # Fertigungslöhne sind Personalaufwand, auch wenn der Export sie unter
    # Manufacturing führt. Die Gesamtkostenrechnung trennt nach Kostenart,
    # nicht nach Kostenstelle.
    "labour": f"{_G}/Personalaufwand",
    "mft direct labour": f"{_G}/Personalaufwand",
    "r&d labour": f"{_G}/Personalaufwand",

    # --- Betriebliche Aufwendungen -----------------------------------------
    "employment costs": f"{_G}/Personalaufwand",
    "administration": f"{_G}/Sonstige betriebliche Aufwendungen",
    "operations": f"{_G}/Sonstige betriebliche Aufwendungen",
    "r&d": f"{_G}/Sonstige betriebliche Aufwendungen",
    "legal costs": f"{_G}/Sonstige betriebliche Aufwendungen",
    "china entity": f"{_G}/Sonstige betriebliche Aufwendungen",
    "us office": f"{_G}/Sonstige betriebliche Aufwendungen",
    "bti conference": f"{_G}/Sonstige betriebliche Aufwendungen",
    "financing": f"{_G}/Zinsaufwendungen",
    "sales & marketing/pr costs": f"{_G}/Vertriebskosten",

    # --- Übriges ------------------------------------------------------------
    "other income": f"{_G}/Sonstige betriebliche Ertraege",
    "other expenses": f"{_G}/Sonstige betriebliche Aufwendungen",
    "interest payable": f"{_G}/Zinsaufwendungen",
    "interest expense - leases": f"{_G}/Zinsaufwendungen",
    "income tax expense/(benefit)": f"{_G}/Ertragssteuern",
}

#: Einzelne GuV-Konten, deren Gruppe zu grob ist. Ohne sie verschwänden
#: Abschreibungen, Zinsen und Raumkosten in der Sammelgruppe
#: ``Administration`` — und das EBITDA der Vorlage wäre um die
#: Abschreibungen falsch.
GUV_KONTEN: dict[str, str] = {
    # Abschreibungen — sie stehen unterhalb des EBITDA und dürfen nicht in
    # den sonstigen betrieblichen Aufwendungen hängen.
    "7-10210": f"{_G}/Abschreibungen",           # Depreciation
    "7-10211": f"{_G}/Abschreibungen",           # Depreciation - Leases
    "7-10220": f"{_G}/Abschreibungen",           # Amortisation

    # Raumkosten
    "7-10560": f"{_G}/Raumkosten",               # Rent
    "7-10565": f"{_G}/Raumkosten",               # Rent - Financial Leases
    "7-10850": f"{_G}/Raumkosten",               # Utilities
    "7-10130": f"{_G}/Raumkosten",               # Cleaning
    "7-10600": f"{_G}/Raumkosten",               # Security alarm
    "7-10306": f"{_G}/Raumkosten",               # Property Insurance

    # Reparaturen
    "7-10580": f"{_G}/Reparaturen und Instandhaltungen",

    # Versicherungen, Beiträge und Abgaben
    "7-10302": f"{_G}/Versicherungen, Beitraege und Abgaben",   # Travel Ins.
    "7-10303": f"{_G}/Versicherungen, Beitraege und Abgaben",   # D&O
    "7-10304": f"{_G}/Versicherungen, Beitraege und Abgaben",   # Prof. Indemnity
    "7-10305": f"{_G}/Versicherungen, Beitraege und Abgaben",   # Marine
    "7-10020": f"{_G}/Versicherungen, Beitraege und Abgaben",   # ASIC
    "7-11015": f"{_G}/Versicherungen, Beitraege und Abgaben",   # Duty
    "7-59020": f"{_G}/Versicherungen, Beitraege und Abgaben",   # Workers Comp

    # Werbe- und Reisekosten
    "7-10801": f"{_G}/Werbe- und Reisekosten",   # Travel - Domestic Air/Accom
    "7-10802": f"{_G}/Werbe- und Reisekosten",   # Travel - Domestic Other
    "7-10803": f"{_G}/Werbe- und Reisekosten",   # Travel - International
    "7-10250": f"{_G}/Werbe- und Reisekosten",   # Entertainment deductible
    "7-10260": f"{_G}/Werbe- und Reisekosten",   # Entertainment non deductible
    "7-20010": f"{_G}/Werbe- und Reisekosten",   # Advertising
    "7-25200": f"{_G}/Werbe- und Reisekosten",   # S&M Support - Travel
    "7-25520": f"{_G}/Werbe- und Reisekosten",   # SNEC (Messe)
    "7-25530": f"{_G}/Werbe- und Reisekosten",   # Workshops, Seminars
    "7-20130": f"{_G}/Werbe- und Reisekosten",   # Market Reports

    # Zinsen. Die Kreditbearbeitungsgebühr ist Finanzierungsaufwand und
    # gehört unter das EBIT, nicht in die sonstigen betrieblichen Kosten.
    "7-10221": f"{_G}/Zinsaufwendungen",         # Interest - Lease
    "7-20620": f"{_G}/Zinsaufwendungen",         # Interest Expense
    "7-10460": f"{_G}/Zinsaufwendungen",         # Loan Administration Fees

    # Zinserträge
    "4-51200": f"{_G}/Zinsertraege",             # Interest
    "4-51300": f"{_G}/Zinsertraege",             # Interest Revenue, Interco AST

    # Personalaufwand aus der Sammelgruppe der Auslandsgesellschaft
    "7-20512": f"{_G}/Personalaufwand",          # China Entity, Employee Salary

    # Frachten unter "Operations"
    "7-40200": f"{_G}/Kosten der Warenabgabe",
    "7-40201": f"{_G}/Kosten der Warenabgabe",
    "7-40202": f"{_G}/Kosten der Warenabgabe",
    "7-30170": f"{_G}/Kosten der Warenabgabe",   # Freight on R&D Imports
}

#: Kontonummern, die bewusst KEINEN Pfad bekommen und in die Review-Queue
#: laufen. Ein Sammel-/Verrechnungskonto ohne erkennbaren Inhalt gehört dorthin
#: und nicht in eine Bilanzposition, die jemand später für belegt hält.
OHNE_PFAD = {
    "9-99999": "Suspense account — clearing account with no identifiable content.",
}

#: Einzelne Konten, deren Gruppe zu grob ist.
#:
#: Für die Positionszeile der Vorlage macht das nichts — sie führt eine Zeile
#: ``Vorräte``. Für den HGB-Pfad im Mastersheet schon, und ein Aufriss der
#: Vorräte nach § 266 lebt genau von dieser Trennung.
KONTEN: dict[str, str] = {
    # Payroll Liabilities, inhaltlich aber Rückstellungen.
    "2-70350": f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen",
    "2-53000": f"{_P}/B Rueckstellungen/Steuerrueckstellungen",
    "2-54000": f"{_P}/B Rueckstellungen/Steuerrueckstellungen",
    # 'Total Inventory' fasst Roh-, Hilfs- und Betriebsstoffe, unfertige und
    # fertige Erzeugnisse in einer Gruppe zusammen.
    "1-51000": f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen",
    "1-51500": f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen",
    "1-52000": f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen",
    "1-60000": f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren",
    "1-40300": f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren",
}


def _norm(text) -> str:
    """Gruppenbezeichnung vergleichbar machen. Der Export schneidet lange
    Namen ab (``Provision for Bad & Doubtful`` statt ``... Debts``), deshalb
    wird zusätzlich auf Präfix geprüft."""
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


def ist_guv(klasse: str) -> bool:
    return _norm(klasse) in GUV_KLASSEN


def finde_pfad(gruppe: str, konto: str, klasse: str = "") -> Optional[str]:
    """HGB- bzw. GuV-Pfad einer Kontogruppe.

    Kontospezifisches schlägt die Gruppe, und die Klasse des Exports
    entscheidet, in welcher der beiden Tabellen gesucht wird. Ohne diese
    Trennung träfe eine Gruppe wie ``Other Income``, die es in der GuV zweimal
    gibt, womöglich einen Bilanzpfad.
    """
    if konto in OHNE_PFAD:
        return None
    konten, gruppen = ((GUV_KONTEN, GUV_GRUPPEN) if ist_guv(klasse)
                       else (KONTEN, GRUPPEN))
    if konto in konten:
        return konten[konto]
    g = _norm(gruppe)
    if g in gruppen:
        return gruppen[g]
    # Abgeschnittene Gruppennamen: der gespeicherte Schlüssel beginnt mit dem
    # gelesenen Namen oder umgekehrt.
    for schluessel, pfad in gruppen.items():
        if g and (schluessel.startswith(g) or g.startswith(schluessel)):
            return pfad
    return None


def _stichtag(bis) -> date:
    """``PeriodTo`` nennt den ERSTEN Tag des letzten Periodenmonats.

    ``2023-03-01`` ist die Periode bis einschließlich März 2023, also der
    Stichtag 31.03.2023. Wer den Wert direkt übernimmt, legt den Abschluss
    einen Monat zu früh und verschiebt die ganze Zeitachse der Vorlage.
    """
    d = bis if isinstance(bis, datetime) else datetime.fromisoformat(str(bis))
    naechster = date(d.year + (d.month == 12), d.month % 12 + 1, 1)
    return date.fromordinal(naechster.toordinal() - 1)


@dataclass
class MyobDiagnose:
    """Was der Reader über die Quelle sagen kann."""

    perioden: list[str] = field(default_factory=list)
    stichtage: dict[str, date] = field(default_factory=dict)
    monate: dict[str, int] = field(default_factory=dict)
    kontozeilen: dict[str, int] = field(default_factory=dict)
    #: Periode -> (Summe Schlussbestände, Summe Bewegungen)
    spaltensummen: dict[str, tuple[float, float]] = field(default_factory=dict)
    #: Konten, die in mehreren Zeilen (Kostenstellen) geführt werden.
    mehrfach: dict[str, int] = field(default_factory=dict)
    #: Kontonummern, die in BEIDEN Exporten vorkommen. Sie würden addiert und
    #: müssen deshalb sichtbar sein.
    doppelte_konten: set = field(default_factory=set)
    #: Konten, die bewusst nicht ins Mastersheet gehen, mit Begründung.
    eliminiert: dict = field(default_factory=dict)
    #: Periode -> Betrag, den die Elimination aus dem Rohstand herausnimmt.
    doppelt_gezaehlt: dict = field(default_factory=dict)
    #: Ob eine GuV mitgeladen wurde.
    hat_guv: bool = False
    ohne_pfad: list[tuple[str, str, str]] = field(default_factory=list)
    gruppen_ohne_zuordnung: set = field(default_factory=set)

    def spaltenprobe(self) -> list[tuple[str, bool, str]]:
        """Belegt, dass Spalte O der Schlussbestand ist und nicht die Bewegung.

        Beide Spalten summieren sich je Periode auf null — eine Saldenliste
        tut das, und eine Bewegungsspalte auch. Unterscheiden lassen sie sich
        an der Größenordnung und daran, dass Schlussbestand der Vorperiode
        plus Bewegung den Schlussbestand ergibt.
        """
        proben: list[tuple[str, bool, str]] = []
        for p in self.perioden:
            saldo, bewegung = self.spaltensummen[p]
            # Liegt die GuV daneben, trägt der Rohstand das Ergebnis doppelt:
            # einmal auf den GuV-Konten und einmal auf dem Eigenkapitalkonto,
            # das die Elimination herausnimmt.
            saldo -= self.doppelt_gezaehlt.get(p, 0.0)
            proben.append((f"{p}: Bilanzidentität", abs(saldo) <= 1.0,
                           f"Summe aller Schlussbestände {saldo:,.2f}"))
            # Ohne GuV summiert auch die Bewegungsspalte auf null, und genau
            # das macht sie zur Falle. Liegt die GuV daneben, gilt das nicht
            # mehr: das Eigenkapitalkonto trägt dort den Vortrag des
            # Vorjahresergebnisses mit. Die Zahl wird dann nur berichtet.
            proben.append((f"{p}: Bewegungsspalte", 
                           self.hat_guv or abs(bewegung) <= 1.0,
                           f"Summe aller Bewegungen {bewegung:,.2f}"
                           + (" (mit GuV keine Nullprobe)" if self.hat_guv
                              else " — summiert ebenfalls auf null")))
        return proben


class MyobSusaReader(Reader):
    """Liest alle Jahrestabs einer MYOB-Saldenliste in ein Ledger."""

    name = "myob_susa"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            ws = wb.worksheets[0]
            kopf = [str(ws.cell(1, c).value or "").strip()
                    for c in range(1, 18)]
            return "AccountNo" in kopf and "PeriodMovementTYD" in kopf
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        ledger, _ = lies_myob(pfad)
        return ledger


def lies_myob(pfad: str, guv_pfad: Optional[str] = None,
              entity: str = "Projekt Luma"
              ) -> tuple[NormalizedLedger, MyobDiagnose]:
    """Bilanz- und GuV-Export -> ein Ledger plus Diagnose.

    Beide Dateien haben denselben Aufbau und dieselben Perioden; sie
    unterscheiden sich nur in den Klassen, die sie führen. Zusammengeführt
    wird über die Kontonummer, und die Perioden werden **nach Stichtag
    sortiert** — die Tabs des GuV-Exports stehen in der Reihenfolge FY2023,
    FY2026, FY2024, FY2025, und die Blattreihenfolge zu übernehmen hieße, die
    Zeitachse zu verdrehen.
    """
    diag = MyobDiagnose()
    warnungen: list[str] = []
    salden: dict[str, dict[str, float]] = collections.defaultdict(dict)
    stamm: dict[str, tuple[str, str, str]] = {}     # konto -> (bez, gruppe, klasse)
    zeilen_je_konto: collections.Counter = collections.Counter()
    quellen = [p for p in (pfad, guv_pfad) if p]
    diag.hat_guv = bool(guv_pfad)

    for datei in quellen:
        wb = openpyxl.load_workbook(datei, data_only=True)
        for ws in wb.worksheets:
            periode = str(ws.title).replace(" annual", "").strip()
            summe_saldo = summe_bewegung = 0.0
            zeilen = 0
            je_konto: dict[str, float] = collections.defaultdict(float)

            for r in range(2, ws.max_row + 1):
                konto = ws.cell(r, SP_KONTO).value
                if not konto:
                    continue
                konto = str(konto).strip()
                saldo = ws.cell(r, SP_SALDO).value or 0.0
                bewegung = ws.cell(r, SP_BEWEGUNG).value or 0.0
                # Ein Konto steht je Kostenstelle einmal in der Liste. Für das
                # Mastersheet ist die Kostenstelle keine eigene Zeile — sie
                # wäre ein zweiter Schlüssel neben der Kontonummer und liefe
                # der Single-Source-Regel zuwider.
                je_konto[konto] += float(saldo)
                summe_saldo += float(saldo)
                summe_bewegung += float(bewegung)
                zeilen += 1
                zeilen_je_konto[(datei, konto)] += 1
                stamm.setdefault(konto, (
                    str(ws.cell(r, SP_BEZEICHNUNG).value or "").strip(),
                    str(ws.cell(r, SP_GRUPPE).value or "").strip(),
                    str(ws.cell(r, SP_KLASSE).value or "").strip()))

                if diag.stichtage.get(periode) is None:
                    bis, von = ws.cell(r, SP_BIS).value, ws.cell(r, SP_VON).value
                    if bis:
                        diag.stichtage[periode] = _stichtag(bis)
                        if von:
                            v = von if isinstance(von, datetime) else \
                                datetime.fromisoformat(str(von))
                            diag.monate[periode] = (
                                (diag.stichtage[periode].year - v.year) * 12
                                + diag.stichtage[periode].month - v.month + 1)

            for konto, wert in je_konto.items():
                if konto in salden and periode in salden[konto]:
                    diag.doppelte_konten.add(konto)
                salden[konto][periode] = salden[konto].get(periode, 0.0) + wert
            diag.kontozeilen[periode] = diag.kontozeilen.get(periode, 0) + zeilen
            vorher = diag.spaltensummen.get(periode, (0.0, 0.0))
            diag.spaltensummen[periode] = (vorher[0] + summe_saldo,
                                           vorher[1] + summe_bewegung)
        wb.close()

    # Perioden nach Stichtag, nicht nach Blattreihenfolge.
    diag.perioden = sorted(diag.stichtage, key=lambda p: diag.stichtage[p])

    diag.mehrfach = {k: v // len(diag.perioden)
                     for (_, k), v in zeilen_je_konto.items()
                     if v > len(diag.perioden)}

    accounts: list[Account] = []
    for konto in sorted(salden):
        if konto == ERGEBNIS_EIGENKAPITAL and guv_pfad:
            diag.eliminiert[konto] = (
                f"{stamm[konto][0]} — the P&L in a single figure. With the "
                "P&L loaded it would be counted twice.")
            diag.doppelt_gezaehlt = {p: salden[konto].get(p, 0.0)
                                     for p in diag.perioden}
            continue
        bez, gruppe, klasse = stamm[konto]
        guv = ist_guv(klasse)
        fs_pfad = finde_pfad(gruppe, konto, klasse)
        if fs_pfad is None:
            diag.ohne_pfad.append((konto, bez, OHNE_PFAD.get(
                konto, f"Account group '{gruppe}' has no mapping.")))
            if konto not in OHNE_PFAD:
                diag.gruppen_ohne_zuordnung.add(gruppe)
        if guv:
            kontotyp = "guv"
        elif klasse.lower().endswith("assets"):
            kontotyp = "bilanz_aktiv"
        else:
            kontotyp = "bilanz_passiv"
        accounts.append(Account(
            konto=konto, bezeichnung=bez,
            salden=tuple(PeriodBalance(p, round(salden[konto].get(p, 0.0), 2))
                         for p in diag.perioden),
            entity=entity, fs_pfad=fs_pfad, kontotyp=kontotyp))

    if diag.gruppen_ohne_zuordnung:
        warnungen.append("Kontogruppen ohne Zuordnung: "
                         + ", ".join(sorted(diag.gruppen_ohne_zuordnung)))
    if diag.doppelte_konten:
        warnungen.append("Kontonummern in beiden Exporten: "
                         + ", ".join(sorted(diag.doppelte_konten)))
    # Die Probe zählt die Konten, die tatsächlich ins Mastersheet gehen —
    # ``spaltensummen`` hält daneben den Rohstand der Exporte fest.
    for periode in diag.perioden:
        saldo = sum(a.saldo(periode) for a in accounts)
        if abs(saldo) > 1.0:
            warnungen.append(f"{periode}: Summe aller Konten {saldo:,.2f}")

    ledger = NormalizedLedger(
        accounts=accounts, perioden=list(diag.perioden), entity=entity,
        quelle_datei=" + ".join(quellen), hat_kontennachweis=True,
        fingerprint=fingerprint(pfad), warnungen=warnungen)
    return ledger, diag
