"""Reader für den MYOB-Kontenexport (Projekt Luma, australischer Mandant).

Ein Blatt je Periode, 17 Spalten, Bilanz und GuV in getrennten Dateien. Der
Export führt Jahres- **und** Quartalsblätter nebeneinander; welches Blatt
welche Periode trägt, steht nicht im Blattnamen, sondern in ``PeriodFrom`` und
``PeriodTo``.

Fünf Eigenheiten, an denen ein naiver Reader scheitert — jede einzelne
erzeugt ein Databook, das plausibel aussieht und falsch ist:

**``Year`` ist der Schlussbestand, ``PeriodMovementTYD`` die Bewegung.** Der
Spaltenkopf ``Year`` legt etwas anderes nahe. Beide Spalten summieren in der
Bilanz je Periode auf null, sehen also gleich plausibel aus. Wer die Bewegung
nimmt, baut eine Bilanz aus Veränderungen.

**In der GuV ist ``Year`` kumuliert.** Im Quartalsblatt Oct-Dec25 steht dort
das Ergebnis von April bis Dezember, nicht das des Quartals. Für die GuV gilt
deshalb durchgängig ``PeriodMovementTYD``: im Jahresblatt ist er das Jahr, im
Quartalsblatt das Quartal. Für die Bilanz gilt durchgängig ``Year``.

**``PeriodTo`` nennt den ERSTEN Tag des letzten Periodenmonats.** Aus dem
01.03. wird der 31.03., sonst liegt der Abschluss einen Monat zu früh und aus
dem Geschäftsjahr wird ein Rumpfjahr.

**Die Blattreihenfolge ist nicht die Zeitachse.** In der GuV-Datei steht
FY2026 an zweiter Stelle. Sortiert wird nach Stichtag.

**Ein Konto steht mehrfach je Blatt**, einmal je Kostenstelle
(``CostCentreName``). Die Zeilen gehören summiert; wer die erste nimmt,
verliert den Rest lautlos.

Zurückgegeben wird ein Ledger, dessen Konten **beide** Periodenraster tragen:
die Quartale und die Jahre. Das Databook führt das Mastersheet quartalsweise
und alle übrigen Blätter zum Jahresende — beides muss aus derselben Quelle
kommen, sonst laufen sie auseinander.
"""

from __future__ import annotations

import collections
import os
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import fingerprint

#: Spalten, die der Reader braucht. Gesucht wird über die Kopfzeile, nicht
#: über die Position — der Export trägt 17 Spalten in nicht zugesicherter
#: Reihenfolge.
_KONTO = "AccountNo"
_BEZ = "GLDescription"
_GRUPPE = "AccountGroupDesc"
_KLASSE = "ClassDescription"
_BESTAND = "Year"
_BEWEGUNG = "PeriodMovementTYD"
_VON = "PeriodFrom"
_BIS = "PeriodTo"

#: ``ClassDescription`` -> Kontotyp. Sie ist die einzige Strukturangabe der
#: Quelle: Bilanzseite und Fristigkeit, sonst nichts. Welche Position ein
#: Konto trägt, sagt sie nicht — das leistet der Kontenrahmen.
KLASSEN = {
    "Current Assets": ("bilanz_aktiv", "Current"),
    "Non-Current Assets": ("bilanz_aktiv", "Non-current"),
    "Current Liabilities": ("bilanz_passiv", "Current"),
    "Non-Current Liabilities": ("bilanz_passiv", "Non-current"),
    "Equity": ("bilanz_passiv", ""),
    "Income": ("guv", ""),
    "Other Income": ("guv", ""),
    "Cost of Goods Sold": ("guv", ""),
    "Manufacturing": ("guv", ""),
    "Expenses": ("guv", ""),
    "Other Expenses": ("guv", ""),
}


@dataclass
class Periodenraster:
    """Welche Perioden es gibt und wie sie zusammenhängen.

    ``spalten`` ist die Reihenfolge des Mastersheets: je Geschäftsjahr erst
    seine Quartale, dann die Jahresspalte. Dadurch steht die Jahressumme
    unmittelbar rechts neben den Quartalen, aus denen sie entsteht, und die
    Summenformel im Blatt läuft über einen zusammenhängenden Bereich.
    """

    jahre: list[str] = field(default_factory=list)
    quartale_je_jahr: dict[str, list[str]] = field(default_factory=dict)
    stichtage: dict[str, date] = field(default_factory=dict)
    #: Periodenlabel -> Blattname, für das Laufprotokoll.
    blaetter: dict[str, str] = field(default_factory=dict)

    @property
    def spalten(self) -> list[str]:
        out: list[str] = []
        for jahr in self.jahre:
            out += self.quartale_je_jahr.get(jahr, [])
            out.append(jahr)
        return out

    @property
    def quartale(self) -> list[str]:
        return [q for j in self.jahre for q in self.quartale_je_jahr.get(j, [])]

    def ist_jahr(self, periode: str) -> bool:
        return periode in self.jahre


@dataclass
class MyobDiagnose:
    """Was beim Einlesen aufgefallen ist — Grundlage der QA-Eingangsdiagnose."""

    dateien: list[str] = field(default_factory=list)
    #: Periode -> Summe aller Bilanzsalden. Muss null sein.
    bilanzidentitaet: dict[str, float] = field(default_factory=dict)
    #: Jahr -> (Summe der Quartalsbewegungen GuV, Jahreswert). Muss gleich sein.
    guv_aufriss: dict[str, tuple[float, float]] = field(default_factory=dict)
    #: Jahr -> (Bilanzsumme im letzten Quartal, im Jahresblatt). Muss gleich sein.
    bilanz_jahresende: dict[str, tuple[float, float]] = field(default_factory=dict)
    kontozeilen: dict[str, int] = field(default_factory=dict)
    #: Geschäftsjahre, denen Quartale fehlen (Rumpfjahr oder Lücke im Export).
    unvollstaendige_jahre: dict[str, int] = field(default_factory=dict)
    uebersprungene_blaetter: list[str] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)


# ---- Blattebene ----------------------------------------------------------

@dataclass
class _Blatt:
    name: str
    von: date
    bis: date
    monate: int
    #: Konto -> (Bezeichnung, Gruppe, ClassDescription, Bestand, Bewegung)
    konten: dict[str, tuple[str, str, str, float, float]]

    @property
    def ist_jahr(self) -> bool:
        """Ein Blatt über sechs Monate oder mehr ist ein Jahresblatt. Die
        Grenze liegt nicht bei zwölf: FY2023 ist im Export ein Rumpfjahr über
        neun Monate und trotzdem das Jahresblatt."""
        return self.monate >= 6


def _als_datum(tag) -> date:
    """Datum aus der Zelle. Je nachdem, wie der Export erzeugt wurde, steht
    dort ein echtes Datum oder dessen Text — beides muss der Reader lesen,
    sonst hängt das Einlesen an der Einstellung des Exportwerkzeugs."""
    if isinstance(tag, datetime):
        return tag.date()
    if isinstance(tag, date):
        return tag
    text = str(tag).strip()[:19]
    for muster in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, muster).date()
        except ValueError:
            continue
    raise ValueError(f"Unlesbares Periodendatum: {tag!r}")


def _monatsende(tag) -> date:
    tag = _als_datum(tag)
    folgemonat = date(tag.year + (tag.month == 12), tag.month % 12 + 1, 1)
    return date.fromordinal(folgemonat.toordinal() - 1)


def _erster(tag) -> date:
    tag = _als_datum(tag)
    return date(tag.year, tag.month, 1)


def _zahl(wert) -> float:
    if wert is None or wert == "":
        return 0.0
    try:
        return float(str(wert).replace(",", ""))
    except ValueError:
        return 0.0


def _lies_blatt(ws) -> Optional[_Blatt]:
    """Ein Blatt oder ``None``, wenn es keine Kopfzeile trägt.

    Die Trennblätter (``24 25 >>``) sind leer. Sie werden nicht als Fehler
    behandelt, aber in der Diagnose benannt — ein Blatt zu übersehen, weil es
    einen unerwarteten Aufbau hat, wäre derselbe Vorgang.
    """
    zeilen = ws.iter_rows(values_only=True)
    try:
        kopf = [str(h or "").strip() for h in next(zeilen)]
    except StopIteration:
        return None
    if _KONTO not in kopf or _BESTAND not in kopf:
        return None
    idx = {name: kopf.index(name) for name in
           (_KONTO, _BEZ, _GRUPPE, _KLASSE, _BESTAND, _BEWEGUNG, _VON, _BIS)}

    konten: dict[str, tuple[str, str, str, float, float]] = {}
    von = bis = None
    for reihe in zeilen:
        nummer = str(reihe[idx[_KONTO]] or "").strip()
        if not nummer:
            continue
        if von is None and reihe[idx[_VON]]:
            von, bis = _erster(reihe[idx[_VON]]), _monatsende(reihe[idx[_BIS]])
        vorher = konten.get(nummer)
        bestand = _zahl(reihe[idx[_BESTAND]])
        bewegung = _zahl(reihe[idx[_BEWEGUNG]])
        if vorher is None:
            konten[nummer] = (str(reihe[idx[_BEZ]] or "").strip(),
                              str(reihe[idx[_GRUPPE]] or "").strip(),
                              str(reihe[idx[_KLASSE]] or "").strip(),
                              bestand, bewegung)
        else:
            # Kostenstellen: dasselbe Konto mehrfach im Blatt.
            konten[nummer] = (vorher[0], vorher[1], vorher[2],
                              vorher[3] + bestand, vorher[4] + bewegung)
    if von is None:
        return None
    monate = (bis.year - von.year) * 12 + bis.month - von.month + 1
    return _Blatt(ws.title, von, bis, monate, konten)


def _lies_datei(pfad: str, diag: MyobDiagnose) -> list[_Blatt]:
    wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
    blaetter: list[_Blatt] = []
    for name in wb.sheetnames:
        blatt = _lies_blatt(wb[name])
        if blatt is None:
            diag.uebersprungene_blaetter.append(f"{os.path.basename(pfad)}:{name}")
        else:
            blaetter.append(blatt)
    wb.close()
    return blaetter


# ---- Periodenraster ------------------------------------------------------

def _gj_endmonat(blaetter: list[_Blatt]) -> int:
    """Der Monat, in dem das Geschäftsjahr endet — aus den Jahresblättern.

    Nicht geraten und nicht aus dem Blattnamen gelesen: bei einem Mandanten
    mit abweichendem Geschäftsjahr entscheidet diese Zahl, welchem Jahr ein
    Quartal zufällt.
    """
    monate = collections.Counter(b.bis.month for b in blaetter if b.ist_jahr)
    if not monate:
        raise ValueError("Kein Jahresblatt gefunden — das Geschäftsjahr ist "
                         "nicht bestimmbar.")
    return monate.most_common(1)[0][0]


def _jahr_von(stichtag: date, endmonat: int) -> str:
    """Geschäftsjahr eines Stichtags. Der 30.09.2022 gehört bei einem am
    31.03. endenden Geschäftsjahr zu FY2023."""
    return f"FY{stichtag.year + (1 if stichtag.month > endmonat else 0)}"


def _quartal_von(stichtag: date, endmonat: int) -> int:
    """Quartalsnummer innerhalb des Geschäftsjahres, 1 bis 4."""
    versatz = (stichtag.month - endmonat - 1) % 12
    return versatz // 3 + 1


# ---- öffentliche API -----------------------------------------------------

def lies_myob_export(bilanz: list[str], guv: list[str],
                     entity: str = "Single-Entity"
                     ) -> tuple[NormalizedLedger, Periodenraster, MyobDiagnose]:
    """Bilanz- und GuV-Dateien zu einem Ledger mit Quartals- und Jahresspalten.

    Die Trennung der Dateien ist keine Formalie: die Bilanz wird als Bestand
    gelesen, die GuV als Bewegung. Wer beide über denselben Weg einliest,
    bekommt in einem der beiden Fälle die falsche Spalte.
    """
    diag = MyobDiagnose(dateien=[os.path.basename(p) for p in bilanz + guv])
    bs = [b for p in bilanz for b in _lies_datei(p, diag)]
    pl = [b for p in guv for b in _lies_datei(p, diag)]
    if not bs:
        raise ValueError("Keine lesbaren Bilanzblätter gefunden.")
    endmonat = _gj_endmonat(bs)

    raster = Periodenraster()
    # Jahresblatt -> Jahresspalte, Quartalsblatt -> Quartalsspalte.
    for blatt in sorted(bs + pl, key=lambda b: b.bis):
        jahr = _jahr_von(blatt.bis, endmonat)
        if blatt.ist_jahr:
            label = jahr
            if jahr not in raster.jahre:
                raster.jahre.append(jahr)
        else:
            label = f"{jahr} Q{_quartal_von(blatt.bis, endmonat)}"
            je_jahr = raster.quartale_je_jahr.setdefault(jahr, [])
            if label not in je_jahr:
                je_jahr.append(label)
        raster.stichtage[label] = blatt.bis
        raster.blaetter.setdefault(label, blatt.name)

    for jahr, quartale in raster.quartale_je_jahr.items():
        if len(quartale) != 4:
            raster.quartale_je_jahr[jahr] = sorted(quartale)
            diag.unvollstaendige_jahre[jahr] = len(quartale)
            diag.warnungen.append(
                f"{jahr} führt {len(quartale)} statt vier Quartale "
                f"({', '.join(sorted(quartale))}). Die Jahresspalte der GuV "
                f"deckt damit nur diesen Zeitraum ab.")

    accounts = _baue_accounts(bs, pl, raster, endmonat, entity, diag)
    _pruefe(bs, pl, raster, endmonat, diag)

    ledger = NormalizedLedger(
        accounts=accounts, perioden=list(raster.spalten), entity=entity,
        quelle_datei=", ".join(diag.dateien), hat_kontennachweis=False,
        fingerprint=fingerprint(bilanz[0]) if bilanz else "",
        warnungen=list(diag.warnungen))
    return ledger, raster, diag


def _baue_accounts(bs, pl, raster, endmonat, entity, diag) -> list[Account]:
    """Ein Account je Kontonummer, mit einem Saldo für JEDE Spalte.

    Ein Konto, das in einer Periode nicht vorkommt, trägt dort null und nicht
    etwa gar keinen Eintrag — sonst rutschen die Spalten gegeneinander.
    """
    stamm: dict[str, tuple[str, str, str]] = {}
    werte: dict[str, dict[str, float]] = collections.defaultdict(dict)

    for blatt, ist_bilanz in [(b, True) for b in bs] + [(b, False) for b in pl]:
        jahr = _jahr_von(blatt.bis, endmonat)
        label = (jahr if blatt.ist_jahr
                 else f"{jahr} Q{_quartal_von(blatt.bis, endmonat)}")
        for konto, (bez, gruppe, klasse, bestand, bewegung) in blatt.konten.items():
            stamm.setdefault(konto, (bez, gruppe, klasse))
            # Bilanz: Bestand. GuV: Bewegung — ``Year`` ist dort kumuliert.
            werte[konto][label] = bestand if ist_bilanz else bewegung

    spalten = raster.spalten
    accounts: list[Account] = []
    for konto in sorted(stamm):
        bez, gruppe, klasse = stamm[konto]
        typ, frist = KLASSEN.get(klasse, (None, ""))
        salden = tuple(PeriodBalance(p, round(werte[konto].get(p, 0.0), 2))
                       for p in spalten)
        accounts.append(Account(konto=konto, bezeichnung=bez, salden=salden,
                                entity=entity, kontotyp=typ, gruppe=gruppe,
                                fristigkeit=frist))
        if klasse and klasse not in KLASSEN:
            diag.warnungen.append(
                f"Konto {konto}: unbekannte ClassDescription '{klasse}' — "
                f"weder Bilanzseite noch GuV zuordenbar.")
    diag.kontozeilen = {p: sum(1 for a in accounts if a.saldo(p)) for p in spalten}
    return accounts


def _pruefe(bs, pl, raster, endmonat, diag) -> None:
    """Drei Proben, die den Reader rechnerisch belegen.

    Sie prüfen nicht die Zahlen des Mandanten, sondern das Einlesen: eine
    vertauschte Spalte, ein übersehenes Blatt oder eine verlorene Kostenstelle
    fällt hier auf und nicht erst im fertigen Databook.
    """
    for blatt in bs:
        jahr = _jahr_von(blatt.bis, endmonat)
        label = (jahr if blatt.ist_jahr
                 else f"{jahr} Q{_quartal_von(blatt.bis, endmonat)}")
        diag.bilanzidentitaet[label] = round(
            sum(v[3] for v in blatt.konten.values()), 2)

    # GuV: die Summe der Quartalsbewegungen muss den Jahreswert treffen.
    je_jahr_quartal: dict[str, float] = collections.defaultdict(float)
    je_jahr_ganz: dict[str, float] = {}
    for blatt in pl:
        jahr = _jahr_von(blatt.bis, endmonat)
        summe = round(sum(v[4] for v in blatt.konten.values()), 2)
        if blatt.ist_jahr:
            je_jahr_ganz[jahr] = summe
        else:
            je_jahr_quartal[jahr] += summe
    for jahr, ganz in je_jahr_ganz.items():
        diag.guv_aufriss[jahr] = (round(je_jahr_quartal.get(jahr, 0.0), 2), ganz)

    # Bilanz: das letzte Quartal eines Jahres muss das Jahresblatt treffen.
    letzte: dict[str, float] = {}
    ganze: dict[str, float] = {}
    for blatt in bs:
        jahr = _jahr_von(blatt.bis, endmonat)
        summe = round(sum(abs(v[3]) for v in blatt.konten.values()), 2)
        if blatt.ist_jahr:
            ganze[jahr] = summe
        elif _quartal_von(blatt.bis, endmonat) == 4:
            letzte[jahr] = summe
    for jahr, ganz in ganze.items():
        if jahr in letzte:
            diag.bilanz_jahresende[jahr] = (letzte[jahr], ganz)
