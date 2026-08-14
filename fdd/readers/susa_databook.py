"""Reader für die Databook-SuSa (Projekt Kitchenstories / AJNS New Media).

Ein Excel-Tab "Trial Balance (Susa)" mit DATEV-SKR03-Konten, vier
Berichtsspalten (FY2022, FY2023, FY2024, YTD Jul25) und einer Monatsmatrix.

Drei Eigenheiten der Datei, die der Reader ausdrücklich abfängt:

1. **Der Kontenblock endet vor den Nebenrechnungen.** Auf die letzte Kontozeile
   folgt eine Kontrollzeile ("Trial Balance"), die je Spalte auf null summiert;
   darunter stehen freie Nebenrechnungen, in denen in Spalte A erneut
   Kontonummern auftauchen (Delta-Zeilen, Eigenkapitalentwicklung). Gelesen
   wird nur bis zur Kontrollzeile; ``ist_ausgeglichen`` prüft danach, dass
   jede Berichtsspalte über alle Konten null ergibt.

2. **Die Monatsspalten sind kumulierte Salden, keine Periodenwerte.** Die
   FY-Spalte entspricht dem Dezemberwert. Der Reader nimmt deshalb die
   FY-Spalten ab und rührt die Monatsmatrix nicht an; wer Monatswerte braucht,
   ruft ``monatsdelta`` und bekommt echte Bewegungen.

3. **Kontonummern kommen doppelt vor**, weil Konten umbenannt wurden. Tragen
   die beiden Zeilen ihre Werte in disjunkten Perioden, werden sie
   konsolidiert; überschneiden sie sich, wandert der Fall in die Review-Queue
   statt still addiert zu werden.

Spalte E ("Adjusted Categorization") ist eine frühere **menschliche**
Klassifizierung in fremder Syntax. Sie wird bewusst NICHT als Eingabe für das
Mapping benutzt, sondern nur mitgeführt, damit der Benchmark-Tab die eigene
Klassifizierung dagegenstellen kann.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint

TAB = "Trial Balance (Susa)"

#: Die vier Berichtsspalten. LTM/YTD-Vorjahresspalten sind abgeleitete
#: Kennzahlen und gehören nicht in den Ledger.
BERICHTSSPALTEN = ("FY2022", "FY2023", "FY2024", "YTD Jul25")

#: Statistikkonten der DATEV-Nebenrechnung. Sie saldieren gegeneinander auf
#: null, dürfen aber keine Bilanzposition erzeugen.
_STATISTIK = {"9140 0", "9199 0"}

#: Bewertungskorrektur (EWB) zu den Forderungen aus L+L. Gehört als
#: Korrekturzeile unter die Forderungen, nicht als eigene Position.
_EWB = "9960 0"
_EWB_PFAD = ("/Aktiva/B Umlaufvermoegen/II Forderungen und sonstige "
             "Vermoegensgegenstaende/Forderungen aus Lieferungen und Leistungen")

_MONAT_RE = re.compile(r"^(20\d{2})/(\d{2})$")


@dataclass
class Duplikat:
    """Zwei Zeilen mit derselben Kontonummer."""

    konto: str
    zeilen: list[int]
    bezeichnungen: list[str]
    ueberschneidung: list[str] = field(default_factory=list)

    @property
    def konsolidierbar(self) -> bool:
        return not self.ueberschneidung


@dataclass
class SusaDiagnose:
    kontozeilen: int
    spaltensummen: dict[str, float]
    duplikate: list[Duplikat]
    kontrollzeile: Optional[int] = None
    #: konto -> {periode: kumulierter Monatswert}
    monate: dict[str, dict[str, float]] = field(default_factory=dict)
    #: konto -> menschliche Klassifizierung aus Spalte E
    benchmark: dict[str, str] = field(default_factory=dict)

    @property
    def ist_ausgeglichen(self) -> bool:
        return all(abs(v) < 0.01 for v in self.spaltensummen.values())

    def monatsdelta(self, konto: str) -> dict[str, float]:
        """Echte Monatsbewegungen aus der kumulierten Reihe. Innerhalb eines
        Jahres ist die Bewegung die Differenz zum Vormonat; der Januarwert ist
        bei GuV-Konten bereits die Periodenbewegung, weil die Kumulation zum
        Jahreswechsel neu ansetzt."""
        reihe = self.monate.get(konto, {})
        out, vorjahr, vormonat = {}, None, 0.0
        for p in sorted(reihe):
            jahr = p[:4]
            if jahr != vorjahr:
                vormonat, vorjahr = 0.0, jahr
            out[p] = reihe[p] - vormonat
            vormonat = reihe[p]
        return out


def _zahl(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", ""))
    except ValueError:
        return 0.0


class SusaDatabookReader(Reader):
    name = "susa_databook"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            if TAB not in wb.sheetnames:
                return False
            kopf = {str(c) for c in next(wb[TAB].iter_rows(
                min_row=1, max_row=1, values_only=True)) if c}
            return {"FY2022", "FY2023"} <= kopf
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        ledger, _ = self.lesen_mit_diagnose(pfad)
        return ledger

    def lesen_mit_diagnose(self, pfad: str) -> tuple[NormalizedLedger, SusaDiagnose]:
        wb = openpyxl.load_workbook(pfad, data_only=True)
        ws = wb[TAB]
        warnungen: list[str] = []

        kopf = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)
                if ws.cell(1, c).value is not None}
        spalten = {p: kopf[p] for p in BERICHTSSPALTEN if p in kopf}
        monatsspalten = {k: c for k, c in kopf.items() if _MONAT_RE.match(k)}
        sp_benchmark = kopf.get("Adjusted Categorization")

        ende = self._kontrollzeile(ws)
        rohzeilen = self._lies_zeilen(ws, ende, spalten, monatsspalten, sp_benchmark)

        gruppen: dict[str, list[dict]] = {}
        for z in rohzeilen:
            gruppen.setdefault(z["konto"], []).append(z)

        duplikate = [self._pruefe_duplikat(k, zs, spalten)
                     for k, zs in gruppen.items() if len(zs) > 1]

        accounts: list[Account] = []
        monate: dict[str, dict[str, float]] = {}
        benchmark: dict[str, str] = {}
        for konto, zs in gruppen.items():
            dup = next((d for d in duplikate if d.konto == konto), None)
            strittig = dup is not None and not dup.konsolidierbar
            salden = {p: sum(z["werte"].get(p, 0.0) for z in zs) for p in spalten}
            bez = self._bezeichnung(zs)
            fs_pfad, kontotyp = self._sonderrolle(konto)
            if strittig:
                warnungen.append(
                    f"Konto {konto} kommt {len(zs)} mal vor und die Werte "
                    f"überschneiden sich in {', '.join(dup.ueberschneidung)} "
                    f"({' / '.join(d[:34] for d in dup.bezeichnungen)}) — nicht "
                    "konsolidiert, sondern zur Klärung in der Review-Queue.")
            elif dup is not None:
                warnungen.append(
                    f"Konto {konto} kommt {len(zs)} mal vor (Umbenennung: "
                    f"{' -> '.join(d[:34] for d in dup.bezeichnungen)}); die "
                    "Werte liegen in disjunkten Perioden und wurden konsolidiert.")
            accounts.append(Account(
                konto=konto, bezeichnung=bez, entity="AJNS New Media GmbH",
                fs_pfad=fs_pfad, kontotyp="strittig" if strittig else kontotyp,
                salden=tuple(PeriodBalance(p, salden[p]) for p in spalten)))
            for z in zs:
                for p, w in z["monate"].items():
                    monate.setdefault(konto, {})[p] = monate.get(konto, {}).get(p, 0.0) + w
                if z["benchmark"]:
                    benchmark[konto] = z["benchmark"]

        wb.close()
        summen = {p: sum(a.saldo(p) for a in accounts) for p in spalten}
        diag = SusaDiagnose(kontozeilen=len(rohzeilen), spaltensummen=summen,
                            duplikate=duplikate, kontrollzeile=ende,
                            monate=monate, benchmark=benchmark)
        if not diag.ist_ausgeglichen:
            warnungen.append(
                "Die Trial Balance summiert nicht auf null: "
                + ", ".join(f"{p}={v:,.2f}" for p, v in summen.items() if abs(v) >= 0.01)
                + " — vermutlich wurde über den Kontenblock hinaus gelesen.")

        ledger = NormalizedLedger(
            accounts=accounts, perioden=list(spalten), entity="AJNS New Media GmbH",
            quelle_datei=pfad, hat_kontennachweis=False,
            fingerprint=fingerprint(pfad), warnungen=warnungen)
        return ledger, diag

    # ---- Hilfen -----------------------------------------------------------
    @staticmethod
    def _kontrollzeile(ws) -> int:
        """Zeilennummer der Kontrollzeile. Sie schließt den Kontenblock ab —
        alles darunter sind freie Nebenrechnungen und keine Konten."""
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None and str(ws.cell(r, 3).value or "").strip():
                return r
            if ws.cell(r, 1).value is None and ws.cell(r, 2).value is None:
                return r
        return ws.max_row + 1

    @staticmethod
    def _lies_zeilen(ws, ende: int, spalten, monatsspalten, sp_benchmark) -> list[dict]:
        zeilen = []
        for r in range(2, ende):
            nr = ws.cell(r, 1).value
            if nr is None:
                continue
            roh = str(ws.cell(r, 2).value or "").strip()
            teile = roh.split(None, 1)
            sub = teile[0] if teile and teile[0].isdigit() else "0"
            bez = teile[1].strip() if len(teile) > 1 else roh
            zeilen.append({
                "zeile": r,
                "konto": f"{str(nr).strip()} {sub}",
                "bezeichnung": bez,
                "werte": {p: _zahl(ws.cell(r, c).value) for p, c in spalten.items()},
                "monate": {k: _zahl(ws.cell(r, c).value) for k, c in monatsspalten.items()},
                "benchmark": (str(ws.cell(r, sp_benchmark).value).strip()
                              if sp_benchmark and ws.cell(r, sp_benchmark).value else ""),
            })
        return zeilen

    @staticmethod
    def _pruefe_duplikat(konto: str, zs: list[dict], spalten) -> Duplikat:
        """Überschneidung heißt: mehr als eine Zeile trägt in derselben Periode
        einen Wert. Nur dann ist unklar, welcher Wert gilt."""
        ueberschneidung = [
            p for p in spalten
            if sum(1 for z in zs if abs(z["werte"].get(p, 0.0)) > 0.005) > 1]
        return Duplikat(konto=konto, zeilen=[z["zeile"] for z in zs],
                        bezeichnungen=[z["bezeichnung"] for z in zs],
                        ueberschneidung=ueberschneidung)

    @staticmethod
    def _bezeichnung(zs: list[dict]) -> str:
        """Bei Umbenennung die Bezeichnung der zuletzt bebuchten Zeile."""
        mit_wert = [z for z in zs if any(abs(v) > 0.005 for v in z["werte"].values())]
        return (mit_wert or zs)[-1]["bezeichnung"]

    @staticmethod
    def _sonderrolle(konto: str) -> tuple[Optional[str], Optional[str]]:
        if konto in _STATISTIK:
            return None, "technisch"
        if konto == _EWB:
            # EWB gehört unter die Forderungen aus L+L, nicht in eine eigene
            # Position und auch nicht in den technischen Ausschuss.
            return _EWB_PFAD, "bilanz_aktiv"
        return None, None
