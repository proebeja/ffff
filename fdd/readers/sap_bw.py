"""Reader für SAP-BW HGB-Bilanz-Export mit eingebetteter FS-Hierarchie
(Buchungskreis 4756).

Der Export trägt seine eigene HGB-Gliederung als Baum: Sektions-Knoten
(AKTB, AKTC, PASA…), Römisch-Ebenen (AKTBI, AKTCII…) und numerische
Positionsgruppen (202, 214, 251, 335…), darunter die Blattkonten
(Kontonummer gefüllt). Diese Hierarchie IST die maßgebliche Strukturquelle
(Abschlusstreue) -> ``hat_kontennachweis = True``; die Engine übernimmt den
gelieferten ``fs_pfad`` in Stufe 1.

Fallstrick (verifiziert): Nicht-Blatt-Zeilen tragen **leere Strings** ('')
statt None in den Konto-Spalten; die Positions-Überschriften sind über mehrere
Zeilen umgebrochen. Deshalb wird nicht die (fragmentierte) Anzeigeschrift
geparst, sondern die stabile numerische Positionsgruppe aus Spalte A.

Der Crosswalk unten kodiert die FS-Gruppierung DIESES Exports (DE01-Hierarchie).
Für weitere SAP-Mandanten würde er nach ``config`` wandern; als format-
spezifisches Wissen lebt er hier im Reader.
"""

from __future__ import annotations

import re
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint, parse_deutsche_zahl

# Numerische SAP-Positionsgruppe (Präfix von Spalte A) -> kanonischer HGB-Pfad.
# Reihenfolge: längere/spezifischere Präfixe zuerst prüfen.
_A = "/Aktiva/A Anlagevermoegen"
_UV = "/Aktiva/B Umlaufvermoegen"
_FORD = f"{_UV}/II Forderungen und sonstige Vermoegensgegenstaende"
_P = "/Passiva"

_NODE_CROSSWALK: list[tuple[str, str, Optional[str]]] = [
    # (Präfix, HGB-Pfad, kontotyp)
    ("202", f"{_A}/I Immaterielle Vermoegensgegenstaende/Entgeltlich erworbene Konzessionen", "bilanz_aktiv"),
    ("214", f"{_A}/II Sachanlagen/Andere Anlagen Betriebs- und Geschaeftsausstattung", "bilanz_aktiv"),
    ("244", f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen", "bilanz_aktiv"),
    ("2651", f"{_UV}/I Vorraete/Geleistete Anzahlungen", "bilanz_aktiv"),
    ("251", f"{_FORD}/Forderungen aus Lieferungen und Leistungen", "bilanz_aktiv"),
    ("253", f"{_FORD}/Forderungen gegen verbundene Unternehmen", "bilanz_aktiv"),
    ("262", f"{_FORD}/Forderungen gegen verbundene Unternehmen", "bilanz_aktiv"),
    ("265", f"{_FORD}/Sonstige Vermoegensgegenstaende", "bilanz_aktiv"),
    ("276", f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten", "bilanz_aktiv"),
    ("287", "/Aktiva/C Rechnungsabgrenzungsposten", "bilanz_aktiv"),
    ("300", f"{_P}/A Eigenkapital", "bilanz_passiv"),
    ("335", f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen", "bilanz_passiv"),
    ("348", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten aus Lieferungen und Leistungen", "bilanz_passiv"),
    ("354", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen", "bilanz_passiv"),
    ("370", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen", "bilanz_passiv"),
    ("378", f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten", "bilanz_passiv"),
    ("390", f"{_P}/D Rechnungsabgrenzungsposten", "bilanz_passiv"),
]


class SapBwReader(Reader):
    name = "sap_bw"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            for ws in wb.worksheets:
                if str(ws.title).startswith("Header"):
                    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                        if any("Bilanz/GuV-Position" == str(c).strip() for c in row if c):
                            return True
            return False
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        wb = openpyxl.load_workbook(pfad, data_only=True)
        ws = self._bilanz_sheet(wb)
        warnungen: list[str] = []

        hdr_row, cols = self._finde_spalten(ws)
        p_cur, p_prev = self._perioden(ws)
        perioden = [p_cur, p_prev]

        accounts: list[Account] = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            konto = ws.cell(r, cols["konto"]).value
            if konto is None or str(konto).strip() == "":
                continue
            konto = str(konto).strip()
            node = str(ws.cell(r, cols["node"]).value or "").strip()
            text = str(ws.cell(r, cols["text"]).value or "").strip()
            bez = self._bezeichnung(text, konto)
            fs_pfad, kontotyp = self._node_pfad(node)
            if fs_pfad is None:
                warnungen.append(f"Konto {konto}: SAP-Knoten '{node}' ohne Crosswalk — Review.")
            v_cur = parse_deutsche_zahl(ws.cell(r, cols["cur"]).value)
            v_prev = parse_deutsche_zahl(ws.cell(r, cols["prev"]).value)
            accounts.append(Account(
                konto=konto, bezeichnung=bez,
                salden=(PeriodBalance(p_cur, v_cur), PeriodBalance(p_prev, v_prev)),
                entity="Buchungskreis 4756", fs_pfad=fs_pfad, kontotyp=kontotyp,
            ))

        wb.close()
        return NormalizedLedger(
            accounts=accounts, perioden=perioden, entity="Buchungskreis 4756",
            quelle_datei=pfad, hat_kontennachweis=True,
            fingerprint=fingerprint(pfad), warnungen=warnungen,
        )

    # ---- Hilfen -----------------------------------------------------------
    @staticmethod
    def _bilanz_sheet(wb):
        for ws in wb.worksheets:
            if str(ws.title).startswith("Header"):
                return ws
        return wb.worksheets[0]

    @staticmethod
    def _finde_spalten(ws) -> tuple[int, dict[str, int]]:
        for r in range(1, 15):
            labels = {str(ws.cell(r, c).value).strip(): c
                      for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
            if "Bilanz/GuV-Position" in labels:
                return r, {
                    "node": labels.get("Bilanz/GuV-Position", 1),
                    "text": labels.get("Text Bilanz/GuV-Position", 2),
                    "konto": labels.get("Kontonummer", 5),
                    "cur": labels.get("Summe der Berichtsperiode", 6),
                    "prev": labels.get("Summe der Vergleichsperiode", 7),
                }
        return 9, {"node": 1, "text": 2, "konto": 5, "cur": 6, "prev": 7}

    @staticmethod
    def _perioden(ws) -> tuple[str, str]:
        cur, prev = "Berichtsperiode", "Vergleichsperiode"
        for r in range(1, 9):
            v = ws.cell(r, 1).value
            if v and re.match(r"^\d{4}\.\d{2}", str(v)):
                jahr = str(v)[:4]
                if cur == "Berichtsperiode":
                    cur = jahr
                else:
                    prev = jahr
        return cur, prev

    @staticmethod
    def _bezeichnung(text: str, konto: str) -> str:
        """Konto-Bezeichnung aus dem Text extrahieren: der Text beginnt oft mit
        der Kontonummer (mit/ohne führende 0/H), die wir entfernen — aber nur,
        wenn das führende Token tatsächlich die Kontonummer ist. Sonst bliebe
        z.B. bei '3M Klebeband' ein sinntragendes erstes Wort auf der Strecke."""
        t = text.strip()
        teile = t.split(None, 1)
        if len(teile) > 1 and _ist_kontonummer(teile[0], konto):
            return teile[1].strip()
        return t

    @staticmethod
    def _node_pfad(node: str) -> tuple[Optional[str], Optional[str]]:
        for praefix, pfad, kontotyp in _NODE_CROSSWALK:
            if node.startswith(praefix):
                return pfad, kontotyp
        return None, None


def _ist_kontonummer(token: str, konto: str) -> bool:
    """Ob ``token`` die Kontonummer ``konto`` bezeichnet — tolerant gegen
    führende Nullen und ein 'H'-Präfix (SAP-Anzeige vs. Feldwert)."""
    def norm(s: str) -> str:
        return s.strip().lstrip("H").lstrip("0") or "0"
    return norm(token) == norm(konto)
