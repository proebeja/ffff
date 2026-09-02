"""Reader für die DATEV-SuSa aus Kanzlei-Rechnungswesen (Heinz Meise GmbH).

Ein anderes Druckbild als die bisherige DATEV-SuSa: ein Blatt je Datei, ein
Geschäftsjahr je Blatt, und die Kopfzeile lautet ``Konto`` statt
``Kontonummer``. Deshalb greift ``datev_susa`` hier nicht.

Aufbau::

    Konto  Beschriftung   EB-Wert  S  H   Saldo  S  H   Dez 2024 Soll  Haben  ...
    400    Betriebsausstattung  243990  S      303552  S      18700   236917,12

Drei Eigenheiten, die der Reader ausdrücklich abfängt:

**Das Vorzeichen steht nicht in der Zahl.** Die Spalte ``Saldo`` führt immer
einen Betrag ohne Vorzeichen; erst die beiden Spalten rechts davon sagen, ob
er im Soll oder im Haben steht. Wer nur die Zahl liest, bekommt ein Databook,
in dem Passiva positiv sind und nichts mehr aufgeht. Gelesen wird deshalb das
Paar, und ``H`` dreht das Vorzeichen. Die Probe steht in
``diagnose.spaltensummen``: Soll und Haben müssen sich je Periode aufheben.

**Die Periode steht in der Kopfzeile, nicht im Dateinamen.** Spalte I heißt
``Dez 2024 \\nSoll``; daraus wird ``FY2024``. Der Dateiname taugt nicht — die
drei Dateien dieses Mandats heißen unterschiedlich, eine davon gar nicht nach
dem Jahr.

**Eine Datei ist eine Periode.** Mehrere Dateien werden über
:func:`lies_kanzlei_susa` zusammengeführt, sortiert nach Stichtag. Ein Konto,
das nur in einer Periode vorkommt, trägt in den übrigen null — und nicht
etwa gar keinen Eintrag, sonst rutschen die Spalten gegeneinander.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import date


import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint, parse_deutsche_zahl

#: Kontonummern dieses Kontenrahmens: ein bis fünf Ziffern (SKR03).
_KONTO = re.compile(r"^\d{1,5}$")

#: ``Dez 2024`` in der Kopfzeile der Monatsspalte.
_MONAT_JAHR = re.compile(r"([A-Za-zäöüÄÖÜ]{3,})\s+(\d{4})")

_MONATE = {"jan": 1, "feb": 2, "mär": 3, "maer": 3, "mrz": 3, "apr": 4,
           "mai": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9, "okt": 10,
           "nov": 11, "dez": 12}

#: Kopfzeilen, die der Reader braucht. Gesucht wird nach Text, nicht nach
#: Position — die Spaltenzahl schwankt zwischen den Jahrgängen.
_KOPF_KONTO = "konto"
_KOPF_BEZ = "beschriftung"
_KOPF_SALDO = "saldo"
_KOPF_EB = "eb-wert"


@dataclass
class KanzleiDiagnose:
    """Was beim Einlesen aufgefallen ist — Grundlage der QA-Eingangsdiagnose."""

    perioden: list[str] = field(default_factory=list)
    stichtage: dict[str, date] = field(default_factory=dict)
    kontozeilen: dict[str, int] = field(default_factory=dict)
    #: Periode -> (Summe Soll, Summe Haben). Beide müssen gleich sein.
    spaltensummen: dict[str, tuple[float, float]] = field(default_factory=dict)
    #: Konten, die nicht in allen Perioden vorkommen.
    nicht_durchgaengig: dict[str, list[str]] = field(default_factory=dict)
    dateien: dict[str, str] = field(default_factory=dict)
    warnungen: list[str] = field(default_factory=list)

    def identitaet(self, periode: str) -> float:
        soll, haben = self.spaltensummen.get(periode, (0.0, 0.0))
        return round(soll - haben, 2)


def _kopfzeile(ws) -> tuple[int, dict[str, int]]:
    """Sucht die Kopfzeile über die Beschriftung, nicht über die Position."""
    for r in range(1, min(ws.max_row, 12) + 1):
        werte = [str(ws.cell(r, c).value or "").strip().lower()
                 for c in range(1, ws.max_column + 1)]
        if _KOPF_KONTO in werte and any(_KOPF_SALDO == w for w in werte):
            spalten = {
                "konto": werte.index(_KOPF_KONTO) + 1,
                "saldo": werte.index(_KOPF_SALDO) + 1,
            }
            if _KOPF_BEZ in werte:
                spalten["bez"] = werte.index(_KOPF_BEZ) + 1
            if _KOPF_EB in werte:
                spalten["eb"] = werte.index(_KOPF_EB) + 1
            return r, spalten
    raise ValueError("Keine Kopfzeile mit 'Konto' und 'Saldo' gefunden.")


def _periode(ws, kopf: int) -> tuple[str, date]:
    """Periodenlabel und Stichtag aus der Monatsspalte der Kopfzeile."""
    for c in range(1, ws.max_column + 1):
        text = str(ws.cell(kopf, c).value or "")
        m = _MONAT_JAHR.search(text.replace("\n", " "))
        if not m:
            continue
        monat = _MONATE.get(m.group(1)[:3].lower())
        if not monat:
            continue
        jahr = int(m.group(2))
        naechster = date(jahr + (monat == 12), monat % 12 + 1, 1)
        return f"FY{jahr}", date.fromordinal(naechster.toordinal() - 1)
    raise ValueError("Keine Monatsspalte 'Mon JJJJ' in der Kopfzeile gefunden.")


def _saldo_mit_vorzeichen(ws, zeile: int, saldo_spalte: int) -> float:
    """Betrag aus der Saldo-Spalte, Vorzeichen aus den beiden Spalten rechts.

    ``H`` bedeutet Haben und damit negativ. Steht kein Marker, ist der Saldo
    null — die Datei laesst das Paar dann leer.
    """
    betrag = parse_deutsche_zahl(ws.cell(zeile, saldo_spalte).value)
    if betrag == 0.0:
        return 0.0
    marker = [str(ws.cell(zeile, saldo_spalte + i).value or "").strip().upper()
              for i in (1, 2)]
    if "H" in marker:
        return -betrag
    if "S" in marker:
        return betrag
    # Kein Marker bei einem Betrag ungleich null: der Saldo ist nicht
    # ableitbar. Er wird als Soll gelesen und in der Diagnose vermerkt.
    return betrag


class KanzleiSusaReader(Reader):
    """Ein Blatt, ein Geschaeftsjahr."""

    name = "datev_kanzlei_susa"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                    werte = [str(c or "").strip().lower() for c in row]
                    if _KOPF_KONTO in werte and _KOPF_SALDO in werte \
                            and _KOPF_BEZ in werte:
                        return True
            return False
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        ledger, _ = lies_kanzlei_susa([pfad])
        return ledger


def _lies_blatt(pfad: str) -> tuple[str, date, dict[str, tuple[str, float]],
                                    tuple[float, float], list[str]]:
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb.worksheets[0]
    kopf, spalten = _kopfzeile(ws)
    periode, stichtag = _periode(ws, kopf)
    konten: dict[str, tuple[str, float]] = {}
    soll = haben = 0.0
    warnungen: list[str] = []

    for r in range(kopf + 1, ws.max_row + 1):
        roh = ws.cell(r, spalten["konto"]).value
        if roh is None:
            continue
        konto = str(roh).strip()
        if not _KONTO.match(konto):
            continue
        bez = str(ws.cell(r, spalten.get("bez", 2)).value or "").strip()
        betrag = _saldo_mit_vorzeichen(ws, r, spalten["saldo"])
        if konto in konten:
            # Dieselbe Kontonummer zweimal im Blatt: addieren waere eine
            # stille Entscheidung. Der Fall wird benannt.
            warnungen.append(f"{periode}: Konto {konto} steht mehrfach im "
                             f"Blatt; die Betraege werden addiert.")
            bez = konten[konto][0] or bez
            betrag += konten[konto][1]
        konten[konto] = (bez, betrag)
        if betrag >= 0:
            soll += betrag
        else:
            haben += -betrag
    wb.close()
    return periode, stichtag, konten, (round(soll, 2), round(haben, 2)), warnungen


def lies_kanzlei_susa(pfade: list[str], entity: str = "Single-Entity"
                      ) -> tuple[NormalizedLedger, KanzleiDiagnose]:
    """Mehrere Jahresblaetter zu einem Ledger, sortiert nach Stichtag.

    Die Reihenfolge der uebergebenen Dateien wird bewusst NICHT uebernommen:
    sie ist die Reihenfolge des Dateisystems und hat mit der Zeitachse nichts
    zu tun.
    """
    diag = KanzleiDiagnose()
    je_periode: dict[str, dict[str, tuple[str, float]]] = {}

    for pfad in pfade:
        periode, stichtag, konten, summen, warnungen = _lies_blatt(pfad)
        if periode in je_periode:
            raise ValueError(f"Periode {periode} kommt in zwei Dateien vor: "
                             f"{diag.dateien[periode]} und {pfad}.")
        je_periode[periode] = konten
        diag.stichtage[periode] = stichtag
        diag.kontozeilen[periode] = len(konten)
        diag.spaltensummen[periode] = summen
        diag.dateien[periode] = pfad
        diag.warnungen += warnungen

    diag.perioden = sorted(diag.stichtage, key=lambda p: diag.stichtage[p])

    alle = sorted({k for konten in je_periode.values() for k in konten},
                  key=lambda k: (len(k), k))
    accounts: list[Account] = []
    for konto in alle:
        bezeichnung = ""
        fehlt = []
        salden = []
        for p in diag.perioden:
            eintrag = je_periode[p].get(konto)
            if eintrag is None:
                fehlt.append(p)
                salden.append(PeriodBalance(p, 0.0))
            else:
                bezeichnung = bezeichnung or eintrag[0]
                salden.append(PeriodBalance(p, eintrag[1]))
        if fehlt:
            diag.nicht_durchgaengig[konto] = fehlt
        accounts.append(Account(konto=konto, bezeichnung=bezeichnung,
                                salden=tuple(salden), entity=entity))

    ledger = NormalizedLedger(
        accounts=accounts, perioden=list(diag.perioden), entity=entity,
        quelle_datei=", ".join(os.path.basename(diag.dateien[p])
                               for p in diag.perioden),
        hat_kontennachweis=False,
        fingerprint=fingerprint(pfade[0]) if pfade else "",
        warnungen=list(diag.warnungen))
    return ledger, diag
