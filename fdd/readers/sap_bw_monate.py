"""Reader für den monatlich geschnittenen SAP-BW-Export (Buchungskreis 4756,
Hierarchie DE01) — eine Datei je Geschäftsjahr, ein Tab je Monatsscheibe.

**Die Monatstabs sind kumulierte Jahresscheiben, keine Einzelmonate.** Tab
"1-2" führt in der Berichtsperiode den Stand 01.01–31.01 und in der
Vergleichsperiode den Stand 01.01–28.02; Tab "2-3" beginnt wieder bei
01.01–28.02. Aufeinanderfolgende Tabs überlappen sich also um genau eine
Spalte. Wer die Tabs aufsummierte, vervielfachte das Jahr — die Zusammen-
führung nimmt deshalb die kumulierte Reihe ab und prüft die Überlappung als
Kontrolle (``ueberlappungs_bruch``). Der Jahreswert ist die
Vergleichsperiode des letzten Tabs (schließt die Sonderperioden 13–16 ein).

Die eingebettete FS-Hierarchie ist die Strukturquelle (``fs_pfad`` gesetzt,
``hat_kontennachweis = True``): Der Export trägt seine HGB-Gliederung als
Baum aus Sektions-, Buchstaben-, Römisch- und numerischen Positionsknoten.
Geparst wird — wie im Schwesterreader ``sap_bw`` — die stabile numerische
Positionsgruppe aus Spalte A, nicht die über mehrere Zeilen umgebrochene
Anzeigeschrift.

Der Knoten ``AUS`` ("ausgesonderte Konten, nicht für diese Bilanz und GuV
relevant") trägt IFRS-Brutto- und Steuerbilanzkonten. Sie gehören nicht ins
HGB-Databook und werden als technisch markiert.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

from ..core.model import Account, NormalizedLedger, PeriodBalance
from .base import Reader, fingerprint, parse_deutsche_zahl

# Numerische SAP-Positionsgruppe -> kanonischer HGB-Pfad. Deckt die
# DE01-Hierarchie dieses Buchungskreises für Bilanz UND GuV ab; die Pfade
# folgen dem Vokabular der Hausconvention (reklassifizierung).
_A = "/Aktiva/A Anlagevermoegen"
_UV = "/Aktiva/B Umlaufvermoegen"
_FORD = f"{_UV}/II Forderungen und sonstige Vermoegensgegenstaende"
_P = "/Passiva"
_G = "/GuV"

_NODE_CROSSWALK: list[tuple[str, str, Optional[str]]] = [
    # ---- Bilanz ----------------------------------------------------------
    ("202", f"{_A}/I Immaterielle Vermoegensgegenstaende/Entgeltlich erworbene Konzessionen", "bilanz_aktiv"),
    ("214", f"{_A}/II Sachanlagen/Andere Anlagen Betriebs- und Geschaeftsausstattung", "bilanz_aktiv"),
    ("244", f"{_UV}/I Vorraete/Unfertige Erzeugnisse und Leistungen", "bilanz_aktiv"),
    ("248", f"{_UV}/I Vorraete/Fertige Erzeugnisse und Waren", "bilanz_aktiv"),
    ("2651", f"{_UV}/I Vorraete/Geleistete Anzahlungen", "bilanz_aktiv"),
    ("251", f"{_FORD}/Forderungen aus Lieferungen und Leistungen", "bilanz_aktiv"),
    ("253", f"{_FORD}/Forderungen gegen verbundene Unternehmen", "bilanz_aktiv"),
    ("262", f"{_FORD}/Forderungen gegen verbundene Unternehmen", "bilanz_aktiv"),
    ("265", f"{_FORD}/Sonstige Vermoegensgegenstaende", "bilanz_aktiv"),
    ("276", f"{_UV}/IV Kassenbestand und Guthaben bei Kreditinstituten", "bilanz_aktiv"),
    ("287", "/Aktiva/C Rechnungsabgrenzungsposten", "bilanz_aktiv"),
    ("300", f"{_P}/A Eigenkapital", "bilanz_passiv"),
    ("318", f"{_P}/A Eigenkapital", "bilanz_passiv"),
    ("335", f"{_P}/B Rueckstellungen/Sonstige Rueckstellungen", "bilanz_passiv"),
    ("342", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber Kreditinstituten", "bilanz_passiv"),
    ("348", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten aus Lieferungen und Leistungen", "bilanz_passiv"),
    ("354", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen", "bilanz_passiv"),
    ("370", f"{_P}/C Verbindlichkeiten/Verbindlichkeiten gegenueber verbundenen Unternehmen", "bilanz_passiv"),
    ("378", f"{_P}/C Verbindlichkeiten/Sonstige Verbindlichkeiten", "bilanz_passiv"),
    ("390", f"{_P}/D Rechnungsabgrenzungsposten", "bilanz_passiv"),
    # ---- GuV -------------------------------------------------------------
    ("004", f"{_G}/Umsatzerloese", "guv"),
    ("006", f"{_G}/Bestandsveraenderung", "guv"),
    ("008", f"{_G}/Bestandsveraenderung", "guv"),
    ("012", f"{_G}/Sonstige betriebliche Ertraege", "guv"),
    ("020", f"{_G}/Materialaufwand", "guv"),
    ("022", f"{_G}/Materialaufwand", "guv"),
    ("024", f"{_G}/Materialaufwand", "guv"),
    ("026", f"{_G}/Personalaufwand/Loehne und Gehaelter", "guv"),
    ("027", f"{_G}/Personalaufwand/Soziale Abgaben und Altersversorgung", "guv"),
    ("034", f"{_G}/Abschreibungen", "guv"),
    ("042", f"{_G}/Sonstige betriebliche Aufwendungen", "guv"),
    ("056", f"{_G}/Sonstige Zinsen und aehnliche Ertraege", "guv"),
    ("064", f"{_G}/Zinsen und aehnliche Aufwendungen", "guv"),
    ("080", f"{_G}/Steuern vom Einkommen und vom Ertrag", "guv"),
    ("084", f"{_G}/Sonstige Steuern", "guv"),
    ("086", f"{_G}/Ertraege aus Verlustuebernahme", "guv"),
    ("088", f"{_G}/Ertraege aus Verlustuebernahme", "guv"),
    ("101", f"{_P}/A Eigenkapital", "bilanz_passiv"),
]

#: Knoten, die der Export selbst als nicht bilanz-/GuV-relevant ausweist
#: ("AUS — ausgesonderte Konten"): IFRS-Brutto- und Steuerbilanzkonten.
_AUSGESONDERT = ("999", "NICHTZUG")

_TAB_RE = re.compile(r"^\d+-\d+")
_PERIODE_RE = re.compile(r"(\d{4})\.(\d{2})\s*-\s*(\d{4})\.(\d{2})")


@dataclass
class Monatsscheibe:
    """Eine kumulierte Jahresscheibe (Stand seit Jahresbeginn)."""

    tab: str
    bis_monat: int                       # 1..16 (16 = inkl. Sonderperioden)
    salden: dict[str, float] = field(default_factory=dict)   # konto -> YTD


@dataclass
class Monatsdiagnose:
    """Ergebnis der Zusammenführung — Beleg dafür, dass die kumulierte Reihe
    lückenlos ist und der Jahreswert an ihrem Ende steht."""

    jahr: int
    scheiben: list[Monatsscheibe]
    ueberlappungs_bruch: list[str] = field(default_factory=list)

    @property
    def jahreswert(self) -> dict[str, float]:
        return self.scheiben[-1].salden if self.scheiben else {}

    def monatsdelta(self, konto: str) -> list[float]:
        """Echte Monatsbewegungen aus der kumulierten Reihe."""
        vor, out = 0.0, []
        for s in self.scheiben:
            akt = s.salden.get(konto, 0.0)
            out.append(akt - vor)
            vor = akt
        return out


class SapBwMonateReader(Reader):
    name = "sap_bw_monate"

    @classmethod
    def kann_lesen(cls, pfad: str) -> bool:
        if not pfad.lower().endswith((".xlsx", ".xlsm")):
            return False
        try:
            wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            tabs = [t for t in wb.sheetnames if _TAB_RE.match(t)]
            if len(tabs) < 2:
                return False
            ws = wb[tabs[0]]
            for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                if row and str(row[0]).strip() == "Bilanz/GuV-Position":
                    return True
            return False
        finally:
            wb.close()

    def lesen(self, pfad: str) -> NormalizedLedger:
        diag, konten, warnungen = self._lies_datei(pfad)
        periode = f"{diag.jahr}/12"
        werte = diag.jahreswert
        accounts = [
            Account(konto=k, bezeichnung=b, fs_pfad=p, kontotyp=t,
                    entity="Buchungskreis 4756",
                    salden=(PeriodBalance(periode, werte.get(k, 0.0)),))
            for k, (b, p, t) in konten.items()
        ]
        return NormalizedLedger(
            accounts=accounts, perioden=[periode], entity="Buchungskreis 4756",
            quelle_datei=pfad, hat_kontennachweis=True,
            fingerprint=fingerprint(pfad), warnungen=warnungen,
        )

    # ---- Zusammenführung der Monatstabs -----------------------------------
    @staticmethod
    def _positionen_ohne_konten(summen: dict[str, float],
                                mit_konten: dict[str, set]) -> list[tuple[str, float]]:
        """Knoten, die eine Positionssumme ausweisen, zu denen der Export aber
        keine Blattkonten liefert. Solche Positionen fehlen im Databook
        vollständig — ein stiller, materieller Datenverlust."""
        offen = []
        for node, betrag in summen.items():
            if abs(betrag) < 0.005 or node in mit_konten:
                continue
            # Nur Blattgruppen prüfen (numerische Knoten), keine Sammelebenen:
            # deren Summe kommt aus den darunterliegenden Gruppen.
            if not node.isdigit():
                continue
            if any(k != node and k.startswith(node) for k in mit_konten):
                continue
            offen.append((node, betrag))
        return sorted(offen)

    @classmethod
    def _lies_datei(cls, pfad: str):
        """Gibt (Monatsdiagnose, {konto: (bez, fs_pfad, kontotyp)}, warnungen)."""
        wb = openpyxl.load_workbook(pfad, data_only=True)
        warnungen: list[str] = []
        konten: dict[str, tuple[str, Optional[str], Optional[str]]] = {}
        scheiben: list[Monatsscheibe] = []
        jahr = 0
        vorheriger_vergleich: Optional[dict[str, float]] = None
        brueche: list[str] = []
        knoten_summen: dict[str, float] = {}
        knoten_konten: dict[str, set] = {}
        seitenwechsel: dict[str, tuple] = {}

        tabs = [t for t in wb.sheetnames if _TAB_RE.match(t)]
        for tab in sorted(tabs, key=cls._tab_sortierung):
            ws = wb[tab]
            hz = cls._kopfzeile(ws)
            if hz is None:
                warnungen.append(f"Tab '{tab}': keine Kopfzeile gefunden — übersprungen.")
                continue
            j, bis_b, bis_v = cls._perioden(ws)
            jahr = jahr or j
            ist_letzter = tab == sorted(tabs, key=cls._tab_sortierung)[-1]
            bericht, vergleich = cls._lies_tab(
                ws, hz, konten,
                knoten_summen if ist_letzter else None,
                knoten_konten if ist_letzter else None,
                seitenwechsel)

            # Überlappung: die Berichtsperiode dieses Tabs muss dem Stand
            # entsprechen, den der Vortab als Vergleichsperiode auswies.
            if vorheriger_vergleich is not None:
                for k in set(bericht) | set(vorheriger_vergleich):
                    a, b = bericht.get(k, 0.0), vorheriger_vergleich.get(k, 0.0)
                    if abs(a - b) > 0.005:
                        brueche.append(f"{tab}/{k}: {a:.2f} statt {b:.2f}")
            else:
                scheiben.append(Monatsscheibe(tab, bis_b, bericht))
            scheiben.append(Monatsscheibe(tab, bis_v, vergleich))
            vorheriger_vergleich = vergleich

        wb.close()
        for k, (alt, neu) in sorted(seitenwechsel.items()):
            warnungen.append(
                f"Konto {k} wechselt im Jahresverlauf die Bilanzseite "
                f"({str(alt).rsplit('/', 1)[-1]} -> {str(neu).rsplit('/', 1)[-1]}); "
                "maßgeblich ist der Stand am Jahresende.")

        # Integritätsprobe: die Sektionssummen, die der Export selbst
        # ausweist, gegen die Summe der tatsächlich gelieferten Konten. Jede
        # Abweichung bedeutet, dass Konten fehlen (der Export sie also
        # unterdrückt hat) — sonst rechnete das Databook stillschweigend mit
        # einer unvollständigen Position weiter.
        sektion_soll = {k: v for k, v in knoten_summen.items()
                        if k in ("BILAKTIVA", "BILPASSIVA", "GUV")}
        sektion_ist = {"BILAKTIVA": 0.0, "BILPASSIVA": 0.0, "GUV": 0.0}
        for k, (_b, pf, typ) in konten.items():
            if typ == "technisch" or not pf:
                continue
            ziel = ("BILAKTIVA" if pf.startswith("/Aktiva")
                    else "BILPASSIVA" if pf.startswith("/Passiva") else "GUV")
            sektion_ist[ziel] += (vorheriger_vergleich or {}).get(k, 0.0)
        for sekt, soll in sorted(sektion_soll.items()):
            abw = sektion_ist.get(sekt, 0.0) - soll
            if abs(abw) > 0.5:
                warnungen.append(
                    f"Sektion {sekt}: der Export weist {soll:,.2f} EUR aus, die "
                    f"gelieferten Konten summieren sich auf {sektion_ist[sekt]:,.2f} EUR "
                    f"(Differenz {abw:,.2f}). Es fehlen Kontozeilen.")
        for node, betrag in cls._positionen_ohne_konten(knoten_summen, knoten_konten):
            warnungen.append(
                f"Position '{node}' weist {betrag:,.2f} EUR aus, der Export "
                "liefert dazu aber keine einzige Kontozeile — die Position "
                "fehlt damit im Databook. Kontenaufriss nachfordern.")
        if brueche:
            warnungen.append(
                f"Kumulierte Reihe bricht an {len(brueche)} Stelle(n): "
                + "; ".join(brueche[:3]) + (" …" if len(brueche) > 3 else ""))
        return Monatsdiagnose(jahr, scheiben, brueche), konten, warnungen

    @classmethod
    def _lies_tab(cls, ws, hz: int, konten: dict, knoten_summen: dict | None = None,
                  knoten_konten: dict | None = None, seitenwechsel: dict | None = None):
        """Ein Tab: Blattkonten mit beiden Spalten. Nebenbei werden die
        ausgewiesenen Knotensummen und die Zahl der Blattkonten je Knoten
        mitgeschrieben — daraus entsteht die Deckungsprüfung (siehe
        ``_positionen_ohne_konten``)."""
        bericht: dict[str, float] = {}
        vergleich: dict[str, float] = {}

        for r in range(hz + 1, ws.max_row + 1):
            node = str(ws.cell(r, 1).value or "").strip()
            if not node:
                continue
            konto = str(ws.cell(r, 5).value or "").strip()
            if not konto:
                # Knotenzeile. Trägt sie einen Wert, ist es die ausgewiesene
                # Positionssumme des Knotens.
                if knoten_summen is not None:
                    w = ws.cell(r, 7).value
                    if isinstance(w, (int, float)):
                        # Der Export schreibt Öffnungs- und Abschlusszeile mit
                        # demselben Code; die letzte ist die Positionssumme.
                        knoten_summen[node] = float(w)
                continue
            # Die Blattzeile trägt ihre Positionsgruppe selbst in Spalte A —
            # kein Mitführen des Baums nötig, kein Verrutschen möglich.
            gruppe = node
            if knoten_konten is not None:
                knoten_konten.setdefault(node, set()).add(konto)
            # SAP-FS-Hierarchien schalten ein Konto je nach Saldovorzeichen
            # zwischen Forderungs- und Verbindlichkeitsknoten um; dasselbe
            # Konto kann im Jahresverlauf die Seite wechseln. Maßgeblich ist
            # der Stand am Jahresende, deshalb überschreibt der spätere Tab.
            text = str(ws.cell(r, 2).value or "").strip()
            fs_pfad, kontotyp = cls._node_pfad(gruppe)
            vorher = konten.get(konto)
            if vorher and vorher[1] != fs_pfad and seitenwechsel is not None:
                seitenwechsel[konto] = (vorher[1], fs_pfad)
            konten[konto] = (cls._bezeichnung(text, konto), fs_pfad, kontotyp)
            bericht[konto] = bericht.get(konto, 0.0) + parse_deutsche_zahl(ws.cell(r, 6).value)
            vergleich[konto] = vergleich.get(konto, 0.0) + parse_deutsche_zahl(ws.cell(r, 7).value)
        return bericht, vergleich

    # ---- Hilfen -----------------------------------------------------------
    @staticmethod
    def _tab_sortierung(tab: str) -> tuple[int, int]:
        m = re.match(r"^(\d+)-(\d+)", tab)
        return (int(m.group(1)), int(m.group(2))) if m else (99, 99)

    @staticmethod
    def _kopfzeile(ws) -> Optional[int]:
        for r in range(1, 20):
            if str(ws.cell(r, 1).value or "").strip() == "Bilanz/GuV-Position":
                return r
        return None

    @staticmethod
    def _perioden(ws) -> tuple[int, int, int]:
        """(Jahr, Endmonat Berichtsperiode, Endmonat Vergleichsperiode)."""
        jahr, ende = 0, []
        for r in range(1, 12):
            m = _PERIODE_RE.search(str(ws.cell(r, 1).value or ""))
            if m:
                jahr = jahr or int(m.group(1))
                ende.append(int(m.group(4)))
        while len(ende) < 2:
            ende.append(0)
        return jahr, ende[0], ende[1]

    @staticmethod
    def _bezeichnung(text: str, konto: str) -> str:
        teile = text.strip().split(None, 1)
        if len(teile) > 1 and teile[0].strip().lstrip("H").lstrip("0") == konto.lstrip("H").lstrip("0"):
            return teile[1].strip()
        return text.strip()

    @staticmethod
    def _node_pfad(node: str) -> tuple[Optional[str], Optional[str]]:
        if node.startswith(_AUSGESONDERT):
            # Der Export weist diese Konten selbst als nicht relevant aus.
            return None, "technisch"
        for praefix, pfad, kontotyp in _NODE_CROSSWALK:
            if node.startswith(praefix):
                return pfad, kontotyp
        return None, None


def lies_sap_jahre(pfade: list[str]) -> tuple[NormalizedLedger, list[Monatsdiagnose]]:
    """Mehrere Jahresdateien zu einem Ledger verschmelzen.

    Jede Datei liefert genau eine Periode (den Jahreswert am Ende ihrer
    kumulierten Reihe). Konten werden über die Kontonummer zusammengeführt;
    ein Konto, das ein Jahr nicht kennt, steht dort mit 0."""
    diagnosen: list[Monatsdiagnose] = []
    je_jahr: dict[int, dict] = {}
    je_periode: dict[str, dict[str, float]] = {}
    warnungen: list[str] = []

    for p in pfade:
        diag, konten, warn = SapBwMonateReader._lies_datei(p)
        diagnosen.append(diag)
        warnungen.extend(f"[{diag.jahr}] {w}" for w in warn)
        je_jahr[diag.jahr] = konten
        je_periode[f"{diag.jahr}/12"] = diag.jahreswert

    # Ein Konto kann seine Position über die Jahre wechseln (Vorzeichen-
    # umschaltung der FS-Hierarchie). Das Mastersheet führt einen Pfad je
    # Konto; maßgeblich ist das jüngste Jahr, denn darauf führt das Databook.
    stamm: dict[str, tuple[str, Optional[str], Optional[str]]] = {}
    for jahr in sorted(je_jahr):
        for k, v in je_jahr[jahr].items():
            vorher = stamm.get(k)
            if vorher and vorher[1] != v[1]:
                warnungen.append(
                    f"Konto {k} liegt {jahr} unter einer anderen Position als im "
                    f"Vorjahr ({str(vorher[1]).rsplit('/', 1)[-1]} -> "
                    f"{str(v[1]).rsplit('/', 1)[-1]}); das Mastersheet folgt dem "
                    "jüngsten Jahr.")
            stamm[k] = v

    perioden = sorted(je_periode, reverse=True)
    accounts = [
        Account(konto=k, bezeichnung=b, fs_pfad=pf, kontotyp=t,
                entity="Buchungskreis 4756",
                salden=tuple(PeriodBalance(p, je_periode[p].get(k, 0.0))
                             for p in perioden))
        for k, (b, pf, t) in stamm.items()
    ]
    ledger = NormalizedLedger(
        accounts=accounts, perioden=perioden, entity="Buchungskreis 4756",
        quelle_datei=" + ".join(pfade), hat_kontennachweis=True,
        fingerprint=fingerprint(pfade[0]), warnungen=warnungen,
    )
    return ledger, diagnosen
