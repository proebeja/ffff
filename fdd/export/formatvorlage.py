"""Ausgabe nach ``Formatvorlage_Lead.xlsx``.

Die Vorlage ist die Quelle des Formats, nicht die Beschreibung daneben.
Theme-Farben, Zeilenhöhen und Gliederungsebenen lassen sich aus einer
Aufzählung nicht zuverlässig rekonstruieren; deshalb entsteht die Mappe als
``shutil.copy`` der Vorlage, und jede Zeile bekommt ihren Stil per
``copy.copy(muster._style)`` **innerhalb derselben Arbeitsmappe**. Über zwei
Mappen hinweg zeigt ``_style`` in die falsche Stiltabelle und openpyxl wirft
beim Speichern einen ``IndexError``.

Der Aufbau, den die Vorlage vorgibt:

* **Spalte C deutsch, Spalte D englisch**, beide immer befüllt. Umgeschaltet
  wird über die Spaltengruppierung — C liegt auf ``outlineLevel 2`` und ist
  eingeklappt, also ist Englisch sichtbar. Nicht übersetzen, sondern
  umschalten.
* **Ticker in E und F**, eingeklappt, blau. Im Aufriss ist Ticker 1 die
  Kontonummer und Ticker 2 die Klasse; im Lead ist Ticker 1 der Aufriss-Tab.
* **Jede Kontozeile zieht per ``SUMIFS`` aus dem Mastersheet** über beide
  Ticker der eigenen Zeile. Kein Direktverweis auf ein anderes Blatt, und die
  Kriterien werden nie als Text in die Formel geschrieben — ein getippter
  Ticker mit abweichender Schreibweise liefert stumm null.
* **Alles in Tausend, die Division steht in der Formel.** Nur Zeilen, die
  unmittelbar aus dem Mastersheet ziehen, werden geteilt. Eine Summenzeile
  über bereits geteilte Zeilen darf nicht noch einmal geteilt werden — das
  ist die häufigste Falle dieser Umstellung.
"""

from __future__ import annotations

import copy
import os
import shutil
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

VORLAGE = os.path.join(os.path.dirname(__file__), "..", "vorlagen",
                       "Formatvorlage_Lead.xlsx")

#: Das Blatt, aus dem die Stile stammen. Es bleibt bis zum Schluss in der
#: Mappe: wer eine Zeile hinzufügt, braucht die Muster noch.
MUSTERBLATT = "A_Muster"

#: Spalten der Ausgabeblätter (Vorlage).
SP_RAND, SP_SCHMAL, SP_DE, SP_EN, SP_T1, SP_T2 = 1, 2, 3, 4, 5, 6
SP_ERSTE_PERIODE = 7
#: Nach den Periodenspalten: eine schmale Spalte, dann der Kommentar.
BREITE_RAND, BREITE_SCHMAL = 11.44, 0.89
BREITE_DE, BREITE_EN, BREITE_TICKER = 45.33, 37.11, 13.0
BREITE_PERIODE, BREITE_KOMMENTAR = 10.44, 48.0

#: Zeilen der Vorlage: Kopf in 6 und 7, Inhalt ab 9.
Z_KOPF1, Z_KOPF2, Z_ERSTE = 6, 7, 9

#: Zeilenhoehe. Die Anleitung nennt 11,4 — die Vorlagedatei fuehrt 11,25.
#: Massgeblich ist die Datei: sie ist die Quelle des Formats, nicht der Text
#: daneben. Gelesen wird sie deshalb aus der Musterzeile.
HOEHE_ABSTAND = 4.5

#: Spalten des Mastersheets (Vorlage). Es bleibt in EUR.
MS_KONTO, MS_BEZ_DE, MS_BEZ_EN, MS_KLASSE = 1, 2, 3, 4
MS_NA_DE, MS_NA_EN, MS_AUFRISS = 5, 6, 7
MS_ERSTE_PERIODE = 8


@dataclass(frozen=True)
class Muster:
    """Eine Musterzeile der Vorlage, aus der der Stil kopiert wird."""

    blatt: str
    zeile: int


#: Je Zeilentyp die Musterzeile. Wer das Format ändern will, ändert die
#: Vorlage — nicht diese Tabelle und schon gar nicht den Code darunter.
KOPF1 = Muster(MUSTERBLATT, 6)          # Projektzeile, dunkel hinterlegt
KOPF2 = Muster(MUSTERBLATT, 7)          # Einheit
KONTO = Muster(MUSTERBLATT, 9)          # Kontozeile, grau
SUMME = Muster(MUSTERBLATT, 11)         # Summe, fett, Linie oben
KONTROLLE = Muster(MUSTERBLATT, 13)     # Kontrollzeile, Arial 9, zwei Stellen
FUSSNOTE = Muster(MUSTERBLATT, 15)      # kursiv, grau
POSITION = Muster("Lead NA", 10)        # Lead-Position, zieht aus dem Aufriss
BLOCKSUMME = Muster("Lead NA", 25)      # fett, Linie oben und unten


def oeffne(ziel: str) -> openpyxl.Workbook:
    """Neue Mappe als Kopie der Vorlage."""
    os.makedirs(os.path.dirname(os.path.abspath(ziel)), exist_ok=True)
    shutil.copy(VORLAGE, ziel)
    return openpyxl.load_workbook(ziel)


def _muster_zelle(wb, muster: Muster, spalte: int):
    return wb[muster.blatt].cell(muster.zeile, spalte)


def stil(wb, ws, zeile: int, muster: Muster, bis: int) -> None:
    """Überträgt den Stil einer Musterzeile auf eine Zeile des Zielblatts.

    Kopiert wird Spalte für Spalte, weil die Vorlage je Spalte ein anderes
    Zahlenformat und eine andere Schriftfarbe führt: die Ticker sind blau,
    die Beträge tragen das Buchhaltungsformat, die Beschriftung nicht.
    """
    for spalte in range(1, bis + 1):
        quelle = _muster_zelle(wb, muster, min(spalte, 11))
        # Spalten jenseits der Vorlage bekommen den Stil der Periodenspalte.
        if spalte >= SP_ERSTE_PERIODE:
            quelle = _muster_zelle(wb, muster, SP_ERSTE_PERIODE)
        ws.cell(zeile, spalte)._style = copy.copy(quelle._style)
    vorbild = wb[muster.blatt].row_dimensions.get(muster.zeile)
    if vorbild is not None and vorbild.height:
        ws.row_dimensions[zeile].height = vorbild.height


def abstandszeile(ws, zeile: int) -> None:
    ws.row_dimensions[zeile].height = HOEHE_ABSTAND


def spaltenbreiten(ws, anzahl_perioden: int) -> int:
    """Breiten und Gliederung nach der Vorlage. Gibt die letzte Spalte zurück.

    Die Sprachumschaltung ist keine Übersetzung zur Laufzeit, sondern eine
    Gliederungsebene: C ist eingeklappt, D sichtbar. Beide bleiben befüllt.
    """
    ws.column_dimensions[get_column_letter(SP_RAND)].width = BREITE_RAND
    ws.column_dimensions[get_column_letter(SP_SCHMAL)].width = BREITE_SCHMAL

    de = ws.column_dimensions[get_column_letter(SP_DE)]
    de.width, de.outlineLevel, de.hidden = BREITE_DE, 2, True
    ws.column_dimensions[get_column_letter(SP_EN)].width = BREITE_EN

    for spalte in (SP_T1, SP_T2):
        t = ws.column_dimensions[get_column_letter(spalte)]
        t.width, t.outlineLevel, t.hidden = BREITE_TICKER, 1, True

    letzte = SP_ERSTE_PERIODE + anzahl_perioden - 1
    for spalte in range(SP_ERSTE_PERIODE, letzte + 1):
        ws.column_dimensions[get_column_letter(spalte)].width = BREITE_PERIODE
    ws.column_dimensions[get_column_letter(letzte + 1)].width = BREITE_SCHMAL
    ws.column_dimensions[get_column_letter(letzte + 2)].width = BREITE_KOMMENTAR

    ws.sheet_view.showGridLines = False
    # Die Position steht ÜBER ihren Konten, nicht darunter.
    ws.sheet_properties.outlinePr.summaryBelow = False
    return letzte


def kopf(wb, ws, projekt: str, titel_de: str, titel_en: str, perioden,
         zeile2_de: str, zeile2_en: str, ticker_de: str, ticker_en: str,
         letzte_spalte: int) -> None:
    """Die beiden Kopfzeilen. Der Projektname kommt von genau einer Stelle."""
    stil(wb, ws, Z_KOPF1, KOPF1, letzte_spalte + 2)
    stil(wb, ws, Z_KOPF2, KOPF2, letzte_spalte + 2)
    ws.cell(Z_KOPF1, SP_DE, f"Projekt {projekt} — {titel_de}")
    ws.cell(Z_KOPF1, SP_EN, f"Projekt {projekt} — {titel_en}")
    ws.cell(Z_KOPF1, SP_T1, "Ticker 1")
    ws.cell(Z_KOPF1, SP_T2, "Ticker 2")
    ws.cell(Z_KOPF2, SP_DE, zeile2_de)
    ws.cell(Z_KOPF2, SP_EN, zeile2_en)
    ws.cell(Z_KOPF2, SP_T1, ticker_de)
    ws.cell(Z_KOPF2, SP_T2, ticker_en)
    for i, p in enumerate(perioden):
        ws.cell(Z_KOPF1, SP_ERSTE_PERIODE + i, p)
        ws.cell(Z_KOPF2, SP_ERSTE_PERIODE + i, "IST")


def fussnoten(wb, ws, zeile: int, zeilen: list[tuple[str, str]],
              letzte_spalte: int) -> int:
    for de, en in zeilen:
        stil(wb, ws, zeile, FUSSNOTE, letzte_spalte + 2)
        ws.cell(zeile, SP_DE, de)
        ws.cell(zeile, SP_EN, en)
        zeile += 1
    return zeile


# --------------------------------------------------------------------------
# Mastersheet
# --------------------------------------------------------------------------

def schreibe_mastersheet(wb, mapped, perioden: list[str],
                         aufriss_je_konto: dict[str, str]) -> int:
    """Das Mastersheet in der Spaltenordnung der Vorlage, in EUR.

    Gibt die letzte belegte Zeile zurück — jede ``SUMIFS``-Formel der übrigen
    Blätter braucht sie als Bereichsende. Ein zu kurzer Bereich schneidet
    Konten ab, ohne dass irgendetwas rot wird.
    """
    ws = wb["Mastersheet"]
    muster_kopf = [ws.cell(1, c)._style for c in range(1, 11)]
    muster_zeile = [ws.cell(2, c)._style for c in range(1, 11)]

    letzte_spalte = MS_ERSTE_PERIODE + len(perioden) - 1
    kopfzeile = (["Konto", "Bezeichnung DE", "Bezeichnung EN", "Klasse",
                  "NA-Zeile DE", "NA-Zeile EN", "Aufriss-Tab"] + list(perioden))
    for i, titel in enumerate(kopfzeile, start=1):
        z = ws.cell(1, i, titel)
        z._style = copy.copy(muster_kopf[min(i, 10) - 1])

    zeile = 2
    for m in sorted(mapped, key=lambda x: (len(x.konto), x.konto)):
        werte = [m.konto, m.bezeichnung, m.bezeichnung, m.klasse.value,
                 m.na_de, m.na_en, aufriss_je_konto.get(m.konto, "")]
        for i, wert in enumerate(werte, start=1):
            z = ws.cell(zeile, i, wert)
            z._style = copy.copy(muster_zeile[min(i, 10) - 1])
        for i, p in enumerate(perioden):
            z = ws.cell(zeile, MS_ERSTE_PERIODE + i, round(m.saldo(p), 2))
            z._style = copy.copy(muster_zeile[9])
        zeile += 1

    # Reste der Vorlagenzeilen entfernen: die Muster standen bis Zeile 8.
    for r in range(zeile, ws.max_row + 1):
        for c in range(1, letzte_spalte + 1):
            ws.cell(r, c).value = None
    ws.freeze_panes = "B2"
    return zeile - 1


# --------------------------------------------------------------------------
# Aufriss-Tab
# --------------------------------------------------------------------------

def _sumifs_konto(ms_zeilen: int, periodenspalte: int, zeile: int) -> str:
    """Kontozeile: beide Kriterien aus den Ticker-Spalten derselben Zeile."""
    w = get_column_letter(MS_ERSTE_PERIODE + periodenspalte)
    k = get_column_letter(MS_KONTO)
    kl = get_column_letter(MS_KLASSE)
    t1 = get_column_letter(SP_T1)
    t2 = get_column_letter(SP_T2)
    return (f"=SUMIFS(Mastersheet!${w}$2:${w}${ms_zeilen},"
            f"Mastersheet!${k}$2:${k}${ms_zeilen},${t1}{zeile},"
            f"Mastersheet!${kl}$2:${kl}${ms_zeilen},${t2}{zeile})/1000")


def _sumifs_aufriss(ms_zeilen: int, periodenspalte: int, blatt: str) -> str:
    """Kontrolle: dieselbe Summe über die Aufriss-Spalte des Mastersheets.

    Ein zweiter, unabhängiger Weg — er nennt kein einziges Konto, sondern das
    Merkmal. Beide müssen dieselbe Zahl liefern.
    """
    w = get_column_letter(MS_ERSTE_PERIODE + periodenspalte)
    a = get_column_letter(MS_AUFRISS)
    return (f"SUMIFS(Mastersheet!${w}$2:${w}${ms_zeilen},"
            f"Mastersheet!${a}$2:${a}${ms_zeilen},\"{blatt}\")/1000")


def schreibe_aufriss(wb, blatt: str, titel_de: str, titel_en: str, konten,
                     perioden: list[str], ms_zeilen: int,
                     projekt: str) -> dict:
    """Ein Aufriss-Tab nach der Vorlage.

    ``konten`` ist eine Liste von ``(kontonummer, klasse, bezeichnung_de,
    bezeichnung_en)``. Die Ticker werden aus dem Mastersheet übernommen und
    nirgends getippt.
    """
    ws = wb.create_sheet(blatt)
    letzte = spaltenbreiten(ws, len(perioden))
    kopf(wb, ws, projekt, f"Aufriss {titel_de}", f"Breakdown {titel_en}",
         perioden, "in TEUR · Einzelkonten je Periode",
         "in kEUR · accounts by period", "Konto", "Klasse", letzte)

    zeile = Z_ERSTE
    erste_konto = zeile
    for konto, klasse, bez_de, bez_en in konten:
        stil(wb, ws, zeile, KONTO, letzte + 2)
        ws.cell(zeile, SP_DE, f"{konto} {bez_de}")
        ws.cell(zeile, SP_EN, f"{konto} {bez_en}")
        ws.cell(zeile, SP_T1, konto)
        ws.cell(zeile, SP_T2, klasse)
        for i in range(len(perioden)):
            ws.cell(zeile, SP_ERSTE_PERIODE + i,
                    _sumifs_konto(ms_zeilen, i, zeile))
        zeile += 1
    letzte_konto = zeile - 1

    summenzeile = zeile
    stil(wb, ws, summenzeile, SUMME, letzte + 2)
    ws.cell(summenzeile, SP_DE, "Summe Aufriss")
    ws.cell(summenzeile, SP_EN, "Total breakdown")
    for i in range(len(perioden)):
        sp = get_column_letter(SP_ERSTE_PERIODE + i)
        # NICHT noch einmal durch 1.000 teilen: die Kontozeilen sind es schon.
        ws.cell(summenzeile, SP_ERSTE_PERIODE + i,
                f"=SUM({sp}{erste_konto}:{sp}{letzte_konto})")
    zeile += 1
    abstandszeile(ws, zeile)
    zeile += 1

    kontrollzeile = zeile
    stil(wb, ws, kontrollzeile, KONTROLLE, letzte + 2)
    ws.cell(kontrollzeile, SP_DE,
            "Kontrolle (Aufriss ./. Mastersheet je Aufriss-Tab)")
    ws.cell(kontrollzeile, SP_EN,
            "Check (breakdown less Mastersheet by breakdown tab)")
    for i in range(len(perioden)):
        sp = get_column_letter(SP_ERSTE_PERIODE + i)
        ws.cell(kontrollzeile, SP_ERSTE_PERIODE + i,
                f"={sp}{summenzeile}-" + _sumifs_aufriss(ms_zeilen, i, blatt))
    zeile += 2

    zeile = fussnoten(wb, ws, zeile, [
        ("Kontozeilen ziehen per SUMIFS aus dem Mastersheet über Kontonummer "
         "und Klasse.",
         "Account rows pull from the Mastersheet by account number and class."),
        ("Ticker 1 ist die Kontonummer, Ticker 2 die Klasse. Beide Spalten "
         "sind eingeklappt.",
         "Ticker 1 is the account number, Ticker 2 the class. Both columns "
         "are collapsed."),
        ("Kein Direktverweis auf andere Blätter. Nur das Mastersheet ist "
         "Datenquelle.",
         "No direct links to other sheets. The Mastersheet is the only data "
         "source."),
        ("Das Mastersheet führt EUR. Die Division durch 1.000 steht in der "
         "Formel, nicht im Zahlenformat.",
         "The Mastersheet holds EUR. The division by 1,000 sits in the "
         "formula, not the number format."),
    ], letzte)

    ws.freeze_panes = ws.cell(Z_ERSTE, SP_ERSTE_PERIODE)
    return {"blatt": blatt, "erste_konto": erste_konto,
            "letzte_konto": letzte_konto, "summenzeile": summenzeile,
            "kontrollzeile": kontrollzeile, "konten": [k[0] for k in konten]}
