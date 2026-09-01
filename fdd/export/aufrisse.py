"""Option B: die Positionssumme kommt aus dem Aufriss, nicht aus dem Mastersheet.

``databook_architektur.option_b`` verlangt drei Schichten — Mastersheet,
Aufriss-Tabs, Lead-Tabs — und dreht damit die Richtung um, in der die Vorlage
gebaut ist. In der Vorlage zieht der Aufriss aus dem Lead und stellt seine
eigene Summe dagegen; in Option B zieht der **Lead aus dem Aufriss**, und die
Kontrolle läuft über die eingeklappten Kontozeilen, die weiterhin per
``SUMIFS`` am Mastersheet hängen. Das sind zwei unabhängige Wege auf dieselbe
Zahl, und genau darin liegt der Sinn der dritten Schicht.

Nicht jeder Aufriss der Vorlage kann eine Position tragen:

* ``NA_TWC`` und ``NA_Net Debt`` sind **Sammel-Tabs**. Sie holen jede Zeile
  aus dem Lead NA und ergänzen Normalisierungen. Sie zu invertieren hieße, sie
  auf sich selbst zeigen zu lassen.
* ``NA_Ford LuL`` und ``NA_Verb LuL`` gliedern nach **Fälligkeit**. Eine
  Saldenliste kennt keine Fälligkeiten. Ohne offene-Posten-Liste bliebe die
  Altersstruktur leer, und ein Lead, der aus einem leeren Aufriss zieht, zeigt
  null. Diese Positionen bleiben deshalb auf dem Weg aus dem Mastersheet — was
  fehlt, ist eine Datenanforderung und kein Mangel des Laufs.
* ``NA_Vorraete`` und ``NA_Sachanlagen`` gliedern nach Bestandteilen, und die
  stehen im Kontenplan. Sie werden befüllt und tragen ihre Position.

Die Zuordnung Konto -> Aufrisszeile ist eine fachliche Entscheidung je Mandat
und steht deshalb im Runner, nicht hier. Dieses Modul kennt nur die Mechanik.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

#: ``=Cockpit!D39`` -> Zähler 1. Spalte C ist die Eröffnungsspalte (Zähler 0).
_COCKPIT = re.compile(r"=Cockpit!\$?([A-Z]+)\$?39\b")


@dataclass
class Aufrisszeile:
    """Eine Zeile des Aufrisses und die Konten, die sie trägt."""

    zeile: int
    label_de: str
    label_en: str
    konten: list[str] = field(default_factory=list)


@dataclass
class Aufrissplan:
    """Ein Aufriss-Tab, seine Zeilen und die Lead-Position, die er trägt."""

    blatt: str
    #: Zeile im Lead NA, die künftig auf die Summe dieses Aufrisses zeigt.
    lead_zeile: int
    #: Zeile im Aufriss, die die Summe der Position führt.
    summenzeile: int
    zeilen: list[Aufrisszeile]
    hinweis: str = ""


@dataclass
class Aufrissbefund:
    """Kontrolle je Position: Aufriss gegen die Kontozeilen."""

    position: str
    blatt: str
    lead_zeile: int
    aufriss: dict[str, float]
    konten: dict[str, float]
    konten_ohne_zeile: list[str] = field(default_factory=list)

    def differenz(self, periode: str) -> float:
        return round(self.aufriss.get(periode, 0.0)
                     - self.konten.get(periode, 0.0), 2)

    @property
    def ok(self) -> bool:
        return (not self.konten_ohne_zeile
                and all(abs(self.differenz(p)) <= 1.0 for p in self.aufriss))


def periodenspalten(ws) -> dict[int, int]:
    """Zähler der Zeitachse -> Spaltennummer in diesem Aufriss.

    Jeder Aufriss hat eine eigene Spaltenzahl und einen eigenen Anfang: die
    Vorräte beginnen in F mit Zähler 1, die Sachanlagen in E, das TWC in E mit
    Zähler 0. Abgelesen wird die Kopfzeile, nicht geraten.
    """
    spalten: dict[int, int] = {}
    for c in range(1, ws.max_column + 1):
        wert = ws.cell(4, c).value
        if isinstance(wert, str):
            m = _COCKPIT.fullmatch(wert.strip())
            if m:
                # Spalte C des Cockpits ist Zähler 0.
                zaehler = sum((ord(z) - 64) * 26 ** i
                              for i, z in enumerate(reversed(m.group(1)))) - 3
                spalten[zaehler] = c
    return spalten


def _sumifs(ms_spalte: int, konten: list[str]) -> str:
    """Sichtbare Formel: Summe der genannten Konten aus dem Mastersheet.

    Bewusst dieselbe Form wie in den Kontoslots der Lead-Tabs — ein Aufriss,
    der seine Zahlen anders holt als die Kontozeile daneben, wäre keine
    unabhängige Kontrolle, sondern eine zweite Quelle.
    """
    sp = get_column_letter(ms_spalte)
    teile = [f'SUMIFS(Mastersheet!${sp}$2:${sp}$400,'
             f'Mastersheet!$A$2:$A$400,"{k}")' for k in konten]
    return "=(" + "+".join(teile) + ")/1000"


def schreibe_aufrisse(wb, plaene: list[Aufrissplan], ach, ms_spalte,
                      lead_erste: int, zeilen_je_position: dict,
                      perioden: list[str]) -> tuple[int, list[Aufrissbefund]]:
    """Befüllt die Aufrisse und hängt die Lead-Positionen an ihre Summe.

    ``zeilen_je_position`` bildet die Lead-Zeile auf die Mastersheet-Zeilen
    der Position ab. Daraus entsteht die Pflicht-Kontrolle: die Summe der
    Kontozeilen gegen die Summe des Aufrisses. Ein Konto, das in keiner
    Aufrisszeile steht, fällt dabei auf — im Lead sähe man es nicht, weil die
    Positionszeile dann aus dem Aufriss kommt und das Konto schlicht fehlt.
    """
    lead = wb["Lead NA"]
    geschrieben = 0
    befunde: list[Aufrissbefund] = []

    for plan in plaene:
        ws = wb[plan.blatt]
        spalten = periodenspalten(ws)

        for zeile in plan.zeilen:
            if zeile.label_de:
                ws.cell(zeile.zeile, 3, zeile.label_de)
                ws.cell(zeile.zeile, 4, zeile.label_en or zeile.label_de)
                geschrieben += 2
            for periode in perioden:
                spalte = spalten.get(ach._index(periode, guv=False))
                if spalte is None or not zeile.konten:
                    continue
                ws.cell(zeile.zeile, spalte,
                        _sumifs(ms_spalte(periode), zeile.konten))
                geschrieben += 1

        # Die Lead-Position zieht ab jetzt aus dem Aufriss. Das ist der
        # eigentliche Unterschied zu Option A.
        for periode in perioden:
            c_lead = ach.lead_spalte(periode, lead_erste, guv=False)
            spalte = spalten.get(ach._index(periode, guv=False))
            if c_lead is None or spalte is None:
                continue
            lead.cell(plan.lead_zeile, c_lead,
                      f"='{plan.blatt}'!{get_column_letter(spalte)}"
                      f"{plan.summenzeile}")
            geschrieben += 1

        # Pflicht-Kontrolle über zwei unabhängige Wege. Der Aufriss wird dabei
        # **mit Vielfachheit** aufsummiert: ein Konto, das versehentlich in
        # zwei Aufrisszeilen steht, zählt hier zweimal und fällt als Differenz
        # auf. Über eine Menge gerechnet bliebe genau dieser Fehler unsichtbar.
        zugeordnet = {k for z in plan.zeilen for k in z.konten}
        ms_zeilen = zeilen_je_position.get(plan.lead_zeile, [])
        je_schluessel = {z.schluessel: z.werte for z in ms_zeilen}
        befunde.append(Aufrissbefund(
            position=str(lead.cell(plan.lead_zeile, 4).value
                         or lead.cell(plan.lead_zeile, 3).value),
            blatt=plan.blatt, lead_zeile=plan.lead_zeile,
            aufriss={p: round(sum(je_schluessel.get(k, {}).get(p, 0.0)
                                  for zl in plan.zeilen for k in zl.konten), 2)
                     for p in perioden},
            konten={p: round(sum(z.werte.get(p, 0.0) for z in ms_zeilen), 2)
                    for p in perioden},
            konten_ohne_zeile=sorted(z.schluessel for z in ms_zeilen
                                     if z.schluessel not in zugeordnet)))
    return geschrieben, befunde
