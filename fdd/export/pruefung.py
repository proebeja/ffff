"""Abnahme der befüllten Vorlage: Formatvergleich und Recalc.

Beides prüft dieselbe Frage von zwei Seiten. Der **Formatvergleich** stellt
die Ausgabe Zelle für Zelle gegen die Vorlage und listet jede Abweichung in
Fill, Schrift, Rahmen, Einzug, Zahlenformat und Zeilenhöhe. Weil die Ausgabe
eine Kopie ist und nur Werte in vorhandene Zellen geschrieben werden, muss die
Liste leer sein — jeder Eintrag ist ein Beleg dafür, dass doch irgendwo ein
Format erzeugt wurde.

Der **Recalc** rechnet die Mappe außerhalb von Excel nach und zählt die
Fehlerzellen. Die Vorlage bringt Altlasten mit (Add-in-Funktionen, die
außerhalb von Excel nicht existieren); alles darüber hinaus stammt aus der
Befüllung.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

import openpyxl


def _farbe(c) -> str:
    if c is None:
        return "-"
    if c.type == "theme":
        return f"theme{c.theme}/{round(c.tint or 0, 4)}"
    return f"{c.type}:{getattr(c, c.type, None)}"


def _signatur(zelle) -> dict[str, str]:
    """Die Formatmerkmale, die die Abnahmekriterien nennen.

    Bewusst als lesbare Zeichenkette je Merkmal und nicht als Objektvergleich:
    ein ``repr`` enthält die Objektadresse und meldet Abweichungen, wo keine
    sind. Der Vergleich soll das Format prüfen, nicht die Identität.
    """
    f, fi, b, a = zelle.font, zelle.fill, zelle.border, zelle.alignment
    return {
        "Fill": f"{fi.patternType}/{_farbe(fi.fgColor)}/{_farbe(fi.bgColor)}",
        "Schrift": (f"{f.name}/{f.sz}/b={bool(f.b)}/i={bool(f.i)}/"
                    f"u={f.u}/{_farbe(f.color)}"),
        "Rahmen": "|".join(
            f"{seite}:{getattr(b, seite).style}/{_farbe(getattr(b, seite).color)}"
            for seite in ("left", "right", "top", "bottom")),
        "Ausrichtung": (f"{a.horizontal}/{a.vertical}/Einzug={a.indent}/"
                        f"umbruch={bool(a.wrapText)}"),
        "Zahlenformat": zelle.number_format,
    }


@dataclass
class Formatabweichung:
    blatt: str
    zelle: str
    eigenschaft: str
    vorlage: str
    ausgabe: str


@dataclass
class Formatbericht:
    abweichungen: list[Formatabweichung] = field(default_factory=list)
    #: Zellen, für die die Vorlage gar keinen Eintrag hält. Sie haben kein
    #: Format, mit dem sie gleich sein könnten — ihr Stil stammt aus der
    #: Nachbarzelle. Sie werden getrennt ausgewiesen und nicht unterschlagen.
    neue_zellen: list[str] = field(default_factory=list)
    ausgenommen: list[str] = field(default_factory=list)

    @property
    def anzahl(self) -> int:
        return len(self.abweichungen)


#: Vom Formatgrundsatz ausgenommen (``excel_format.geltung``): Mastersheet und
#: Review-Queue sind Arbeitsblätter, keine Ausgabeblätter.
ARBEITSBLAETTER = ("Mastersheet", "Review-Queue", "Zuordnung", "QA",
                   "Status je Spalte", "Verhaltensprüfung",
                   # dieselben Blätter in englischer Berichtssprache
                   "Review queue", "Mapping", "Status by column",
                   "Behaviour check", "Assumptions", "Open items")


def vergleiche_format(vorlage_pfad: str, ausgabe_pfad: str,
                      blaetter: Optional[list[str]] = None) -> Formatbericht:
    """Zellweiser Formatvergleich Ausgabe gegen Vorlage.

    Blätter, die es nur in der Ausgabe gibt, sind Arbeitsblätter und vom
    Formatgrundsatz ausgenommen; sie werden übersprungen.
    """
    v = openpyxl.load_workbook(vorlage_pfad)
    a = openpyxl.load_workbook(ausgabe_pfad)
    bericht = Formatbericht()

    for name in v.sheetnames:
        if blaetter and name not in blaetter:
            continue
        if name in ARBEITSBLAETTER:
            bericht.ausgenommen.append(name)
            continue
        if name not in a.sheetnames:
            bericht.abweichungen.append(Formatabweichung(
                name, "-", "Blatt", "vorhanden", "fehlt"))
            continue
        wv, wa = v[name], a[name]
        zeilen = max(wv.max_row, wa.max_row)
        spalten = max(wv.max_column, wa.max_column)
        for r in range(1, zeilen + 1):
            hv = wv.row_dimensions[r].height
            ha = wa.row_dimensions[r].height
            if hv != ha:
                bericht.abweichungen.append(Formatabweichung(
                    name, f"Zeile {r}", "Zeilenhöhe", str(hv), str(ha)))
            for c in range(1, spalten + 1):
                zv, za = wv.cell(r, c), wa.cell(r, c)
                if zv._style is None and zv.value is None:
                    if za.value is not None:
                        bericht.neue_zellen.append(f"{name}!{zv.coordinate}")
                    continue
                sv, sa = _signatur(zv), _signatur(za)
                for merkmal, wert in sv.items():
                    if wert != sa[merkmal]:
                        bericht.abweichungen.append(Formatabweichung(
                            name, zv.coordinate, merkmal, wert, sa[merkmal]))
    return bericht


@dataclass
class Recalcergebnis:
    fehler: dict[str, int] = field(default_factory=dict)
    je_blatt: dict[str, int] = field(default_factory=dict)
    zellen: dict[str, object] = field(default_factory=dict)
    nicht_lesbar: list[str] = field(default_factory=list)

    @property
    def anzahl(self) -> int:
        return sum(self.fehler.values())


class _RobusteMappe:
    """``ExcelModel``, das an unlesbaren Formeln nicht abbricht.

    Die Vorlage trägt Hilfszellen, deren Text mit ``=`` beginnt, ohne eine
    Formel zu sein (``=technische hilfszelle --> nicht löschen``). Der Parser
    wirft darauf einen ``FormulaError`` und die ganze Mappe lässt sich nicht
    rechnen. Diese Zellen werden als Text behandelt — das ist dasselbe, was
    die Option ``--force`` der Kommandozeile tut.
    """

    @staticmethod
    def modell():
        import formulas
        from formulas.errors import FormulaError

        class Modell(formulas.ExcelModel):
            nicht_lesbar: list[str] = []

            def compile_cell(self, cell, context, references, formula_references):
                try:
                    return super().compile_cell(cell, context, references,
                                                formula_references)
                except FormulaError:
                    Modell.nicht_lesbar.append(
                        f"{context.get('sheet', '?')}!{cell.coordinate}")
                    return self._compile_cell(
                        cell.coordinate, cell.value, context, False, references)

        Modell.nicht_lesbar = []
        return Modell


def recalc(pfad: str) -> Recalcergebnis:
    """Rechnet die Mappe nach und zählt die Fehlerzellen je Art und Blatt."""
    modell = _RobusteMappe.modell()
    xl = modell().loads(pfad).finish(circular=True)
    loesung = xl.calculate()

    erg = Recalcergebnis(nicht_lesbar=list(modell.nicht_lesbar))
    for schluessel, wert in loesung.items():
        try:
            v = wert.value[0, 0]
        except Exception:
            continue
        text = str(v)
        if not text.startswith("#"):
            continue
        blatt = schluessel.split("'")[1] if "'" in schluessel else schluessel
        erg.fehler[text] = erg.fehler.get(text, 0) + 1
        erg.je_blatt[blatt] = erg.je_blatt.get(blatt, 0) + 1
        erg.zellen[schluessel] = text
    return erg


def gerechnete_werte(pfad: str, blatt: str) -> dict[str, object]:
    """Alle gerechneten Werte eines Blattes, nach Zelladresse.

    Der Recalc liefert Schlüssel der Form ``'[DATEI.XLSX]BLATT'!A1``. Für die
    Abnahme interessiert nur das Blatt und die Zelle.
    """
    modell = _RobusteMappe.modell()
    loesung = modell().loads(pfad).finish(circular=True).calculate()
    marke = f"]{blatt.upper()}'!"
    out: dict[str, object] = {}
    for schluessel, wert in loesung.items():
        if marke not in schluessel:
            continue
        zelle = schluessel.split("!", 1)[1]
        try:
            out[zelle] = wert.value[0, 0]
        except Exception:
            continue
    return out
