"""Excel-Export im Hausformat.

Hausformat (Handover Abschnitt 3): Arial, Teal-Palette, Zahlenformat
``#,##0;(#,##0);"-"``, Kontrollzeile. Sichtbare, prüfbare Formeln.

Single Source of Truth: das **Mastersheet** ist das eine Zuhause jeder
Kontozahl. Der Net-Debt-Tab rechnet nichts neu, sondern summiert die
Mastersheet-Werte über sichtbare ``SUMIFS``-Formeln, die auf die
Klasse-/NA-Spalte zeigen. Damit kann keine Zahl an zwei Stellen auseinander-
laufen, und der Prüfer sieht nachvollziehbare Formeln.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core.model import Klasse, MappedAccount
from ..views.net_debt import NetDebtView
from ..views.review_queue import ReviewEintrag
from ..views.schedules import Aufriss, Schedules
from ..views.working_capital import WCView

# ---- Hausformat-Konstanten ------------------------------------------------
ZAHLENFORMAT = '#,##0;(#,##0);"-"'
ZAHLENFORMAT_CHECK = '#,##0.00;(#,##0.00);"-"'   # Kontroll-/Abstimmzeilen
FONT_NAME = "Arial"

# Teal-Hauspalette
TEAL_DK = "005858"        # Titel
TEAL = "008888"           # Header-Bänder, Zwischenüberschriften
TEAL_PRI = "00B0B0"       # Akzent
TINT1 = "E0F8F8"          # helle Zeilen-Bänderung
TINT2 = "C8F0F0"          # zweite Bänderungsstufe
MINT = "88E0D8"           # Hervorhebung
GELB = "FFF2A8"           # Flag / zu prüfen

# Schriftfarben-Konvention: Herkunft des Werts auf einen Blick
FARBE_INPUT = "0000FF"    # blau  = hartcodierter Input
FARBE_FORMEL = "000000"   # schwarz = Formel im selben Blatt
FARBE_LINK = "008000"     # grün  = Querverweis auf ein anderes Blatt

_kopf_fill = PatternFill("solid", fgColor=TEAL)
_sub_fill = PatternFill("solid", fgColor=TINT1)
_grau_fill = PatternFill("solid", fgColor=TINT2)
_gelb_fill = PatternFill("solid", fgColor=GELB)
_kopf_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
_bold = Font(name=FONT_NAME, bold=True, size=10)
_normal = Font(name=FONT_NAME, size=10)
_input = Font(name=FONT_NAME, size=10, color=FARBE_INPUT)
_link = Font(name=FONT_NAME, size=10, color=FARBE_LINK)
_link_bold = Font(name=FONT_NAME, size=10, bold=True, color=FARBE_LINK)
_hinweis = Font(name=FONT_NAME, italic=True, size=9, color="808080")
_duenn = Side(style="thin", color="BFBFBF")
_rahmen = Border(bottom=_duenn)
_top_double = Border(top=Side(style="double", color=TEAL))
_top_thin = Border(top=Side(style="thin", color=TEAL))


def _titel(ws, zeile: int, spalte: int, text: str, groesse: int = 11):
    c = ws.cell(zeile, spalte, text)
    c.font = Font(name=FONT_NAME, bold=True, size=groesse, color=TEAL_DK)
    return c


@dataclass
class MastersheetLayout:
    """Merkt sich, wo im Mastersheet welche Spalte/Zeilen liegen, damit
    Aufrisse und Kontrollzeilen korrekt referenzieren."""

    sheetname: str
    erste_datenzeile: int
    letzte_datenzeile: int
    spalte_klasse: int
    spalte_na: int
    perioden_spalten: dict[str, int]
    # Schlüssel ist die Objekt-Identität des MappedAccount, NICHT die Kontonummer:
    # reale Charts (z.B. Namur) tragen dieselbe Kontonummer mehrfach als
    # verschiedene Unterkonten — ein konto->Zeile-Map würde die verwechseln.
    zeile_je_account: dict[int, int]

    def bereich(self, spalte: int) -> str:
        c = get_column_letter(spalte)
        return f"'{self.sheetname}'!${c}${self.erste_datenzeile}:${c}${self.letzte_datenzeile}"

    def wert_zelle(self, m: MappedAccount, periode: str) -> str:
        """Zellbezug auf den Saldo genau dieses Kontos je Periode."""
        c = get_column_letter(self.perioden_spalten[periode])
        return f"'{self.sheetname}'!${c}${self.zeile_je_account[id(m)]}"

    def klasse_zelle(self, m: MappedAccount) -> str:
        c = get_column_letter(self.spalte_klasse)
        return f"'{self.sheetname}'!${c}${self.zeile_je_account[id(m)]}"


def _ref_formel(aufriss: "Optional[AufrissRef]", periode: str, nd_teil: bool) -> str:
    """Formel einer Lead-Zeile: Verweis auf die passende Aufriss-Summenzelle.
    nd_teil=True -> Net-Debt-Lead (gemischt: thereof ND, sonst Summe);
    nd_teil=False -> Working-Capital-Lead (gemischt: operating, sonst Summe)."""
    if aufriss is None:
        return "0"
    if aufriss.is_mixed:
        zelle = aufriss.thereof_nd[periode] if nd_teil else aufriss.operating[periode]
    else:
        zelle = aufriss.total[periode]
    return "=" + zelle


def _krit(text: str) -> str:
    """Escaping für ein SUMIFS-Text-Kriterium: Doppelquotes verdoppeln (sonst
    bricht der Formel-String), Excel-Wildcards (* ? ~) mit ~ neutralisieren
    (sonst ändern sie stillschweigend die Match-Semantik)."""
    text = text.replace('"', '""')
    for w in ("~", "*", "?"):
        text = text.replace(w, "~" + w)
    return text


@dataclass
class AufrissRef:
    """Zellbezüge auf die Summenzeile eines Aufrisses (Quelle der Lead-Zeilen)."""

    sheetname: str
    is_mixed: bool
    total: dict[str, str]        # nicht-gemischt: Periode -> Summenzelle
    operating: dict[str, str]    # gemischt: Periode -> operating-Summenzelle
    thereof_nd: dict[str, str]   # gemischt: Periode -> thereof-ND-Summenzelle
    # gemischt: Konto -> Periode -> Zelle der thereof-ND-Spalte. Speist die
    # aufklappbaren Detailzeilen des Net-Debt-Leads — dieselbe Spalte, aus der
    # auch die Summenzeile zieht, nur eine Ebene tiefer.
    konto_thereof_nd: dict[str, dict[str, str]] = field(default_factory=dict)


def schreibe_databook(pfad: str, mapped: list[MappedAccount], nd: NetDebtView,
                      review: list[ReviewEintrag], perioden: list[str],
                      entity: str, meta: Optional[dict] = None,
                      wc: "Optional[WCView]" = None,
                      schedules: "Optional[Schedules]" = None,
                      recon=None, setup=None, lead_na=None, lead_pl=None,
                      ja_recon=None) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    layout = _schreibe_mastersheet(wb, mapped, perioden, entity)
    refs = _schreibe_schedules(wb, schedules, layout, perioden, entity) if schedules else {}
    if lead_na is not None and lead_na.bloecke:
        _schreibe_lead_na(wb, lead_na, layout, perioden, refs, setup)
    _schreibe_net_debt(wb, nd, layout, perioden, entity, refs, setup)
    if wc is not None:
        _schreibe_working_capital(wb, wc, layout, perioden, entity, refs, setup)
    if lead_pl is not None and lead_pl.bloecke:
        _schreibe_lead_pl(wb, lead_pl, layout, perioden, refs, setup)
    if recon is not None:
        _schreibe_reconciliation(wb, recon, perioden, entity)
    if ja_recon is not None:
        _schreibe_ja_reconciliation(wb, ja_recon, entity)
    _schreibe_review(wb, review, perioden)
    if meta:
        _schreibe_info(wb, meta)
    wb.save(pfad)


def _vorlaeufig_banner(ws, zeile: int, setup) -> None:
    """Kennzeichnet das Blatt als vorläufig, wenn kein Kontennachweis vorlag."""
    if setup is None or getattr(setup, "ist_abschlusstreu", True):
        return
    c = ws.cell(zeile, 2, "VORLÄUFIG — NICHT ABSCHLUSSTREU: kein Kontennachweis; "
                          "Struktur aus Hausconvention/SKR-Default")
    c.font = Font(name=FONT_NAME, bold=True, size=9, color="9C0006")
    c.fill = _gelb_fill


# ---- Reconciliation SuSa vs. Kontennachweis ------------------------------
def _schreibe_reconciliation(wb, recon, perioden, entity) -> None:
    ws = wb.create_sheet("Recon SuSa-KN")
    ws.sheet_view.showGridLines = False
    _titel(ws, 1, 2, f"Reconciliation SuSa gegen Kontennachweis — {entity}")
    ws.cell(2, 2, "in EUR · Differenz = SuSa − Kontennachweis · Differenzen sind "
                  "Information (Abschlussbuchungen/Umgliederungen), nicht per se Fehler"
            ).font = _hinweis

    hz = 4
    ws.cell(hz, 2, "HGB-Position").font = _kopf_font
    ws.cell(hz, 2).fill = _kopf_fill
    spalten: list[tuple[int, str]] = []
    c = 3
    for p in perioden:
        for lab in (f"{p} SuSa", f"{p} KN", f"{p} Diff"):
            cell = ws.cell(hz, c, lab); cell.font = _kopf_font; cell.fill = _kopf_fill
            spalten.append((c, lab)); c += 1
    letzte_spalte = c - 1

    r = hz + 1
    erste = r
    for z in recon.zeilen:
        ws.cell(r, 2, z.hgb_pfad).font = _normal
        cc = 3
        for p in perioden:
            for wert, fmt in ((z.susa.get(p, 0.0), ZAHLENFORMAT),
                              (z.kn.get(p, 0.0), ZAHLENFORMAT),
                              (z.differenz(p), ZAHLENFORMAT_CHECK)):
                cell = ws.cell(r, cc, round(wert, 2))
                cell.number_format = fmt
                cell.font = _normal
                if fmt == ZAHLENFORMAT_CHECK and abs(wert) > 0.005:
                    cell.fill = _gelb_fill
                    cell.font = _bold
                cc += 1
        r += 1
    letzte = r - 1

    # Summenzeile über alle Positionen
    ws.cell(r, 2, "Summe (muss je Spalte aufgehen)").font = _bold
    for col in range(3, letzte_spalte + 1):
        L = get_column_letter(col)
        cell = ws.cell(r, col, f"=SUM({L}{erste}:{L}{letzte})")
        cell.number_format = ZAHLENFORMAT_CHECK
        cell.font = _bold
        cell.fill = _sub_fill
        cell.border = _top_double
    r += 2

    # Mengen-Differenzen: Konten, die nur eine Seite kennt
    for titel, konten in (("Nur im Kontennachweis (fehlten in der SuSa)", recon.nur_im_kn),
                          ("Nur in der SuSa (im Abschluss nicht nachgewiesen)", recon.nur_in_susa)):
        h = ws.cell(r, 2, f"{titel}: {len(konten)}")
        h.font = _bold
        h.fill = _grau_fill
        r += 1
        for konto in konten:
            bez, salden = recon.details.get(konto, ("", {}))
            ws.cell(r, 2, f"{konto}  {bez}").font = _normal
            cc = 3
            for p in perioden:
                cell = ws.cell(r, cc, round(salden.get(p, 0.0), 2))
                cell.number_format = ZAHLENFORMAT
                cell.font = _input
                cc += 3          # unter der jeweiligen SuSa-Spalte
            r += 1
        if not konten:
            ws.cell(r, 2, "(keine)").font = _hinweis
            r += 1
        r += 1

    _breiten(ws, {2: 62})
    for col, _ in spalten:
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.freeze_panes = ws.cell(hz + 1, 3)


# ---- Aufriss-Schicht (Schedules) -----------------------------------------
def _schreibe_schedules(wb, schedules: Schedules, layout: MastersheetLayout,
                        perioden, entity) -> dict[str, AufrissRef]:
    """Schreibt je Net-Asset-Zeile einen Aufriss-Tab. Jede Kontozeile
    referenziert den Mastersheet-Einzelwert (SSOT); die Summenzeile summiert.
    Gemischte Aufrisse (NA_OA/OL/OP) führen zwei Spalten je Periode: operating
    (Klasse OWC) und thereof ND (Klasse ND), gesteuert durch die Klasse-Zelle
    des Mastersheets. Leere Aufrisse werden ausgeblendet. Gibt je NA-Zeile die
    Summenzell-Bezüge zurück, aus denen die Leads ziehen."""
    refs: dict[str, AufrissRef] = {}
    for a in schedules.aufrisse:
        refs[a.na_de] = (_schreibe_mixed_aufriss(wb, a, layout, perioden)
                         if a.is_mixed
                         else _schreibe_einfacher_aufriss(wb, a, layout, perioden))
    return refs


def _aufriss_kopf(ws, a: Aufriss) -> None:
    ws.cell(1, 2, f"Aufriss {a.sheetname}: {a.na_de} / {a.na_en}").font = Font(
        name=FONT_NAME, bold=True, size=12, color=TEAL)
    speist = {"ND": "Net-Debt-Lead", "WC": "Working-Capital-Lead",
              "beide": "WC-Lead (operating) + ND-Lead (thereof ND)",
              "NA": "Lead NA", "PL": "Lead PL"}[a.speist]
    ws.cell(2, 2, f"in EUR · Einzelkonten je Periode aus dem Mastersheet · speist {speist}").font = Font(
        name=FONT_NAME, italic=True, size=9)


def _schreibe_einfacher_aufriss(wb, a: Aufriss, layout, perioden) -> AufrissRef:
    ws = wb.create_sheet(a.sheetname)
    ws.sheet_view.showGridLines = False
    _aufriss_kopf(ws, a)
    hz = 4
    kopf = ["", "Konto", "Bezeichnung", "Klasse"] + list(perioden)
    for c, t in enumerate(kopf[1:], start=2):
        cell = ws.cell(hz, c, t); cell.font = _kopf_font; cell.fill = _kopf_fill
    p0 = 5
    r = hz + 1
    erste = r
    for m in a.konten:
        ws.cell(r, 2, m.konto).font = _normal
        ws.cell(r, 3, m.bezeichnung).font = _normal
        kc = ws.cell(r, 4, m.klasse.value); kc.font = _bold; kc.fill = _klasse_fill(m.klasse)
        for i, p in enumerate(perioden):
            cell = ws.cell(r, p0 + i, "=" + layout.wert_zelle(m, p))
            cell.number_format = ZAHLENFORMAT; cell.font = _link   # gruen: Querverweis
        r += 1
    letzte = r - 1
    total = {}
    tc = ws.cell(r, 3, "Summe (Quelle Lead)"); tc.font = _bold
    for i, p in enumerate(perioden):
        col = get_column_letter(p0 + i)
        cell = ws.cell(r, p0 + i, f"=SUM({col}{erste}:{col}{letzte})")
        cell.number_format = ZAHLENFORMAT; cell.font = _bold; cell.fill = _sub_fill
        cell.border = _top_double
        total[p] = f"'{a.sheetname}'!${col}${r}"
    _breiten(ws, {2: 12, 3: 40, 4: 8})
    for i in range(len(perioden)):
        ws.column_dimensions[get_column_letter(p0 + i)].width = 14
    if a.ist_leer:
        ws.sheet_state = "hidden"
    return AufrissRef(a.sheetname, False, total, {}, {})


def _schreibe_mixed_aufriss(wb, a: Aufriss, layout, perioden) -> AufrissRef:
    ws = wb.create_sheet(a.sheetname)
    ws.sheet_view.showGridLines = False
    _aufriss_kopf(ws, a)
    hz = 4
    ws.cell(hz, 2, "Konto").font = _kopf_font; ws.cell(hz, 2).fill = _kopf_fill
    ws.cell(hz, 3, "Bezeichnung").font = _kopf_font; ws.cell(hz, 3).fill = _kopf_fill
    ws.cell(hz, 4, "Klasse").font = _kopf_font; ws.cell(hz, 4).fill = _kopf_fill
    # je Periode zwei Spalten: operating | thereof ND
    p0 = 5
    op_col, nd_col = {}, {}
    for i, p in enumerate(perioden):
        oc, nc = p0 + 2 * i, p0 + 2 * i + 1
        op_col[p], nd_col[p] = oc, nc
        for col, lab in ((oc, f"{p} operating"), (nc, f"{p} thereof ND")):
            cell = ws.cell(hz, col, lab); cell.font = _kopf_font; cell.fill = _kopf_fill
    r = hz + 1
    erste = r
    konto_nd: dict[str, dict[str, str]] = {}
    for m in a.konten:
        ws.cell(r, 2, m.konto).font = _normal
        ws.cell(r, 3, m.bezeichnung).font = _normal
        kc = ws.cell(r, 4, m.klasse.value); kc.font = _bold; kc.fill = _klasse_fill(m.klasse)
        konto_nd[m.konto] = {
            p: f"'{a.sheetname}'!${get_column_letter(nd_col[p])}${r}" for p in perioden}
        kz = layout.klasse_zelle(m)
        for p in perioden:
            wert = layout.wert_zelle(m, p)
            # Split ausschließlich über die Klasse-Zelle des Mastersheets:
            op = ws.cell(r, op_col[p], f'=IF({kz}="OWC",{wert},IF({kz}="TWC",{wert},0))')
            nd = ws.cell(r, nd_col[p], f'=IF({kz}="ND",{wert},0)')
            for cell in (op, nd):
                cell.number_format = ZAHLENFORMAT; cell.font = _link   # gruen: Querverweis
        r += 1
    letzte = r - 1
    tc = ws.cell(r, 3, "Summe (operating -> WC-Lead / thereof ND -> ND-Lead)"); tc.font = _bold
    operating, thereof_nd = {}, {}
    for p in perioden:
        for col, ziel in ((op_col[p], operating), (nd_col[p], thereof_nd)):
            L = get_column_letter(col)
            cell = ws.cell(r, col, f"=SUM({L}{erste}:{L}{letzte})")
            cell.number_format = ZAHLENFORMAT; cell.font = _bold; cell.fill = _sub_fill
            cell.border = _top_double
            ziel[p] = f"'{a.sheetname}'!${L}${r}"
    _breiten(ws, {2: 12, 3: 40, 4: 8})
    for i in range(len(perioden) * 2):
        ws.column_dimensions[get_column_letter(p0 + i)].width = 14
    if a.ist_leer:
        ws.sheet_state = "hidden"
    return AufrissRef(a.sheetname, True, {}, operating, thereof_nd, konto_nd)


# ---- Mastersheet (Single Source of Truth) --------------------------------
def _schreibe_mastersheet(wb, mapped, perioden, entity) -> MastersheetLayout:
    ws = wb.create_sheet("Mastersheet")
    kopf = ["Konto", "Bezeichnung", "Entity", "HGB-Pfad (DE)", "HGB-Path (EN)",
            "Klasse", "NA-Zeile (DE)", "NA-Zeile (EN)", "Quelle/Regel", "Review"]
    perioden_start = len(kopf) + 1
    for p in perioden:
        kopf.append(p)
    _schreibe_kopf(ws, kopf, zeile=1)

    r = 2
    zeile_je_account: dict[int, int] = {}
    for m in sorted(mapped, key=lambda x: x.konto):
        zeile_je_account[id(m)] = r
        ws.cell(r, 1, m.konto).font = _normal
        ws.cell(r, 2, m.bezeichnung).font = _normal
        ws.cell(r, 3, m.account.entity).font = _normal
        ws.cell(r, 4, m.hgb_pfad).font = _normal
        ws.cell(r, 5, m.hgb_pfad_en).font = _normal
        c6 = ws.cell(r, 6, m.klasse.value); c6.font = _bold
        c6.fill = _klasse_fill(m.klasse)
        ws.cell(r, 7, m.na_de).font = _normal
        ws.cell(r, 8, m.na_en).font = _normal
        quelle = m.quelle.value + (f" [{m.regel_id}]" if m.regel_id else "")
        ws.cell(r, 9, quelle).font = _normal
        rv = ws.cell(r, 10, "REVIEW" if m.review else ""); rv.font = _bold
        for i, p in enumerate(perioden):
            cell = ws.cell(r, perioden_start + i, round(m.saldo(p), 2))
            cell.number_format = ZAHLENFORMAT
            cell.font = _input          # blau: hartcodierter Input
        r += 1
    letzte = r - 1

    perioden_spalten = {p: perioden_start + i for i, p in enumerate(perioden)}
    _breiten(ws, {1: 10, 2: 34, 3: 18, 4: 46, 5: 46, 6: 8, 7: 30, 8: 30, 9: 26, 10: 9})
    ws.freeze_panes = "A2"
    return MastersheetLayout(
        sheetname="Mastersheet", erste_datenzeile=2, letzte_datenzeile=max(letzte, 2),
        spalte_klasse=6, spalte_na=7, perioden_spalten=perioden_spalten,
        zeile_je_account=zeile_je_account,
    )


# ---- Net-Debt-Tab (zieht aus Aufrissen, Kontrollzeile prüft Mastersheet) --
def _lead_outline(ws) -> None:
    """Outline-Ausrichtung der Lead-Tabs: Summenzeilen stehen ÜBER ihren
    Detailzeilen, die einklappbare Quellenspalte links neben ihrer Gruppe."""
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False


def _quelle_spalte(ws, spalte: int = 4) -> None:
    """Die Spalte mit den Aufriss-Verweisen wird gruppiert: im Arbeitsmodus
    sichtbar, für den Report-Export mit einem Klick ausgeblendet."""
    ws.column_dimensions[get_column_letter(spalte)].outline_level = 1


def _lead_kopf(ws, kopf_zeile: int, p0: int, perioden, kontext: str) -> None:
    for c, txt in ((2, "Ref."), (3, "Net-Asset-Position"), (4, "Quelle"), (5, kontext)):
        ws.cell(kopf_zeile, c, txt)
    for i, p in enumerate(perioden):
        ws.cell(kopf_zeile, p0 + i, p)
    _style_kopf_row(ws, kopf_zeile, 2, p0 + len(perioden) - 1)


def _hgb_position(m: MappedAccount) -> str:
    """Herkunftsposition eines Kontos: das Blatt seines HGB-Pfads."""
    return m.hgb_pfad.rstrip("/").rsplit("/", 1)[-1] if m.hgb_pfad else ""


def _schreibe_net_debt(wb, nd: NetDebtView, layout: MastersheetLayout,
                       perioden, entity, refs: dict[str, AufrissRef],
                       setup=None) -> None:
    ws = wb.create_sheet("Net Debt")
    ws.sheet_view.showGridLines = False
    _lead_outline(ws)
    p0 = 6  # erste Perioden-Spalte (nach Ref./Position/Quelle/Herkunft)

    # Titelblock
    t = ws.cell(1, 2, f"Net Debt — {entity}"); t.font = Font(name=FONT_NAME, bold=True, size=13, color=TEAL)
    ws.cell(2, 2, "in EUR · jede Zeile zieht aus genau einem Aufriss · Umgliederungs"
                  "zeilen links aufklappbar · Spalte 'Quelle' für den Report ausblendbar").font = _hinweis
    _vorlaeufig_banner(ws, 3, setup)

    kopf_zeile = 4
    _lead_kopf(ws, kopf_zeile, p0, perioden, "Herkunft")

    r = kopf_zeile + 1
    ref = 1
    daten_zeilen: list[int] = []

    # Die SUMIFS-Formel filtert auf (Klasse=ND, NA-Zeile). Damit keine NA-Zeile
    # doppelt summiert wird, muss jede NA-Zeile im ganzen Tab genau einmal
    # erscheinen — direkte und Umgliederungs-Zeilen müssen disjunkt sein.
    direkt_namen = {z.na_de for z in nd.direkt}
    umg_namen = {z.na_de for z in nd.umgliederung}
    kollision = direkt_namen & umg_namen
    if kollision:
        raise ValueError(
            "Net-Debt-Export: NA-Zeile(n) in direkter UND Umgliederungs-Gruppe "
            f"({sorted(kollision)}) — SUMIFS würde doppelt zählen. Prüfe die "
            "Reklassifizierung/aus_mixed-Zuordnung."
        )

    def detailzeilen(z, aufriss: Optional[AufrissRef]) -> None:
        """Einzelkonten des thereof-ND-Teils, eingeklappt unter ihrer
        Summenzeile. Die Werte kommen aus derselben thereof-ND-Spalte des
        Aufrisses, aus der auch die Summenzeile zieht — keine zweite Quelle."""
        nonlocal r
        if aufriss is None or not aufriss.konto_thereof_nd:
            return
        for m in sorted(z.konten, key=lambda m: m.konto):
            zellen = aufriss.konto_thereof_nd.get(m.konto)
            if zellen is None:
                continue
            ws.cell(r, 2, m.konto).font = _normal
            ws.cell(r, 3, f"    {m.bezeichnung}").font = _normal
            ws.cell(r, 4, aufriss.sheetname).font = _hinweis
            ws.cell(r, 5, _hgb_position(m)).font = _hinweis
            for i, p in enumerate(perioden):
                cell = ws.cell(r, p0 + i, "=" + zellen[p])
                cell.number_format = ZAHLENFORMAT
                cell.font = _link          # gruen: zieht aus dem Aufriss
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True
            r += 1

    def schreibe_gruppe(titel: str, zeilen, mit_details: bool = False) -> None:
        nonlocal r, ref
        if not zeilen:
            return
        c = ws.cell(r, 3, titel); c.font = _bold; c.fill = _grau_fill
        r += 1
        for z in zeilen:
            ws.cell(r, 2, ref).font = _normal
            aufriss = refs.get(z.na_de)
            ws.cell(r, 3, f"{z.na_de} / {z.na_en}").font = _normal
            if aufriss:
                ws.cell(r, 4, aufriss.sheetname).font = _hinweis
            for i, p in enumerate(perioden):
                # Jede Zeile zieht aus GENAU EINEM Aufriss: gemischte Position ->
                # thereof-ND-Summe, sonst -> Aufriss-Summe. Kein direkter
                # Mastersheet-Zugriff mehr (nur die Kontrollzeile prüft dagegen).
                cell = ws.cell(r, p0 + i)
                cell.value = _ref_formel(aufriss, p, nd_teil=True)
                cell.number_format = ZAHLENFORMAT
                cell.font = _link          # gruen: zieht aus dem Aufriss
            daten_zeilen.append(r)
            ref += 1
            r += 1
            if mit_details:
                detailzeilen(z, aufriss)

    schreibe_gruppe("Direkte Net-Debt-Positionen", nd.direkt)
    schreibe_gruppe("Umgliederung aus NWC in ND (thereof ND)", nd.umgliederung,
                    mit_details=True)

    # Zwischensumme Net Debt — summiert ausschließlich die Summenzeilen, nie
    # die aufgeklappten Detailzeilen (sonst doppelte Zählung).
    sub_zeile = r
    ws.cell(sub_zeile, 3, "Netto-Finanzvermögen / -Verbindlichkeiten (Net cash / Net debt)").font = _bold
    for i, p in enumerate(perioden):
        col = p0 + i
        if daten_zeilen:
            spalte = get_column_letter(col)
            formel = "=" + "+".join(f"{spalte}{z}" for z in daten_zeilen)
        else:
            formel = 0
        cell = ws.cell(sub_zeile, col, formel)
        cell.number_format = ZAHLENFORMAT
        cell.font = _bold
        cell.fill = _sub_fill
        cell.border = _top_double
    ws.cell(sub_zeile, 2).fill = _sub_fill

    # Kontrollzeile: unabhängig ALLE ND-Konten aus dem Mastersheet summieren und
    # gegen die aus den View-Zeilen gebaute Zwischensumme stellen. Differenz != 0
    # deckt auf, dass eine ND-Position in keiner View-Zeile repräsentiert ist.
    k = sub_zeile + 2
    ws.cell(k, 3, "Kontrollzeile (Σ ND Mastersheet − Zwischensumme, muss 0 sein)").font = Font(
        name=FONT_NAME, italic=True, size=9)
    for i, p in enumerate(perioden):
        col = p0 + i
        spalte = get_column_letter(col)
        gesamt_nd = (
            f"SUMIFS({layout.bereich(layout.perioden_spalten[p])},"
            f"{layout.bereich(layout.spalte_klasse)},\"ND\")"
        )
        cell = ws.cell(k, col, f"={gesamt_nd}-{spalte}{sub_zeile}")
        cell.number_format = ZAHLENFORMAT_CHECK
        cell.font = Font(name=FONT_NAME, italic=True, size=9, color=FARBE_LINK)
        cell.fill = _gelb_fill

    _breiten(ws, {2: 10, 3: 52, 4: 14, 5: 30})
    _quelle_spalte(ws)
    for i in range(len(perioden)):
        ws.column_dimensions[get_column_letter(p0 + i)].width = 15
    ws.freeze_panes = ws.cell(kopf_zeile + 1, p0)


# ---- Working-Capital-Tab (zieht aus Aufrissen, Kontrollzeile prüft MS) ----
def _schreibe_working_capital(wb, wc: WCView, layout: MastersheetLayout,
                              perioden, entity, refs: dict[str, AufrissRef],
                              setup=None) -> None:
    """Gliederung nach der klassischen FDD-Schnittrichtung: Block TWC, Block
    OWC, darunter NWC. Die OA/OL-Ableitung bleibt je Zeile als Spalte sichtbar.

    Gemischte Positionen erscheinen in der Reported-Sicht: Bilanzwert, darunter
    nachrichtlich der in den Net Debt umgegliederte Teil, darunter der
    operative Rest — und nur dieser läuft in den Blocksaldo."""
    ws = wb.create_sheet("Working Capital")
    ws.sheet_view.showGridLines = False
    _lead_outline(ws)
    p0 = 6

    ws.cell(1, 2, f"Working Capital (Ist je Periode) — {entity}").font = Font(
        name=FONT_NAME, bold=True, size=13, color=TEAL)
    ws.cell(2, 2, "in EUR · jede Zeile zieht aus genau einem Aufriss · WC-Definition "
                  "über alle Perioden identisch · gemischte Positionen in der "
                  "Reported-Sicht (= operating + thereof ND aus dem Aufriss; Konten "
                  "der Position, die noch in der Review-Queue stehen, sind darin "
                  "nicht enthalten) · noch keine normalisierte Referenz").font = _hinweis
    _vorlaeufig_banner(ws, 3, setup)

    kopf_zeile = 4
    _lead_kopf(ws, kopf_zeile, p0, perioden, "OA/OL")

    r = kopf_zeile + 1
    ref = 1

    def summe_zeile(label: str, rows: list[int], fett=True, fill=None) -> int:
        nonlocal r
        c = ws.cell(r, 3, label); c.font = _bold if fett else _normal
        if fill:
            c.fill = fill
        for i, p in enumerate(perioden):
            col = p0 + i
            sp = get_column_letter(col)
            formel = ("=" + "+".join(f"{sp}{z}" for z in rows)) if rows else 0
            cell = ws.cell(r, col, formel)
            cell.number_format = ZAHLENFORMAT
            cell.font = _bold if fett else _normal
            if fill:
                cell.fill = fill
        zr = r
        r += 1
        return zr

    def position(z) -> int:
        """Schreibt eine WC-Position und gibt die Zeile zurück, die in den
        Blocksaldo läuft. Gemischte Positionen bekommen drei Zeilen
        (reported / davon ND / operativ), alle anderen genau eine."""
        nonlocal r, ref
        aufriss = refs.get(z.na_de)
        quelle = aufriss.sheetname if aufriss else ""
        gemischt = aufriss is not None and aufriss.is_mixed

        ws.cell(r, 2, ref).font = _normal
        ws.cell(r, 3, f"{z.na_de} / {z.na_en}"
                      + (" (reported)" if gemischt else "")).font = _normal
        ws.cell(r, 4, quelle).font = _hinweis
        ws.cell(r, 5, z.seite).font = _hinweis
        ref += 1

        if not gemischt:
            for i, p in enumerate(perioden):
                cell = ws.cell(r, p0 + i, _ref_formel(aufriss, p, nd_teil=False))
                cell.number_format = ZAHLENFORMAT
                cell.font = _link          # gruen: zieht aus dem Aufriss
            zr = r
            r += 1
            return zr

        # Reported = operating + thereof ND, beides aus demselben Aufriss.
        rep = r
        for i, p in enumerate(perioden):
            cell = ws.cell(r, p0 + i,
                           f"={aufriss.operating[p]}+{aufriss.thereof_nd[p]}")
            cell.number_format = ZAHLENFORMAT
            cell.font = _link
        r += 1

        # Nachrichtlicher Abzug — der Betrag selbst wird ausschließlich im
        # Net-Debt-Tab geführt, hier steht er nur als Überleitungsposten.
        dav = r
        ws.cell(r, 3, "    davon ND (umgegliedert in Net Debt)").font = _hinweis
        ws.cell(r, 4, quelle).font = _hinweis
        for i, p in enumerate(perioden):
            cell = ws.cell(r, p0 + i, f"=-{aufriss.thereof_nd[p]}")
            cell.number_format = ZAHLENFORMAT
            cell.font = _hinweis
        r += 1

        # Operativer Rest — nur diese Zeile läuft in den Blocksaldo.
        op = r
        ws.cell(r, 3, "    → operativer Anteil im Working Capital").font = _normal
        ws.cell(r, 5, z.seite).font = _hinweis
        for i, p in enumerate(perioden):
            sp = get_column_letter(p0 + i)
            cell = ws.cell(r, p0 + i, f"={sp}{rep}+{sp}{dav}")
            cell.number_format = ZAHLENFORMAT
            cell.font = _normal
            cell.border = _top_thin
        r += 1
        return op

    def klasse_block(klasse: str, titel: str, saldo_label: str) -> int:
        nonlocal r
        h = ws.cell(r, 3, titel); h.font = Font(name=FONT_NAME, bold=True, color=TEAL)
        h.fill = _grau_fill
        r += 1
        rows = [position(z) for z in wc.zeilen_je_klasse(klasse)]
        return summe_zeile(saldo_label, rows, fett=True, fill=_sub_fill)

    twc_row = klasse_block(
        "TWC", "Block 1 — Trade Working Capital", "Saldo Trade Working Capital (TWC)")
    owc_row = klasse_block(
        "OWC", "Block 2 — Other Working Capital", "Saldo Other Working Capital (OWC)")

    # NWC = Saldo TWC + Saldo OWC (Passiva sind vorzeichenrichtig negativ
    # gespeichert, die Salden addieren sich daher schlicht).
    nwc_row = r
    ws.cell(nwc_row, 3, "Net Working Capital (TWC + OWC)").font = _bold
    for i, p in enumerate(perioden):
        col = p0 + i
        sp = get_column_letter(col)
        cell = ws.cell(nwc_row, col, f"={sp}{twc_row}+{sp}{owc_row}")
        cell.number_format = ZAHLENFORMAT
        cell.font = _bold
        cell.fill = _sub_fill
        cell.border = _top_double
    ws.cell(nwc_row, 2).fill = _sub_fill
    r += 2

    # Kontrollzeile: alle TWC+OWC im Mastersheet minus NWC = 0.
    ws.cell(r, 3, "Kontrollzeile (Σ TWC+OWC Mastersheet − NWC, muss 0 sein)").font = Font(
        name=FONT_NAME, italic=True, size=9)
    for i, p in enumerate(perioden):
        col = p0 + i
        sp = get_column_letter(col)
        rng = layout.bereich(layout.perioden_spalten[p])
        kl = layout.bereich(layout.spalte_klasse)
        gesamt = f"SUMIFS({rng},{kl},\"TWC\")+SUMIFS({rng},{kl},\"OWC\")"
        cell = ws.cell(r, col, f"={gesamt}-{sp}{nwc_row}")
        cell.number_format = ZAHLENFORMAT_CHECK
        cell.font = Font(name=FONT_NAME, italic=True, size=9, color=FARBE_LINK)
        cell.fill = _gelb_fill

    _breiten(ws, {2: 10, 3: 52, 4: 14, 5: 30})
    _quelle_spalte(ws)
    for i in range(len(perioden)):
        ws.column_dimensions[get_column_letter(p0 + i)].width = 15
    ws.freeze_panes = ws.cell(kopf_zeile + 1, p0)


# ---- Lead NA / Lead PL (ziehen aus Aufrissen, Kontrollzeile prüft MS) ----
def _schreibe_uebersichts_lead(wb, view, layout, perioden, refs, setup,
                               tabname: str, titel: str, hinweis: str,
                               gesamt_label: str, klassen: tuple[str, ...],
                               gegenprobe: tuple[str, str] | None = None) -> None:
    """Gemeinsamer Aufbau für Lead NA und Lead PL: Blöcke mit Zwischensumme,
    Gesamtsumme, Kontrollzeile gegen das Mastersheet. Jede Positionszeile zieht
    aus genau einem Aufriss — bei gemischten Positionen aus der Summe beider
    Spalten, denn der Lead zeigt hier den Bilanzwert."""
    ws = wb.create_sheet(tabname)
    ws.sheet_view.showGridLines = False
    _lead_outline(ws)
    p0 = 6

    ws.cell(1, 2, f"{titel} — {view.entity}").font = Font(
        name=FONT_NAME, bold=True, size=13, color=TEAL)
    ws.cell(2, 2, hinweis).font = _hinweis
    _vorlaeufig_banner(ws, 3, setup)

    kopf_zeile = 4
    _lead_kopf(ws, kopf_zeile, p0, perioden, "Klasse")
    r = kopf_zeile + 1
    ref = 1
    block_rows: list[int] = []

    def summe_zeile(label, rows, fill=None, doppel=False) -> int:
        nonlocal r
        c = ws.cell(r, 3, label); c.font = _bold
        if fill:
            c.fill = fill
        for i, p in enumerate(perioden):
            sp = get_column_letter(p0 + i)
            formel = ("=" + "+".join(f"{sp}{z}" for z in rows)) if rows else 0
            cell = ws.cell(r, p0 + i, formel)
            cell.number_format = ZAHLENFORMAT
            cell.font = _bold
            if fill:
                cell.fill = fill
            if doppel:
                cell.border = _top_double
        zr = r
        r += 1
        return zr

    def positions_zeilen(zeilen) -> list[int]:
        nonlocal r, ref
        rows = []
        for z in zeilen:
            aufriss = refs.get(z.na_de)
            ws.cell(r, 2, ref).font = _normal
            ws.cell(r, 3, f"{z.na_de} / {z.na_en}").font = _normal
            if aufriss:
                ws.cell(r, 4, aufriss.sheetname).font = _hinweis
            kc = ws.cell(r, 5, z.klasse); kc.font = _hinweis
            for i, p in enumerate(perioden):
                cell = ws.cell(r, p0 + i, _lead_wert(aufriss, p, z.klasse))
                cell.number_format = ZAHLENFORMAT
                cell.font = _link          # gruen: zieht aus dem Aufriss
            rows.append(r)
            ref += 1
            r += 1
        return rows

    for b in view.bloecke:
        h = ws.cell(r, 3, b.titel); h.font = Font(name=FONT_NAME, bold=True, color=TEAL)
        h.fill = _grau_fill
        r += 1
        rows = positions_zeilen(b.zeilen)
        block_rows.append(summe_zeile(f"Saldo {b.titel}", rows, fill=_sub_fill))

    gesamt = summe_zeile(gesamt_label, block_rows, fill=_sub_fill, doppel=True)
    ws.cell(gesamt, 2).fill = _sub_fill

    # Gegenprobe (Lead NA): Eigenkapital als eigene Zeile, danach muss
    # Net Assets + Eigenkapital null sein — die Bilanz schließt.
    ek_rows: list[int] = []
    if gegenprobe and view.nachrichtlich:
        r += 1
        h = ws.cell(r, 3, gegenprobe[0]); h.font = Font(name=FONT_NAME, bold=True, color=TEAL)
        h.fill = _grau_fill
        r += 1
        ek_rows = positions_zeilen(view.nachrichtlich)

    k = r + 1
    ws.cell(k, 3, f"Kontrollzeile (Σ {'/'.join(klassen)} Mastersheet − {gesamt_label}, muss 0 sein)").font = Font(
        name=FONT_NAME, italic=True, size=9)
    for i, p in enumerate(perioden):
        sp = get_column_letter(p0 + i)
        rng = layout.bereich(layout.perioden_spalten[p])
        kl = layout.bereich(layout.spalte_klasse)
        summe = "+".join(f'SUMIFS({rng},{kl},"{c}")' for c in klassen)
        cell = ws.cell(k, p0 + i, f"={summe}-{sp}{gesamt}")
        cell.number_format = ZAHLENFORMAT_CHECK
        cell.font = Font(name=FONT_NAME, italic=True, size=9, color=FARBE_LINK)
        cell.fill = _gelb_fill

    if ek_rows:
        # Noch nicht klassifizierte Konten (Review-Queue) gehören zur Bilanz,
        # aber in keinen Net-Asset-Block. Ohne sie ginge die Schlussprobe genau
        # um ihren Betrag daneben — deshalb stehen sie hier ausdrücklich.
        rq = k + 1
        ws.cell(rq, 3, "nachrichtlich: noch offen (Review-Queue, in keinem Block)").font = Font(
            name=FONT_NAME, italic=True, size=9)
        for i, p in enumerate(perioden):
            rng = layout.bereich(layout.perioden_spalten[p])
            kl = layout.bereich(layout.spalte_klasse)
            cell = ws.cell(rq, p0 + i, f'=SUMIFS({rng},{kl},"REVIEW")')
            cell.number_format = ZAHLENFORMAT
            cell.font = Font(name=FONT_NAME, italic=True, size=9, color=FARBE_LINK)

        k2 = rq + 1
        ws.cell(k2, 3, f"Kontrollzeile ({gesamt_label} + Eigenkapital + Review, "
                       "muss 0 sein — die Bilanz schließt)").font = Font(
            name=FONT_NAME, italic=True, size=9)
        for i, p in enumerate(perioden):
            sp = get_column_letter(p0 + i)
            ek = "+".join(f"{sp}{z}" for z in ek_rows)
            cell = ws.cell(k2, p0 + i, f"={sp}{gesamt}+{ek}+{sp}{rq}")
            cell.number_format = ZAHLENFORMAT_CHECK
            cell.font = Font(name=FONT_NAME, italic=True, size=9, color=FARBE_LINK)
            cell.fill = _gelb_fill

    _breiten(ws, {2: 10, 3: 52, 4: 14, 5: 30})
    _quelle_spalte(ws)
    for i in range(len(perioden)):
        ws.column_dimensions[get_column_letter(p0 + i)].width = 15
    ws.freeze_panes = ws.cell(kopf_zeile + 1, p0)


def _lead_wert(aufriss, periode: str, klasse: str) -> str:
    """Wert einer Lead-Zeile aus ihrem Aufriss.

    Eine gemischte Position erscheint im Lead NA zweimal — mit ihrem
    operativen Teil im Working-Capital-Block und mit ihrem thereof-ND-Teil im
    Net-Debt-Block. Jede Zeile zieht deshalb genau ihren Teil; zöge sie den
    vollen Bilanzwert, stünde die Position doppelt in den Net Assets."""
    if aufriss is None:
        return "0"
    if aufriss.is_mixed:
        zelle = aufriss.thereof_nd[periode] if klasse == "ND" else aufriss.operating[periode]
        return "=" + zelle
    return "=" + aufriss.total[periode]


def _schreibe_lead_na(wb, view, layout, perioden, refs, setup=None) -> None:
    _schreibe_uebersichts_lead(
        wb, view, layout, perioden, refs, setup, "Lead NA", "Lead NA (Net Assets)",
        "in EUR · jede Zeile zieht aus genau einem Aufriss · Net-Asset-Brücke: "
        "Anlagevermögen + Working Capital + Net Debt + latente Steuern",
        "Net Assets", ("FA", "TWC", "OWC", "ND", "DT"),
        gegenprobe=("Gegenprobe: Eigenkapital", "EQ"))


def _schreibe_lead_pl(wb, view, layout, perioden, refs, setup=None) -> None:
    _schreibe_uebersichts_lead(
        wb, view, layout, perioden, refs, setup, "Lead PL", "Lead PL (Gewinn- und Verlustrechnung)",
        "in EUR · jede Zeile zieht aus genau einem Aufriss · Vorzeichen wie im "
        "Mastersheet (Erträge negativ, Aufwendungen positiv) · Summe = Ergebnis "
        "mit umgekehrtem Vorzeichen",
        "Summe GuV (= Jahresergebnis, Vorzeichen invers)", ("PL",))



# ---- Reconciliation gegen den Jahresabschluss ----------------------------
def _schreibe_ja_reconciliation(wb, rec, entity) -> None:
    """Positionssummen des Databooks gegen die Bilanz/GuV nach Jahresabschluss.
    Der Abschluss ist hier Abstimmziel, nicht Strukturquelle."""
    ws = wb.create_sheet("Recon Databook-JA")
    ws.sheet_view.showGridLines = False
    _titel(ws, 1, 2, f"Reconciliation Databook gegen Jahresabschluss — {entity}")
    ws.cell(2, 2, "in EUR · Differenz = Databook − Jahresabschluss · der Abschluss "
                  "ist Abstimmziel, er ändert kein Mapping").font = _hinweis
    r = 3
    for h in rec.hinweise:
        c = ws.cell(r, 2, h)
        c.font = Font(name=FONT_NAME, bold=True, size=9, color="9C0006")
        c.fill = _gelb_fill
        r += 1

    perioden = rec.perioden
    if not perioden:
        ws.cell(r + 1, 2, "Keine gemeinsame Periode — es gibt nichts abzustimmen.").font = _bold
        _breiten(ws, {2: 96})
        return

    kopf = r + 1
    ws.cell(kopf, 2, "HGB-Position (Jahresabschluss)")
    ws.cell(kopf, 3, "Kanonischer Pfad")
    for i, p in enumerate(perioden):
        for j, lab in enumerate(("Databook", "Jahresabschluss", "Differenz")):
            ws.cell(kopf, 4 + i * 3 + j, f"{p} {lab}")
    _style_kopf_row(ws, kopf, 2, 3 + len(perioden) * 3)

    zeile = kopf + 1
    for z in rec.zeilen:
        ws.cell(zeile, 2, z.label[:70]).font = _normal
        ws.cell(zeile, 3, z.hgb_pfad[-52:]).font = _hinweis
        for i, p in enumerate(perioden):
            for j, wert in enumerate((z.databook.get(p, 0.0), z.ja.get(p, 0.0),
                                      z.differenz(p))):
                c = ws.cell(zeile, 4 + i * 3 + j, wert)
                c.number_format = ZAHLENFORMAT
                c.font = _bold if j == 2 else _normal
                if j == 2 and abs(wert) > 0.005:
                    c.fill = _gelb_fill
        zeile += 1

    zeile += 1
    ws.cell(zeile, 2, "Gesamtdifferenz").font = _bold
    for i, p in enumerate(perioden):
        c = ws.cell(zeile, 6 + i * 3, rec.gesamtdifferenz(p))
        c.number_format = ZAHLENFORMAT
        c.font = _bold
        c.fill = _sub_fill
        c.border = _top_double

    zeile += 2
    ws.cell(zeile, 2, f"Positionen nur im Databook ({len(rec.nur_im_databook)})").font = _bold
    for p in rec.nur_im_databook[:40]:
        zeile += 1
        ws.cell(zeile, 2, p).font = _hinweis
    zeile += 2
    ws.cell(zeile, 2, f"Zeilen des Abschlusses ohne Zuordnung ({len(rec.nur_im_ja)})").font = _bold
    for p in rec.nur_im_ja[:40]:
        zeile += 1
        ws.cell(zeile, 2, p).font = _hinweis

    _breiten(ws, {2: 62, 3: 46})
    for i in range(len(perioden) * 3):
        ws.column_dimensions[get_column_letter(4 + i)].width = 16
    ws.freeze_panes = ws.cell(kopf + 1, 4)


# ---- Review-Queue ---------------------------------------------------------
def _schreibe_review(wb, review: list[ReviewEintrag], perioden) -> None:
    ws = wb.create_sheet("Review-Queue")
    kopf = ["Konto", "Bezeichnung", "HGB-Pfad", "Klasse", "Status",
            "Quelle/Regel", "Grund"]
    p_start = len(kopf) + 1
    kopf += list(perioden)
    _schreibe_kopf(ws, kopf, zeile=1)
    r = 2
    for e in review:
        ws.cell(r, 1, e.konto).font = _normal
        ws.cell(r, 2, e.bezeichnung).font = _normal
        ws.cell(r, 3, e.hgb_pfad).font = _normal
        ws.cell(r, 4, e.klasse).font = _bold
        st = ws.cell(r, 5, e.status); st.font = _bold
        if e.status.startswith("Pflichtfrage"):
            st.fill = PatternFill("solid", fgColor="F9CB9C")
        elif "Verhaltensprüfung" in e.status:
            st.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(r, 6, (e.quelle + (f" [{e.regel_id}]" if e.regel_id else ""))).font = _normal
        ws.cell(r, 7, e.grund).font = _normal
        for i, p in enumerate(perioden):
            cell = ws.cell(r, p_start + i, round(e.salden.get(p, 0.0), 2))
            cell.number_format = ZAHLENFORMAT
            cell.font = _normal
        r += 1
    if not review:
        ws.cell(2, 1, "(keine offenen Fälle)").font = Font(name=FONT_NAME, italic=True)
    _breiten(ws, {1: 10, 2: 40, 3: 44, 4: 8, 5: 26, 6: 24, 7: 52})
    ws.freeze_panes = "A2"


def _schreibe_info(wb, meta: dict) -> None:
    ws = wb.create_sheet("Info")
    ws.cell(1, 1, "Lauf-Metadaten (Reproduzierbarkeit)").font = Font(
        name=FONT_NAME, bold=True, size=12, color=TEAL)
    r = 3
    for k, v in meta.items():
        ws.cell(r, 1, str(k)).font = _bold
        ws.cell(r, 2, str(v)).font = _normal
        r += 1
    _breiten(ws, {1: 28, 2: 70})


# ---- Stil-Helfer ----------------------------------------------------------
def _schreibe_kopf(ws, kopf, zeile) -> None:
    for c, text in enumerate(kopf, start=1):
        cell = ws.cell(zeile, c, text)
        cell.font = _kopf_font
        cell.fill = _kopf_fill
        cell.alignment = Alignment(vertical="center")


def _style_kopf_row(ws, zeile, c_von, c_bis) -> None:
    for c in range(c_von, c_bis + 1):
        cell = ws.cell(zeile, c)
        cell.font = _kopf_font
        cell.fill = _kopf_fill


def _klasse_fill(klasse: Klasse) -> PatternFill:
    farben = {
        Klasse.ND: "F4CCCC", Klasse.TWC: "D9EAD3", Klasse.OWC: "FFF2CC",
        Klasse.FA: "CFE2F3", Klasse.EQ: "D9D2E9", Klasse.DT: "EAD1DC",
        Klasse.REVIEW: "F9CB9C", Klasse.TECH: "EFEFEF", Klasse.PL: "FFFFFF",
        Klasse.MIXED: "FCE5CD",
    }
    return PatternFill("solid", fgColor=farben.get(klasse, "FFFFFF"))


def _breiten(ws, mapping: dict[int, int]) -> None:
    for col, w in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = w
