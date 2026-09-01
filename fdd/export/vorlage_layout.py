"""Liest den Aufbau eines Lead-Tabs **aus der Vorlage**, statt ihn zu kennen.

Zeilennummern sind in dieser Vorlage keine Nebensache: 2.115 Verweise aus den
übrigen Tabs zeigen auf feste Zeilen, und openpyxl zieht sie beim Einfügen
nicht mit. Der Befüllcode darf deshalb weder Zeilen einfügen noch
Zeilennummern kennen — er fragt die Vorlage, wo etwas steht.

Das hat einen zweiten Nutzen: wird die Vorlage zentral erweitert (mehr
Kontoslots, eine zusätzliche Position), folgt der Code ohne Änderung.

Erkannt wird alles an Merkmalen, die die Vorlage selbst trägt:

* Zeilentyp ``POS``/``KTO`` in der dritten Ticker-Spalte,
* Dummy-Zeilen an der Beschriftung ``Dummy n``,
* Summenzeilen an der Formel ``SUMIF(<Zeilentyp>,"<>KTO",...)``; ihr
  Zeilenbereich definiert den Block.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl.utils import get_column_letter

#: ``=SUMIF($R$40:$R$78,"<>KTO",E40:E78)`` -> (40, 78)
_SUMIF = re.compile(r'SUMIF\(\$[A-Z]+\$(\d+):\$[A-Z]+\$(\d+),\s*"<>KTO"', re.I)
_DUMMY = re.compile(r"^\s*Dummy\s*\d*\s*$", re.I)


@dataclass
class Position:
    """Eine Positionszeile mit ihren eingeklappten Kontoslots."""

    zeile: int
    ticker1: str
    ticker2: str
    slots: list[int] = field(default_factory=list)
    aus_dummy: bool = False

    @property
    def schluessel(self) -> tuple[str, str]:
        return (self.ticker1 or "", self.ticker2 or "")


@dataclass
class Block:
    """Ein Summenbereich: von/bis und die Zeile mit der Summe."""

    von: int
    bis: int
    summenzeile: int
    titel: str
    positionen: list[Position] = field(default_factory=list)
    dummies: list[int] = field(default_factory=list)

    @property
    def klassen(self) -> list[str]:
        return list(dict.fromkeys(p.ticker2 for p in self.positionen if p.ticker2))

    def freier_dummy(self) -> Optional[int]:
        return self.dummies[0] if self.dummies else None


@dataclass
class LeadLayout:
    """Der gelesene Aufbau eines Lead-Tabs."""

    blatt: str
    ticker: tuple[str, str, str]        # Spaltenbuchstaben Ticker1/2/Zeilentyp
    erste_spalte: int                   # erste Periodenspalte (E = 5)
    letzte_spalte: int                  # letzte Periodenspalte
    bloecke: list[Block] = field(default_factory=list)

    @property
    def positionen(self) -> list[Position]:
        return [p for b in self.bloecke for p in b.positionen]

    def finde(self, ticker1: str, ticker2: str) -> Optional[Position]:
        for p in self.positionen:
            if p.schluessel == (ticker1, ticker2):
                return p
        return None

    def block_mit_klasse(self, klasse: str) -> Optional[Block]:
        for b in self.bloecke:
            if klasse in b.klassen:
                return b
        return None

    def block_mit_titel(self, titel: str) -> Optional[Block]:
        for b in self.bloecke:
            if (b.titel or "").strip().lower() == titel.strip().lower():
                return b
        return None


def _zeilentyp_spalte(ws, spalte: str, zeile: int) -> Optional[str]:
    v = ws[f"{spalte}{zeile}"].value
    return str(v).strip() if v is not None else None


def lies_layout(ws, ticker: tuple[str, str, str], erste_spalte: int,
                letzte_spalte: int) -> LeadLayout:
    """Baut das Layout eines Lead-Tabs aus der Vorlage."""
    t1, t2, t3 = ticker
    wert = get_column_letter(erste_spalte)
    layout = LeadLayout(ws.title, ticker, erste_spalte, letzte_spalte)

    # 1. Summenzeilen finden -> Blöcke. Ein Block endet an seiner Summenzeile;
    #    verschachtelte Summen (Rohergebnis über Gesamtleistung) beginnen erst
    #    hinter der vorigen Summe, deshalb wird der Bereich beschnitten.
    roh: list[tuple[int, int, int, str]] = []
    for r in range(1, ws.max_row + 1):
        f = ws[f"{wert}{r}"].value
        if not isinstance(f, str):
            continue
        m = _SUMIF.search(f)
        if m:
            roh.append((int(m.group(1)), int(m.group(2)), r,
                        str(ws[f"C{r}"].value or "")))
    roh.sort(key=lambda x: x[2])

    verbraucht = 0
    for von, bis, summenzeile, titel in roh:
        von = max(von, verbraucht + 1)
        if von > bis:
            continue
        layout.bloecke.append(Block(von, bis, summenzeile, titel))
        verbraucht = max(verbraucht, bis)

    # 2. Positionen mit ihren Kontoslots aufnehmen. Die Slots werden bewusst
    #    ohne Rücksicht auf die Blockgrenze gezählt: bei ``Sonstige Steuern``
    #    liegen sie in der Vorlage hinter der Summenzeile. Für die Summe ist
    #    das folgenlos ("<>KTO"), für die Vollständigkeitskontrolle nicht.
    positionen: list[Position] = []
    r = 1
    while r <= ws.max_row:
        if _zeilentyp_spalte(ws, t3, r) == "POS":
            p = Position(r, str(ws[f"{t1}{r}"].value or ""),
                         str(ws[f"{t2}{r}"].value or ""))
            s = r + 1
            while s <= ws.max_row and _zeilentyp_spalte(ws, t3, s) == "KTO":
                p.slots.append(s)
                s += 1
            positionen.append(p)
            r = s
            continue
        r += 1

    # 3. Positionen und Dummy-Zeilen den Blöcken zuordnen.
    for b in layout.bloecke:
        b.positionen = [p for p in positionen if b.von <= p.zeile <= b.bis]
        b.dummies = [r for r in range(b.von, b.bis + 1)
                     if _DUMMY.match(str(ws[f"C{r}"].value or ""))]
    return layout
