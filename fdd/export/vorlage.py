"""Befüllt die Dealtool-Vorlage. Baut kein Databook mehr.

Das ist die eigentliche Umstellung: bisher hat der Code Blätter erzeugt und
formatiert, jetzt kopiert er die Hausvorlage und schreibt Werte in vorhandene
Zellen. Kein ``openpyxl.Workbook()``, keine ``Font``, kein ``PatternFill``,
keine Spaltenbreite. Theme-Farben und Zeilenhöhen lassen sich aus einer
Beschreibung nicht rekonstruieren; eine nachgebaute Kopfzeile wird weiß statt
dunkel. Wo dieser Code doch einen Stil setzt, kopiert er ihn mit
``cell._style = copy.copy(quelle._style)`` aus einer Nachbarzeile **derselben**
Arbeitsmappe.

Drei Regeln, die den Aufbau bestimmen:

1. **Keine Zeileneinschübe.** 2.115 Verweise aus den übrigen Tabs zeigen auf
   feste Zeilennummern. Eine neue Position entsteht durch Umbenennen einer
   Dummy-Zeile, eine überzählige Position ist ein Befund für die Review-Queue.
2. **Ticker nie tippen.** Ticker 1 und die Kontonummern stammen aus dem
   Mastersheet. Eine abweichende Schreibweise liefert stumm null.
3. **Summen immer als** ``SUMIF(<Zeilentyp>;"<>KTO";...)``. Die Kontoslots
   liegen innerhalb der Summenbereiche; ein blankes ``SUM`` zählt sie doppelt.
"""

from __future__ import annotations

import copy
import os
import re
import shutil
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from ..core.hausconvention import normalisiere
from ..core.model import Klasse, MappedAccount
from . import vorlage_zuordnung as zu
from .aufrisse import schreibe_aufrisse
from .vorlage_layout import Block, LeadLayout, Position, lies_layout

VORLAGE = os.path.join(os.path.dirname(__file__), "..", "vorlagen",
                       "Dealtool_Template_v4_1.xlsx")

#: Erste Periodenspalte des Mastersheets (F). Lead NA Spalte E und Lead PL
#: Spalte E entsprechen beide dieser Spalte.
MS_ERSTE = 6
MS_LETZTE = 15                      # O
MS_ERSTE_ZEILE = 2
#: Die SUMIFS der Vorlage reichen bis Mastersheet-Zeile 400.
MS_LETZTE_ZEILE = 400

_TICKER_NA = ("P", "Q", "R")
_TICKER_PL = ("T", "U", "V")


# --------------------------------------------------------------------------
# Zeitachse
# --------------------------------------------------------------------------

@dataclass(frozen=True)
class Periode:
    """Eine unserer Perioden mit ihrem Stichtag.

    ``art`` unterscheidet volle Geschäftsjahre von einer Zwischenperiode. Die
    Vorlage behandelt beide verschieden: das Zwischenjahr ist in der Bilanz
    die CYT-Spalte, in der GuV die YTD-Spalte — und das sind **nicht**
    dieselben Spaltennummern.
    """

    label: str
    stichtag: date
    art: str = "jahr"               # "jahr" | "zwischen"


@dataclass
class Zeitachse:
    """Bildet unsere Perioden auf die Spalten der Vorlage ab.

    Die Vorlage rechnet aus Cockpit C17 (Ende der ersten historischen Periode)
    und C18 (Anzahl) drei Zeitachsen aus. Die Bilanzachse hat neun, die
    GuV-Achse zehn Spalten, und ab der Zwischenperiode laufen sie
    auseinander: in der Bilanz folgt auf die Jahre direkt der CYT-Stichtag, in
    der GuV liegt davor noch die LTM-Spalte. Genau daraus stammte der
    Off-by-one im Equity Roll Forward.
    """

    perioden: list[Periode]

    def __post_init__(self) -> None:
        self.jahre = sorted((p for p in self.perioden if p.art == "jahr"),
                            key=lambda p: p.stichtag)
        self.zwischen = sorted((p for p in self.perioden if p.art == "zwischen"),
                               key=lambda p: p.stichtag)
        if not self.jahre:
            raise ValueError("Die Vorlage braucht mindestens eine volle "
                             "Jahresperiode als historische Basis.")
        if len(self.zwischen) > 1:
            raise ValueError("Die Vorlage hält genau eine Zwischenperiode "
                             "(CYT/YTD) vor, geliefert wurden "
                             f"{len(self.zwischen)}.")
        if len(self.jahre) > 7:
            raise ValueError(
                f"{len(self.jahre)} Jahresperioden passen nicht in die "
                "Vorlage (Bilanz: neun Spalten, davon eine Eröffnungs- und "
                "eine CYT-Spalte). Die Vorlage wird zentral erweitert, nicht "
                "je Mandat.")

    @property
    def anzahl_historisch(self) -> int:
        return len(self.jahre)

    @property
    def ende_erste_periode(self) -> date:
        return self.jahre[0].stichtag

    @property
    def cyt_beginn(self) -> date:
        if self.zwischen:
            return date(self.zwischen[0].stichtag.year, 1, 1)
        return self._platzhalter()[0]

    @property
    def cyt_ende(self) -> date:
        if self.zwischen:
            return self.zwischen[0].stichtag
        return self._platzhalter()[1]

    def _platzhalter(self) -> tuple[date, date]:
        """CYT-Daten, wenn das Mandat keine Zwischenperiode liefert.

        Das Cockpit prüft die beiden Felder selbst: der CYT-Beginn darf nicht
        vor dem historischen Zeitraum liegen und der CYT-Stichtag nicht der
        31.12. sein. Bleiben die Werte der Vorlage stehen, meldet die Mappe
        zwei Fehler, obwohl nichts fehlt — sie gehören deshalb hinter das
        letzte historische Jahr gelegt. Befüllt wird die Spalte nicht.
        """
        letztes = self.jahre[-1].stichtag
        beginn = date.fromordinal(letztes.toordinal() + 1)
        monat = beginn.month % 12 + 1
        jahr = beginn.year + (beginn.month == 12)
        ende = date.fromordinal(date(jahr + (monat == 12),
                                     monat % 12 + 1, 1).toordinal() - 1)
        return beginn, ende

    def _index(self, label: str, guv: bool) -> Optional[int]:
        """Zähler der Vorlage (0 = Eröffnungsspalte) für diese Periode."""
        for i, p in enumerate(self.jahre, start=1):
            if p.label == label:
                return i
        if self.zwischen and self.zwischen[0].label == label:
            # Bilanz: der CYT-Stichtag folgt unmittelbar auf die Jahre.
            # GuV: davor liegt die LTM-Spalte, die wir nicht befüllen.
            return self.anzahl_historisch + (2 if guv else 1)
        return None

    def ms_spalte(self, label: str, guv: bool) -> Optional[int]:
        i = self._index(label, guv)
        if i is None:
            return None
        spalte = MS_ERSTE + i
        if spalte > MS_LETZTE:
            raise ValueError(f"Periode {label} liegt hinter Mastersheet-Spalte O.")
        return spalte

    def lead_spalte(self, label: str, erste: int, guv: bool) -> Optional[int]:
        """Spalte im Lead-Tab. Lead-Spalte E entspricht Mastersheet F."""
        i = self._index(label, guv)
        return None if i is None else erste + i


def stichtag(label: str) -> tuple[date, str]:
    """Leitet Stichtag und Art aus einem Periodenlabel ab.

    Erkannt werden die Schreibweisen, die unsere Reader erzeugen: ``FY2024``,
    ``YTD 07/2026``, ``2024/12`` und ``31.12.2024``. Ein Stichtag zum 31.12.
    ist eine Jahresperiode, jeder andere eine Zwischenperiode.
    """
    def ende(jahr: int, monat: int) -> tuple[date, str]:
        naechster = date(jahr + (monat == 12), monat % 12 + 1, 1)
        tag = date.fromordinal(naechster.toordinal() - 1)
        return tag, "jahr" if monat == 12 else "zwischen"

    t = label.strip()
    m = re.fullmatch(r"FY\s*(\d{4})", t, re.I)
    if m:
        return ende(int(m.group(1)), 12)
    m = re.fullmatch(r"(?:YTD|YTG)\s*(\d{1,2})[/.](\d{4})", t, re.I)
    if m:
        return ende(int(m.group(2)), int(m.group(1)))
    m = re.fullmatch(r"(\d{4})[/.-](\d{1,2})", t)
    if m:
        return ende(int(m.group(1)), int(m.group(2)))
    m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", t)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1))), \
            ("jahr" if (m.group(1), m.group(2)) == ("31", "12") else "zwischen")
    raise ValueError(f"Periodenlabel nicht lesbar: {label!r}")


def zeitachse(perioden: list[str]) -> Zeitachse:
    return Zeitachse([Periode(p, *stichtag(p)) for p in perioden])


# --------------------------------------------------------------------------
# Mastersheet
# --------------------------------------------------------------------------

@dataclass
class MSZeile:
    """Eine Zeile des Mastersheets — die einzige Datenquelle der Mappe."""

    schluessel: str                 # Spalte A, zugleich Ticker 1 des Kontoslots
    konto: str
    bezeichnung: str
    klasse: str                     # Ticker 2 der Position
    na_zeile: str                   # Ticker 1 der Position (Vokabular der Vorlage)
    review: bool
    guv: bool
    werte: dict[str, float] = field(default_factory=dict)
    hinweis: str = ""
    #: Unser eigenes Positionsvokabular. Es wird für die Zuordnungstabelle
    #: mitgeführt: ohne es steht dort zweimal derselbe Name und die Abweichung,
    #: um die es geht, ist unsichtbar.
    na_de: str = ""
    na_en: str = ""

    @property
    def gruppe(self) -> tuple[str, str]:
        return (self.na_zeile, self.klasse)


def baue_mastersheet_zeilen(mapped: list[MappedAccount], perioden: list[str]
                            ) -> tuple[list[MSZeile], list[MSZeile]]:
    """Übersetzt das Datenmodell in Mastersheet-Zeilen.

    Zwei Dinge passieren hier, und nur hier:

    * **Vorzeichen der GuV.** Unsere Konvention ist Soll positiv, damit eine
      Summen- und Saldenliste zu null aufgeht. Die Vorlage rechnet
      ``Rohergebnis = Gesamtleistung + Materialaufwand`` und braucht deshalb
      Erträge positiv und Aufwendungen negativ. Die GuV wird gedreht, die
      Bilanz nicht. Ohne diese Drehung stimmt jede Zwischensumme der GuV im
      Vorzeichen nicht.
    * **Seitenwechsel.** Ein Konto, das je Periode auf einer anderen
      Bilanzseite steht, bekommt je Position eine eigene Zeile. Die zweite
      Zeile trägt einen ergänzten Schlüssel, damit der Kontoslot der einen
      Position nicht die Werte der anderen mitzählt.

    Zurück kommen zwei Listen: die Zeilen fürs Mastersheet und die Konten
    **ohne Saldo in irgendeiner Periode**. Letztere tragen zu keiner Position
    bei und belegen nur Zeilen — bei Luma sind es 192 von 434, und die
    SUMIFS der Vorlage reichen bis Zeile 400. Sie verschwinden nicht, sondern
    werden ausgewiesen; die Vorlage wird deswegen nicht angefasst.
    """
    zeilen: list[MSZeile] = []
    for m in mapped:
        # Welche Position belegt das Konto in welcher Periode?
        je_gruppe: dict[str, dict[str, float]] = {}
        namen: dict[str, str] = {}
        for p in perioden:
            na, na_en = m.na_in(p)
            je_gruppe.setdefault(na, {})[p] = m.saldo(p)
            namen[na] = na_en

        klasse = zu.ziel_klasse(m)
        guv = m.klasse is Klasse.PL
        for i, (na_de, werte) in enumerate(je_gruppe.items()):
            ziel_na, _, _ = zu.ziel_position(na_de)
            schluessel = m.konto if i == 0 else f"{m.konto} ({na_de})"
            zeilen.append(MSZeile(
                schluessel=schluessel, konto=m.konto, bezeichnung=m.bezeichnung,
                klasse=klasse, na_zeile=ziel_na, review=m.review, guv=guv,
                werte={p: (-v if guv else v) for p, v in werte.items()},
                hinweis=("Seitenwechsel: eigene Zeile je Bilanzseite"
                         if len(je_gruppe) > 1 else ""),
                na_de=na_de, na_en=namen.get(na_de, "")))

    belegt = [z for z in zeilen
              if any(abs(v) > 0.005 for v in z.werte.values())]
    leer = [z for z in zeilen if z not in belegt]
    return belegt, leer


def _schreibe_mastersheet(ws, zeilen: list[MSZeile], ach: Zeitachse) -> int:
    if len(zeilen) > MS_LETZTE_ZEILE - MS_ERSTE_ZEILE + 1:
        raise ValueError(
            f"{len(zeilen)} Mastersheet-Zeilen, die SUMIFS der Vorlage reichen "
            f"nur bis Zeile {MS_LETZTE_ZEILE}. Die Vorlage wird zentral "
            "erweitert, nicht je Mandat.")

    geschrieben = 0
    for i, z in enumerate(zeilen):
        r = MS_ERSTE_ZEILE + i
        ws.cell(r, 1, z.schluessel)
        ws.cell(r, 2, z.bezeichnung)
        ws.cell(r, 3, z.klasse)
        ws.cell(r, 4, z.na_zeile)
        ws.cell(r, 5, "Review" if z.review else None)
        for spalte in range(MS_ERSTE, MS_LETZTE + 1):
            ws.cell(r, spalte).value = None
        for periode, wert in z.werte.items():
            spalte = ach.ms_spalte(periode, z.guv)
            if spalte is not None:
                ws.cell(r, spalte, round(wert, 2))
        geschrieben += 5 + (MS_LETZTE - MS_ERSTE + 1)

    # Reste einer früheren Befüllung entfernen — die Vorlage bringt eine
    # Platzhalterzeile mit.
    for r in range(MS_ERSTE_ZEILE + len(zeilen), MS_LETZTE_ZEILE + 1):
        if any(ws.cell(r, c).value is not None for c in range(1, MS_LETZTE + 1)):
            for c in range(1, MS_LETZTE + 1):
                ws.cell(r, c).value = None
    return geschrieben


# --------------------------------------------------------------------------
# Lead-Tabs
# --------------------------------------------------------------------------

@dataclass
class Slotbefund:
    """Eine Position, deren Kontoslots nicht reichen — Befund, kein Einschub.

    Zwei Ausprägungen, die auseinanderzuhalten sind: einer Position der
    Vorlage reichen die acht Slots nicht, oder eine über eine Dummy-Zeile
    angelegte Position hat gar keine Slots. Das zweite ist keine Frage der
    Anzahl, sondern eine Lücke der Vorlage.
    """

    position: str
    klasse: str
    zeile: int
    konten: int
    slots: int
    aus_dummy: bool = False
    #: Die Kontoschlüssel, für die kein Slot mehr da war. Ohne sie ließe sich
    #: nicht beziffern, wie viel Detail tatsächlich fehlt — die Position selbst
    #: zeigt ja weiterhin den vollen Betrag.
    ohne_slot: list[str] = field(default_factory=list)

    @property
    def fehlend(self) -> int:
        return self.konten - self.slots

    @property
    def art(self) -> str:
        return ("Dummy-Position ohne Kontoslots" if self.aus_dummy
                else "Kontoslots reichen nicht")


@dataclass
class Zuordnungszeile:
    """Protokoll: welche unserer Positionen landet auf welcher Vorlagenzeile."""

    na_de: str
    klasse: str
    ziel_na: str
    zeile: Optional[int]
    art: str                    # "Vorlage" | "Dummy" | "ohne Zeile"
    konten: int
    grund: str = ""
    #: Die Kontoschlüssel der Gruppe. Eine Position ohne Kontoslots verliert
    #: sonst ihr Detail, und genau dort steckt bei Brehna das meiste Geld.
    schluessel: str = ""


def _periodenspalten(ws) -> tuple[int, int]:
    """Erste und letzte Periodenspalte, abgelesen an der Kopfzeile.

    Die Kopfzeile verweist je Spalte auf die Zeitachse im Cockpit. Damit
    ergibt sich die Spaltenzahl aus der Vorlage und nicht aus einer Annahme —
    Bilanz und GuV haben unterschiedlich viele.
    """
    erste, letzte = 5, 4
    for c in range(erste, ws.max_column + 1):
        v = ws.cell(4, c).value
        if isinstance(v, str) and v.startswith("=Cockpit!"):
            letzte = c
        elif letzte >= erste:
            break
    return erste, letzte


#: Woran sich eine neu angelegte Position von der gleichnamigen in der Vorlage
#: unterscheidet. Beispiel Brehna: ``Sonstige Verbindlichkeiten`` gibt es in
#: den sonstigen Passiva und — als thereof-ND-Anteil — noch einmal im Net
#: Debt. Ohne den Zusatz stünde derselbe Name zweimal im selben Blatt.
_KLASSENZUSATZ = {
    "FA": ("Anlagevermögen", "fixed assets"),
    "TWC": ("Trade Working Capital", "trade working capital"),
    "OA": ("Sonstige Aktiva", "other assets"),
    "OL": ("Sonstige Passiva", "other liabilities"),
    "ND": ("Net Debt", "net debt"),
    "DT": ("Latente Steuern", "deferred tax"),
}


def _beschriftung(ws, layout: LeadLayout, ziel_na: str, klasse: str,
                  muster: MSZeile) -> tuple[str, str]:
    """Beschriftung für eine über eine Dummy-Zeile angelegte Position.

    Unser Positionsvokabular ist durchgehend transliteriert
    (``Vermoegensgegenstaende``, ``ggue.``) und taugt nicht als Berichtszeile.
    Deshalb wird zuerst gesucht, ob die Vorlage dieselbe Position auf einer
    anderen Seite bereits führt — dann steht die richtige Schreibweise schon
    da, und der Zusatz sagt, worin sich die neue Zeile unterscheidet.
    """
    for p in layout.positionen:
        if p.ticker1 != ziel_na or p.ticker2 == klasse or p.aus_dummy:
            continue
        de, en = ws.cell(p.zeile, 3).value, ws.cell(p.zeile, 4).value
        zusatz = _KLASSENZUSATZ.get(klasse)
        if de and zusatz:
            return f"{de} ({zusatz[0]})", f"{en or de} ({zusatz[1]})"

    ab = zu.ABWEICHUNGEN.get(muster.na_de)
    if ab and ab.label_de:
        return ab.label_de, ab.label_en or ab.label_de
    return muster.na_de or ziel_na, muster.na_en or ziel_na


def _aktiviere_dummy(ws, layout: LeadLayout, block: Block, dummy: int,
                     ziel_na: str, klasse: str, label_de: str, label_en: str
                     ) -> Position:
    """Macht aus einer Dummy-Zeile eine Position. Ohne Zeileneinschub.

    Die Formel wird aus einer vorhandenen Position desselben Blocks
    übernommen und nur der Zeilenbezug der beiden Ticker umgehängt. So kann
    sie gar nicht von der Form der Vorlage abweichen.

    Das Format der Dummy-Zeile bleibt unangetastet. Sie ist bereits als
    Positionszeile gestaltet — ein Stilübertrag von der Nachbarposition hat im
    Brehna-Lauf die Abschlusslinie der letzten Zeile vor der Summe gelöscht.
    """
    t1, t2, t3 = layout.ticker
    referenz = block.positionen[0]

    for c in range(layout.erste_spalte, layout.letzte_spalte + 1):
        formel = ws.cell(referenz.zeile, c).value
        if isinstance(formel, str) and formel.startswith("="):
            for spalte in (t1, t2):
                formel = re.sub(rf"\${spalte}\$?{referenz.zeile}\b",
                                f"${spalte}{dummy}", formel)
            ws.cell(dummy, c, formel)

    ws.cell(dummy, 3, label_de)
    ws.cell(dummy, 4, label_en)
    ws[f"{t1}{dummy}"] = ziel_na
    ws[f"{t2}{dummy}"] = klasse
    ws[f"{t3}{dummy}"] = "POS"
    ws.row_dimensions[dummy].hidden = False

    # Für Ticker 2 und den Zeilentyp hält die Vorlage in der Dummy-Zeile gar
    # keine Zelle vor; eine neu angelegte bekäme die Standardschrift statt der
    # der übrigen Tickerzellen. Nur dieser eine Fall überträgt einen Stil —
    # aus der Nachbarposition und innerhalb derselben Arbeitsmappe.
    for spalte in (t2, t3):
        quelle = ws[f"{spalte}{referenz.zeile}"]
        ziel = ws[f"{spalte}{dummy}"]
        if quelle._style is not None:
            ziel._style = copy.copy(quelle._style)

    block.dummies.remove(dummy)
    pos = Position(dummy, ziel_na, klasse, aus_dummy=True)
    block.positionen.append(pos)
    return pos


#: Meldung, wenn eine Position weder eine Zeile der Vorlage noch eine freie
#: Dummy-Zeile findet. Sie landet in der Zuordnungstabelle und in einer
#: Kontrollzeile, deshalb steht sie in der Berichtssprache.
_OHNE_ZEILE = {
    "de": ("Kein Block bzw. keine freie Dummy-Zeile — Befund für die "
           "Review-Queue, kein Zeileneinschub."),
    "en": ("No block or no free dummy row — a finding for the review queue, "
           "not a row insert."),
}


def _fuelle_lead(ws, zeilen: list[MSZeile], guv: bool, sprache: str = "de"
                 ) -> tuple[LeadLayout, list[Zuordnungszeile], list[Slotbefund], int]:
    """Trägt Positionen und Kontoslots in einen Lead-Tab ein."""
    ticker = _TICKER_PL if guv else _TICKER_NA
    erste, letzte = _periodenspalten(ws)
    layout = lies_layout(ws, ticker, erste, letzte)
    t1 = ticker[0]

    # Die Positionszeile steht oben, die Kontoslots eingeklappt darunter.
    ws.sheet_properties.outlinePr.summaryBelow = False

    ohne_position = {k.value for k in zu.OHNE_POSITION}
    gruppen: dict[tuple[str, str], list[MSZeile]] = {}
    for z in zeilen:
        # Eigenkapital, technische und ungelöste Konten haben in der
        # Net-Asset-Sicht keine Position. Sie stehen im Mastersheet, damit es
        # vollständig bleibt, und laufen bewusst in keine Lead-Zeile.
        if z.klasse in ohne_position:
            continue
        gruppen.setdefault(z.gruppe, []).append(z)

    protokoll: list[Zuordnungszeile] = []
    befunde: list[Slotbefund] = []
    geschrieben = 0

    for (ziel_na, klasse), konten in sorted(gruppen.items()):
        pos = layout.finde(ziel_na, klasse)
        art = "Vorlage"
        grund = ""
        if pos is None:
            # Keine Zeile in der Vorlage: über eine Dummy-Zeile anlegen.
            _, block_hinweis, grund = zu.ziel_position(konten[0].na_de)
            block = (layout.block_mit_titel(block_hinweis) if block_hinweis
                     else layout.block_mit_klasse(klasse))
            dummy = block.freier_dummy() if block else None
            if dummy is None:
                protokoll.append(Zuordnungszeile(
                    konten[0].na_de, klasse, ziel_na, None, "ohne Zeile",
                    len(konten),
                    _OHNE_ZEILE["de" if sprache.lower().startswith("d")
                                else "en"]))
                continue
            label_de, label_en = _beschriftung(ws, layout, ziel_na, klasse,
                                               konten[0])
            pos = _aktiviere_dummy(ws, layout, block, dummy, ziel_na, klasse,
                                   label_de, label_en)
            art = "Dummy"
            geschrieben += 3 + (letzte - erste + 1)

        # Reichen die Slots nicht, entscheidet die Größe, welche Konten
        # sichtbar bleiben. Die Reihenfolge des Mastersheets wäre die
        # Kontonummer, und dann verschwände womöglich das größte Konto der
        # Position hinter acht kleinen.
        konten.sort(key=lambda z: max((abs(v) for v in z.werte.values()),
                                      default=0.0), reverse=True)
        for i, z in enumerate(konten[:len(pos.slots)]):
            slot = pos.slots[i]
            ws[f"{t1}{slot}"] = z.schluessel
            ws.cell(slot, 3, z.bezeichnung)
            ws.cell(slot, 4, z.bezeichnung)
            geschrieben += 3
        for slot in pos.slots[len(konten):]:
            ws[f"{t1}{slot}"] = None
            ws.cell(slot, 3).value = None
            ws.cell(slot, 4).value = None

        if len(konten) > len(pos.slots):
            befunde.append(Slotbefund(
                ziel_na, klasse, pos.zeile, len(konten), len(pos.slots),
                pos.aus_dummy,
                [z.schluessel for z in konten[len(pos.slots):]]))
        protokoll.append(Zuordnungszeile(
            konten[0].na_de, klasse, ziel_na, pos.zeile, art, len(konten),
            grund, ", ".join(z.schluessel for z in konten)))
    return layout, protokoll, befunde, geschrieben


# --------------------------------------------------------------------------
# Equity Roll Forward
# --------------------------------------------------------------------------

#: Zuordnung der Eigenkapitalkonten auf die Bewegungszeilen des Roll Forward.
#:
#: Die englischen Stichworte stehen gleichberechtigt daneben: ein
#: MYOB- oder NetSuite-Kontenplan trägt ``Ordinary Shares`` und ``Retained
#: Earnings``, und eine Zuordnung, die nur deutsche Bezeichnungen kennt,
#: liefert für ein solches Mandat eine leere Fortschreibung.
_EK_ZEILEN = (
    ("∆ Gezeichnetes Kapital", ("gezeichnetes kapital", "stammkapital",
                                "kapitalkonto", "festkapital",
                                "share capital", "shares", "warrant",
                                "equity interest")),
    ("∆ Kapitalrücklage", ("kapitalruecklage", "ruecklage",
                           "equity costs", "share premium")),
    ("Gewinnausschüttung", ("ausschuettung", "entnahme", "dividende",
                            "distribution", "dividend")),
)

#: Die Klassen, aus denen sich das Nettovermögen zusammensetzt.
_NETTO = (Klasse.FA, Klasse.TWC, Klasse.OWC, Klasse.ND, Klasse.DT)

#: Konten, deren Bewegung das Periodenergebnis ist. Sie bleiben hier außen
#: vor, weil das Ergebnis über das Lead PL in den Roll Forward kommt.
#:
#: Der Gewinn-/Verlustvortrag gehört ausdrücklich dazu: seine Bewegung ist
#: nicht die Ausschüttung, sondern die Umbuchung des Vorjahresergebnisses
#: innerhalb des Eigenkapitals. Wer ihn als Bewegung mitzählt, zieht das
#: Vorjahresergebnis ein zweites Mal ab — im Brehna-Lauf ging die Check-Zeile
#: dadurch Jahr für Jahr genau um das Vorjahresergebnis daneben. Steckt in der
#: Bewegung des Vortragskontos doch eine echte Ausschüttung, zeigt die
#: Check-Zeile sie als Rest an; sie gehört dann in die manuelle Zeile.
_EK_ERGEBNIS = ("jahresueberschuss", "jahresfehlbetrag", "periodenergebnis",
                "jahresergebnis", "gewinnvortrag", "verlustvortrag",
                "bilanzgewinn", "bilanzverlust",
                "retained earnings", "current earnings", "accumulated losses",
                "historical balancing")


@dataclass
class RollForward:
    """Ergebnis der Eigenkapitalfortschreibung, unabhängig von Excel gerechnet."""

    bewegungen: dict[str, dict[str, float]] = field(default_factory=dict)
    nicht_zugeordnet: list[str] = field(default_factory=list)
    hinweise: list[str] = field(default_factory=list)
    #: Was die benannten Bewegungen und das Periodenergebnis nicht erklären.
    #: Nur belegt, wenn ein Periodenergebnis mitgegeben wurde.
    rest: dict[str, float] = field(default_factory=dict)
    #: Nettovermögen zu Beginn der ersten Periode, sofern das Mandat eine
    #: Vorgeschichte hat. ``None``, wenn die erste Periode zugleich die erste
    #: des Unternehmens ist.
    anfangsbestand: Optional[float] = None
    #: Bewegung je ergebnisbezogenem Eigenkapitalkonto (Vortrag, laufendes
    #: Ergebnis). Aus ihnen setzt sich der ``rest`` zusammen — ohne diese
    #: Aufstellung bliebe er eine Zahl ohne Adresse.
    ergebniskonten: dict[str, dict[str, float]] = field(default_factory=dict)

    @property
    def hat_rest(self) -> bool:
        return any(abs(v) > 1.0 for v in self.rest.values())


def _roll_forward(mapped: list[MappedAccount], perioden: list[str],
                  periodenergebnis: Optional[dict[str, float]] = None
                  ) -> RollForward:
    """Eigenkapitalbewegungen je Periode.

    Nettovermögen ist das Eigenkapital mit umgekehrtem Vorzeichen, deshalb
    wird die Bewegung negiert. Für die erste Periode gibt es keine Vorperiode;
    die Bewegung wird dann gegen null gerechnet und das ausdrücklich vermerkt.

    ``periodenergebnis`` ist für Quellen gedacht, die **keine GuV** liefern
    und das Ergebnis auf einem Eigenkapitalkonto führen. Dann kann die
    Fortschreibung nicht aus dem Lead PL kommen, und was Kapitalbewegungen und
    Ergebnis zusammen nicht erklären, wird als ``rest`` ausgewiesen, statt die
    Check-Zeile ohne Erklärung auseinanderlaufen zu lassen.
    """
    rf = RollForward()
    ek = [m for m in mapped if m.klasse is Klasse.EQ]
    for titel, _ in _EK_ZEILEN:
        rf.bewegungen[titel] = {p: 0.0 for p in perioden}

    for m in ek:
        text = normalisiere(f"{m.bezeichnung} {m.hgb_pfad}")
        if any(s in text for s in _EK_ERGEBNIS):
            # Kommt über das Lead PL. Die Bewegung wird trotzdem festgehalten,
            # damit ein Rest später eine Adresse hat.
            bewegung, vorher = {}, 0.0
            for p in perioden:
                jetzt = m.saldo(p)
                bewegung[p] = round(-(jetzt - vorher), 2)
                vorher = jetzt
            if any(abs(v) > 0.005 for v in bewegung.values()):
                rf.ergebniskonten[f"{m.konto} {m.bezeichnung}"] = bewegung
            continue
        ziel = next((titel for titel, stichworte in _EK_ZEILEN
                     if any(s in text for s in stichworte)), None)
        if ziel is None:
            rf.nicht_zugeordnet.append(f"{m.konto} {m.bezeichnung}")
            continue
        vorher = 0.0
        for p in perioden:
            jetzt = m.saldo(p)
            rf.bewegungen[ziel][p] += -(jetzt - vorher)
            vorher = jetzt

    if periodenergebnis is not None:
        # Der Rest ist die Veränderung des NETTOVERMÖGENS, die weder eine
        # benannte Kapitalbewegung noch das Periodenergebnis erklärt: in aller
        # Regel Umbuchungen auf den Gewinnvortrag und Vorjahresberichtigungen.
        # Er wird ausgewiesen, nicht verteilt.
        #
        # Angesetzt wird am Nettovermögen und nicht am Eigenkapital, weil
        # beides nur ohne GuV dasselbe ist. Liegt eine GuV vor, steht das
        # Ergebnis auf den GuV-Konten und nicht mehr im Eigenkapital — der
        # Anker am Eigenkapital ginge dann Jahr für Jahr um das Ergebnis
        # daneben.
        vorher_na = 0.0
        for p in perioden:
            jetzt_na = sum(m.saldo(p) for m in mapped if m.klasse in _NETTO)
            erklaert = (sum(w[p] for w in rf.bewegungen.values())
                        + periodenergebnis.get(p, 0.0))
            rf.rest[p] = round((jetzt_na - vorher_na) - erklaert, 2)
            vorher_na = jetzt_na

        # Bleibt in der ERSTEN Periode ein Rest, dann ist das kein
        # unerklärter Vorgang, sondern schlicht die Vorgeschichte: Kapital und
        # aufgelaufene Ergebnisse, die vor dem ersten Stichtag entstanden
        # sind. Sie als Bewegung der ersten Periode auszuweisen, hieße zu
        # behaupten, das Kapital sei in diesem Jahr eingezahlt worden. Die
        # erste Spalte wird deshalb zum Anfangsbestand zusammengezogen.
        erste = perioden[0]
        if abs(rf.rest.get(erste, 0.0)) > 1.0:
            rf.anfangsbestand = round(
                sum(m.saldo(erste) for m in mapped if m.klasse in _NETTO)
                - periodenergebnis.get(erste, 0.0), 2)
            for werte in rf.bewegungen.values():
                werte[erste] = 0.0
            rf.rest[erste] = 0.0
            rf.hinweise.append(
                f"Erste Spalte ({erste}): das Mandat hat eine Vorgeschichte. "
                f"Der Anfangsbestand von {rf.anfangsbestand:,.2f} steht als "
                "Nettovermögen zu Periodenbeginn; Bewegungen weist die erste "
                "Spalte keine aus, weil die Quelle die Vorperiode nicht "
                "enthält.")

    if perioden and rf.anfangsbestand is None:
        rf.hinweise.append(
            f"Erste Spalte ({perioden[0]}): es liegt keine Vorperiode vor, "
            "die Eigenkapitalbewegung wird gegen null gerechnet.")
    return rf


def _schreibe_roll_forward(ws, layout: LeadLayout, rf: RollForward,
                           ach: Zeitachse, pl_erste: int,
                           periodenergebnis: Optional[dict[str, float]] = None,
                           aus_lead_pl: bool = True) -> int:
    """Bewegungszeilen füllen und den Verweis aufs Jahresergebnis geraderücken.

    Der Verweis in der Zeile ``Jahresergebnis`` zeigt in der Vorlage Spalte
    für Spalte auf dieselbe Spalte des Lead PL. Das stimmt für die vollen
    Jahre und ist ab der Zwischenperiode falsch: in der Bilanz steht dort der
    CYT-Stichtag, in der GuV die LTM-Spalte. Die Zuordnung wird deshalb aus
    der Zeitachse gerechnet, nicht abgezählt.
    """
    zeilen = {str(ws.cell(r, 3).value or "").strip(): r
              for r in range(1, ws.max_row + 1)}
    geschrieben = 0

    # Anfangsbestand der ersten Spalte. Die Vorlage kettet ihn aus der
    # Eröffnungsspalte heran; die ist hier leer, weil die Quelle die
    # Vorperiode nicht enthält.
    r_bop = zeilen.get("Nettovermögen (Periodenbeginn)")
    if rf.anfangsbestand is not None and r_bop is not None and ach.perioden:
        c = ach.lead_spalte(ach.jahre[0].label, layout.erste_spalte, guv=False)
        if c is not None:
            ws.cell(r_bop, c, round(rf.anfangsbestand / 1000.0, 6))
            geschrieben += 1

    for titel, werte in rf.bewegungen.items():
        r = zeilen.get(titel)
        if r is None:
            continue
        for periode, wert in werte.items():
            c = ach.lead_spalte(periode, layout.erste_spalte, guv=False)
            if c is not None:
                ws.cell(r, c, round(wert / 1000.0, 6))
                geschrieben += 1

    r = zeilen.get("Jahresergebnis")
    if r is not None:
        for periode in [p.label for p in ach.perioden]:
            c = ach.lead_spalte(periode, layout.erste_spalte, guv=False)
            c_pl = ach.lead_spalte(periode, pl_erste, guv=True)
            if c is None:
                continue
            if aus_lead_pl and c_pl is not None:
                ws.cell(r, c, f"='Lead PL'!{get_column_letter(c_pl)}210")
                geschrieben += 1
            elif periodenergebnis is not None:
                # Quelle ohne GuV: das Ergebnis steht auf einem
                # Eigenkapitalkonto. Der Verweis aufs Lead PL zeigte auf eine
                # leere Zeile und die Fortschreibung liefe um das ganze
                # Ergebnis daneben.
                ws.cell(r, c, round(periodenergebnis.get(periode, 0.0) / 1000.0, 6))
                geschrieben += 1

    # Der Roll Forward endet mit der letzten belegten Periode. Die Vorlage
    # kettet den Anfangsbestand über alle neun Spalten durch; ohne Plandaten
    # trüge die erste leere Spalte das Nettovermögen der Vorperiode gegen eine
    # leere Bilanz — und die Check-Zeile zeigte genau diesen Betrag an, obwohl
    # nichts fehlt. Auch der Verweis aufs Jahresergebnis muss weg: er zeigt in
    # den ungenutzten Bilanzspalten auf GuV-Spalten, die durchaus befüllt sind.
    belegt = {ach.lead_spalte(p.label, layout.erste_spalte, guv=False)
              for p in ach.perioden} | {layout.erste_spalte}
    bewegungszeilen = [zeilen[t] for t, _ in _EK_ZEILEN if t in zeilen]
    for zeile in [zeilen.get("Nettovermögen (Periodenbeginn)"), r,
                  *bewegungszeilen]:
        if zeile is None:
            continue
        for c in range(layout.erste_spalte, layout.letzte_spalte + 1):
            if c not in belegt:
                ws.cell(zeile, c).value = None
                geschrieben += 1
    return geschrieben


# --------------------------------------------------------------------------
# Cockpit und Sprache
# --------------------------------------------------------------------------

@dataclass
class Mandat:
    """Was der Setup-Dialog abfragt. Es landet im Cockpit, sonst nirgends."""

    projekt: str
    waehrung: str = "EUR"
    quelle_de: str = "Datenraum/ eigene Analysen"
    quelle_en: str = "Virtual Data Room/ Company Information"
    sprache: str = "de"                 # "de" | "en"
    #: ``databook_architektur`` der Hausconvention kennt zwei Formen: Option A
    #: (zwei Schichten, die Lead-Zeile zieht per ``SUMIFS`` aus dem
    #: Mastersheet) und Option B (drei Schichten, die Lead-Zeile zieht aus
    #: einem Aufriss-Tab). Beide sind verdrahtet. Vorgabewert bleibt A, weil B
    #: einen Aufrissplan je Mandat braucht — welches Konto in welche
    #: Aufrisszeile gehört, ist eine fachliche Entscheidung und kann nicht
    #: aus dem Kontenplan geraten werden.
    architektur: str = "option_a"


def _schreibe_cockpit(ws, mandat: Mandat, ach: Zeitachse) -> int:
    """Perioden und Beschriftungen stehen nur hier. Kein Tab hartcodiert sie."""
    werte = {
        "C4": mandat.projekt,
        "C5": mandat.waehrung,
        "C7": mandat.quelle_de,
        "C8": mandat.quelle_en,
        "C17": ach.ende_erste_periode,
        "C18": ach.anzahl_historisch,
        "C21": ach.cyt_beginn,
        "C22": ach.cyt_ende,
    }
    for koordinate, wert in werte.items():
        ws[koordinate] = wert
    return len(werte)


#: Die Kopfzeile eines Sprachblocks: Projektname aus dem Cockpit.
_KOPFZELLE = "=Cockpit!$C$4"


def _sprachbloecke(ws) -> Optional[tuple[range, range]]:
    """Findet die beiden Kopfblöcke eines Blattes.

    Nicht an festen Zeilennummern: der deutsche Block beginnt in der Zeile,
    deren Spalte C den Projektnamen zieht, der englische in der Zeile, deren
    Spalte D dasselbe tut. Der Abstand der beiden ist die Blocklänge. In den
    meisten Tabs stehen sie in den Zeilen 4 und 6, in ``PL_YTD`` eine Zeile
    tiefer und drei Zeilen lang — eine feste Annahme hätte dieses Blatt
    stillschweigend übersprungen.
    """
    de = en = None
    for r in range(1, 15):
        if de is None and ws.cell(r, 3).value == _KOPFZELLE:
            de = r
        if en is None and ws.cell(r, 4).value == _KOPFZELLE:
            en = r
    if de is None or en is None or en <= de:
        return None
    laenge = en - de
    return range(de, de + laenge), range(en, en + laenge)


def _sprache(wb, sprache: str) -> int:
    """Blendet den nicht gewählten Kopfblock und die Bezeichnungsspalte aus.

    Nicht übersetzen: beide Fassungen bleiben in der Datei.
    """
    deutsch = sprache.lower().startswith("d")
    betroffen = 0
    for ws in wb.worksheets:
        bloecke = _sprachbloecke(ws)
        if bloecke is None:
            continue
        block_de, block_en = bloecke
        for r in block_de:
            ws.row_dimensions[r].hidden = not deutsch
        for r in block_en:
            ws.row_dimensions[r].hidden = deutsch
        ws.column_dimensions["C"].hidden = not deutsch
        ws.column_dimensions["D"].hidden = deutsch
        betroffen += 1
    return betroffen


# --------------------------------------------------------------------------
# Arbeitsblätter (vom Formatgrundsatz ausgenommen)
# --------------------------------------------------------------------------

def _arbeitsblatt(wb, titel: str, kopf: list[str], zeilen: list[list]) -> int:
    if titel in wb.sheetnames:
        del wb[titel]
    ws = wb.create_sheet(titel)
    ws.append(kopf)
    for z in zeilen:
        ws.append(z)
    ws.freeze_panes = "A2"
    return (len(zeilen) + 1) * len(kopf)


#: ``art`` ist intern ein Schlüssel und wird erst beim Schreiben übersetzt.
_ART_EN = {"Vorlage": "template", "Dummy": "dummy row",
           "ohne Zeile": "no row"}


# --------------------------------------------------------------------------
# Orchestrierung
# --------------------------------------------------------------------------

@dataclass
class Befuellergebnis:
    pfad: str
    zeilen: list[MSZeile]
    zuordnung: list[Zuordnungszeile]
    slotbefunde: list[Slotbefund]
    roll_forward: RollForward
    zellen: int
    zeitachse: Zeitachse
    #: Konten ohne Saldo in irgendeiner Periode. Sie stehen nicht im
    #: Mastersheet, aber auf einem eigenen Blatt.
    nullzeilen: list[MSZeile] = field(default_factory=list)
    #: Option B: je Position der Vergleich Aufriss gegen Kontozeilen.
    aufrissbefunde: list = field(default_factory=list)

    @property
    def ohne_zeile(self) -> list[Zuordnungszeile]:
        """Positionen, die in keinem Lead-Tab gelandet sind.

        Der gefährlichste Zustand der ganzen Befüllung: die Konten stehen im
        Mastersheet, die Bilanz geht im Datenmodell auf, und im Lead fehlt der
        Betrag trotzdem. Er gehört deshalb in eine Kontrollzeile und nicht nur
        in eine Zeile des Zuordnungsblatts.
        """
        return [z for z in self.zuordnung if z.art == "ohne Zeile"]


def schreibe_dealtool(ziel: str, *, mapped: list[MappedAccount],
                      perioden: list[str], mandat: Mandat,
                      ergebnis_lt_quelle: Optional[dict[str, float]] = None,
                      review: Optional[list] = None,
                      zusatzblaetter: Optional[dict] = None,
                      achse: Optional[Zeitachse] = None,
                      periodenergebnis: Optional[dict[str, float]] = None,
                      aufrisse: Optional[list] = None
                      ) -> Befuellergebnis:
    """Kopiert die Vorlage und befüllt sie.

    ``achse`` übergibt die Zeitachse ausdrücklich. Ohne sie werden die
    Stichtage aus den Periodenlabels gelesen, und das geht nur gut, solange
    das Geschäftsjahr am 31.12. endet. ``FY2023`` heißt in einem australischen
    Kontenplan den 31.03.2023.

    ``periodenergebnis`` ist für Quellen ohne GuV gedacht — siehe
    :func:`_roll_forward`.
    """
    if mandat.architektur not in ("option_a", "option_b"):
        raise ValueError(
            f"Architektur {mandat.architektur!r} kennt die Hausconvention "
            "nicht. Vorgesehen sind option_a (zwei Schichten) und option_b "
            "(drei Schichten mit Aufriss-Tabs).")
    if mandat.architektur == "option_b" and not aufrisse:
        raise ValueError(
            "Option B ohne Aufrissplan: die Lead-Positionen sollen ihre Summe "
            "aus einem Aufriss ziehen, aber es ist keiner übergeben. Ein "
            "leerer Aufriss ergäbe eine Position von null.")

    os.makedirs(os.path.dirname(os.path.abspath(ziel)), exist_ok=True)
    shutil.copy(VORLAGE, ziel)
    wb = openpyxl.load_workbook(ziel)

    ach = achse or zeitachse(perioden)
    zellen = _schreibe_cockpit(wb["Cockpit"], mandat, ach)

    zeilen, nullzeilen = baue_mastersheet_zeilen(mapped, perioden)
    zellen += _schreibe_mastersheet(wb["Mastersheet"], zeilen, ach)

    na = wb["Lead NA"]
    pl = wb["Lead PL"]
    layout_na, zuo_na, slots_na, n1 = _fuelle_lead(
        na, [z for z in zeilen if not z.guv], guv=False, sprache=mandat.sprache)
    layout_pl, zuo_pl, slots_pl, n2 = _fuelle_lead(
        pl, [z for z in zeilen if z.guv], guv=True, sprache=mandat.sprache)
    zellen += n1 + n2

    # Liegt eine GuV vor, holt sich die Fortschreibung das Ergebnis über den
    # Verweis aufs Lead PL — so ist es in der Vorlage gedacht. Ohne GuV bleibt
    # nur der Wert vom Eigenkapitalkonto.
    hat_guv = any(z.guv for z in zeilen)
    # Option B: die Positionssumme kommt aus dem Aufriss. Das muss VOR dem
    # Roll Forward geschehen, damit die Fortschreibung auf den fertigen
    # Lead-Zeilen aufsetzt.
    aufrissbefunde: list = []
    if aufrisse:
        je_position = {}
        for z in zeilen:
            pos = layout_na.finde(z.na_zeile, z.klasse)
            if pos is not None:
                je_position.setdefault(pos.zeile, []).append(z)
        n3, aufrissbefunde = schreibe_aufrisse(
            wb, aufrisse, ach, lambda p: ach.ms_spalte(p, guv=False),
            layout_na.erste_spalte, je_position, perioden)
        zellen += n3

    rf = _roll_forward(mapped, perioden, periodenergebnis)
    zellen += _schreibe_roll_forward(na, layout_na, rf, ach,
                                     layout_pl.erste_spalte, periodenergebnis,
                                     aus_lead_pl=hat_guv)

    # Gegenprobe der GuV gegen die Quelldatei (Zeile "Jahresergebnis lt.
    # Quelldatei"). Ohne sie ist die Check-Zeile darunter sinnlos.
    if ergebnis_lt_quelle:
        r = next((r for r in range(1, pl.max_row + 1)
                  if str(pl.cell(r, 3).value or "").startswith(
                      "Jahresergebnis lt.")), None)
        if r is not None:
            for periode, wert in ergebnis_lt_quelle.items():
                c = ach.lead_spalte(periode, layout_pl.erste_spalte, guv=True)
                if c is not None:
                    pl.cell(r, c, round(wert / 1000.0, 6))
                    zellen += 1

    zellen += _sprache(wb, mandat.sprache)

    # Auch die Arbeitsblätter folgen der Berichtssprache. Sie sind zwar
    # Arbeitspapiere, gehen aber mit derselben Datei an denselben Leser.
    deutsch = mandat.sprache.lower().startswith("d")
    zuordnung = zuo_na + zuo_pl
    zellen += _arbeitsblatt(
        wb, "Zuordnung" if deutsch else "Mapping",
        (["Position (unser Vokabular)", "Klasse", "Position (Vorlage)",
          "Zeile", "Art", "Konten", "Kontoschlüssel",
          "Begründung der Abweichung"] if deutsch else
         ["Line item (our vocabulary)", "Class", "Line item (template)",
          "Row", "Type", "Accounts", "Account keys",
          "Reason for the deviation"]),
        [[z.na_de, z.klasse, z.ziel_na, z.zeile,
          z.art if deutsch else _ART_EN.get(z.art, z.art), z.konten,
          z.schluessel, z.grund] for z in zuordnung])

    if review:
        zellen += _arbeitsblatt(
            wb, "Review-Queue" if deutsch else "Review queue",
            (["Konto", "Bezeichnung", "Klasse", "Status", "Grund"] if deutsch
             else ["Account", "Description", "Class", "Status", "Reason"]),
            [[r.konto, r.bezeichnung, getattr(r, "klasse", ""),
              getattr(r, "status", ""), getattr(r, "grund", "")]
             for r in review])

    # Arbeitsblätter (QA, Status, Verhaltensprüfung …). Sie sind vom
    # Formatgrundsatz ausgenommen und werden deshalb schlicht angelegt.
    for titel, (kopf, inhalt) in (zusatzblaetter or {}).items():
        zellen += _arbeitsblatt(wb, titel, kopf, inhalt)

    # Pflichtkontrolle der Option B. Sie steht hier und nicht im Runner, weil
    # sie zur Architektur gehört: wer die Positionssumme aus dem Aufriss
    # zieht, muss zeigen, dass der Aufriss die Konten der Position vollständig
    # und ohne Doppelung trägt. Sonst zeigt der Lead eine runde Zahl, der ein
    # Konto fehlt, und die Bilanz geht trotzdem auf.
    if aufrissbefunde:
        kopf = (["Position", "Aufriss", "Zeile im Lead NA",
                 "Konten ohne Aufrisszeile"] if deutsch else
                ["Line item", "Schedule", "Row in Lead NA",
                 "Accounts with no schedule row"])
        for p in perioden:
            kopf += ([f"{p} lt. Aufriss", f"{p} lt. Kontozeilen",
                      f"{p} Differenz"] if deutsch else
                     [f"{p} per schedule", f"{p} per account rows",
                      f"{p} difference"])
        inhalt = []
        for b in aufrissbefunde:
            z = [b.position, b.blatt, b.lead_zeile,
                 ", ".join(b.konten_ohne_zeile)]
            for p in perioden:
                z += [b.aufriss.get(p, 0.0), b.konten.get(p, 0.0),
                      b.differenz(p)]
            inhalt.append(z)
        zellen += _arbeitsblatt(
            wb, "Aufriss-Kontrolle" if deutsch else "Schedule control",
            kopf, inhalt)

    if nullzeilen:
        zellen += _arbeitsblatt(
            wb, "Nullkonten" if deutsch else "Nil accounts",
            (["Konto", "Bezeichnung", "Klasse", "Position"] if deutsch else
             ["Account", "Description", "Class", "Line item"]),
            [[z.konto, z.bezeichnung, z.klasse, z.na_zeile]
             for z in nullzeilen])

    wb.save(ziel)
    return Befuellergebnis(ziel, zeilen, zuordnung, slots_na + slots_pl, rf,
                           zellen, ach, nullzeilen, aufrissbefunde)
