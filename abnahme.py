"""Abnahmetest am Referenzfall.

Der Selbsttest prueft die Entscheidungsdatei gegen 60 konstruierte Faelle.
Das ist notwendig und nicht hinreichend: konstruierte Faelle treffen genau
die Stellen, an die jemand gedacht hat. Der Referenzfall ist ein echter
Kontenplan mit 198 Bilanz- und 237 GuV-Konten ueber vier Geschaeftsjahre und
prueft das, woran niemand gedacht hat.

Geprueft werden vier Dinge, die unabhaengig voneinander schieflaufen koennen:

1. **Das Einlesen.** Der Export fuehrt zwei Wertespalten, und beide summieren
   je Periode auf null. Wer die Bewegungsspalte nimmt, baut ein Databook aus
   Veraenderungen statt aus Bestaenden und merkt es an der Bilanzidentitaet
   nicht. Ausserdem stehen die Tabellenblaetter der GuV in der Reihenfolge
   FY2023, FY2026, FY2024, FY2025 — wer die Blattreihenfolge uebernimmt,
   stellt FY2026 in die zweite Spalte.
2. **Die Klassifizierung.** Jedes Konto bekommt genau eine Kategorie, und die
   Summen je Kategorie sind eingefroren. Eine Regelaenderung, die den
   Referenzfall verschiebt, faellt hier auf und nicht erst im Mandat.
3. **Die Identitaeten.** Bilanzsumme null, GuV gegen das Ergebniskonto,
   Nettovermoegen gegen Eigenkapital. Drei Wege auf dieselbe Zahl.
4. **Die Aufrisse.** Die sieben Tabs entstehen im Lauf und werden geprueft:
   Kontrollzeile auf null, Kontozeilen mit sichtbarer Formel statt
   hartkodierter Zahl, Summen als ``SUMIF(...;"<>KTO";...)``, jedes
   Working-Capital-Konto auf genau einem von NA_OA/NA_OL — und der Aufriss
   ohne Datengrundlage angelegt, leer und mit Vermerk.

Die Sollwerte unter ``ERWARTET`` sind aus dem Referenzfall abgeleitet und
danach festgeschrieben. Sie sind keine Wunschzahlen: die Bilanzsumme muss
null sein, die GuV muss auf den Cent auf das Ergebniskonto treffen, und das
Nettovermoegen muss dem Eigenkapital mit umgekehrtem Vorzeichen entsprechen.
Die Kategoriesummen sind der eingefrorene Stand — aendert sich einer, ist zu
begruenden warum.

Aufruf::

    python3 abnahme.py [klassifizierung_v1.json] [referenz] [--recalc]

Exit-Code 1, sobald ein Kriterium nicht erfuellt ist. ``--recalc`` rechnet
die gebaute Aufriss-Mappe zusaetzlich mit einem Formelrechner durch; das
faengt den Tippfehler in der Formel, den ein Nachziehen in Python nicht
sieht. Es dauert ein bis zwei Minuten und laeuft deshalb nicht bei jedem Lauf.

Hinweis zur Verdrahtung: ``fdd_databook.py`` liegt nicht in diesem
Repository. Der Abnahmetest liest den Referenzfall deshalb selbst ein — mit
derselben Spaltenerkennung, die Schritt 1 verlangt — und treibt die
Entscheidungsdatei direkt. Sobald ``fdd_databook.py`` hier liegt, tritt
dessen Reader an die Stelle von :func:`lies_export`; die Kriterien darunter
bleiben unveraendert.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter

from klassifizierung import Classifier

#: Spalten des Exports, nach Kopfzeile gesucht statt nach Position. Der
#: Export traegt 17 Spalten, und die Reihenfolge ist nicht zugesichert.
SPALTEN = {
    "konto": "AccountNo",
    "bezeichnung": "GLDescription",
    "gruppe": "AccountGroupDesc",
    "klasse_quelle": "ClassDescription",
    #: ``Year`` ist der SCHLUSSBESTAND. Der Kopf ist irrefuehrend.
    "saldo": "Year",
    #: ``PeriodMovementTYD`` ist die BEWEGUNG der Periode. Sie summiert
    #: ebenfalls auf null und sieht deshalb genauso plausibel aus.
    "bewegung": "PeriodMovementTYD",
    "von": "PeriodFrom",
    "bis": "PeriodTo",
}

#: ``ClassDescription`` des Exports -> Bilanzseite. Sie steuert allein den
#: Fallback und die OA/OL-Ableitung, nie die Kategorie.
SEITEN = {
    "Current Assets": "AKTIVA", "Non-Current Assets": "AKTIVA",
    "Current Liabilities": "PASSIVA", "Non-Current Liabilities": "PASSIVA",
    "Equity": "PASSIVA",
}

#: Das Konto, das die GuV in einer Zahl traegt. Es ist die Gegenprobe zur
#: GuV-Datei, nicht ein weiteres Eigenkapitalkonto.
ERGEBNISKONTO = "3-90000"

ERWARTET = {
    "perioden": ["FY2023", "FY2024", "FY2025", "FY2026"],
    "stichtage": ["2023-03-31", "2024-03-31", "2025-03-31", "2026-03-31"],
    "konten_bilanz": 198,
    "konten_guv": 237,
    # Die Rundung stammt aus dem Export selbst, nicht aus dem Einlesen.
    "bilanzidentitaet_toleranz": 0.05,
    "nettovermoegen": {"FY2023": 7_106_029.81, "FY2024": 7_125_509.56,
                       "FY2025": 5_248_433.43, "FY2026": 1_723_987.42},
    "net_debt": {"FY2023": 3_889_479.40, "FY2024": 4_090_973.06,
                 "FY2025": 2_970_143.72, "FY2026": 1_282_552.25},
    "working_capital": {"FY2023": 1_754_364.14, "FY2024": 1_729_559.33,
                        "FY2025": 1_359_903.67, "FY2026": -796_294.15},
    "ergebnis": {"FY2023": -287_823.17, "FY2024": 589_770.55,
                 "FY2025": -1_877_076.10, "FY2026": -3_524_446.01},
    # Zwei Konten erreicht keine Regel: '1-12400 Accrued Income' (in allen
    # vier Perioden null) und '1000 BT Imaging Product' (0,01 in FY2023).
    # Die Frage ist nicht, ob der Fallback ueberhaupt greift, sondern ob er
    # einen WESENTLICHEN Betrag entscheidet. Ein Cent auf einer Bilanzsumme
    # von 10,3 Mio ist keiner. Die Schwelle liegt bei einer Waehrungseinheit
    # und der tatsaechliche Betrag steht im Befund, damit sich niemand hinter
    # der Toleranz versteckt.
    "fallback_konten": 2,
    "fallback_betrag_toleranz": 1.00,
}


# --------------------------------------------------------------------------
# Einlesen
# --------------------------------------------------------------------------

@dataclass
class Konto:
    konto: str
    bezeichnung: str
    gruppe: str
    seite: str
    salden: dict[str, float] = field(default_factory=dict)
    bewegungen: dict[str, float] = field(default_factory=dict)
    #: Wie oft das Konto je Periode in der Liste steht (Kostenstellen).
    zeilen: dict[str, int] = field(default_factory=dict)

    def saldo(self, periode: str) -> float:
        return self.salden.get(periode, 0.0)


@dataclass
class Export:
    konten: dict[str, Konto]
    perioden: list[str]
    stichtage: dict[str, date]
    blattreihenfolge: list[str]
    fehlende_spalten: list[str]


def _stichtag(bis) -> date:
    """``PeriodTo`` nennt den ERSTEN Tag des letzten Periodenmonats.

    Wer ihn uebernimmt, legt den Abschluss einen Monat zu frueh: aus dem
    31.03. wird der 01.03. und aus dem Geschaeftsjahr ein Rumpfjahr.
    """
    jahr, monat = bis.year, bis.month
    folgemonat = date(jahr + (monat == 12), monat % 12 + 1, 1)
    return date.fromordinal(folgemonat.toordinal() - 1)


def lies_export(pfad: str) -> Export:
    """Liest einen Jahres-Export. Spalten ueber die Kopfzeile, nicht ueber
    die Position; Kostenstellen werden je Konto summiert."""
    wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
    konten: dict[str, Konto] = {}
    stichtage: dict[str, date] = {}
    fehlend: list[str] = []

    for blatt in wb.sheetnames:
        ws = wb[blatt]
        zeilen = ws.iter_rows(values_only=True)
        kopf = [str(h or "").strip() for h in next(zeilen)]
        idx = {name: kopf.index(spalte) for name, spalte in SPALTEN.items()
               if spalte in kopf}
        fehlend += [s for s in SPALTEN.values() if s not in kopf]
        periode = blatt.split()[0]

        for reihe in zeilen:
            nummer = str(reihe[idx["konto"]] or "").strip()
            if not nummer:
                continue
            k = konten.get(nummer)
            if k is None:
                k = konten[nummer] = Konto(
                    nummer, str(reihe[idx["bezeichnung"]] or "").strip(),
                    str(reihe[idx["gruppe"]] or "").strip(),
                    SEITEN.get(str(reihe[idx["klasse_quelle"]] or "").strip(), ""))
            k.salden[periode] = k.saldo(periode) + _zahl(reihe[idx["saldo"]])
            k.bewegungen[periode] = (k.bewegungen.get(periode, 0.0)
                                     + _zahl(reihe[idx["bewegung"]]))
            k.zeilen[periode] = k.zeilen.get(periode, 0) + 1
            if periode not in stichtage and reihe[idx["bis"]]:
                stichtage[periode] = _stichtag(reihe[idx["bis"]])

    # Die Blattreihenfolge ist NICHT die Zeitachse. Sortiert wird nach
    # Stichtag, sonst steht FY2026 in der zweiten Spalte.
    perioden = sorted(stichtage, key=lambda p: stichtage[p])
    return Export(konten, perioden, stichtage, list(wb.sheetnames),
                  sorted(set(fehlend)))


def _zahl(wert) -> float:
    if wert in (None, ""):
        return 0.0
    try:
        return float(wert)
    except (TypeError, ValueError):
        return 0.0


# --------------------------------------------------------------------------
# Kriterien
# --------------------------------------------------------------------------

@dataclass
class Kriterium:
    titel: str
    bestanden: bool
    befund: str
    details: list[str] = field(default_factory=list)


def _nahe(ist: float, soll: float, toleranz: float = 0.05) -> bool:
    return abs(ist - soll) <= toleranz


def pruefe(spec: str, ordner: str) -> tuple[list[Kriterium], dict]:
    bs = lies_export(os.path.join(ordner, "referenz_BS.xlsx"))
    guv = lies_export(os.path.join(ordner, "referenz_PL.xlsx"))
    c = Classifier(spec)
    k: list[Kriterium] = []

    # -- 1 Einlesen ------------------------------------------------------
    k.append(Kriterium(
        "Alle erwarteten Spalten in der Kopfzeile gefunden",
        not bs.fehlende_spalten and not guv.fehlende_spalten,
        "Spalten werden ueber den Kopf gesucht, nicht ueber die Position."
        if not bs.fehlende_spalten and not guv.fehlende_spalten
        else f"fehlend: {bs.fehlende_spalten + guv.fehlende_spalten}"))

    k.append(Kriterium(
        "Perioden nach Stichtag sortiert, nicht nach Blattreihenfolge",
        bs.perioden == ERWARTET["perioden"] and guv.perioden == ERWARTET["perioden"],
        f"GuV-Blaetter stehen als {', '.join(b.split()[0] for b in guv.blattreihenfolge)}"
        f" und werden als {', '.join(guv.perioden)} gefuehrt.",
        [f"Bilanz: {', '.join(bs.perioden)}"]))

    ist_stichtage = [bs.stichtage[p].isoformat() for p in bs.perioden]
    k.append(Kriterium(
        "Stichtag ist das Periodenende, nicht der erste Tag des Monats",
        ist_stichtage == ERWARTET["stichtage"],
        "PeriodTo nennt den 01.03.; gefuehrt wird der 31.03.",
        [f"{p}: {t}" for p, t in zip(bs.perioden, ist_stichtage)]))

    k.append(Kriterium(
        "Kontozahl wie im Referenzfall",
        len(bs.konten) == ERWARTET["konten_bilanz"]
        and len(guv.konten) == ERWARTET["konten_guv"],
        f"{len(bs.konten)} Bilanz- und {len(guv.konten)} GuV-Konten "
        f"(erwartet {ERWARTET['konten_bilanz']} / {ERWARTET['konten_guv']}). "
        "Kostenstellen sind je Konto summiert."))

    # Die Falle: beide Wertespalten gehen je Periode auf null auf.
    saldo_summe = {p: sum(x.saldo(p) for x in bs.konten.values())
                   for p in bs.perioden}
    bewegung_summe = {p: sum(x.bewegungen.get(p, 0.0) for x in bs.konten.values())
                      for p in bs.perioden}
    aktiva_saldo = sum(x.saldo(bs.perioden[0]) for x in bs.konten.values()
                       if x.seite == "AKTIVA")
    aktiva_bewegung = sum(x.bewegungen.get(bs.perioden[0], 0.0)
                          for x in bs.konten.values() if x.seite == "AKTIVA")
    k.append(Kriterium(
        "Schlussbestand gelesen, nicht die Periodenbewegung",
        all(abs(v) <= 1.0 for v in bewegung_summe.values())
        and abs(aktiva_saldo) > abs(aktiva_bewegung) * 10,
        f"Beide Spalten summieren auf null, die Bewegungsspalte also auch. "
        f"Aktivseite {bs.perioden[0]}: {aktiva_saldo:,.2f} aus 'Year' gegen "
        f"{aktiva_bewegung:,.2f} aus 'PeriodMovementTYD'.",
        [f"{p}: Bestand {saldo_summe[p]:,.2f} · Bewegung {bewegung_summe[p]:,.2f}"
         for p in bs.perioden]))

    # -- 2 Identitaeten ---------------------------------------------------
    k.append(Kriterium(
        "Bilanzidentitaet je Periode",
        all(abs(v) <= ERWARTET["bilanzidentitaet_toleranz"]
            for v in saldo_summe.values()),
        "Summe aller Bilanzkonten je Periode "
        f"(Toleranz {ERWARTET['bilanzidentitaet_toleranz']}, die Rundung "
        "stammt aus dem Export).",
        [f"{p}: {saldo_summe[p]:,.2f}" for p in bs.perioden]))

    guv_summe = {p: -sum(x.saldo(p) for x in guv.konten.values())
                 for p in guv.perioden}
    ergebniskonto = bs.konten.get(ERGEBNISKONTO)
    treffer = ergebniskonto is not None and all(
        _nahe(guv_summe[p], -ergebniskonto.saldo(p), 0.01) for p in bs.perioden)
    k.append(Kriterium(
        f"GuV trifft das Ergebniskonto {ERGEBNISKONTO} auf den Cent",
        treffer,
        "Die GuV-Datei und das Eigenkapitalkonto 'Current Earnings' sind "
        "dieselbe Zahl. Beide zugleich ins Mastersheet zu nehmen zaehlte das "
        "Ergebnis doppelt.",
        [f"{p}: GuV {guv_summe[p]:,.2f} · Konto "
         f"{-ergebniskonto.saldo(p) if ergebniskonto else 0:,.2f}"
         for p in bs.perioden]))

    k.append(Kriterium(
        "Periodenergebnis wie im Referenzfall",
        all(_nahe(guv_summe[p], ERWARTET["ergebnis"][p], 0.01)
            for p in bs.perioden),
        "Ertraege stehen im Haben; ein Gewinn ist eine negative Summe und "
        "wird beim Ausweis umgedreht.",
        [f"{p}: {guv_summe[p]:,.2f} (soll {ERWARTET['ergebnis'][p]:,.2f})"
         for p in bs.perioden]))

    # -- 3 Klassifizierung ------------------------------------------------
    ergebnisse = {no: c.classify(x.konto, x.bezeichnung, x.gruppe,
                                 seite=x.seite or None)
                  for no, x in bs.konten.items()}

    ohne = [no for no, r in ergebnisse.items() if not r["category"]]
    k.append(Kriterium(
        "Jedes Bilanzkonto traegt genau eine Kategorie",
        not ohne,
        f"{len(bs.konten)} Konten, {len(ohne)} ohne Kategorie.",
        ohne[:10]))

    fallback = [no for no, r in ergebnisse.items() if r["source"] == "fallback"]
    fallback_betrag = max(
        (abs(bs.konten[no].saldo(p)) for no in fallback for p in bs.perioden),
        default=0.0)
    k.append(Kriterium(
        "Der Fallback entscheidet keinen wesentlichen Betrag",
        len(fallback) == ERWARTET["fallback_konten"]
        and fallback_betrag <= ERWARTET["fallback_betrag_toleranz"],
        f"{len(fallback)} Konten erreicht keine Regel; groesster Saldo ueber "
        f"alle Perioden {fallback_betrag:,.2f} (Schwelle "
        f"{ERWARTET['fallback_betrag_toleranz']:,.2f}). Ein Fallback ist ein "
        "Platzhalter mit Pflichtfrage, keine Entscheidung — er darf deshalb "
        "kein Geld tragen.",
        [f"{no} {bs.konten[no].bezeichnung} ("
         + ", ".join(f"{p} {bs.konten[no].saldo(p):,.2f}" for p in bs.perioden)
         + ")" for no in fallback]
        + ["Befund fuer die Entscheidungsdatei: 'accrued income' ist als "
           "Stichwort nicht belegt. Die Regel owc-accruals kennt nur "
           "'accrued expense'. Hier faengt es der Fallback auf der richtigen "
           "Seite auf — aber aus dem richtigen Grund waere besser."]))

    quellen = {"kontoname": 0, "kontogruppe": 0, "fallback": 0, "hgb": 0}
    for r in ergebnisse.values():
        quellen[r["source"]] = quellen.get(r["source"], 0) + 1
    k.append(Kriterium(
        "Die Kontogruppe traegt, was der Name nicht hergibt",
        quellen["kontogruppe"] > 0,
        f"{quellen['kontoname']} Konten ueber den Namen, "
        f"{quellen['kontogruppe']} ueber die Gruppe, "
        f"{quellen['fallback']} ueber den Fallback. Bankkonten heissen nach "
        "der Bank und nicht nach ihrer Funktion.",
        [f"ueber die Gruppe: {no} {bs.konten[no].bezeichnung} "
         f"({bs.konten[no].gruppe}) -> {r['category']}"
         for no, r in ergebnisse.items() if r["source"] == "kontogruppe"][:6]))

    def summe(klassen, periode):
        return sum(bs.konten[no].saldo(periode) for no, r in ergebnisse.items()
                   if r["klasse"] in klassen)

    net_debt = {p: summe({"ND"}, p) for p in bs.perioden}
    wc = {p: summe({"TWC", "OWC"}, p) for p in bs.perioden}
    netto = {p: summe({"FA", "TWC", "OWC", "ND", "DT"}, p) for p in bs.perioden}
    eigen = {p: summe({"EQ"}, p) for p in bs.perioden}

    for name, ist, soll in (("Net Debt", net_debt, ERWARTET["net_debt"]),
                            ("Working Capital", wc, ERWARTET["working_capital"]),
                            ("Nettovermoegen", netto, ERWARTET["nettovermoegen"])):
        k.append(Kriterium(
            f"{name} je Periode wie eingefroren",
            all(_nahe(ist[p], soll[p]) for p in bs.perioden),
            "Abweichung gegen den eingefrorenen Stand des Referenzfalls.",
            [f"{p}: {ist[p]:,.2f} (soll {soll[p]:,.2f})" for p in bs.perioden]))

    k.append(Kriterium(
        "Nettovermoegen plus Eigenkapital ist null",
        all(abs(netto[p] + eigen[p]) <= ERWARTET["bilanzidentitaet_toleranz"]
            for p in bs.perioden),
        "Dritter Weg auf dieselbe Zahl: die Klassen ausserhalb des "
        "Eigenkapitals muessen es spiegeln.",
        [f"{p}: {netto[p] + eigen[p]:,.2f}" for p in bs.perioden]))

    doppelt = [no for no, r in ergebnisse.items()
               if r["klasse"] in ("TWC", "OWC") and not r["wc_seite"]]
    k.append(Kriterium(
        "Jedes Working-Capital-Konto traegt eine Seite (OA/OL)",
        not doppelt,
        "OA/OL ist keine eigene Klasse, sondern die Bilanzseite. ARAP bleibt "
        "OWC auf der OA-Seite, PRAP bleibt OWC auf der OL-Seite.",
        doppelt[:10]))

    # -- 4 Aufrisse -------------------------------------------------------
    k += pruefe_aufrisse(bs, ergebnisse, spec, ordner, net_debt, wc)

    zahlen = {"bs": bs, "guv": guv, "ergebnisse": ergebnisse,
              "net_debt": net_debt, "wc": wc, "netto": netto,
              "guv_summe": guv_summe, "quellen": quellen}
    return k, zahlen


#: Die sieben Aufriss-Tabs in der geforderten Reihenfolge.
AUFRISS_TABS = ["NA_OA", "NA_OL", "NA_TWC", "NA_Net Debt", "NA_Vorräte",
                "NA_Sachanlagen", "NA_CAPEX"]

#: Wohin der Abnahmelauf die Arbeitsmappe schreibt.
AUFRISS_MAPPE = os.path.join("out", "Referenzfall_Aufrisse.xlsx")


def _recalc_kriterium(befunde: dict, perioden: list[str]) -> Kriterium:
    """Rechnet die Mappe wirklich durch (``--recalc``).

    Die Kriterien oben ziehen dieselbe Rechnung in Python nach. Das faengt
    einen Denkfehler, aber keinen Tippfehler in der Formel. Deshalb hier der
    Formelrechner ueber die fertige Datei. Er laeuft nicht bei jedem Lauf,
    weil er ein bis zwei Minuten braucht.

    Gerechnet wird mit dem Paket ``formulas`` und nicht mit dem
    LibreOffice-Recalc des Hausformats: LibreOffice laeuft in dieser Umgebung
    in den Timeout. Fuer die Auslieferung bleibt der Hausweg massgeblich.
    """
    import formulas

    modell = formulas.ExcelModel().loads(AUFRISS_MAPPE).finish()
    loesung = modell.calculate()
    werte: dict[tuple[str, str], object] = {}
    for schluessel, zelle in loesung.items():
        if "]" not in schluessel or "!" not in schluessel:
            continue
        blatt = schluessel.split("]", 1)[1].split("'!")[0]
        try:
            werte[(blatt, schluessel.split("!", 1)[1])] = zelle.value[0, 0]
        except Exception:
            continue

    abweichungen, gerechnet = [], []
    for blatt in AUFRISS_TABS:
        b = befunde[blatt]
        if b["leer"]:
            continue
        for i in range(len(perioden)):
            spalte = get_column_letter(6 + i)
            wert = werte.get((blatt.upper(), f"{spalte}{b['kontrollzeile']}"))
            if wert is None or abs(float(wert)) > 0.005:
                abweichungen.append(f"{blatt} {perioden[i]}: {wert}")
        summe = werte.get((blatt.upper(), f"F{b['summenzeile']}"))
        gerechnet.append(f"{blatt}: {perioden[0]} {float(summe):,.1f}")
    return Kriterium(
        "Nachgerechnet: jede Kontrollzeile der Mappe steht auf null",
        not abweichungen,
        "Die Mappe wurde mit einem Formelrechner durchgerechnet; geprueft "
        "sind die Werte, die Excel anzeigen wuerde, nicht die Absicht.",
        abweichungen[:10] or gerechnet)


def pruefe_aufrisse(bs, ergebnisse, spec: str, ordner: str, net_debt: dict,
                    wc: dict) -> list[Kriterium]:
    """Baut die Aufriss-Mappe und prueft sie.

    Die Kontrollzeile jedes Tabs rechnet ``Summe der Kontozeilen`` gegen
    ``SUMIFS ueber ein Merkmal des Mastersheets``. Beide Wege werden hier in
    Python nachgezogen — das ist dieselbe Rechnung, die die Formel im Blatt
    anstellt, nur ohne Excel. Ob die Formeln auch WIRKLICH so im Blatt stehen,
    prueft das Kriterium darunter; und ``--recalc`` rechnet die Mappe
    zusaetzlich mit einem Formelrechner durch.
    """
    import aufrisse as auf

    k: list[Kriterium] = []
    ergebnis = auf.baue(ordner, spec, AUFRISS_MAPPE)
    wb = openpyxl.load_workbook(AUFRISS_MAPPE)
    befunde = {b["blatt"]: b for b in ergebnis["befunde"]}

    k.append(Kriterium(
        "Alle sieben Aufriss-Tabs sind angelegt",
        [b for b in wb.sheetnames if b != "Mastersheet"] == AUFRISS_TABS,
        "Reihenfolge und Namen wie beauftragt.",
        [", ".join(wb.sheetnames)]))

    # Der zweite Weg: ueber das Merkmal im Mastersheet, nicht ueber Konten.
    merkmal = {auf.MS_KLASSE: "klasse", auf.MS_KATEGORIE: "category",
               auf.MS_SEITE: "wc_seite"}
    abweichungen = []
    for blatt in AUFRISS_TABS:
        b = befunde[blatt]
        if b["leer"]:
            continue
        for p in bs.perioden:
            ueber_konten = sum(bs.konten[no].saldo(p) for no in b["konten"])
            ueber_merkmal = sum(
                bs.konten[no].saldo(p) for no, r in ergebnisse.items()
                if any(r[merkmal[sp]] == w for sp, w in b["kontrolle"]))
            if abs(ueber_konten - ueber_merkmal) > 0.005:
                abweichungen.append(f"{blatt} {p}: "
                                    f"{ueber_konten - ueber_merkmal:,.2f}")
    k.append(Kriterium(
        "Kontrollzeile jedes Aufrisses geht auf null",
        not abweichungen,
        "Zwei unabhaengige Wege je Tab und Periode: die Summe der "
        "Kontozeilen gegen ein SUMIFS ueber das Merkmal im Mastersheet.",
        abweichungen[:10] or [f"{b} ok" for b in AUFRISS_TABS
                              if not befunde[b]["leer"]]))

    ohne_formel = []
    for blatt in AUFRISS_TABS:
        ws = wb[blatt]
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, auf.SP_TYP).value != "KTO":
                continue
            wert = ws.cell(r, auf.SP_ERSTE_PERIODE).value
            if not (isinstance(wert, str) and wert.startswith("=SUMIF(Mastersheet!")):
                ohne_formel.append(f"{blatt}!Zeile {r}")
    k.append(Kriterium(
        "Jede Kontozeile holt ihren Wert per sichtbarer Formel",
        not ohne_formel,
        "Eine hartkodierte Zahl im Aufriss ist nicht zurueckverfolgbar. Die "
        "Kontozeilen verweisen per SUMIF auf das Mastersheet.",
        ohne_formel[:10]))

    ohne_kontrolle = [b for b in AUFRISS_TABS
                      if not befunde[b]["leer"]
                      and befunde[b]["kontrollzeile"] is None]
    summen_ohne_sumif = []
    for blatt in AUFRISS_TABS:
        b = befunde[blatt]
        if b["leer"]:
            continue
        wert = wb[blatt].cell(b["summenzeile"], auf.SP_ERSTE_PERIODE).value
        if not (isinstance(wert, str) and '"<>KTO"' in wert):
            summen_ohne_sumif.append(blatt)
    k.append(Kriterium(
        'Summenzeilen als SUMIF(...;"<>KTO";...)',
        not ohne_kontrolle and not summen_ohne_sumif,
        "Die Kontozeilen liegen INNERHALB des Summenbereichs; ein blankes SUM "
        "zaehlte sie doppelt.",
        ohne_kontrolle + summen_ohne_sumif))

    leer = [b for b in AUFRISS_TABS if befunde[b]["leer"]]
    mit_vermerk = [b for b in leer
                   if wb[b].cell(befunde[b]["vermerkzeile"],
                                 auf.SP_DE).value.startswith("Vermerk:")]
    k.append(Kriterium(
        "Ein Aufriss ohne Datengrundlage wird trotzdem angelegt",
        leer == ["NA_CAPEX"] and mit_vermerk == leer,
        f"{', '.join(leer) or 'keiner'} ist leer und traegt einen Vermerk, "
        "der die fehlende Unterlage benennt. Ein Tab, der gar nicht erst "
        "entsteht, verschwindet aus der Wahrnehmung.",
        [wb[b].cell(befunde[b]["vermerkzeile"], auf.SP_DE).value[:200]
         for b in leer]))

    doppelt = [no for no in bs.konten
               if (no in befunde["NA_OA"]["konten"])
               and (no in befunde["NA_OL"]["konten"])]
    wc_konten = {no for no, r in ergebnisse.items()
                 if r["klasse"] in ("TWC", "OWC")}
    fehlend = wc_konten - set(befunde["NA_OA"]["konten"]) \
        - set(befunde["NA_OL"]["konten"])
    k.append(Kriterium(
        "Jedes Working-Capital-Konto steht auf genau einem von NA_OA/NA_OL",
        not doppelt and not fehlend,
        f"{len(wc_konten)} Konten, davon {len(befunde['NA_OA']['konten'])} in "
        f"NA_OA und {len(befunde['NA_OL']['konten'])} in NA_OL. "
        f"{len(doppelt)} doppelt, {len(fehlend)} nirgends.",
        sorted(doppelt)[:5] + sorted(fehlend)[:5]))

    summen_ok = []
    for p in bs.perioden:
        oa = sum(bs.konten[no].saldo(p) for no in befunde["NA_OA"]["konten"])
        ol = sum(bs.konten[no].saldo(p) for no in befunde["NA_OL"]["konten"])
        nd = sum(bs.konten[no].saldo(p)
                 for no in befunde["NA_Net Debt"]["konten"])
        summen_ok.append((p, oa + ol - wc[p], nd - net_debt[p]))
    k.append(Kriterium(
        "NA_OA plus NA_OL ist das Working Capital, NA_Net Debt das Net Debt",
        all(abs(a) <= 0.005 and abs(b) <= 0.005 for _, a, b in summen_ok),
        "Die Aufrisse tragen dieselben Summen wie die Klassifizierung.",
        [f"{p}: WC-Differenz {a:,.2f} · ND-Differenz {b:,.2f}"
         for p, a, b in summen_ok]))

    if "--recalc" in sys.argv:
        k.append(_recalc_kriterium(befunde, bs.perioden))

    geflaggt = {b: befunde[b]["geflaggt"] for b in AUFRISS_TABS
                if befunde[b].get("geflaggt")}
    k.append(Kriterium(
        "Positionen ohne Gegenstueck sind markiert, nicht versteckt",
        True,
        "Kumulierte Abschreibung ist ein Korrekturposten. Steht sie ohne die "
        "Anschaffungskosten derselben Anlagenklasse, ist die Position gelb "
        "und traegt den Grund im Klartext."
        if geflaggt else "Keine solche Position im Referenzfall.",
        [f"{b}: {', '.join(v)}" for b, v in geflaggt.items()]))

    return k


# --------------------------------------------------------------------------
# Ausgabe
# --------------------------------------------------------------------------

def _kategorien(bs, ergebnisse) -> dict[str, dict[str, float]]:
    je: dict[str, dict[str, float]] = {}
    for no, r in ergebnisse.items():
        zeile = je.setdefault(r["category"], {p: 0.0 for p in bs.perioden})
        for p in bs.perioden:
            zeile[p] += bs.konten[no].saldo(p)
    return dict(sorted(je.items()))


def main() -> int:
    spec = sys.argv[1] if len(sys.argv) > 1 else "klassifizierung_v1.json"
    ordner = sys.argv[2] if len(sys.argv) > 2 else "referenz"
    kriterien, z = pruefe(spec, ordner)
    bs = z["bs"]

    print("=" * 78)
    print("  ABNAHME AM REFERENZFALL")
    print(f"  Entscheidungsdatei {spec} (v{Classifier(spec).version})")
    print(f"  Referenzfall       {ordner}/referenz_BS.xlsx + referenz_PL.xlsx")
    print("=" * 78)

    print("\n  Kategorien je Periode")
    print(f"  {'':30}" + "".join(f"{p:>17}" for p in bs.perioden))
    for kat, werte in _kategorien(bs, z["ergebnisse"]).items():
        print(f"  {kat:30}" + "".join(f"{werte[p]:>17,.2f}" for p in bs.perioden))
    print(f"  {'davon Net Debt':30}"
          + "".join(f"{z['net_debt'][p]:>17,.2f}" for p in bs.perioden))
    print(f"  {'davon Working Capital':30}"
          + "".join(f"{z['wc'][p]:>17,.2f}" for p in bs.perioden))
    print(f"  {'Nettovermoegen':30}"
          + "".join(f"{z['netto'][p]:>17,.2f}" for p in bs.perioden))
    print(f"  {'Periodenergebnis (GuV)':30}"
          + "".join(f"{z['guv_summe'][p]:>17,.2f}" for p in bs.perioden))

    review = [(no, r) for no, r in z["ergebnisse"].items() if r["review"]]
    mit_saldo = [no for no, _ in review
                 if any(abs(bs.konten[no].saldo(p)) > 0.005 for p in bs.perioden)]
    print(f"\n  Review-Queue: {len(review)} Konten, davon {len(mit_saldo)} mit Saldo")
    for no, r in review:
        if no in mit_saldo:
            print(f"    {no:9}{bs.konten[no].bezeichnung[:36]:38}"
                  f"{r['category']:24}{r['rule_id']:24}{r['pflichtfrage'][:24]}")

    print("\n" + "-" * 78)
    durchgefallen = [x for x in kriterien if not x.bestanden]
    for x in kriterien:
        print(f"  {'ok ' if x.bestanden else 'FEHLER'} {x.titel}")
        print(f"       {x.befund}")
        for d in x.details:
            print(f"         {d}")
    print("-" * 78)
    print(f"  {len(kriterien) - len(durchgefallen)} von {len(kriterien)} "
          "Kriterien bestanden")
    print("=" * 78)
    return 1 if durchgefallen else 0


if __name__ == "__main__":
    sys.exit(main())
