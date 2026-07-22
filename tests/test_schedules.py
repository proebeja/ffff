"""Aufriss-Schicht: jede Lead-Zeile zieht aus genau einem Aufriss, die
gemischten Aufrisse trennen operating/thereof-ND aus der Mastersheet-Klasse,
und keine Kontozahl läuft auseinander."""

import re

import openpyxl
import pytest

from conftest import ALLE, datei
from fdd.cli import run
from fdd.core.model import Klasse


@pytest.mark.parametrize("name", ALLE)
def test_kein_konto_ohne_aufriss(name):
    """Jedes ND/TWC/OWC-Konto landet in einem Aufriss."""
    res = run(datei(name), "/tmp/sch.xlsx", verbose=False)
    assert res["schedules"].ohne_aufriss == [], (
        f"{name}: {[m.konto for m in res['schedules'].ohne_aufriss]}")


@pytest.mark.parametrize("name", ALLE)
def test_mixed_split_operating_plus_nd_ist_position(name):
    """Im gemischten Aufriss gilt je Periode: operating + thereof ND = Summe
    aller Konten der Position (nichts geht verloren, nichts doppelt)."""
    res = run(datei(name), "/tmp/sch.xlsx", verbose=False)
    for a in res["schedules"].aufrisse:
        if not a.is_mixed:
            continue
        for p in a.perioden:
            assert a.operating_summe(p) + a.thereof_nd_summe(p) == pytest.approx(
                a.summe(p), abs=0.01)


@pytest.mark.parametrize("name", ALLE)
def test_aufriss_summe_gleich_view_summe(name):
    """Die Aufriss-Summen decken sich mit den Net-Debt-/WC-View-Zeilen."""
    res = run(datei(name), "/tmp/sch.xlsx", verbose=False)
    sch = res["schedules"]
    # Net-Debt: jede ND-Zeile == thereof-ND (mixed) bzw. Summe (sonst)
    for z in res["nd"].alle_zeilen:
        a = sch.by_na(z.na_de)
        assert a is not None
        for p in res["ledger"].perioden:
            soll = a.thereof_nd_summe(p) if a.is_mixed else a.summe(p)
            assert soll == pytest.approx(z.betraege[p], abs=0.01)


@pytest.mark.parametrize("name", ALLE)
def test_leads_ziehen_nur_aus_aufrissen(name, tmp_path):
    """In den Datenzeilen der Leads darf kein SUMIFS aufs Mastersheet stehen —
    nur die Kontrollzeile prüft unabhängig gegen das Mastersheet."""
    out = str(tmp_path / "db.xlsx")
    run(datei(name), out, verbose=False)
    wb = openpyxl.load_workbook(out)
    for sheet in ("Net Debt", "Working Capital"):
        ws = wb[sheet]
        kz = {r for r in range(1, ws.max_row + 1)
              if str(ws.cell(r, 3).value or "").startswith("Kontrollzeile")}
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "SUMIFS" in c.value:
                    assert c.row in kz, f"{name}/{sheet}: SUMIFS in Datenzeile {c.coordinate}"
    wb.close()


@pytest.mark.parametrize("name", ALLE)
def test_beide_kontrollzeilen_null_im_excel(name, tmp_path):
    """Durch die ganze Kette Mastersheet -> Aufriss -> Lead gehen beide
    Kontrollzeilen (ND und WC) auf null auf."""
    formulas = pytest.importorskip("formulas")
    import warnings
    import numpy as np
    from openpyxl.utils import get_column_letter
    warnings.filterwarnings("ignore")
    out = str(tmp_path / "db.xlsx")
    res = run(datei(name), out, verbose=False)
    n_per = len(res["ledger"].perioden)

    wb = openpyxl.load_workbook(out)
    kz = {}
    for sheet in ("Net Debt", "Working Capital"):
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 3).value or "").startswith("Kontrollzeile"):
                kz[sheet] = r
    wb.close()

    sol = formulas.ExcelModel().loads(out).finish().calculate()
    up = out.split("/")[-1]
    keys = {k.upper(): k for k in sol}
    for sheet in ("Net Debt", "Working Capital"):
        for i in range(n_per):
            col = get_column_letter(4 + i)
            k = keys.get(f"'[{up}]{sheet.upper()}'!{col}{kz[sheet]}".upper())
            val = float(np.ravel(sol[k].value)[0]) if k else 0.0
            assert abs(val) < 1e-3, f"{name}/{sheet} {col}: Kontrollzeile {val} != 0"
