"""Die Vorlagen-Mechanik: was beim Befüllen nicht passieren darf.

Die Prüfungen hier sind bewusst nicht auf Zahlen gerichtet, sondern auf die
drei Regeln, deren Verletzung still bleibt und erst in der fertigen Mappe
auffällt: verschobene Zeilennummern, ein Ticker mit abweichender Schreibweise
und ein blankes ``SUM`` über einen Bereich mit Kontoslots.
"""

import openpyxl
import pytest

from conftest import datei
from fdd.brehna import Quellen, run
from fdd.export import vorlage
from fdd.export.pruefung import vergleiche_format
from fdd.export.vorlage_layout import lies_layout

DATEIEN = [f"Testdaten_Brehna_JA_{j}.pdf"
           for j in ("2023", "2024", "2025", "2026_Juli")]
LEADS = (("Lead NA", ("P", "Q", "R"), 13), ("Lead PL", ("T", "U", "V"), 14))


@pytest.fixture(scope="module")
def ausgabe(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("vorlage") / "dealtool.xlsx")
    run(Quellen(jahresabschluesse=[datei(n) for n in DATEIEN]), out,
        verbose=False)
    return out


@pytest.fixture(scope="module")
def wb(ausgabe):
    return openpyxl.load_workbook(ausgabe)


@pytest.fixture(scope="module")
def vorlage_wb():
    return openpyxl.load_workbook(vorlage.VORLAGE)


# ---- Die harten Regeln ----------------------------------------------------
@pytest.mark.parametrize("blatt", ["Lead NA", "Lead PL"])
def test_keine_zeile_eingefuegt(wb, vorlage_wb, blatt):
    """2.115 Verweise aus den übrigen Tabs zeigen auf feste Zeilennummern.

    Ein Einschub verschiebt sie, und openpyxl zieht sie nicht mit. Die
    Zeilenzahl muss deshalb Zeile für Zeile die der Vorlage sein.
    """
    assert wb[blatt].max_row == vorlage_wb[blatt].max_row
    for r in range(1, vorlage_wb[blatt].max_row + 1):
        erwartet = vorlage_wb[blatt].cell(r, 3).value
        if isinstance(erwartet, str) and not erwartet.startswith(("Dummy", "=")):
            assert wb[blatt].cell(r, 3).value == erwartet, r


@pytest.mark.parametrize("blatt,ticker,letzte", LEADS)
def test_keine_blanke_summe_ueber_kontoslots(wb, blatt, ticker, letzte):
    """Die Kontoslots liegen in den Summenbereichen. Ein blankes ``SUM``
    zählt sie doppelt."""
    ws = wb[blatt]
    for r in range(1, ws.max_row + 1):
        for c in range(5, letzte + 1):
            f = ws.cell(r, c).value
            if not isinstance(f, str) or "SUM(" not in f.upper():
                continue
            pytest.fail(f"{blatt}!{ws.cell(r, c).coordinate}: {f}")


@pytest.mark.parametrize("blatt,ticker,letzte", LEADS)
def test_jeder_positionsticker_findet_das_mastersheet(wb, blatt, ticker, letzte):
    """Eine abweichende Schreibweise liefert stumm null.

    Deshalb wird jede befüllte Position gegen das Mastersheet geprüft: das
    Paar aus NA-Zeile und Klasse muss dort vorkommen. Genau dieser Fall — ein
    Ticker hieß ``verbundene`` statt ``verbundenen`` — ist im Brehna-Lauf
    schon einmal passiert.
    """
    ms = wb["Mastersheet"]
    paare = {(ms.cell(r, 4).value, ms.cell(r, 3).value)
             for r in range(2, ms.max_row + 1) if ms.cell(r, 1).value}
    layout = lies_layout(wb[blatt], ticker, 5, letzte)
    ticker_paare = {(p.ticker1, p.ticker2) for p in layout.positionen}
    for paar in {(p.ticker1, p.ticker2) for p in layout.positionen if p.aus_dummy}:
        assert paar in paare, f"{blatt}: Ticker {paar} steht in keinem Mastersheet"

    # Und die Gegenrichtung, die eigentlich gefährliche: eine Klasse im
    # Mastersheet, für die es keine Positionszeile gibt, verschwindet
    # geräuschlos aus dem Lead.
    ohne_position = {k.value for k in __import__(
        "fdd.export.vorlage_zuordnung", fromlist=["x"]).OHNE_POSITION}
    guv = blatt == "Lead PL"
    for na, klasse in paare:
        if klasse in ohne_position or (klasse == "PL") != guv:
            continue
        assert (na, klasse) in ticker_paare, \
            f"{blatt}: Mastersheet führt {na}/{klasse}, der Lead-Tab nicht"


@pytest.mark.parametrize("blatt,ticker,letzte", LEADS)
def test_kontoslots_verweisen_auf_vorhandene_konten(wb, blatt, ticker, letzte):
    """Ein Kontoslot mit unbekanntem Schlüssel zeigt still null."""
    ms = wb["Mastersheet"]
    schluessel = {ms.cell(r, 1).value for r in range(2, ms.max_row + 1)}
    layout = lies_layout(wb[blatt], ticker, 5, letzte)
    for p in layout.positionen:
        for slot in p.slots:
            wert = wb[blatt][f"{ticker[0]}{slot}"].value
            if wert is not None:
                assert wert in schluessel, f"{blatt}!{slot}: {wert}"


# ---- Darstellung ----------------------------------------------------------
@pytest.mark.parametrize("blatt", ["Lead NA", "Lead PL"])
def test_position_oben_konten_eingeklappt(wb, blatt):
    """``summaryBelow = False`` setzt das Gliederungssymbol neben die
    Positionszeile statt neben die Zeile unter der Gruppe."""
    ws = wb[blatt]
    assert ws.sheet_properties.outlinePr.summaryBelow is False


@pytest.mark.parametrize("blatt,ticker,letzte", LEADS)
def test_kontoslots_bleiben_eingeklappt_und_grau(wb, blatt, ticker, letzte):
    ws = wb[blatt]
    layout = lies_layout(ws, ticker, 5, letzte)
    for p in layout.positionen:
        for slot in p.slots:
            assert ws.row_dimensions[slot].hidden, f"{blatt}!{slot}"
            assert ws.row_dimensions[slot].outline_level == 3, f"{blatt}!{slot}"
            assert ws.cell(slot, 3).alignment.indent == 2, f"{blatt}!{slot}"


def test_sprache_blendet_den_anderen_block_aus(wb):
    """Sprachwahl durch Ein- und Ausblenden, nicht durch Übersetzen."""
    ws = wb["Lead NA"]
    assert not ws.row_dimensions[4].hidden and not ws.row_dimensions[5].hidden
    assert ws.row_dimensions[6].hidden and ws.row_dimensions[7].hidden
    assert not ws.column_dimensions["C"].hidden
    assert ws.column_dimensions["D"].hidden


def test_cockpit_traegt_die_zeitachse(wb):
    """Perioden werden in keinem Tab hartcodiert, sondern im Cockpit gesetzt."""
    c = wb["Cockpit"]
    assert c["C17"].value.strftime("%d.%m.%Y") == "31.12.2023"
    assert c["C18"].value == 3
    assert c["C22"].value.strftime("%d.%m.%Y") == "31.07.2026"
    assert wb["Lead NA"]["E4"].value == "=Cockpit!C39"


def test_format_bleibt_das_der_vorlage(ausgabe):
    """Ein Ausgabeblatt ist zellweise formatgleich mit der Vorlage.

    Weil die Ausgabe eine Kopie ist und nur Werte in vorhandene Zellen
    geschrieben werden, muss die Liste leer sein. Jeder Eintrag wäre der
    Beleg, dass doch irgendwo ein Format erzeugt wurde.
    """
    bericht = vergleiche_format(vorlage.VORLAGE, ausgabe)
    assert bericht.abweichungen == []


# ---- Zeitachse ------------------------------------------------------------
def test_bilanz_und_guv_spalte_laufen_ab_der_zwischenperiode_auseinander():
    """Bilanz und GuV haben unterschiedlich viele Spalten.

    In der Bilanz folgt auf die Jahre unmittelbar der CYT-Stichtag, in der GuV
    liegt davor noch die LTM-Spalte. Aus dieser Asymmetrie stammte der
    Off-by-one im Equity Roll Forward.
    """
    ach = vorlage.zeitachse(["FY2023", "FY2024", "FY2025", "YTD 07/2026"])
    assert ach.anzahl_historisch == 3
    for p in ("FY2023", "FY2024", "FY2025"):
        assert ach.ms_spalte(p, guv=False) == ach.ms_spalte(p, guv=True)
    assert ach.ms_spalte("YTD 07/2026", guv=False) == 10
    assert ach.ms_spalte("YTD 07/2026", guv=True) == 11


def test_roll_forward_verweist_auf_die_richtige_guv_spalte(wb):
    """Die Zwischenperiode steht in der Bilanz in Spalte I, in der GuV in J."""
    ws = wb["Lead NA"]
    zeile = next(r for r in range(1, ws.max_row + 1)
                 if ws.cell(r, 3).value == "Jahresergebnis")
    assert ws.cell(zeile, 8).value == "='Lead PL'!H210"      # FY2025
    assert ws.cell(zeile, 9).value == "='Lead PL'!J210"      # YTD 07/2026


def test_roll_forward_endet_mit_der_letzten_periode(wb):
    """Die Vorlage kettet den Anfangsbestand über alle neun Spalten durch.

    Ohne Plandaten trüge die erste leere Spalte das Nettovermögen der
    Vorperiode gegen eine leere Bilanz, und die Check-Zeile zeigte genau
    diesen Betrag an, obwohl nichts fehlt.
    """
    ws = wb["Lead NA"]
    zeile = next(r for r in range(1, ws.max_row + 1)
                 if ws.cell(r, 3).value == "Nettovermögen (Periodenbeginn)")
    # E ist die Eröffnungsspalte, F bis I sind die vier Perioden.
    assert [ws.cell(zeile, c).value for c in range(10, 14)] == [None] * 4
    assert ws.cell(zeile, 9).value == "=H231"


def test_platzhalterzeile_der_vorlage_ist_ueberschrieben(wb):
    ms = wb["Mastersheet"]
    assert ms["A2"].value and "Mastersheet wird" not in str(ms["B2"].value)


@pytest.mark.parametrize("label,tag,art", [
    ("FY2024", "31.12.2024", "jahr"),
    ("YTD 07/2026", "31.07.2026", "zwischen"),
    ("2024/12", "31.12.2024", "jahr"),
    ("31.03.2025", "31.03.2025", "zwischen"),
])
def test_periodenlabel_wird_gelesen(label, tag, art):
    t, a = vorlage.stichtag(label)
    assert t.strftime("%d.%m.%Y") == tag and a == art


def test_zu_viele_jahresperioden_werden_abgelehnt():
    """Die Vorlage wird zentral erweitert, nicht je Mandat."""
    with pytest.raises(ValueError, match="zentral erweitert"):
        vorlage.zeitachse([f"FY{j}" for j in range(2015, 2024)])
