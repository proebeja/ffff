"""End-to-end: läuft gegen alle vier Datensätze ohne Absturz; das erzeugte
Excel rechnet und die Kontrollzeile geht auf. Plus Namur-Regression gegen die
vorhandene Kategorisierung."""

import os
import warnings

import openpyxl
import pytest

from conftest import ALLE, NAMUR, datei
from fdd.cli import run
from fdd.core.model import Klasse
from fdd.readers.namur_databook import kategorie_map


# ---- Abnahmekriterium: läuft gegen alle vier ohne Absturz -----------------
@pytest.mark.parametrize("name", ALLE)
def test_pipeline_laeuft_ohne_absturz(name, tmp_path):
    out = str(tmp_path / "db.xlsx")
    res = run(datei(name), out, verbose=False)
    assert os.path.exists(out)
    assert len(res["mapped"]) > 0
    # jede Zahl bis zum Einzelkonto rückverfolgbar: jede ND-View-Zeile trägt Konten
    for z in res["nd"].alle_zeilen:
        assert z.konten, f"{name}: NA-Zeile {z.na_de} ohne Konten-Referenz"


@pytest.mark.parametrize("name", ALLE)
def test_excel_struktur_vorhanden(name, tmp_path):
    out = str(tmp_path / "db.xlsx")
    run(datei(name), out, verbose=False)
    wb = openpyxl.load_workbook(out)
    assert {"Mastersheet", "Net Debt", "Review-Queue"}.issubset(set(wb.sheetnames))
    wb.close()


# ---- Formeln rechnen; Kontrollzeile geht auf (echter Formel-Engine) -------
@pytest.mark.parametrize("name", ALLE)
def test_formeln_rechnen_und_kontrollzeile_null(name, tmp_path):
    formulas = pytest.importorskip("formulas")
    import numpy as np
    warnings.filterwarnings("ignore")
    out = str(tmp_path / "db.xlsx")
    res = run(datei(name), out, verbose=False)

    sol = formulas.ExcelModel().loads(out).finish().calculate()
    fname = os.path.basename(out).upper()

    # Kontrollzeilen-Zeile finden: Label in Spalte C
    wb = openpyxl.load_workbook(out)
    ws = wb["Net Debt"]
    kz_row = sub_row = None
    for r in range(1, ws.max_row + 1):
        c = str(ws.cell(r, 3).value or "")
        if c.startswith("Kontrollzeile"):
            kz_row = r
        if c.startswith("Netto-Finanzvermögen"):
            sub_row = r
    wb.close()
    assert kz_row and sub_row

    from openpyxl.utils import get_column_letter
    for i in range(len(res["ledger"].perioden)):
        col = get_column_letter(4 + i)
        key = f"'[{fname}]NET DEBT'!{col}{kz_row}"
        v = sol.get(key)
        val = np.ravel(v.value)[0] if v is not None else 0.0
        assert abs(float(val)) < 1e-3, f"{name} {col}: Kontrollzeile {val} != 0"


# ---- Namur-Regression: eigene Klasse vs. vorhandene Kategorisierung --------
def test_namur_regression_gegen_vorhandene_kategorisierung(tmp_path):
    res = run(datei(NAMUR), str(tmp_path / "db.xlsx"), verbose=False)
    kat = kategorie_map(datei(NAMUR))
    # Mapping der Namur-Analysetaxonomie auf unsere Klasse (nur eindeutige Fälle)
    erwartet = {
        "/fixed assets/": Klasse.FA,
        "/equity/": Klasse.EQ,
        "/net debt/liabilities to credit institutions": Klasse.ND,
        "/net debt/provisions for pensions": Klasse.ND,
        "/net debt/tax provisions": Klasse.ND,
    }
    treffer = gesamt = 0
    for m in res["mapped"]:
        k = kat.get(m.konto, "")
        for prefix, kl in erwartet.items():
            if k.startswith(prefix):
                gesamt += 1
                if m.klasse == kl:
                    treffer += 1
                break
    assert gesamt >= 15, f"zu wenige eindeutige Vergleichsfälle: {gesamt}"
    quote = treffer / gesamt
    # Determinierbare Positionen (FA/EQ/ND-Kern) sollen sehr gut treffen
    assert quote >= 0.85, f"Namur-Trefferquote nur {quote:.0%} ({treffer}/{gesamt})"
