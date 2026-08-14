"""Projekt Kitchenstories (AJNS New Media GmbH): SuSa mit drei Defekten,
Kontennachweis als Strukturquelle nur für die Bilanz, spaltenweiser Status,
zweistufige Abstimmung und Benchmark gegen eine manuelle Klassifizierung."""

import openpyxl
import pytest

from conftest import datei
from fdd.kitchenstories import Quellen, run
from fdd.readers.datev_ja_pdf import lies_datev_ja
from fdd.readers.datev_kontenplan_pdf import lies_kontenplan
from fdd.readers.susa_databook import SusaDatabookReader
from fdd.views.benchmark import uebersetze

SUSA = "Testdaten_Kitchenstories_SuSa.xlsx"
JA = "Testdaten_Kitchenstories_JA_2023.pdf"
PLAN = "Testdaten_Kitchenstories_Kontenplan_2025.pdf"
BERICHT = "Testdaten_Kitchenstories_Pruefbericht_2022.pdf"
PERIODEN = ["FY2022", "FY2023", "FY2024", "YTD Jul25"]


@pytest.fixture(scope="module")
def diagnose():
    _, d = SusaDatabookReader().lesen_mit_diagnose(datei(SUSA))
    return d


@pytest.fixture(scope="module")
def res(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("ks") / "ks.xlsx")
    return run(Quellen(susa=datei(SUSA), jahresabschluss=datei(JA),
                       kontenplan=datei(PLAN), pruefbericht=datei(BERICHT)),
               out, verbose=False)


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("ks2") / "ks.xlsx")
    run(Quellen(susa=datei(SUSA), jahresabschluss=datei(JA),
                kontenplan=datei(PLAN), pruefbericht=datei(BERICHT)),
        out, verbose=False)
    return openpyxl.load_workbook(out)


# ---- Defekt a): Kontenblock endet vor den Nebenrechnungen ----------------
def test_genau_317_kontozeilen(diagnose):
    """Unter dem Kontenblock stehen freie Nebenrechnungen, die in Spalte A
    erneut Kontonummern führen. Wer zu weit liest, bekommt mehr Zeilen."""
    assert diagnose.kontozeilen == 317


def test_alle_periodenspalten_summieren_auf_null(diagnose):
    assert diagnose.ist_ausgeglichen
    for p in PERIODEN:
        assert abs(diagnose.spaltensummen[p]) < 0.01, p


# ---- Defekt b): Monatsspalten sind kumuliert ----------------------------
def test_monatsspalten_sind_kumuliert_und_delta_ist_die_bewegung(diagnose):
    reihe = diagnose.monate["410 0"]
    assert reihe["2022/12"] == pytest.approx(48_505.00, abs=0.01)
    delta = diagnose.monatsdelta("410 0")
    assert sum(delta[p] for p in delta if p.startswith("2022")) == pytest.approx(
        reihe["2022/12"], abs=0.01)


# ---- Defekt c): doppelte Kontonummern -----------------------------------
def test_neun_duplikate_davon_eines_strittig(diagnose):
    assert len(diagnose.duplikate) == 9
    strittig = [d for d in diagnose.duplikate if not d.konsolidierbar]
    assert [d.konto for d in strittig] == ["1204 0"]
    assert strittig[0].ueberschneidung == ["FY2024"]


def test_strittiges_duplikat_landet_in_der_review_queue(res):
    konten = {m.konto for m in res["review"]} if res["review"] and hasattr(
        res["review"][0], "konto") else {e.account.konto for e in res["review"]}
    assert "1204 0" in konten


def test_kontenplan_loest_das_strittige_duplikat_auf(res):
    """Die zweite Zeile trägt die Bezeichnung von Konto 1205 0."""
    assert any("1205 0" in w for w in res["ledger"].warnungen)


# ---- Besonderheiten ------------------------------------------------------
def test_statistikkonten_erzeugen_keine_bilanzposition(res):
    by = {m.konto: m for m in res["mapped"]}
    for k in ("9140 0", "9199 0"):
        assert by[k].klasse.value == "TECH"


def test_ewb_liegt_unter_den_forderungen(res):
    by = {m.konto: m for m in res["mapped"]}
    assert "Forderungen aus Lieferungen und Leistungen" in by["9960 0"].hgb_pfad


def test_kontenplan_ist_keine_strukturquelle_sondern_lueckenfueller(res):
    """Der Kontenplan darf nur greifen, wo der Kontennachweis schweigt."""
    kn_konten = set(res["kn"].konten)
    plan = res["plan"]
    for konto in plan.mit_funktion():
        if konto in kn_konten:
            m = next((m for m in res["mapped"] if m.konto == konto), None)
            if m:
                assert m.hgb_pfad == res["kn"].konten[konto].hgb_pfad


# ---- Kontennachweis als Strukturquelle ----------------------------------
def test_kontennachweis_stimmt_gegen_die_bilanz_derselben_datei():
    """Gegenprobe für den PDF-Parser."""
    ja = lies_datev_ja(datei(JA))
    kn, bil = {}, {b.hgb_pfad: b for b in ja.bilanz if b.hgb_pfad}
    for e in ja.eintraege:
        if e.hgb_pfad:
            kn[e.hgb_pfad] = kn.get(e.hgb_pfad, 0.0) + e.gj
    assert len(kn) >= 18
    for pfad, summe in kn.items():
        assert summe == pytest.approx(bil[pfad].gj, abs=0.01), pfad


def test_saldenspaltung_wird_erkannt():
    ja = lies_datev_ja(datei(JA))
    assert set(ja.gespaltene_konten()) == {"1400 0", "1600 0", "701 0"}


def test_verlustvortrag_ist_ein_abzugsposten():
    """Er steht auf der Passivseite, ist aber ein Sollsaldo."""
    ja = lies_datev_ja(datei(JA))
    e = next(x for x in ja.eintraege if x.konto == "868 0")
    assert e.ist_abzugsposten
    assert e.vorzeichenrichtig() == pytest.approx(17_865_402.25, abs=0.01)


# ---- Spaltenweiser Status ------------------------------------------------
def test_status_gilt_je_spalte_nicht_pauschal(res):
    st = res["status"]
    assert st.gemischt
    assert st.fuer("FY2023").ist_abschlusstreu
    assert st.fuer("FY2022").ist_abschlusstreu
    assert not st.fuer("FY2024").ist_abschlusstreu
    assert not st.fuer("YTD Jul25").ist_abschlusstreu


def test_guv_ist_nirgends_abschlusstreu(res):
    for s in res["status"].spalten:
        assert not s.guv.startswith("abschlusstreu")


# ---- Abstimmung ----------------------------------------------------------
def test_fy2023_nur_eine_echte_differenz(res):
    r = res["recon_kn"]
    assert [(k.konto, round(k.differenz, 2)) for k in r.echte_differenzen] == [
        ("410 0", 172.00)]
    assert r.rest_gesamt() == pytest.approx(172.00, abs=0.01)


def test_saldenspaltung_ist_erklaert_nicht_offen(res):
    for z in res["recon_kn"].positionen:
        if abs(z.erklaert) > 0.005:
            assert abs(z.rest) < 0.01, z.hgb_pfad


def test_verrechnung_1789_in_1780_ist_erklaert(res):
    erklaert = {k.konto for k in res["recon_kn"].erklaerte_posten}
    assert {"1780 0", "1789 0"} <= erklaert


def test_fy2022_geht_nach_ueberleitung_auf(res):
    """Gegen den Prüfbericht bleibt nach den erklärten Posten nichts stehen."""
    assert res["recon_agg"].rest_gesamt() == pytest.approx(0.0, abs=0.01)
    assert res["recon_agg"].mit_rest == []


# ---- Benchmark -----------------------------------------------------------
@pytest.mark.parametrize("roh,klasse", [
    ("/netdebt/cashlike/shopify", "ND"),
    ("/profit and loss/income/sales/B2B", "PL"),
    ("/Gehalt/MA", "PL"),
    ("/QoE/Verkauf Sachanlagen", "PL"),
])
def test_fremde_syntax_wird_uebersetzt(roh, klasse):
    assert uebersetze(roh).klasse == klasse


def test_wc_bereinigung_bleibt_unuebersetzt():
    """Sagt Working Capital, lässt aber TWC/OWC offen — das darf nicht
    geraten werden."""
    assert uebersetze("/WC-Bereinigung/UL").klasse is None


def test_benchmark_beeinflusst_das_mapping_nicht(res):
    """Die manuelle Spalte ist Vergleich, nicht Eingabe."""
    from fdd.core.model import Quelle
    quellen = {m.quelle for m in res["mapped"]}
    assert Quelle.OVERRIDE not in quellen
    assert len(res["benchmark"].zeilen) == 79


def test_benchmark_weist_abweichungen_aus(res):
    bm = res["benchmark"]
    assert bm.abweichungen
    assert all(z.mensch_klasse != z.eigene_klasse for z in bm.abweichungen)


# ---- Databook-Struktur ---------------------------------------------------
@pytest.mark.parametrize("tab", ["Mastersheet", "Lead NA", "Net Debt",
                                 "Working Capital", "Lead PL", "Recon FY2023",
                                 "Recon FY2022 aggregiert", "Benchmark",
                                 "Status je Spalte", "Review-Queue"])
def test_tab_vorhanden(wb, tab):
    assert tab in wb.sheetnames


def test_status_steht_an_der_spalte(wb):
    ws = wb["Net Debt"]
    kopf = [str(ws.cell(4, c).value) for c in range(6, 10)]
    assert all(p in k for p, k in zip(PERIODEN, kopf))
    assert "abschlusstreu" in kopf[1]
    assert "vorläufig" in kopf[2]


def test_summenformeln_zeigen_auf_die_datenzeilen(wb):
    """Der Statuseintrag darf keine Zeile verschieben — sonst zeigten die
    Summenformeln ins Leere."""
    ws = wb["Net Debt"]
    sub = next(r for r in range(5, ws.max_row + 1)
               if str(ws.cell(r, 3).value).startswith("Netto-Finanzvermögen"))
    formel = ws.cell(sub, 6).value
    for r in range(5, sub):
        ist_daten = (isinstance(ws.cell(r, 6).value, str)
                     and "!" in str(ws.cell(r, 6).value)
                     and ws.row_dimensions[r].outline_level != 1)
        if ist_daten:
            assert f"F{r}" in formel, f"Datenzeile {r} fehlt in der Summe"
