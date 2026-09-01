"""Projekt Luma: Databook aus einer jährlichen MYOB-Saldenliste, Option B.

Die Quelle ist ein System-Export ohne Abschluss, ohne Kontennachweis und ohne
GuV. Die Prüfungen hier richten sich deshalb auf genau die Stellen, an denen
so eine Quelle stillschweigend falsch gelesen wird: die falsche Wertespalte,
der falsche Stichtag und ein Kontenplan, den die deutschen Regeln nicht
treffen.
"""

import openpyxl
import pytest

from conftest import datei
from fdd.core.model import Klasse
from fdd.luma import ERGEBNISKONTO, Quellen, run
from fdd.export.vorlage_layout import lies_layout
from fdd.readers.myob_susa import finde_pfad, lies_myob

DATEI = "Testdaten_Luma_SuSa_jaehrlich.xlsx"
GUV = "Testdaten_Luma_PL_jaehrlich.xlsx"
GEPRUEFT = "Testdaten_Luma_Abschluss_3Jahre.pdf"
ENTWURF = "Testdaten_Luma_Abschluss_EntwurfFY26.pdf"
PERIODEN = ["FY2023", "FY2024", "FY2025", "FY2026"]


def _quellen():
    return Quellen(saldenliste=datei(DATEI), guv=datei(GUV),
                   abschluss_geprueft=datei(GEPRUEFT),
                   abschluss_entwurf=datei(ENTWURF))


@pytest.fixture(scope="module")
def res(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("luma") / "luma.xlsx")
    return run(_quellen(), out, verbose=False)


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("luma2") / "luma.xlsx")
    run(_quellen(), out, verbose=False)
    return openpyxl.load_workbook(out)


# ---- Reader ---------------------------------------------------------------
def test_schlussbestand_statt_bewegung():
    """Der Export führt beide Spalten, und beide summieren auf null.

    Wer die Bewegungsspalte nimmt, baut ein Databook aus Veränderungen: die
    Aktivseite läge in FY2023 bei 260 TAUD statt bei 10,4 Mio.

    Die 10.436.102,56 liegen um 176.547,82 über der Aktivsumme des Exports
    (10.259.554,74). Das ist das Konto ``1-12500 Unearned Revenue``, ein
    Aktivkonto mit Habensaldo, das abgegrenzte Umsatzerlöse trägt und deshalb
    auf die Passivseite umgegliedert ist.
    """
    led, d = lies_myob(datei(DATEI))
    aktiva = sum(a.saldo("FY2023") for a in led.accounts
                 if (a.fs_pfad or "").startswith("/Aktiva"))
    assert aktiva == pytest.approx(10_436_102.56, abs=1.0)
    # Die Bewegungsspalte derselben Periode summiert ebenfalls auf null und
    # wäre damit auf den ersten Blick genauso plausibel.
    assert abs(d.spaltensummen["FY2023"][1]) <= 1.0


def test_stichtag_ist_der_letzte_tag_des_geschaeftsjahres():
    """``PeriodTo`` nennt den ersten Tag des letzten Periodenmonats. Wer ihn
    übernimmt, legt den Abschluss einen Monat zu früh."""
    _, d = lies_myob(datei(DATEI))
    assert [d.stichtage[p].strftime("%d.%m.%Y") for p in PERIODEN] == [
        "31.03.2023", "31.03.2024", "31.03.2025", "31.03.2026"]
    assert d.monate == {"FY2023": 9, "FY2024": 12, "FY2025": 12, "FY2026": 12}


def test_bilanzidentitaet_je_periode():
    led, _ = lies_myob(datei(DATEI))
    for p in PERIODEN:
        assert abs(sum(a.saldo(p) for a in led.accounts)) <= 1.0, p


def test_kostenstellen_werden_je_konto_summiert():
    """Ein Konto steht je Kostenstelle einmal in der Liste. Im Mastersheet ist
    es eine Zeile — die Kostenstelle wäre ein zweiter Schlüssel."""
    led, d = lies_myob(datei(DATEI))
    assert d.mehrfach, "Testdatei sollte Mehrfachzeilen enthalten"
    assert len(led.accounts) == len({a.konto for a in led.accounts})


def test_jede_kontogruppe_ist_zugeordnet():
    """Eine Gruppe ohne Zuordnung landet lautlos ohne Pfad in der Bilanz."""
    _, d = lies_myob(datei(DATEI))
    assert d.gruppen_ohne_zuordnung == set()
    assert [k for k, _, _ in d.ohne_pfad] == ["9-99999"]


@pytest.mark.parametrize("gruppe,erwartet", [
    ("Trade Debtors", "Forderungen aus Lieferungen und Leistungen"),
    ("Trade Creditors", "Verbindlichkeiten aus Lieferungen und Leistungen"),
    ("efine", "IV Kassenbestand und Guthaben bei Kreditinstituten"),
    # Der abgeschnittene Gruppenname des Exports muss trotzdem treffen.
    ("Provision for Bad & Doubtful", "Forderungen aus Lieferungen und Leistungen"),
])
def test_gruppenzuordnung(gruppe, erwartet):
    assert erwartet in finde_pfad(gruppe, "1-00000")


def test_finanzierung_ist_net_debt(res):
    """Trade Facility, R&D finance, Leasing und die Firmenkreditkarte stehen
    im Export unter den kurzfristigen Verbindlichkeiten. Es sind
    Finanzierungen — der Inhalt entscheidet, nicht die Rubrik."""
    by = {m.konto: m for m in res["mapped"]}
    for konto in ("2-16000", "2-70500", "2-80000", "2-90000", "2-15050",
                  "2-70400", "2-80100"):
        assert by[konto].klasse is Klasse.ND, f"{konto} {by[konto].bezeichnung}"


# ---- Ausgabe --------------------------------------------------------------
def test_cockpit_traegt_das_maerz_geschaeftsjahr(wb):
    """Aus dem Label ``FY2023`` würde der 31.12. abgeleitet. Die Zeitachse
    kommt deshalb aus den Stichtagen des Exports."""
    c = wb["Cockpit"]
    assert c["C17"].value.strftime("%d.%m.%Y") == "31.03.2023"
    assert c["C18"].value == 4
    assert c["C5"].value == "AUD"


def test_berichtssprache_englisch(wb):
    """Sprachwahl durch Ein- und Ausblenden, nicht durch Übersetzen: der
    deutsche Kopfblock und die deutsche Bezeichnungsspalte verschwinden,
    beide Fassungen bleiben in der Datei."""
    ws = wb["Lead NA"]
    assert ws.row_dimensions[4].hidden and ws.row_dimensions[5].hidden
    assert not ws.row_dimensions[6].hidden and not ws.row_dimensions[7].hidden
    assert ws.column_dimensions["C"].hidden
    assert not ws.column_dimensions["D"].hidden
    # Die Arbeitsblätter dieses Laufs folgen derselben Sprache.
    for blatt in ("Mapping", "Review queue", "Status by column"):
        assert blatt in wb.sheetnames


def test_cyt_felder_erzeugen_keine_fehlermeldung(wb):
    """Das Cockpit prüft die CYT-Felder selbst. Ohne Zwischenperiode müssen
    sie hinter dem letzten historischen Jahr liegen, sonst meldet die Mappe
    zwei Fehler, obwohl nichts fehlt."""
    c = wb["Cockpit"]
    assert c["C21"].value.strftime("%d.%m.%Y") == "01.04.2026"
    assert c["C22"].value.month != 12


def test_guv_wird_geladen_und_trifft_jede_position(res, wb):
    """Jede GuV-Position muss eine Zeile im Lead PL finden. Eine, die keine
    findet, verschwände lautlos aus der Ergebnisrechnung."""
    guv = [m for m in res["mapped"] if m.klasse is Klasse.PL]
    assert len(guv) > 200
    ms = wb["Mastersheet"]
    paare = {(ms.cell(r, 4).value, ms.cell(r, 3).value)
             for r in range(2, ms.max_row + 1)
             if ms.cell(r, 1).value and ms.cell(r, 3).value == "PL"}
    layout = lies_layout(wb["Lead PL"], ("T", "U", "V"), 5, 14)
    ticker = {(p.ticker1, p.ticker2) for p in layout.positionen}
    assert paare - ticker == set()


def test_ergebniskonto_faellt_bei_geladener_guv_heraus(res):
    """``3-90000 Current Earnings`` ist die GuV in einer Zahl. Bleibt es
    stehen, steht das Ergebnis zweimal in der Mappe und die Summe aller
    Konten geht nicht mehr auf null."""
    assert ERGEBNISKONTO not in {m.konto for m in res["mapped"]}
    assert ERGEBNISKONTO in res["diagnose"].eliminiert


def test_ergebnis_kommt_aus_der_guv(res):
    """Vorzeichen: Erträge stehen im Haben, ein Gewinn ist eine negative
    Summe und erhöht das Nettovermögen."""
    assert res["ergebnis"]["FY2024"] == pytest.approx(589_770.55, abs=0.01)
    assert res["ergebnis"]["FY2023"] == pytest.approx(-287_823.17, abs=0.01)


def test_perioden_nach_stichtag_nicht_nach_blattreihenfolge():
    """Die Tabs des GuV-Exports stehen als FY2023, FY2026, FY2024, FY2025."""
    import openpyxl as _o
    blaetter = _o.load_workbook(datei(GUV), read_only=True).sheetnames
    assert blaetter != sorted(blaetter), "Testdatei sollte unsortiert sein"
    led, _ = lies_myob(datei(DATEI), datei(GUV))
    assert led.perioden == PERIODEN


def test_abschluss_wird_gelesen_und_nicht_getippt(res):
    """Die Abstimmzahlen stammen aus den PDF, nicht aus einer Konstante."""
    a = res["abschluss"]
    from datetime import date
    assert a.wert("bilanz", date(2025, 3, 31), "NET ASSETS") == 5_466_698
    assert a.wert("guv", date(2024, 3, 31),
                  "Profit (Loss) for the year") == 619_069
    assert a.wert("guv", date(2026, 3, 31),
                  "Profit (Loss) for the year") == -3_024_663


def test_ueberleitung_weist_die_differenz_aus(res):
    """Die Abschlüsse sind konsolidiert, die Saldenliste ist eine Division.
    Die Differenz gehört ausgewiesen, nicht wegdefiniert."""
    zeilen = {u.position: u for u in res["ueberleitung"]}
    assert set(zeilen) == {"Net assets", "Revenue", "Result for the period"}
    ergebnis = zeilen["Result for the period"]
    assert ergebnis.differenz("FY2024") == pytest.approx(-29_298, abs=1)
    assert ergebnis.differenz("FY2025") == pytest.approx(22_985, abs=1)


def test_erste_periode_ist_anfangsbestand_keine_bewegung(res):
    """Das Mandat hat eine Vorgeschichte: 7,4 Mio Nettovermögen bestanden vor
    dem ersten Stichtag. Als Bewegung der ersten Periode ausgewiesen, hieße
    das, das Kapital sei in FY2023 eingezahlt worden."""
    rf = res["befuellt"].roll_forward
    assert rf.anfangsbestand == pytest.approx(7_393_852.99, abs=0.01)
    assert all(w["FY2023"] == 0.0 for w in rf.bewegungen.values())


def test_rest_wird_ausgewiesen_nicht_verteilt(res):
    """FY2024 trägt eine Eigenkapitalbewegung, die weder Kapital noch Ergebnis
    erklärt. Sie gehört in die manuelle Zeile des Roll Forward, sobald der
    Mandant sie erklärt — und bis dahin sichtbar in die Kontrollzeile."""
    rf = res["befuellt"].roll_forward
    assert rf.rest["FY2024"] == pytest.approx(268_375.20, abs=0.01)
    assert rf.rest["FY2023"] == 0.0
    assert "3-80000 Retained Earnings" in rf.ergebniskonten


def test_gemischte_positionen_landen_im_lead(res, wb):
    """Die drei gemischten Positionen lösen die Typ-2-Regeln auf englischen
    Bezeichnungen nicht auf. Sie bekommen über QA A6 eine vorläufige Klasse —
    und dabei muss auch die WC-Seite gesetzt werden.

    Ohne sie trüge der Ticker ``OWC``, den die Vorlage nicht kennt: 34 Konten
    standen im Mastersheet und in keiner Position des Lead NA, während die
    Bilanz im Datenmodell aufging.
    """
    vorlaeufig = [m for m in res["mapped"] if "QA A6" in (m.begruendung or "")
                  and not m.hgb_pfad.endswith("Noch nicht zugeordnet")]
    assert vorlaeufig, "Testdatei sollte gemischte Positionen enthalten"
    for m in vorlaeufig:
        if m.klasse in (Klasse.OWC, Klasse.TWC):
            assert m.seite in ("OA", "OL"), f"{m.konto} {m.bezeichnung}"

    # Im Mastersheet darf kein ``OWC`` mehr stehen — bis auf den Platzhalter
    # "Noch nicht zugeordnet", der keine Bilanzseite hat, aus der sich eine
    # WC-Seite ableiten ließe.
    ms = wb["Mastersheet"]
    owc = [ms.cell(r, 4).value for r in range(2, ms.max_row + 1)
           if ms.cell(r, 3).value == "OWC"]
    assert owc == ["Noch nicht zugeordnet"]


def test_keine_position_geht_im_lead_verloren(res):
    """Bis auf das Suspense-Konto, das bewusst keine Position hat."""
    verloren = [z.ziel_na for z in res["befuellt"].ohne_zeile]
    assert verloren == ["Noch nicht zugeordnet"]


def test_kontrollen_bis_auf_die_benannten_befunde(res):
    # Die Kontrollnamen stehen in der Berichtssprache des Mandats.
    benannt = ("residual", "account slot")
    offen = [k.name for k in res["kontrollen"]
             if not k.ok and not any(b in k.name for b in benannt)]
    assert offen == []
    # Die Rundungsdifferenz von bis zu 4 Cent stammt aus dem Export selbst.
    bilanz = next(k for k in res["kontrollen"]
                  if k.name.startswith("Balance sheet identity"))
    assert all(abs(v) <= 0.05 for v in bilanz.je_periode.values())


def test_grosse_konten_bekommen_die_slots(wb):
    """Reichen die acht Slots nicht, entscheidet die Größe. Sonst verschwindet
    das größte Konto der Position hinter acht kleinen."""
    ws = wb["Lead NA"]
    slots = {ws[f"P{r}"].value for r in range(19, 27)}
    assert "1-70610" in slots        # Prototype R&D, größtes Sachanlagenkonto
    assert "1-90200" in slots        # Right of Use Assets


# ---- Option B: drei Schichten ---------------------------------------------
def test_option_b_ohne_aufriss_wird_abgelehnt():
    """Option B ohne Aufrissplan ergäbe Positionen von null.

    Die Lead-Zeile zieht dann aus einem leeren Aufriss, und das Databook zeigt
    stumm eine runde, falsche Zahl. Eine Fehlermeldung ist besser."""
    from fdd.export import vorlage
    with pytest.raises(ValueError, match="Aufriss"):
        vorlage.schreibe_dealtool(
            "/tmp/unbenutzt.xlsx", mapped=[], perioden=["FY2024"],
            mandat=vorlage.Mandat(projekt="x", architektur="option_b"))


def test_unbekannte_architektur_wird_abgelehnt():
    from fdd.export import vorlage
    with pytest.raises(ValueError, match="option_a"):
        vorlage.schreibe_dealtool(
            "/tmp/unbenutzt.xlsx", mapped=[], perioden=["FY2024"],
            mandat=vorlage.Mandat(projekt="x", architektur="option_c"))


def test_lead_zieht_die_position_aus_dem_aufriss(wb):
    """Der Kern der Option B: die Positionszeile verweist auf den Aufriss.

    Unter Option A stünde hier ein SUMIFS aufs Mastersheet. Die erste
    Periode liegt im Lead in Spalte F (E ist die Eröffnungsspalte), in den
    Sachanlagen in E und in den Vorräten in F. Genau deshalb wird die
    Kopfzeile jedes Aufrisses gelesen und nicht abgezählt.
    """
    ws = wb["Lead NA"]
    assert ws["F18"].value == "='NA_Sachanlagen'!E20"
    assert ws["I18"].value == "='NA_Sachanlagen'!H20"
    assert ws["F40"].value == "='NA_Vorräte'!F22"
    assert ws["I40"].value == "='NA_Vorräte'!I22"


def test_eroeffnungsspalte_bleibt_am_mastersheet(wb):
    """Die Aufrisse der Vorlage führen keine Eröffnungsspalte.

    Spalte E des Leads hat im Aufriss also kein Gegenstück. Sie behält die
    Formel der Vorlage — dieselbe Quelle, nur ohne Zwischenschicht."""
    ws = wb["Lead NA"]
    assert "Mastersheet" in ws["E18"].value
    assert "Mastersheet" in ws["E40"].value


def test_kontoslots_haengen_weiter_am_mastersheet(wb):
    """Die zweite Schicht des Wegs darf nicht mitwandern.

    Zöge auch der Kontoslot aus dem Aufriss, gäbe es nur noch einen Weg auf
    die Zahl und die Kontrolle wäre eine Tautologie."""
    ws = wb["Lead NA"]
    assert "Mastersheet" in ws["F19"].value
    assert "Mastersheet" in ws["F41"].value


def test_aufriss_summiert_die_konten_der_position(res):
    """Pflichtkontrolle: Aufriss gegen Kontozeilen, je Position und Periode."""
    befunde = res["befuellt"].aufrissbefunde
    assert {b.blatt for b in befunde} == {"NA_Vorräte", "NA_Sachanlagen"}
    for b in befunde:
        assert b.konten_ohne_zeile == [], b.konten_ohne_zeile
        assert all(b.differenz(p) == 0.0 for p in PERIODEN), b.aufriss


def test_aufriss_kontrolle_steht_in_der_mappe(wb):
    """Die Kontrolle gehört ins Databook, nicht nur ins Protokoll."""
    assert "Schedule control" in wb.sheetnames
    ws = wb["Schedule control"]
    assert ws.max_row == 3           # Kopfzeile plus zwei Aufrisse
    assert "FY2026 difference" in [c.value for c in ws[1]]


def test_aufriss_traegt_die_beschriftung_der_quelle(wb):
    """Die Vorlage gliedert Vorräte nach Produkten, die Quelle nach
    Bestandteilen. Beschriftet wird, was tatsächlich in der Zeile steht."""
    ws = wb["NA_Vorräte"]
    assert ws["D13"].value == "Work in progress – materials and parts"
    assert ws["D21"].value == "Valuation allowance for inventories"
    assert "1-51000" in ws["F13"].value
