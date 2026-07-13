"""Working-Capital-View: Aufteilung OA/OL × TWC/OWC, NWC-Saldo, keine
Raster-Löcher, und Kontrollzeile == 0 im erzeugten Excel."""

import os
import warnings

import openpyxl
import pytest

from conftest import ALLE, datei
from fdd.cli import run
from fdd.core.model import Klasse
from fdd.views.working_capital import baue_working_capital


@pytest.mark.parametrize("name", ALLE)
def test_nwc_ist_summe_aller_wc_zeilen(name):
    res = run(datei(name), "/tmp/wc_probe.xlsx", verbose=False)
    wc = res["wc"]
    for p in wc.perioden:
        # NWC = OA + OL (vorzeichenrichtig) = TWC + OWC
        assert wc.net_working_capital(p) == pytest.approx(
            wc.operating_assets(p) + wc.operating_liabilities(p), abs=0.01)
        assert wc.net_working_capital(p) == pytest.approx(
            wc.twc(p) + wc.owc(p), abs=0.01)


@pytest.mark.parametrize("name", ALLE)
def test_keine_wc_konten_ohne_na_zeile(name):
    """Jedes als TWC/OWC klassifizierte Konto landet in einer NA-Zeile —
    keins fällt durchs Raster."""
    res = run(datei(name), "/tmp/wc_probe.xlsx", verbose=False)
    assert res["wc"].ohne_na_zeile == [], (
        f"{name}: WC-Konten ohne NA-Zeile: "
        f"{[(m.konto, m.klasse.value) for m in res['wc'].ohne_na_zeile]}")


@pytest.mark.parametrize("name", ALLE)
def test_wc_seiten_und_klassen_konsistent(name):
    """Operating Assets nur aus /Aktiva, Operating Liabilities nur aus /Passiva."""
    res = run(datei(name), "/tmp/wc_probe.xlsx", verbose=False)
    for z in res["wc"].zeilen:
        for m in z.konten:
            seite = "OA" if m.hgb_pfad.startswith("/Aktiva") else "OL"
            assert z.seite == seite
            assert m.klasse in (Klasse.TWC, Klasse.OWC)


@pytest.mark.parametrize("name", ALLE)
def test_wc_kontrollzeile_null_im_excel(name, tmp_path):
    formulas = pytest.importorskip("formulas")
    import numpy as np
    warnings.filterwarnings("ignore")
    out = str(tmp_path / "db.xlsx")
    run(datei(name), out, verbose=False)

    wb = openpyxl.load_workbook(out)
    assert "Working Capital" in wb.sheetnames
    ws = wb["Working Capital"]
    kz_row = n_per = 0
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 3).value or "").startswith("Kontrollzeile"):
            kz_row = r
    for c in range(4, ws.max_column + 1):
        if ws.cell(4, c).value:
            n_per += 1
    wb.close()
    assert kz_row and n_per

    sol = formulas.ExcelModel().loads(out).finish().calculate()
    up = os.path.basename(out).upper()
    from openpyxl.utils import get_column_letter
    for i in range(n_per):
        col = get_column_letter(4 + i)
        v = sol.get(f"'[{up}]WORKING CAPITAL'!{col}{kz_row}")
        val = np.ravel(v.value)[0] if v is not None else 0.0
        assert abs(float(val)) < 1e-3, f"{name} {col}: WC-Kontrollzeile {val} != 0"
