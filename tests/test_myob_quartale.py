"""MYOB-Export und quartalsweises Mastersheet.

Die Fälle prüfen genau die Stellen, an denen ein Databook plausibel aussieht
und falsch ist: die vertauschte Wertespalte, das kumulierte GuV-Feld, die
Blattreihenfolge, die verlorene Kostenstelle — und im Export die Frage, ob
eine Jahresspalte ein Bestand oder eine Summe ist.
"""

from __future__ import annotations

from datetime import date

import openpyxl
import pytest

from fdd.core.model import Klasse, MappedAccount, PeriodBalance, Quelle
from fdd.core.model import Account
from fdd.export import excel
from fdd.readers.myob_export import (Periodenraster, _jahr_von, _lies_blatt,
                                     _quartal_von, lies_myob_export)

#: Kopfzeile des Exports. Der Reader sucht die Spalten über sie, nicht über
#: die Position — deshalb steht sie hier in einer anderen Reihenfolge als in
#: der Originaldatei.
KOPF = ["AccountNo", "GLDescription", "ClassDescription", "AccountGroupDesc",
        "CostCentreName", "PeriodMovementTYD", "Year", "PeriodFrom", "PeriodTo"]


def blatt(ws, zeilen, von, bis):
    ws.append(KOPF)
    for konto, bez, klasse, gruppe, kst, bewegung, bestand in zeilen:
        ws.append([konto, bez, klasse, gruppe, kst, bewegung, bestand, von, bis])


@pytest.fixture
def export(tmp_path):
    """Ein Minimalexport mit zwei Jahren, vier Quartalen und Kostenstellen."""
    bs = openpyxl.Workbook()
    bs.remove(bs.active)
    # FY2025: zwei Quartale (Rumpfjahr), FY2026: zwei Quartale.
    # Bestand des Kontos 1-1000 wächst, das Gegenkonto 2-1000 spiegelt es.
    stand = {"Q3-25": 300.0, "Q4-25": 400.0, "Q3-26": 500.0, "Q4-26": 600.0}
    perioden = {
        "Q3-25": ("2024-10-01", "2024-12-01"), "Q4-25": ("2025-01-01", "2025-03-01"),
        "Q3-26": ("2025-10-01", "2025-12-01"), "Q4-26": ("2026-01-01", "2026-03-01"),
    }
    for name, (von, bis) in perioden.items():
        blatt(bs.create_sheet(name), [
            # Dasselbe Konto zweimal, einmal je Kostenstelle: 3/4 und 1/4.
            ("1-1000", "CBA Main Cheque Acct", "Current Assets", "Bank", "A",
             0.0, stand[name] * 0.75),
            ("1-1000", "CBA Main Cheque Acct", "Current Assets", "Bank", "B",
             0.0, stand[name] * 0.25),
            ("3-1000", "Retained Earnings", "Equity", "Equity", "",
             0.0, -stand[name]),
        ], von, bis)
    # Jahresblätter ABSICHTLICH am Ende und in falscher Reihenfolge.
    blatt(bs.create_sheet("Jahr 26"), [
        ("1-1000", "CBA Main Cheque Acct", "Current Assets", "Bank", "", 0.0, 600.0),
        ("3-1000", "Retained Earnings", "Equity", "Equity", "", 0.0, -600.0),
    ], "2025-04-01", "2026-03-01")
    blatt(bs.create_sheet("Jahr 25"), [
        ("1-1000", "CBA Main Cheque Acct", "Current Assets", "Bank", "", 0.0, 400.0),
        ("3-1000", "Retained Earnings", "Equity", "Equity", "", 0.0, -400.0),
    ], "2024-04-01", "2025-03-01")
    bs_pfad = tmp_path / "bs.xlsx"
    bs.save(bs_pfad)

    pl = openpyxl.Workbook()
    pl.remove(pl.active)
    # ``Year`` ist in der GuV KUMULIERT: Q4 trägt Q3+Q4. Wer ihn liest,
    # bekommt für Q4 den doppelten Betrag.
    guv = {"Q3-25": (10.0, 10.0), "Q4-25": (20.0, 30.0),
           "Q3-26": (40.0, 40.0), "Q4-26": (50.0, 90.0)}
    for name, (bewegung, kumuliert) in guv.items():
        von, bis = perioden[name]
        blatt(pl.create_sheet(name), [
            ("6-1000", "Salaries & Wages", "Expenses", "Employment Costs", "",
             bewegung, kumuliert)], von, bis)
    blatt(pl.create_sheet("Jahr 25"), [
        ("6-1000", "Salaries & Wages", "Expenses", "Employment Costs", "",
         30.0, 30.0)], "2024-04-01", "2025-03-01")
    blatt(pl.create_sheet("Jahr 26"), [
        ("6-1000", "Salaries & Wages", "Expenses", "Employment Costs", "",
         90.0, 90.0)], "2025-04-01", "2026-03-01")
    pl_pfad = tmp_path / "pl.xlsx"
    pl.save(pl_pfad)
    return lies_myob_export([str(bs_pfad)], [str(pl_pfad)], entity="Test")


# ---- Periodenlogik -------------------------------------------------------

def test_geschaeftsjahr_folgt_dem_stichtag_nicht_dem_kalenderjahr():
    """Bei einem am 31.03. endenden Geschäftsjahr gehört der 30.09.2022 zu
    FY2023 — und der 31.03.2023 noch zu FY2023."""
    assert _jahr_von(date(2022, 9, 30), 3) == "FY2023"
    assert _jahr_von(date(2023, 3, 31), 3) == "FY2023"
    assert _jahr_von(date(2023, 6, 30), 3) == "FY2024"


def test_quartalsnummer_zaehlt_ab_geschaeftsjahresbeginn():
    for monat, erwartet in ((6, 1), (9, 2), (12, 3), (3, 4)):
        assert _quartal_von(date(2025, monat, 28), 3) == erwartet


def test_periodto_ist_der_erste_tag_des_letzten_monats(export):
    """Aus dem 01.03. wird der 31.03. Wer ihn übernimmt, legt den Abschluss
    einen Monat zu früh."""
    _, raster, _ = export
    assert raster.stichtage["FY2026"] == date(2026, 3, 31)
    assert raster.stichtage["FY2026 Q4"] == date(2026, 3, 31)


def test_blattreihenfolge_ist_nicht_die_zeitachse(export):
    """Die Jahresblätter stehen in der Testdatei am Ende und verkehrt herum."""
    _, raster, _ = export
    assert raster.jahre == ["FY2025", "FY2026"]
    assert raster.spalten == ["FY2025 Q3", "FY2025 Q4", "FY2025",
                              "FY2026 Q3", "FY2026 Q4", "FY2026"]


def test_leere_trennblaetter_werden_benannt_nicht_verschluckt(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "24 25 >>"
    wb.create_sheet("leer2")
    blatt(wb.create_sheet("Jahr"), [
        ("1-1000", "Cash", "Current Assets", "Bank", "", 0.0, 0.0)],
        "2024-04-01", "2025-03-01")
    p = tmp_path / "x.xlsx"
    wb.save(p)
    _, _, diag = lies_myob_export([str(p)], [], entity="T")
    assert len(diag.uebersprungene_blaetter) == 2


# ---- die beiden Wertespalten --------------------------------------------

def test_bilanz_liest_den_bestand_nicht_die_bewegung(export):
    ledger, raster, _ = export
    bank = next(a for a in ledger.accounts if a.konto == "1-1000")
    assert bank.saldo("FY2026 Q3") == 500.0
    assert bank.saldo("FY2026") == 600.0


def test_guv_liest_die_bewegung_nicht_das_kumulierte_feld(export):
    """``Year`` trägt in der GuV Q3+Q4. Der Reader muss das Quartal liefern."""
    ledger, _, _ = export
    lohn = next(a for a in ledger.accounts if a.konto == "6-1000")
    assert lohn.saldo("FY2026 Q3") == 40.0
    assert lohn.saldo("FY2026 Q4") == 50.0     # nicht 90.0
    assert lohn.saldo("FY2026") == 90.0


def test_kostenstellen_werden_summiert_nicht_ueberschrieben(export):
    """Das Konto steht je Blatt zweimal, 3/4 und 1/4. Wer die erste Zeile
    nimmt, verliert ein Viertel der Bilanz."""
    ledger, _, _ = export
    bank = next(a for a in ledger.accounts if a.konto == "1-1000")
    assert bank.saldo("FY2025 Q4") == 400.0


def test_reader_belegt_sich_selbst(export):
    """Die drei Proben der Diagnose müssen aufgehen, sonst ist eine Spalte
    vertauscht oder ein Blatt übersehen."""
    _, _, diag = export
    assert all(abs(w) < 0.005 for w in diag.bilanzidentitaet.values())
    for jahr, (aus_quartalen, jahreswert) in diag.guv_aufriss.items():
        assert abs(aus_quartalen - jahreswert) < 0.005, jahr
    for jahr, (q4, ganz) in diag.bilanz_jahresende.items():
        assert abs(q4 - ganz) < 0.005, jahr


def test_rumpfjahr_wird_gemeldet(export):
    """Beide Jahre führen nur zwei Quartale — das ist kein Fehler des
    Readers, aber es muss dranstehen."""
    _, _, diag = export
    assert set(diag.unvollstaendige_jahre) == {"FY2025", "FY2026"}
    assert diag.warnungen


def test_kontotyp_und_gruppe_kommen_aus_der_quelle(export):
    ledger, _, _ = export
    bank = next(a for a in ledger.accounts if a.konto == "1-1000")
    lohn = next(a for a in ledger.accounts if a.konto == "6-1000")
    assert (bank.kontotyp, bank.fristigkeit) == ("bilanz_aktiv", "Current")
    assert bank.gruppe == "Bank"
    assert lohn.kontotyp == "guv"


# ---- das quartalsweise Mastersheet ---------------------------------------

def _mapped(klasse: Klasse, salden: dict[str, float]) -> MappedAccount:
    a = Account(konto="1-1000", bezeichnung="X",
                salden=tuple(PeriodBalance(p, v) for p, v in salden.items()))
    return MappedAccount(account=a, hgb_pfad="/AASB/Aktiva/Cash",
                         hgb_pfad_en="/AASB/Aktiva/Cash", klasse=klasse,
                         na_de="Cash", na_en="Cash",
                         quelle=Quelle.AASB_STICHWORT)


@pytest.fixture
def raster() -> Periodenraster:
    r = Periodenraster(jahre=["FY2026"])
    r.quartale_je_jahr["FY2026"] = [f"FY2026 Q{i}" for i in (1, 2, 3, 4)]
    return r


def _schreibe(tmp_path, mapped, raster):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    excel._schreibe_mastersheet(wb, mapped, ["FY2026"], "Test", raster)
    p = tmp_path / "ms.xlsx"
    wb.save(p)
    return openpyxl.load_workbook(p)["Mastersheet"]


def test_mastersheet_fuehrt_quartale_und_jahr(tmp_path, raster):
    werte = {f"FY2026 Q{i}": 100.0 * i for i in (1, 2, 3, 4)}
    ws = _schreibe(tmp_path, [_mapped(Klasse.TWC, {**werte, "FY2026": 400.0})],
                   raster)
    kopf = [ws.cell(1, c).value for c in range(12, 17)]
    assert kopf[:4] == list(werte)
    assert kopf[4].startswith("FY2026")


def test_bilanzjahr_ist_der_bestand_des_letzten_quartals(tmp_path, raster):
    """Die Verwechslung ist der klassische Fehler: wer die Bilanz über die
    Quartale summiert, vervierfacht sie."""
    werte = {f"FY2026 Q{i}": 100.0 * i for i in (1, 2, 3, 4)}
    ws = _schreibe(tmp_path, [_mapped(Klasse.TWC, {**werte, "FY2026": 400.0})],
                   raster)
    assert ws.cell(2, 16).value == "=O2"


def test_guvjahr_ist_die_summe_der_quartale(tmp_path, raster):
    werte = {f"FY2026 Q{i}": 100.0 * i for i in (1, 2, 3, 4)}
    ws = _schreibe(tmp_path, [_mapped(Klasse.PL, {**werte, "FY2026": 1000.0})],
                   raster)
    assert ws.cell(2, 16).value == "=SUM(L2:O2)"


def test_widerspruch_bleibt_stehen_statt_hinter_der_formel_zu_verschwinden(
        tmp_path, raster):
    """Trifft die Formel den Wert der Quelle nicht, bleibt die Zahl der
    Quelle stehen und die Zelle wird markiert."""
    werte = {f"FY2026 Q{i}": 100.0 for i in (1, 2, 3, 4)}
    ws = _schreibe(tmp_path, [_mapped(Klasse.TWC, {**werte, "FY2026": 999.0})],
                   raster)
    zelle = ws.cell(2, 16)
    assert zelle.value == 999.0
    assert zelle.fill.start_color.rgb not in (None, "00000000")


def test_ohne_raster_bleibt_das_mastersheet_wie_bisher(tmp_path):
    """Die Quartalsspalten sind ein Zusatz, keine Umstellung: ohne ``raster``
    schreibt das Mastersheet unverändert die übergebenen Perioden."""
    ws = _schreibe(tmp_path, [_mapped(Klasse.TWC, {"FY2026": 400.0})], None)
    assert ws.cell(1, 12).value == "FY2026"
    assert ws.cell(2, 12).value == 400.0
    assert ws.cell(1, 13).value is None
