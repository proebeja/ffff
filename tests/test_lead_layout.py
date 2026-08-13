"""Darstellung und Gliederung der Lead-Tabs: aufklappbare Umgliederungs-
Details, einklappbare Quellenspalte, TWC/OWC-Blöcke, Reported-Sicht."""

import openpyxl
import pytest

from conftest import ECKART, datei
from fdd.cli import run

LEADS = ("Net Debt", "Working Capital")


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("lead") / "eckart.xlsx")
    run(datei(ECKART), out, verbose=False,
        kontennachweis=datei("Kontennachweis_Eckart_2022-2024.xlsx"))
    return openpyxl.load_workbook(out)


def _zeilen(ws):
    return {r: ws.cell(r, 3).value for r in range(5, ws.max_row + 1)
            if ws.cell(r, 3).value}


# ---- 1) Umgliederungszeilen aufklappbar ----------------------------------
def test_umgliederung_hat_eingeklappte_detailzeilen(wb):
    ws = wb["Net Debt"]
    details = [r for r in range(5, ws.max_row + 1)
               if ws.row_dimensions[r].outline_level == 1]
    assert details, "keine gruppierten Detailzeilen im Net-Debt-Tab"
    for r in details:
        assert ws.row_dimensions[r].hidden, f"Zeile {r} ist nicht eingeklappt"
        assert ws.cell(r, 2).value, "Detailzeile ohne Kontonummer"
        assert ws.cell(r, 5).value, "Detailzeile ohne Herkunftsposition"


def test_summenzeile_steht_ueber_den_details(wb):
    assert wb["Net Debt"].sheet_properties.outlinePr.summaryBelow is False


def test_detailzeilen_ziehen_aus_demselben_aufriss_wie_die_summe(wb):
    """Kein zweiter Datenpfad: Detail- und Summenzeile referenzieren beide den
    gemischten Aufriss der Position."""
    ws = wb["Net Debt"]
    letzte_summe = None
    for r in range(5, ws.max_row + 1):
        if not ws.cell(r, 3).value:
            continue
        quelle = ws.cell(r, 4).value
        if ws.row_dimensions[r].outline_level == 1:
            assert quelle == letzte_summe, (
                f"Detailzeile {r} zieht aus '{quelle}', Summenzeile aus "
                f"'{letzte_summe}'")
            for c in range(6, ws.max_column + 1):
                f = ws.cell(r, c).value
                if isinstance(f, str) and f.startswith("="):
                    assert quelle in f, f"Detailwert {r}/{c} nicht aus {quelle}"
        elif quelle:
            letzte_summe = quelle


def test_detailzeilen_laufen_nicht_in_die_zwischensumme(wb):
    """Die Zwischensumme addiert ausschließlich Summenzeilen — sonst würde der
    thereof-ND-Teil doppelt gezählt."""
    ws = wb["Net Debt"]
    sub = next(r for r, t in _zeilen(ws).items()
               if str(t).startswith("Netto-Finanzvermögen"))
    formel = ws.cell(sub, 6).value
    details = {r for r in range(5, ws.max_row + 1)
               if ws.row_dimensions[r].outline_level == 1}
    for r in details:
        assert f"F{r}" not in formel, f"Detailzeile {r} steckt in der Zwischensumme"


# ---- 2) Quellenverweise in eigener, einklappbarer Spalte -----------------
@pytest.mark.parametrize("tab", LEADS)
def test_quellenspalte_ist_gruppiert_und_zeilentext_frei_von_verweisen(wb, tab):
    ws = wb[tab]
    assert ws.cell(4, 4).value == "Quelle"
    assert ws.column_dimensions["D"].outline_level == 1, "Spalte nicht gruppiert"
    for r, text in _zeilen(ws).items():
        assert "[" not in str(text), f"{tab} Zeile {r} trägt noch einen Tab-Verweis"


@pytest.mark.parametrize("tab", LEADS)
def test_quellenspalte_ist_gefuellt(wb, tab):
    ws = wb[tab]
    assert any(ws.cell(r, 4).value for r in range(5, ws.max_row + 1))


# ---- 3) WC-Lead: TWC-Block / OWC-Block / NWC -----------------------------
def test_wc_ist_nach_twc_und_owc_gegliedert(wb):
    texte = list(_zeilen(wb["Working Capital"]).values())
    reihenfolge = [t for t in texte if str(t).startswith(("Block ", "Saldo ", "Net Working"))]
    assert reihenfolge == [
        "Block 1 — Trade Working Capital",
        "Saldo Trade Working Capital (TWC)",
        "Block 2 — Other Working Capital",
        "Saldo Other Working Capital (OWC)",
        "Net Working Capital (TWC + OWC)",
    ]


def test_nwc_ist_summe_der_beiden_blocksalden(wb):
    ws = wb["Working Capital"]
    z = _zeilen(ws)
    twc = next(r for r, t in z.items() if t == "Saldo Trade Working Capital (TWC)")
    owc = next(r for r, t in z.items() if t == "Saldo Other Working Capital (OWC)")
    nwc = next(r for r, t in z.items() if t == "Net Working Capital (TWC + OWC)")
    assert ws.cell(nwc, 6).value == f"=F{twc}+F{owc}"


def test_oa_ol_bleibt_als_spalte_erhalten(wb):
    ws = wb["Working Capital"]
    assert ws.cell(4, 5).value == "OA/OL"
    seiten = {ws.cell(r, 5).value for r in range(5, ws.max_row + 1)}
    assert {"OA", "OL"} <= seiten


# ---- 4) Reported-Sicht mit davon-ND --------------------------------------
def test_gemischte_positionen_zeigen_reported_davon_nd_operativ(wb):
    ws = wb["Working Capital"]
    z = _zeilen(ws)
    reported = [r for r, t in z.items() if "(reported)" in str(t)]
    assert len(reported) == 3, "drei gemischte Positionen erwartet"
    for r in reported:
        assert "davon ND" in str(z[r + 1])
        assert "operativer Anteil" in str(z[r + 2])
        # operativ = reported + (negativ gezeigter) ND-Abzug
        assert ws.cell(r + 2, 6).value == f"=F{r}+F{r + 1}"
        # beide Zahlen aus demselben Aufriss
        quelle = ws.cell(r, 4).value
        assert quelle in ws.cell(r, 6).value
        assert quelle in ws.cell(r + 1, 6).value


def test_nur_der_operative_anteil_laeuft_in_den_owc_saldo(wb):
    """Der ND-Betrag bleibt dem Net-Debt-Tab vorbehalten: weder die
    Reported- noch die davon-Zeile darf im Blocksaldo auftauchen."""
    ws = wb["Working Capital"]
    z = _zeilen(ws)
    owc = next(r for r, t in z.items() if t == "Saldo Other Working Capital (OWC)")
    formel = ws.cell(owc, 6).value
    for r, t in z.items():
        if "(reported)" in str(t) or "davon ND" in str(t):
            assert f"F{r}" not in formel, f"Zeile {r} läuft in den OWC-Saldo"
        if "operativer Anteil" in str(t):
            assert f"F{r}" in formel, f"operative Zeile {r} fehlt im OWC-Saldo"


# ---- unveränderte Invarianten -------------------------------------------
@pytest.mark.parametrize("tab", LEADS)
def test_keine_lead_zeile_zieht_direkt_aus_dem_mastersheet(wb, tab):
    """Single Source: nur die Kontrollzeile darf gegen das Mastersheet prüfen."""
    ws = wb[tab]
    for r in range(5, ws.max_row + 1):
        if "ontrollzeile" in str(ws.cell(r, 3).value):
            continue
        for c in range(2, ws.max_column + 1):
            f = ws.cell(r, c).value
            if isinstance(f, str) and f.startswith("="):
                assert "Mastersheet" not in f, f"{tab} {r}/{c} zieht aus dem Mastersheet"


@pytest.mark.parametrize("tab", LEADS)
def test_kontrollzeile_ist_weiterhin_vorhanden(wb, tab):
    assert any("ontrollzeile" in str(t) for t in _zeilen(wb[tab]).values())
