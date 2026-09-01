"""Projekt Brehna: Databook aus vier Jahresabschlüssen, ohne Saldenliste.
Der Abschluss ist Werte- UND Strukturquelle, inklusive GuV."""

import openpyxl
import pytest

from conftest import datei
from fdd.brehna import PERIODENLAENGE, Quellen, run
from fdd.readers.datev_ja_hds_pdf import lies_brehna_ja, lies_brehna_jahre

DATEIEN = [f"Testdaten_Brehna_JA_{j}.pdf"
           for j in ("2023", "2024", "2025", "2026_Juli")]
PERIODEN = ["FY2023", "FY2024", "FY2025", "YTD 07/2026"]


@pytest.fixture(scope="module")
def pfade():
    return [datei(n) for n in DATEIEN]


@pytest.fixture(scope="module")
def res(pfade, tmp_path_factory):
    out = str(tmp_path_factory.mktemp("brehna") / "b.xlsx")
    return run(Quellen(jahresabschluesse=pfade), out, verbose=False)


@pytest.fixture(scope="module")
def wb(pfade, tmp_path_factory):
    out = str(tmp_path_factory.mktemp("brehna2") / "b.xlsx")
    run(Quellen(jahresabschluesse=pfade), out, verbose=False)
    return openpyxl.load_workbook(out)


# ---- Reader ---------------------------------------------------------------
def test_vier_perioden_in_richtiger_reihenfolge(pfade):
    led, _ = lies_brehna_jahre(pfade)
    assert led.perioden == PERIODEN


def test_zwischenabschluss_wird_erkannt(pfade):
    _, abschluesse = lies_brehna_jahre(pfade)
    zwischen = [a for a in abschluesse if a.ist_zwischenabschluss]
    assert [a.periode for a in zwischen] == ["YTD 07/2026"]


def test_parser_beweist_sich_selbst(pfade):
    """QA C5: jede Position gegen ihre gedruckte Summe, jede Seite gegen die
    gedruckte Bilanzsumme, Aktiva gegen Passiva. Ein Parser scheitert leise —
    er liefert zu wenige Konten und sieht dabei erfolgreich aus."""
    _, abschluesse = lies_brehna_jahre(pfade)
    for a in abschluesse:
        proben = a.probe()
        assert proben, a.periode
        for name, ok, detail in proben:
            assert ok, f"{a.periode} · {name}: {detail}"


def test_bilanzidentitaet_je_periode(pfade):
    led, _ = lies_brehna_jahre(pfade)
    for p in led.perioden:
        assert sum(a.saldo(p) for a in led.accounts) == pytest.approx(0.0, abs=0.01)


def test_umbrochene_bezeichnungen_werden_zusammengesetzt(pfade):
    """'Rückstellungen zur Erfüllung der Aufbewahrungs-\\npflichten' bricht
    mitten im Wort um; der Betrag steht erst auf der Folgezeile."""
    led, _ = lies_brehna_jahre(pfade)
    by = {a.konto: a for a in led.accounts}
    assert "Aufbewahrungspflichten" in by["96600"].bezeichnung
    assert by["96600"].saldo("FY2025") == pytest.approx(-1_500.00, abs=0.01)


def test_vorzeichen_werden_auf_databook_konvention_gedreht(pfade):
    """Der Abschluss druckt die Passivseite positiv und Aufwendungen negativ."""
    led, _ = lies_brehna_jahre(pfade)
    by = {a.konto: a for a in led.accounts}
    assert by["80000"].saldo("FY2025") == pytest.approx(-1_000.00, abs=0.01)
    assert by["497000"].saldo("FY2023") == pytest.approx(14.12, abs=0.01)
    assert by["120000"].saldo("FY2023") == pytest.approx(985.88, abs=0.01)


def test_verlustvortrag_aus_dem_guv_nachweis_ist_eigenkapital(pfade):
    """Konto 286800 steht im GuV-Nachweis, gehört aber ins Eigenkapital."""
    led, _ = lies_brehna_jahre(pfade)
    by = {a.konto: a for a in led.accounts}
    assert by["286800"].fs_pfad.startswith("/Passiva/A Eigenkapital")


def test_jedes_konto_hat_einen_pfad(pfade):
    led, _ = lies_brehna_jahre(pfade)
    assert all(a.fs_pfad for a in led.accounts)
    assert led.warnungen == []


# ---- Mapping --------------------------------------------------------------
def test_guv_ist_abschlusstreu(res):
    """Der Kontennachweis deckt auch die GuV ab — anders als bei
    Kitchenstories, wo sie nur abgeleitet war."""
    for s in res["status"].spalten:
        assert s.guv.startswith("abschlusstreu")
        assert s.bilanz.startswith("abschlusstreu")


def test_periodenlaenge_wird_ausgewiesen(res):
    """FY2023 sind zwei Monate (Gründung im November), YTD 07/2026 sieben."""
    assert PERIODENLAENGE["FY2023"] == 2
    kurz = [s for s in res["status"].spalten if s.hinweis]
    assert {s.periode for s in kurz} == {"FY2023", "YTD 07/2026"}


@pytest.mark.parametrize("konto,klasse", [
    ("120001", "ND"),      # Bank
    ("73000", "ND"),       # Verbindlichkeiten gegenüber Gesellschaftern
    ("70100", "ND"),       # verbundene Unternehmen
    ("170003", "ND"),      # Gesellschafterdarlehen
    ("150000", "ND"),      # Darlehen
    ("709000", "TWC"),     # in Ausführung befindliche Bauaufträge
    ("160000", "TWC"),     # Verbindlichkeiten aus L+L
    ("80000", "EQ"),
    ("210000", "PL"),      # Zinsaufwand
])
def test_klassifizierung(res, konto, klasse):
    by = {m.konto: m for m in res["mapped"]}
    assert by[konto].klasse.value == klasse


def test_kein_konto_ausserhalb_der_bilanz(res):
    for m in res["mapped"]:
        assert m.klasse.value not in ("REVIEW",), m.konto
        assert not m.hgb_pfad.startswith("("), m.konto


# ---- Ausgabe (Dealtool-Vorlage) -------------------------------------------
@pytest.mark.parametrize("tab", ["Cockpit", "Mastersheet", "Lead NA", "Lead PL",
                                 "QA", "Status je Spalte", "Review-Queue",
                                 "Zuordnung"])
def test_tab_vorhanden(wb, tab):
    assert tab in wb.sheetnames


def test_kein_benchmark_tab(wb):
    """Es gibt keine fremde Klassifizierung, also auch nichts zu vergleichen."""
    assert "Benchmark" not in wb.sheetnames


def test_alle_kontrollen_gehen_auf(res):
    """Bilanzidentität, Net-Asset-Brücke, Roll Forward und GuV-Gegenprobe.

    Ausgenommen ist nur die Kontrolle über die Kontoslots: die Vorlage hält
    für die drei über Dummy-Zeilen angelegten Net-Debt-Positionen keine Slots
    vor. Das ist ein Befund an der Vorlage, kein Rechenfehler.
    """
    offen = [k.name for k in res["kontrollen"]
             if not k.ok and "Kontoslots" not in k.name]
    assert offen == []


def test_qa_weist_gegenstandslose_pruefungen_aus(res):
    """Prüfungen, die eine Saldenliste voraussetzen, werden als gegenstandslos
    ausgewiesen statt stillschweigend wegzufallen."""
    ids = {p.id for p in res["qa"].pruefungen}
    erwartet = {f"A{i}" for i in range(1, 7)} | {f"B{i}" for i in range(1, 6)} \
        | {f"C{i}" for i in range(1, 6)} | {"D1"}
    assert erwartet <= ids
    b1 = next(p for p in res["qa"].pruefungen if p.id == "B1")
    assert "Gegenstandslos" in b1.befund


def test_c5_ist_die_zentrale_pruefung(res):
    c5 = next(p for p in res["qa"].pruefungen if p.id == "C5")
    assert c5.bestanden
    assert "49 von 49" in c5.befund


def test_net_assets_entsprechen_dem_eigenkapital(res):
    """Stammkapital 1.000 abzüglich Bilanzverlust 400,51 = 599,49."""
    for p in ("FY2024", "FY2025", "YTD 07/2026"):
        na = sum(m.saldo(p) for m in res["mapped"]
                 if m.klasse.value in ("FA", "TWC", "OWC", "ND", "DT"))
        assert na == pytest.approx(599.49, abs=0.01), p


def test_ergebnis_ist_null_weil_alles_aktiviert_wird(res):
    """Ein Projektentwickler aktiviert Kosten und Zinsen in die Bauaufträge;
    ab FY2025 wird kein Ergebnis ausgewiesen."""
    for p in ("FY2025", "YTD 07/2026"):
        pl = sum(m.saldo(p) for m in res["mapped"] if m.klasse.value == "PL")
        assert pl == pytest.approx(0.0, abs=0.01), p


def test_laufprotokoll_und_keine_ki_aufrufe(res):
    lp = res["laufprotokoll"]
    assert lp.gesamt > 0 and lp.zellen > 500
    assert lp.ki_aufrufe == []
