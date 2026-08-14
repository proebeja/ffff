"""Projekt Chiara: monatlich geschnittener SAP-BW-Export als Strukturquelle,
mehrjährige Zusammenführung, Lead NA / Lead PL, Reconciliation gegen den
Jahresabschluss."""

import openpyxl
import pytest

from conftest import datei
from fdd.cli import run
from fdd.readers.sap_bw_monate import SapBwMonateReader, lies_sap_jahre

CHIARA = [f"Testdaten_Chiara_SAP_4756_Monate_{j}.xlsx" for j in (2022, 2023, 2024)]
CHIARA_JA = "Testdaten_Chiara_Datenbuch_JA.xlsx"
CHIARA_2025 = "Testdaten_Chiara_SAP_4756_Monate_2025.xlsx"


@pytest.fixture(scope="module")
def pfade():
    return [datei(n) for n in CHIARA]


@pytest.fixture(scope="module")
def res(pfade, tmp_path_factory):
    out = str(tmp_path_factory.mktemp("chiara") / "chiara.xlsx")
    return run(pfade, out, verbose=False, jahresabschluss=datei(CHIARA_JA))


@pytest.fixture(scope="module")
def wb(res, pfade, tmp_path_factory):
    out = str(tmp_path_factory.mktemp("chiara2") / "chiara.xlsx")
    run(pfade, out, verbose=False, jahresabschluss=datei(CHIARA_JA))
    return openpyxl.load_workbook(out)


# ---- Reader: kumulierte Monatsscheiben -----------------------------------
def test_reader_erkennt_monatsexport(pfade):
    assert all(SapBwMonateReader.kann_lesen(p) for p in pfade)


def test_monatstabs_sind_kumuliert_und_lueckenlos(pfade):
    """Die Tabs überlappen sich um eine Spalte; jede Grenze muss exakt
    anschließen, sonst wäre die Reihe gebrochen."""
    _, diagnosen = lies_sap_jahre(pfade)
    assert len(diagnosen) == 3
    for d in diagnosen:
        assert d.ueberlappungs_bruch == [], f"{d.jahr}: {d.ueberlappungs_bruch[:3]}"
        assert len(d.scheiben) == 12
        assert d.scheiben[-1].bis_monat == 16      # inkl. Sonderperioden


def test_summe_der_monatsdeltas_ergibt_den_jahreswert(pfade):
    _, diagnosen = lies_sap_jahre(pfade)
    for d in diagnosen:
        for konto, jahr in d.jahreswert.items():
            assert sum(d.monatsdelta(konto)) == pytest.approx(jahr, abs=0.005)


def test_tabs_duerfen_nicht_aufsummiert_werden(pfade):
    """Gegenprobe zur häufigsten Fehlbedienung: die Tabs sind kumuliert, ihre
    Summe ist ein Vielfaches des Jahres."""
    _, diagnosen = lies_sap_jahre(pfade)
    d = next(x for x in diagnosen if x.jahr == 2022)
    konto = "4100000000"        # GEHÄLTER, läuft übers Jahr monoton hoch
    naiv = sum(s.salden.get(konto, 0.0) for s in d.scheiben)
    assert naiv > 5 * d.jahreswert[konto]


def test_jahre_werden_zu_einem_ledger_verschmolzen(pfade):
    led, _ = lies_sap_jahre(pfade)
    assert led.perioden == ["2024/12", "2023/12", "2022/12"]
    assert led.hat_kontennachweis, "die FS-Hierarchie ist die Strukturquelle"


def test_ausgesonderte_konten_werden_technisch_markiert(pfade):
    """Der Export weist IFRS-Brutto- und Steuerbilanzkonten selbst als nicht
    relevant aus ("AUS"); sie dürfen das Databook nicht erreichen."""
    led, _ = lies_sap_jahre(pfade)
    tech = [a for a in led.accounts if a.kontotyp == "technisch"]
    assert len(tech) > 50
    assert all(a.fs_pfad is None for a in tech)


def test_fehlende_kontozeilen_werden_gemeldet(pfade):
    """Knoten 004 (Umsatzerlöse) weist eine Positionssumme aus, der Export
    liefert dazu keine Konten — das darf nicht stillschweigend durchgehen."""
    led, _ = lies_sap_jahre(pfade)
    assert any("004" in w and "keine einzige Kontozeile" in w for w in led.warnungen)
    assert any("Sektion GUV" in w for w in led.warnungen)


def test_seitenwechsel_wird_gemeldet(pfade):
    """Die FS-Hierarchie schaltet Konten je nach Saldovorzeichen zwischen
    Forderungs- und Verbindlichkeitsknoten um."""
    led, _ = lies_sap_jahre(pfade)
    assert any("wechselt im Jahresverlauf die Bilanzseite" in w for w in led.warnungen)


def test_2025_export_ist_ein_anderes_format():
    """Die 2025er Datei ist eine andere Query-Ausprägung (englische
    Leaf-Item-Struktur ohne Kontonummern) und wird bewusst nicht gelesen."""
    assert not SapBwMonateReader.kann_lesen(datei(CHIARA_2025))


# ---- Databook-Modus und Struktur -----------------------------------------
def test_databook_ist_abschlusstreu(res):
    """Die eingebettete FS-Hierarchie ist eine vollwertige Strukturquelle —
    auch ohne separaten Kontennachweis."""
    s = res["setup"]
    assert s.ist_abschlusstreu
    assert "NICHT ABSCHLUSSTREU" not in s.databook_kennzeichen
    assert s.abdeckung == pytest.approx(1.0)


def test_alle_abschlussrelevanten_konten_haben_struktur_aus_dem_abschluss(res):
    from fdd.core.model import Quelle
    fremd = [m for m in res["mapped"]
             if m.klasse.value != "TECH" and m.quelle == Quelle.SKR_DEFAULT]
    assert not fremd, f"SKR-Default greift bei {[m.konto for m in fremd][:5]}"


def test_guv_wird_gemappt(res):
    pl = [m for m in res["mapped"] if m.klasse.value == "PL"]
    assert len(pl) > 100
    assert all(m.hgb_pfad.startswith("/GuV") for m in pl)


# ---- Lead NA / Lead PL ---------------------------------------------------
def test_lead_na_und_lead_pl_existieren(wb):
    assert "Lead NA" in wb.sheetnames
    assert "Lead PL" in wb.sheetnames


def test_lead_na_zaehlt_gemischte_positionen_nicht_doppelt(wb):
    """Eine gemischte Position steht im WC- und im ND-Block. Zöge sie beide
    Male den vollen Bilanzwert, wären die Net Assets zu hoch."""
    ws = wb["Lead NA"]
    for r in range(5, ws.max_row + 1):
        quelle, klasse = ws.cell(r, 4).value, ws.cell(r, 5).value
        if quelle in ("NA_OA", "NA_OL", "NA_OP"):
            formel = ws.cell(r, 6).value
            assert "+" not in formel, f"Zeile {r} zieht den vollen Bilanzwert"
            assert quelle in formel


@pytest.mark.parametrize("tab", ["Lead NA", "Lead PL"])
def test_neue_leads_ziehen_nicht_aus_dem_mastersheet(wb, tab):
    ws = wb[tab]
    for r in range(5, ws.max_row + 1):
        lab = str(ws.cell(r, 3).value)
        # Kontroll- und nachrichtliche Zeilen dürfen gegen das Mastersheet
        # prüfen — nur die Positionszeilen selbst nicht.
        if "ontrollzeile" in lab or lab.startswith("nachrichtlich"):
            continue
        for c in range(2, ws.max_column + 1):
            f = ws.cell(r, c).value
            if isinstance(f, str) and f.startswith("="):
                assert "Mastersheet" not in f, f"{tab} {r}/{c}"


@pytest.mark.parametrize("tab", ["Lead NA", "Net Debt", "Working Capital", "Lead PL"])
def test_jeder_lead_hat_eine_kontrollzeile(wb, tab):
    ws = wb[tab]
    assert any("ontrollzeile" in str(ws.cell(r, 3).value)
               for r in range(5, ws.max_row + 1))


# ---- Reconciliation gegen den Jahresabschluss ----------------------------
def test_ja_recon_laeuft_und_ist_kein_strukturgeber(res):
    """Der Jahresabschluss ist Abstimmziel — er darf kein Mapping ändern."""
    from fdd.core.model import Quelle
    rec = res["ja_recon"]
    assert rec is not None and rec.zeilen
    assert all(m.quelle != Quelle.REVIEW or m.klasse.value == "REVIEW"
               for m in res["mapped"])
    # Alle Pfade der abschlussrelevanten Konten stammen weiterhin aus der
    # SAP-Hierarchie — der Jahresabschluss hat nichts umgehängt.
    quellen = {m.quelle for m in res["mapped"] if m.klasse.value != "TECH"}
    assert quellen <= {Quelle.KONTENNACHWEIS, Quelle.HAUSCONVENTION, Quelle.REVIEW}


def test_ja_recon_nutzt_nur_gemeinsame_perioden(res):
    """Der Abschluss deckt 2020–2023 ab, das Databook 2022–2024."""
    assert res["ja_recon"].perioden == ["2023/12", "2022/12"]
