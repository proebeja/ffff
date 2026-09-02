"""Kitchenstories, zweiter Lauf: die sechs Änderungen aus v2.8, die
Einfrierung gegen Lauf 1 und die sieben geforderten Prüfungen."""

import json

import openpyxl
import pytest

from conftest import datei
from fdd.core.hausconvention import Hausconvention
from fdd.kitchenstories import Quellen, run

SUSA = "Testdaten_Kitchenstories_SuSa.xlsx"
JA = "Testdaten_Kitchenstories_JA_2023.pdf"
PLAN = "Testdaten_Kitchenstories_Kontenplan_2025.pdf"
BERICHT = "Testdaten_Kitchenstories_Pruefbericht_2022.pdf"
PERIODEN = ["FY2022", "FY2023", "FY2024", "YTD Jul25"]


@pytest.fixture(scope="module")
def snapshot(tmp_path_factory):
    """Lauf 1 als eingefrorenes Entscheidungsprotokoll."""
    from fdd.engine.einfrierung import schreibe_snapshot
    out = tmp_path_factory.mktemp("l1")
    res = run(_quellen(), str(out / "l1.xlsx"), verbose=False)
    pfad = str(out / "snapshot.json")
    schreibe_snapshot(pfad, res["mapped"], lauf=1,
                      hausconvention="2.8", fingerprint=res["ledger"].fingerprint)
    return pfad


def _quellen():
    return Quellen(susa=datei(SUSA), jahresabschluss=datei(JA),
                   kontenplan=datei(PLAN), pruefbericht=datei(BERICHT))


@pytest.fixture(scope="module")
def res(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("l2") / "l2.xlsx")
    return run(_quellen(), out, verbose=False)


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    out = str(tmp_path_factory.mktemp("l2b") / "l2.xlsx")
    run(_quellen(), out, verbose=False)
    return openpyxl.load_workbook(out)


# ---- Hausconvention v2.8 --------------------------------------------------
def test_hausconvention_ist_v29():
    """Die geltende Konvention ist v2.9 (ROLLBACK.md).

    Der Pin steht hier, damit ein stiller Wechsel der Konfigurationsdatei
    auffaellt. Die v2.8-Mechaniken (Seitenwechsel je Periode, Saldenvortrag,
    Verhaltenspruefung) gelten unveraendert weiter — v2.9 ergaenzt sie nur um
    ``databook_architektur``, ``excel_format`` und ``lead_tab_konvention``.
    """
    assert Hausconvention.laden().version == "2.9"


def test_zahlungsverkehr_regeln_greifen_in_allen_gemischten_positionen():
    hc = Hausconvention.laden()
    for rs in ("sonstige_vermoegensgegenstaende", "sonstige_verbindlichkeiten",
               "sonstige_rueckstellungen"):
        ids = {r.id for r in hc.typ2_regeln(rs)}
        assert {"zdl-verrechnungskonto", "zdl-kreditkarte",
                "gesellschafter-gf-forderung"} <= ids


# ---- 1) Seitenwechsel -----------------------------------------------------
def test_seitenwechsel_nur_bei_echtem_wechsel(res):
    """Ein Konto mit durchgehend gleichem Vorzeichen wechselt nichts. Die
    Vorsteuerkonten stehen dauerhaft im Umsatzsteuerblock der sonstigen
    Verbindlichkeiten und dürfen nicht umgegliedert werden."""
    betroffen = {f.konto for f in res["seitenwechsel"]}
    assert {"701 0", "1600 0", "1731 2"} <= betroffen
    assert not betroffen & {"1571 0", "1576 0", "1577 0", "1781 0"}


def test_konto_701_wechselt_die_bilanzseite(res):
    m = next(m for m in res["mapped"] if m.konto == "701 0")
    assert m.pfad_in("FY2022").startswith("/Passiva")
    assert m.pfad_in("FY2023").startswith("/Aktiva")
    assert m.klasse.value == "ND", "die Klasse bleibt periodenfest"


def test_kontennachweis_schlaegt_die_vorzeichenableitung(res):
    """1789 0 wechselt das Vorzeichen, der Abschluss führt es aber in beiden
    Jahren unter den sonstigen Verbindlichkeiten."""
    m = next(m for m in res["mapped"] if m.konto == "1789 0")
    assert "FY2022" not in m.pfad_je_periode


def test_jeder_seitenwechsel_erzeugt_eine_pflichtfrage(res):
    for f in res["seitenwechsel"]:
        assert f.konto in f.pflichtfrage and "Hintergrund" in f.pflichtfrage
    assert all(any(f.pflichtfrage == b for b in res["qa"].offene_befunde)
               for f in res["seitenwechsel"])


# ---- 2) Verhaltensprüfung -------------------------------------------------
def test_vorzeichenwechsel_wird_geflaggt(res):
    vz = [b for b in res["verhalten"] if b.kriterium == "vorzeichenwechsel"]
    assert "701 0" in {b.konto for b in vz}
    assert any(b.wesentlich for b in vz), "der 3,6-Mio-Wechsel ist wesentlich"


def test_verhaltenspruefung_klassifiziert_nicht_um(res):
    """Sie markiert und fragt — sie entscheidet nie über die Klasse."""
    m = next(m for m in res["mapped"] if m.konto == "701 0")
    assert m.klasse.value == "ND"


# ---- 3) Lead NA: Eigenkapital inkl. Jahresergebnis ------------------------
def test_eigenkapital_fy2023_entspricht_dem_abschluss(res):
    """Abschluss: 6.079,23. Restdifferenz sind die bekannten 172 aus Konto 410."""
    eq = sum(m.saldo("FY2023") for m in res["mapped"] if m.klasse.value == "EQ")
    pl = sum(m.saldo("FY2023") for m in res["mapped"] if m.klasse.value == "PL")
    assert eq + pl == pytest.approx(-(6_079.23 + 172.00), abs=0.01)


def test_eigenkapital_ytd_ist_fortgeschrieben(res):
    def ek(p):
        return sum(m.saldo(p) for m in res["mapped"]
                   if m.klasse.value in ("EQ", "PL"))
    assert ek("YTD Jul25") != pytest.approx(ek("FY2024"), abs=1.0)


def test_kontrollzeile_ohne_ausgleichsposten(wb):
    """QA A6: keine Kontrollzeile darf eine Review-, Ergebnis- oder
    Technikzeile als Ausgleichsposten verwenden."""
    ws = wb["Lead NA"]
    for r in range(5, ws.max_row + 1):
        lab = str(ws.cell(r, 3).value or "")
        if "Kontrollzeile" not in lab:
            continue
        assert "Review" not in lab and "technische" not in lab
        assert "Periodenergebnis" not in lab


# ---- 4) Saldenvorträge ----------------------------------------------------
def test_saldenvortragskonten_sind_nicht_tech(res):
    by = {m.konto: m for m in res["mapped"]}
    for k in ("9000 0", "9008 0", "9009 0"):
        assert by[k].klasse.value != "TECH"
        assert by[k].hgb_pfad.startswith("/Passiva/A Eigenkapital")


def test_statistikkonten_bleiben_tech(res):
    by = {m.konto: m for m in res["mapped"]}
    assert by["9140 0"].klasse.value == "TECH"
    assert by["9199 0"].klasse.value == "TECH"


def test_wertberichtigung_hat_vorrang_vor_der_bereichsregel(res):
    """9960 wird von einer Typ-1-Regel gefangen und ist deshalb nicht TECH."""
    m = next(m for m in res["mapped"] if m.konto == "9960 0")
    assert m.klasse.value != "TECH"
    assert "Forderungen aus Lieferungen" in m.hgb_pfad


# ---- 5) Regelgruppe Zahlungsverkehr ---------------------------------------
def test_zahlungsdienstleister_werden_net_debt(res):
    by = {m.konto: m for m in res["mapped"]}
    for k in ("1731 5", "1731 6", "1731 7"):      # Spendesk, Shopify, Finway
        assert by[k].klasse.value == "ND", k


def test_kreditkarte_wird_net_debt(res):
    assert next(m for m in res["mapped"] if m.konto == "1731 2").klasse.value == "ND"


def test_ermessensfaelle_bleiben_bewusst_unkodiert(res):
    """AAG-Erstattung, erhaltene Anzahlungen und das Verrechnungskonto sind
    Ermessensfragen — sie dürfen nicht als ND eingefroren werden."""
    by = {m.konto: m for m in res["mapped"]}
    assert by["1520 0"].klasse.value != "ND"      # Krankenkasse AAG
    assert by["1718 0"].klasse.value != "ND"      # erhaltene Anzahlungen


# ---- 6) QA A6 -------------------------------------------------------------
def test_kein_konto_ausserhalb_der_bilanz(res):
    for m in res["mapped"]:
        if m.klasse.value == "TECH":
            continue
        assert m.hgb_pfad and not m.hgb_pfad.startswith("("), m.konto
        assert m.klasse.value != "REVIEW", m.konto


def test_review_status_bleibt_neben_der_zuordnung(res):
    """Die Zuordnung ersetzt die Klärung nicht."""
    assert res["review"], "die Review-Queue darf nicht leer laufen"


# ---- QA-Tab ---------------------------------------------------------------
def test_qa_tab_weist_jede_einzelpruefung_aus(wb, res):
    assert "QA" in wb.sheetnames
    ids = {p.id for p in res["qa"].pruefungen}
    erwartet = {f"A{i}" for i in range(1, 7)} | {f"B{i}" for i in range(1, 6)} \
        | {f"C{i}" for i in range(1, 6)} | {"D1"}
    assert erwartet <= ids


def test_b1_wird_berichtet(res):
    b1 = next(p for p in res["qa"].pruefungen if p.id == "B1")
    assert b1.bestanden and "umuliert" in b1.befund


def test_qa_weist_nicht_zugeordnete_je_periode_aus(res):
    assert set(res["qa"].nicht_zugeordnet) == set(PERIODEN)


# ---- Abstimmungen ---------------------------------------------------------
def test_fy2022_geht_auf(res):
    """Die Differenz von 3.605.431 auf beiden Seiten ist verschwunden.

    Das Eigenkapital trägt weiterhin eine große Differenz gegen den Bericht —
    sie ist aber vollständig überge­leitet (Ergebnis noch auf den GuV-Konten,
    nicht gedeckter Fehlbetrag nach § 268 III HGB). Maßgeblich ist der Rest."""
    r = res["recon_agg"]
    assert r.rest_gesamt() == pytest.approx(0.0, abs=0.01)
    assert r.mit_rest == []
    for z in r.zeilen:
        if z.label.startswith(("II. Forderungen", "C. Verbindlichkeiten")):
            assert abs(z.differenz) < 100_000, (
                f"{z.label}: der Seitenwechsel hätte die 3,6 Mio auflösen müssen")


def test_fy2023_nur_die_bekannten_172(res):
    r = res["recon_kn"]
    assert [(k.konto, round(k.differenz, 2)) for k in r.echte_differenzen] == [
        ("410 0", 172.00)]


# ---- Einfrierung ----------------------------------------------------------
def test_einfrierung_erkennt_keine_defekte(snapshot, tmp_path):
    erg = run(_quellen(), str(tmp_path / "l2.xlsx"), verbose=False,
              snapshot=snapshot)["einfrierung"]
    assert erg is not None
    assert erg.defekte == [], [d.konto for d in erg.defekte]


def test_lauf_ist_reproduzierbar(snapshot, tmp_path):
    """Zweimal derselbe Input, zweimal dasselbe Ergebnis."""
    a = run(_quellen(), str(tmp_path / "a.xlsx"), verbose=False, snapshot=snapshot)
    b = run(_quellen(), str(tmp_path / "b.xlsx"), verbose=False, snapshot=snapshot)
    fa = {m.konto: (m.klasse.value, m.hgb_pfad) for m in a["mapped"]}
    fb = {m.konto: (m.klasse.value, m.hgb_pfad) for m in b["mapped"]}
    assert fa == fb


def test_jede_aenderung_hat_einen_ausloeser(snapshot, tmp_path):
    erg = run(_quellen(), str(tmp_path / "l2.xlsx"), verbose=False,
              snapshot=snapshot)["einfrierung"]
    for d in erg.delta:
        assert d.ausloeser.startswith(("1 ", "3 ", "4 ", "5 ")), d.ausloeser


# ---- Laufprotokoll --------------------------------------------------------
def test_laufprotokoll_misst_alle_phasen(res):
    lp = res["laufprotokoll"]
    namen = {p.name for p in lp.phasen}
    assert {"Einlesen SuSa", "Mapping (Kaskade)", "Aufrisse (Schedules)",
            "Excel-Ausgabe"} <= namen
    assert lp.gesamt > 0 and lp.zellen > 1000


def test_keine_ki_aufrufe(res):
    """Die KI-Schicht ist eine Schnittstelle ohne registrierten Provider."""
    assert res["laufprotokoll"].ki_aufrufe == []
